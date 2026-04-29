import os
import asyncio
import base64
import qrcode
import sqlite3
import threading
import time
import queue
import logging
import hashlib
import functools
from collections import Counter
from io import BytesIO
from datetime import datetime, timedelta
from telethon.errors import AuthKeyUnregisteredError
from flask import (
    Flask, session, request, render_template_string,
    jsonify, g, redirect, url_for, send_from_directory, render_template, send_file
)
from telethon import TelegramClient, events, utils
import json
from flask_socketio import SocketIO, emit
from telethon.events import MessageDeleted


import os
import uuid
import json
from werkzeug.utils import secure_filename
from flask import send_from_directory


import secrets
import math
import pandas as pd
import tempfile
import pdfplumber
import re
import html
try:
    from dotenv import load_dotenv
except Exception:
    load_dotenv = None
sqlite3.register_adapter(datetime, lambda d: d.strftime("%Y-%m-%d %H:%M:%S"))

try:
    from google import genai as google_genai_sdk
    from google.genai import types as google_genai_types
except Exception:
    google_genai_sdk = None
    google_genai_types = None






# Настройка логирования
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

if load_dotenv:
    for env_name in (".env", ".env.local"):
        env_path = os.path.join(BASE_DIR, env_name)
        if os.path.exists(env_path):
            load_dotenv(env_path, override=False)

app = Flask(__name__)
app.secret_key = 'supersecretkey_for_demo'
app.config['DATABASE'] = os.path.join(BASE_DIR, 'app.db')
app.json.ensure_ascii = False
# ПЕРЕНЕСТИ СЮДА (строки, которые сейчас перед шаблонами)
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading', ping_timeout=120, ping_interval=25) 
PRODUCT_SEARCH_CACHE = {}
ACCESS_TREE_CACHE = {}
INTERACTION_CLEANUP_RUN_MAX_ENTRIES = 1000
INTERACTION_RESPONSE_START_WINDOW_SECONDS = 5 * 60
INTERACTION_RESPONSE_IDLE_SECONDS = 15 * 60
INTERACTION_RESPONSE_BURST_GAP_SECONDS = 2 * 60
INTERACTION_COMMAND_AUTOMATION_FALLBACK_SECONDS = 15
SQLITE_BUSY_TIMEOUT_MS = 30000
AUTO_BINDING_MAINTENANCE_ENABLED = os.environ.get("ASTORE_AUTO_BINDING_MAINTENANCE", "").strip().lower() in {
    "1", "true", "yes", "on"
}
SUPPLIER_LEARNING_BACKFILL_ENABLED = os.environ.get("ASTORE_SUPPLIER_LEARNING_BACKFILL", "").strip().lower() in {
    "1", "true", "yes", "on"
}
GOOGLE_SHEETS_SCHEDULER_ENABLED = os.environ.get("ASTORE_GOOGLE_SHEETS_SCHEDULER", "").strip().lower() in {
    "1", "true", "yes", "on"
}
interaction_response_cleanup_runs = set()
interaction_response_cleanup_lock = threading.Lock()
sqlite_write_lock = threading.RLock()

@app.after_request
def add_no_cache_headers(response):
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


# Настройка папки для хранения фото
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# НОВЫЙ МАРШРУТ: Отдаем загруженные файлы по ссылке
@app.route('/uploads/<filename>')
def uploaded_file(filename):
    response = send_from_directory(app.config['UPLOAD_FOLDER'], filename, max_age=31536000)
    response.cache_control.public = True
    response.cache_control.max_age = 31536000
    return response

def invalidate_product_search_cache(user_id=None):
    if user_id is None:
        PRODUCT_SEARCH_CACHE.clear()
        return
    PRODUCT_SEARCH_CACHE.pop(int(user_id), None)

def invalidate_access_tree_cache(user_id=None):
    if user_id is None:
        ACCESS_TREE_CACHE.clear()
        return
    ACCESS_TREE_CACHE.pop(int(user_id), None)

def notify_clients(data_type="general"):
    invalidate_product_search_cache()
    invalidate_access_tree_cache()
    socketio.emit('db_updated', {'type': data_type})

def normalize_incoming_text(text):
    if not text:
        return ''
    text = str(text).replace('\r\n', '\n').replace('\r', '\n')
    text = text.replace('\u200b', '').replace('\ufeff', '')
    text = text.replace('`', '')
    text = re.sub(r'\*\*(.*?)\*\*', r'\1', text)
    text = re.sub(r'__(.*?)__', r'\1', text)
    text = re.sub(r'(?m)^\s*[▪◾◼⬛]+\s*', '', text)
    text = re.sub(r'[ \t]+\n', '\n', text)
    return text.strip()

def normalize_interval_minutes(value, default=60):
    try:
        return max(0, int(value))
    except (TypeError, ValueError):
        return default

def get_env_int(name, default, minimum=0):
    try:
        return max(minimum, int(os.environ.get(name, default) or default))
    except (TypeError, ValueError):
        return default

def execute_db_with_lock_retry(db, sql, params=(), retries=6, retry_delay=0.75):
    for attempt in range(retries):
        try:
            return db.execute(sql, params)
        except sqlite3.OperationalError as e:
            if 'locked' not in str(e).lower():
                raise
            if attempt + 1 >= retries:
                raise
            time.sleep(retry_delay * (attempt + 1))

def ensure_tracked_source_row(db, user_id, chat_id, chat_title=None, custom_name=None):
    if chat_id in (None, ""):
        return None
    title = str(chat_title or custom_name or chat_id).strip()
    custom = str(custom_name or "").strip() or None
    db.execute(
        """
        INSERT INTO tracked_chats (user_id, chat_id, chat_title, custom_name)
        VALUES (?, ?, ?, ?)
        ON CONFLICT(user_id, chat_id) DO UPDATE SET
            chat_title = CASE
                WHEN excluded.chat_title IS NOT NULL AND excluded.chat_title != '' THEN excluded.chat_title
                ELSE tracked_chats.chat_title
            END,
            custom_name = CASE
                WHEN COALESCE(tracked_chats.custom_name, '') = '' THEN excluded.custom_name
                ELSE tracked_chats.custom_name
            END
        """,
        (user_id, chat_id, title, custom),
    )
    return db.execute(
        "SELECT id FROM tracked_chats WHERE user_id = ? AND CAST(chat_id AS TEXT) = CAST(? AS TEXT)",
        (user_id, chat_id),
    ).fetchone()

def cleanup_source_before_new_snapshot(db, user_id, chat_id):
    execute_db_with_lock_retry(db, """
        UPDATE product_messages
        SET is_actual = 0
        WHERE message_id IN (
            SELECT id FROM messages
            WHERE user_id = ? AND CAST(chat_id AS TEXT) = CAST(? AS TEXT)
        )
    """, (user_id, chat_id))
    execute_db_with_lock_retry(db, """
        DELETE FROM product_messages
        WHERE status != 'confirmed'
          AND message_id IN (
              SELECT id FROM messages
              WHERE user_id = ? AND CAST(chat_id AS TEXT) = CAST(? AS TEXT)
          )
    """, (user_id, chat_id))
    execute_db_with_lock_retry(db, """
        UPDATE messages
        SET is_blocked = 1
        WHERE user_id = ?
          AND CAST(chat_id AS TEXT) = CAST(? AS TEXT)
          AND id IN (SELECT message_id FROM product_messages WHERE status = 'confirmed')
    """, (user_id, chat_id))
    execute_db_with_lock_retry(db, """
        DELETE FROM messages
        WHERE user_id = ?
          AND CAST(chat_id AS TEXT) = CAST(? AS TEXT)
        AND id NOT IN (SELECT message_id FROM product_messages)
	    """, (user_id, chat_id))


def mark_source_closed_for_sales(db, user_id, chat_id, preserve_message_id=None):
    # Stop/status bot replies are informational. They must not wipe the previous real price snapshot.
    return


def cleanup_source_type_before_new_snapshot(db, user_id, chat_id, msg_type):
    execute_db_with_lock_retry(db, """
        UPDATE product_messages
        SET is_actual = 0
        WHERE message_id IN (
            SELECT id FROM messages
            WHERE user_id = ? AND CAST(chat_id AS TEXT) = CAST(? AS TEXT) AND type = ?
        )
    """, (user_id, chat_id, msg_type))
    execute_db_with_lock_retry(db, """
        DELETE FROM product_messages
        WHERE status != 'confirmed'
          AND message_id IN (
              SELECT id FROM messages
              WHERE user_id = ? AND CAST(chat_id AS TEXT) = CAST(? AS TEXT) AND type = ?
          )
    """, (user_id, chat_id, msg_type))
    execute_db_with_lock_retry(db, """
        UPDATE messages
        SET is_blocked = 1
        WHERE user_id = ?
          AND CAST(chat_id AS TEXT) = CAST(? AS TEXT)
          AND type = ?
          AND id IN (SELECT message_id FROM product_messages WHERE status = 'confirmed')
    """, (user_id, chat_id, msg_type))
    execute_db_with_lock_retry(db, """
        DELETE FROM messages
        WHERE user_id = ?
          AND CAST(chat_id AS TEXT) = CAST(? AS TEXT)
          AND type = ?
          AND id NOT IN (SELECT message_id FROM product_messages)
    """, (user_id, chat_id, msg_type))


CHANNEL_SNAPSHOT_INTERVAL_MINUTES = 5
CHANNEL_SNAPSHOT_MESSAGE_LIMIT = get_env_int("ASTORE_CHANNEL_SNAPSHOT_MESSAGE_LIMIT", 50, minimum=1)
CHANNEL_SNAPSHOT_RESULT_TIMEOUT_SECONDS = get_env_int("ASTORE_CHANNEL_SNAPSHOT_TIMEOUT_SECONDS", 60, minimum=15)
CHANNEL_MESSAGE_PARSE_TIMEOUT_SECONDS = get_env_int("ASTORE_CHANNEL_MESSAGE_PARSE_TIMEOUT_SECONDS", 30, minimum=10)
AUTO_BINDING_SWEEP_INTERVAL_SECONDS = get_env_int("ASTORE_AUTO_BINDING_SWEEP_INTERVAL_SECONDS", 5 * 60, minimum=30)
AUTO_BINDING_SWEEP_MESSAGE_LIMIT = get_env_int("ASTORE_AUTO_BINDING_SWEEP_MESSAGE_LIMIT", 250, minimum=10)
AUTO_BINDING_SWEEP_SOURCE_WINDOW_MINUTES = get_env_int("ASTORE_AUTO_BINDING_SWEEP_SOURCE_WINDOW_MINUTES", 10, minimum=1)
AUTO_BINDING_SWEEP_RETRY_SECONDS = get_env_int("ASTORE_AUTO_BINDING_SWEEP_RETRY_SECONDS", 60 * 60, minimum=60)
AUTO_CONFIRM_MESSAGE_MAX_SECONDS = get_env_int("ASTORE_AUTO_CONFIRM_MESSAGE_MAX_SECONDS", 45, minimum=5)
GENERIC_FILE_PARSE_MAX_LINES = get_env_int("ASTORE_GENERIC_FILE_PARSE_MAX_LINES", 5000, minimum=50)

auto_binding_sweep_attempts = {}


def current_moscow_snapshot_slot(interval_minutes=CHANNEL_SNAPSHOT_INTERVAL_MINUTES):
    now = datetime.now(timezone.utc).replace(tzinfo=None) + timedelta(hours=3)
    slot_minute = now.minute - (now.minute % interval_minutes)
    return now.replace(minute=slot_minute, second=0, microsecond=0)


def build_snapshot_sender_name(sender, fallback_title):
    if not sender:
        return fallback_title or "Unknown"
    username = getattr(sender, 'username', None)
    if username:
        return f"@{username}"
    title = getattr(sender, 'title', None)
    if title:
        return title
    first_name = getattr(sender, 'first_name', '') or ''
    last_name = getattr(sender, 'last_name', '') or ''
    full_name = f"{first_name} {last_name}".strip()
    return full_name or fallback_title or "Unknown"


def build_chat_id_variants(chat_id):
    raw = str(chat_id or '').strip()
    variants = {raw}
    if raw.startswith('-100'):
        variants.add(raw[4:])
    elif raw and raw.lstrip('-').isdigit():
        variants.add(f"-100{raw.lstrip('-')}")
        variants.add(raw.lstrip('-'))
    return {item for item in variants if item}


def entity_matches_chat_id(entity, chat_id):
    variants = build_chat_id_variants(chat_id)
    entity_ids = set()
    entity_id = getattr(entity, 'id', None)
    if entity_id is not None:
        entity_ids.update(build_chat_id_variants(entity_id))
    try:
        peer_id = utils.get_peer_id(entity)
        entity_ids.update(build_chat_id_variants(peer_id))
    except Exception:
        pass
    return bool(variants & entity_ids)


async def resolve_tracked_entity(client, chat_id):
    last_error = None
    for candidate in build_chat_id_variants(chat_id):
        try:
            lookup = int(candidate) if str(candidate).lstrip('-').isdigit() else candidate
            entity = await client.get_entity(lookup)
            if entity_matches_chat_id(entity, chat_id):
                return entity
        except Exception as e:
            last_error = e

    async for dialog in client.iter_dialogs():
        entity = getattr(dialog, 'entity', None)
        if entity and entity_matches_chat_id(entity, chat_id):
            return entity

    raise last_error or ValueError(f"entity not found for chat_id={chat_id}")


def snapshot_message_type_for_entity(entity):
    if getattr(entity, 'broadcast', False):
        return 'channel'
    if getattr(entity, 'megagroup', False) or getattr(entity, 'gigagroup', False):
        return 'group'
    if getattr(entity, 'title', None):
        return 'group'
    return 'private'


async def collect_tracked_channel_snapshot(client, user_id, chat_id, fallback_title, limit=CHANNEL_SNAPSHOT_MESSAGE_LIMIT):
    try:
        entity = await resolve_tracked_entity(client, chat_id)
    except Exception as e:
        return {'status': 'error', 'error': repr(e), 'messages': []}

    if getattr(entity, 'bot', False):
        return {'status': 'skip', 'reason': 'bot', 'messages': []}

    msg_type = snapshot_message_type_for_entity(entity)
    first = getattr(entity, 'first_name', '') or ''
    last = getattr(entity, 'last_name', '') or ''
    chat_title = getattr(entity, 'title', None) or f"{first} {last}".strip() or fallback_title or str(chat_id)
    snapshot_messages = []
    seen_messages = 0
    skipped_messages = 0
    incomplete = False

    async for message in client.iter_messages(entity, limit=limit):
        if not message:
            continue
        if not message.text and not message.document:
            continue
        seen_messages += 1

        try:
            parsed_text = await asyncio.wait_for(
                parse_excel_message(client, message, chat_id, user_id),
                timeout=CHANNEL_MESSAGE_PARSE_TIMEOUT_SECONDS,
            )
        except asyncio.TimeoutError:
            skipped_messages += 1
            incomplete = True
            logger.warning(
                "Channel snapshot skipped slow message user %s chat %s message_id=%s",
                user_id,
                chat_id,
                getattr(message, 'id', None),
            )
            continue
        except Exception as e:
            skipped_messages += 1
            incomplete = True
            logger.warning(
                "Channel snapshot failed to parse message user %s chat %s message_id=%s: %s",
                user_id,
                chat_id,
                getattr(message, 'id', None),
                e,
            )
            continue
        parsed_text = normalize_incoming_text(parsed_text)
        if not parsed_text:
            continue

        try:
            sender = await message.get_sender()
        except Exception:
            sender = None

        sender_name = build_snapshot_sender_name(sender, chat_title)
        msg_date = telegram_message_date_to_local_naive(message.date).strftime("%Y-%m-%d %H:%M:%S")
        snapshot_messages.append({
            'telegram_message_id': message.id,
            'type': msg_type,
            'text': parsed_text,
            'date': msg_date,
            'chat_id': chat_id,
            'chat_title': chat_title,
            'sender_name': sender_name,
        })

    snapshot_messages.sort(key=lambda item: (item['date'], int(item['telegram_message_id'] or 0)))
    return {
        'status': 'ok',
        'chat_title': chat_title,
        'messages': snapshot_messages,
        'seen_messages': seen_messages,
        'skipped_messages': skipped_messages,
        'incomplete': incomplete,
    }


def build_readonly_session_copy(session_file, label):
    raw_path = str(session_file or '').strip()
    if not raw_path:
        raise ValueError("session_file is empty")
    source_path = raw_path if os.path.isabs(raw_path) else os.path.join(BASE_DIR, raw_path)
    if not source_path.endswith('.session'):
        source_path = f"{source_path}.session"
    if not os.path.exists(source_path):
        raise FileNotFoundError(source_path)

    copy_base = os.path.join('/tmp', f"astore_snapshot_{label}_{os.getpid()}_{time.time_ns()}")
    source_db = sqlite3.connect(f"file:{source_path}?mode=ro", uri=True, timeout=10)
    copy_db = sqlite3.connect(f"{copy_base}.session", timeout=10)
    try:
        source_db.backup(copy_db)
    finally:
        copy_db.close()
        source_db.close()
    return copy_base


def cleanup_session_copy(copy_base):
    for suffix in ('.session', '.session-journal', '.session-wal', '.session-shm'):
        try:
            os.remove(f"{copy_base}{suffix}")
        except FileNotFoundError:
            pass
        except Exception:
            logger.debug("Не удалось удалить временную Telegram-сессию %s%s", copy_base, suffix)


async def collect_tracked_channel_snapshot_with_session_copy(session_row, user_id, chat_id, fallback_title):
    copy_base = build_readonly_session_copy(session_row['session_file'], f"{session_row['id']}_{chat_id}")
    client = TelegramClient(copy_base, int(session_row['api_id']), session_row['api_hash'])
    try:
        await client.connect()
        if not await client.is_user_authorized():
            return {'status': 'error', 'error': 'unauthorized', 'messages': []}
        return await asyncio.wait_for(
            collect_tracked_channel_snapshot(client, user_id, chat_id, fallback_title),
            timeout=CHANNEL_SNAPSHOT_RESULT_TIMEOUT_SECONDS,
        )
    finally:
        try:
            await client.disconnect()
        finally:
            cleanup_session_copy(copy_base)


def persist_tracked_channel_snapshot(db, user_id, chat_id, snapshot_payload):
    snapshot_messages = snapshot_payload.get('messages') or []
    chat_title = snapshot_payload.get('chat_title') or str(chat_id)
    incomplete = bool(snapshot_payload.get('incomplete'))
    persisted = 0
    message_db_ids = []

    if not snapshot_messages:
        logger.warning(
            "Channel snapshot for user %s chat %s returned 0 messages; keeping existing DB state",
            user_id,
            chat_id,
        )
        return 0

    with sqlite_write_lock:
        # Only a complete Telegram scan is allowed to remove old rows. If one
        # message times out or fails to parse, keep the previous snapshot and
        # only upsert the rows that were read successfully.
        if incomplete:
            logger.warning(
                "Channel snapshot for user %s chat %s is incomplete: seen=%s parsed=%s skipped=%s; keeping stale rows",
                user_id,
                chat_id,
                snapshot_payload.get('seen_messages'),
                len(snapshot_messages),
                snapshot_payload.get('skipped_messages'),
            )
        else:
            cleanup_source_before_new_snapshot(db, user_id, chat_id)
        for item in snapshot_messages:
            existing = db.execute(
                """
                SELECT id
                FROM messages
                WHERE user_id = ?
                  AND CAST(chat_id AS TEXT) = CAST(? AS TEXT)
                  AND telegram_message_id = ?
                """,
                (user_id, chat_id, item['telegram_message_id'])
            ).fetchone()

            if existing:
                db.execute(
                    """
                    UPDATE messages
                    SET type = ?,
                        text = ?,
                        date = ?,
                        chat_title = ?,
                        sender_name = ?,
                        is_blocked = 0,
                        is_delayed = 0
                    WHERE id = ?
                    """,
                    (
                        item['type'],
                        item['text'],
                        item['date'],
                        chat_title,
                        item['sender_name'],
                        existing['id'],
                    ),
                )
                message_db_id = existing['id']
            else:
                cursor = db.execute(
                    """
                    INSERT INTO messages
                    (user_id, telegram_message_id, type, text, date, chat_id, chat_title, sender_name, is_delayed, is_blocked)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, 0, 0)
                    """,
                    (
                        user_id,
                        item['telegram_message_id'],
                        item['type'],
                        item['text'],
                        item['date'],
                        chat_id,
                        chat_title,
                        item['sender_name'],
                    ),
                )
                message_db_id = cursor.lastrowid

            message_db_ids.append(message_db_id)
            persisted += 1

        db.commit()

    for message_db_id in message_db_ids:
        enqueue_auto_confirm_source_message(user_id, message_db_id, f"Channel {chat_id}")

    notify_clients()
    return persisted


def channel_snapshot_scheduler():
    time.sleep(15)
    last_processed_slots = {}

    with app.app_context():
        while True:
            try:
                slot = current_moscow_snapshot_slot()
                db = get_db()
                tracked_rows = db.execute(
                    """
                    SELECT DISTINCT tc.user_id, tc.chat_id, tc.chat_title, tc.custom_name
                    FROM tracked_chats tc
                    JOIN user_telegram_sessions uts
                      ON uts.user_id = tc.user_id
                     AND uts.status = 'active'
                    WHERE CAST(tc.chat_id AS TEXT) NOT LIKE 'gs_supplier_%'
                      AND NOT EXISTS (
                          SELECT 1
                          FROM interaction_bots ib
                          WHERE ib.user_id = tc.user_id
                            AND ib.status = 'active'
                            AND CAST(ib.resolved_chat_id AS TEXT) = CAST(tc.chat_id AS TEXT)
                      )
                    ORDER BY tc.user_id, tc.id
                    """
                ).fetchall()

                for row in tracked_rows:
                    try:
                        user_id = int(row['user_id'])
                        chat_id = row['chat_id']
                        slot_key = (user_id, str(chat_id))
                        if last_processed_slots.get(slot_key) == slot:
                            continue
                        active_sessions = [
                            session_row
                            for session_row in db.execute(
                                """
                                SELECT id, api_id, api_hash, session_file
                                FROM user_telegram_sessions
                                WHERE user_id = ? AND status = 'active'
                                ORDER BY id
                                """,
                                (user_id,),
                            ).fetchall()
                            if is_parsing_allowed(session_row['id'])
                        ]
                        if not active_sessions:
                            continue

                        fallback_title = row['custom_name'] or row['chat_title'] or str(chat_id)
                        payload = None

                        for session_row in active_sessions:
                            future = None
                            try:
                                candidate = asyncio.run(
                                    collect_tracked_channel_snapshot_with_session_copy(
                                        session_row,
                                        user_id,
                                        chat_id,
                                        fallback_title,
                                    )
                                )
                            except Exception as e:
                                logger.warning(
                                    "Channel snapshot failed for user %s chat %s via session %s: %s",
                                    user_id,
                                    chat_id,
                                    session_row['id'],
                                    repr(e),
                                )
                                continue

                            if candidate.get('status') == 'ok':
                                payload = candidate
                                break
                            if candidate.get('status') == 'skip' and candidate.get('reason') in {'not_channel', 'bot'}:
                                payload = candidate
                                break

                        if not payload:
                            logger.warning(
                                "Channel snapshot skipped: no active session could read chat %s for user %s",
                                chat_id,
                                user_id,
                            )
                            continue
                        if payload.get('status') == 'skip':
                            last_processed_slots[slot_key] = slot
                            continue

                        persisted = persist_tracked_channel_snapshot(db, user_id, chat_id, payload)
                        last_processed_slots[slot_key] = slot
                        logger.info(
                            "Channel snapshot refreshed for user %s chat %s: %s messages",
                            user_id,
                            chat_id,
                            persisted,
                        )
                    except Exception as row_e:
                        logger.error(
                            "Ошибка snapshot-цикла для user %s chat %s: %s",
                            row['user_id'],
                            row['chat_id'],
                            row_e,
                            exc_info=True,
                        )

            except Exception as e:
                logger.error("Ошибка планировщика snapshot-каналов: %s", e, exc_info=True)

            time.sleep(30)


def parse_stored_datetime(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    raw = str(value).split('+', 1)[0].strip()
    for fmt in ('%Y-%m-%d %H:%M:%S.%f', '%Y-%m-%d %H:%M:%S'):
        try:
            return datetime.strptime(raw.split('.')[0] if fmt.endswith('%S') else raw, fmt)
        except ValueError:
            continue
    return None

def telegram_message_date_to_local_naive(message_date):
    if isinstance(message_date, datetime):
        if message_date.tzinfo:
            return message_date.astimezone(timezone.utc).replace(tzinfo=None) + timedelta(hours=3)
        return message_date.replace(tzinfo=None)
    return datetime.now(timezone.utc).replace(tzinfo=None) + timedelta(hours=3)

def should_accept_interaction_bot_message(db, user_id, userbot_id, chat_id, bot_username, message_date, reply_to_msg_id=None, telegram_message_id=None):
    row = get_interaction_bot_for_userbot(db, user_id, userbot_id, chat_id, bot_username)
    if not row:
        return None
    last_run = parse_stored_datetime(row['last_run'])
    if not last_run:
        return None
    msg_date = telegram_message_date_to_local_naive(message_date)
    if msg_date + timedelta(seconds=5) < last_run:
        return None

    manual_intervention_at = parse_stored_datetime(row['manual_intervention_at'] if 'manual_intervention_at' in row.keys() else None)
    if manual_intervention_at and manual_intervention_at >= last_run and msg_date + timedelta(seconds=5) >= manual_intervention_at:
        return None

    last_command_message_id = row['last_command_message_id'] if 'last_command_message_id' in row.keys() else None
    if reply_to_msg_id and last_command_message_id and str(reply_to_msg_id) == str(last_command_message_id):
        return row

    # If this bot message has already been accepted once and stored in the DB,
    # all later edits of the same Telegram message must still be processed.
    if telegram_message_id:
        existing_message = db.execute(
            """
            SELECT id
            FROM messages
            WHERE user_id = ?
              AND CAST(chat_id AS TEXT) = CAST(? AS TEXT)
              AND telegram_message_id = ?
            LIMIT 1
            """,
            (user_id, chat_id, telegram_message_id),
        ).fetchone()
        if existing_message:
            return row

    last_response_at = parse_stored_datetime(row['last_response_at'] if 'last_response_at' in row.keys() else None)
    if last_response_at:
        if msg_date <= last_response_at + timedelta(seconds=INTERACTION_RESPONSE_BURST_GAP_SECONDS):
            return row
        return None

    try:
        interval_seconds = max(0, int(row['interval_minutes'] or 0)) * 60
    except (TypeError, ValueError):
        interval_seconds = 0
    start_window_seconds = max(INTERACTION_RESPONSE_START_WINDOW_SECONDS, min(interval_seconds, 30 * 60) if interval_seconds else 0)
    if msg_date <= last_run + timedelta(seconds=start_window_seconds):
        return row
    return None

def mark_interaction_bot_response_seen(db, interaction_row, message_date):
    if not interaction_row or 'id' not in interaction_row.keys():
        return
    seen_at = telegram_message_date_to_local_naive(message_date).strftime('%Y-%m-%d %H:%M:%S')
    db.execute(
        "UPDATE interaction_bots SET last_response_at = ? WHERE id = ?",
        (seen_at, interaction_row['id'])
    )

def build_telegram_ref_variants(value):
    clean = normalize_telegram_ref(value)
    if not clean:
        return ['', '', '', '']
    return [clean, f"@{clean}", f"https://t.me/{clean}", f"t.me/{clean}"]

def get_interaction_bot_for_userbot(db, user_id, userbot_id, chat_id, bot_username):
    variants = build_telegram_ref_variants(bot_username)
    return db.execute("""
        SELECT id, userbot_id, bot_username, commands, last_run, interval_minutes,
               last_command_message_id, last_response_at, manual_intervention_at,
               custom_name, tracked_chat_id, resolved_chat_id
        FROM interaction_bots
        WHERE user_id = ?
          AND userbot_id = ?
          AND status = 'active'
          AND (
              CAST(resolved_chat_id AS TEXT) = CAST(? AS TEXT)
              OR LOWER(bot_username) IN (?, ?, ?, ?)
          )
        ORDER BY last_run DESC
        LIMIT 1
    """, (user_id, userbot_id, chat_id, *variants)).fetchone()

def is_automated_outgoing_interaction_command(interaction_row, message):
    if not interaction_row:
        return False
    message_id = getattr(message, 'id', None)
    last_command_message_id = interaction_row['last_command_message_id'] if 'last_command_message_id' in interaction_row.keys() else None
    if message_id and last_command_message_id and str(message_id) == str(last_command_message_id):
        return True

    last_run = parse_stored_datetime(interaction_row['last_run'])
    msg_dt = telegram_message_date_to_local_naive(getattr(message, 'date', None))
    if not last_run or abs((msg_dt - last_run).total_seconds()) > INTERACTION_COMMAND_AUTOMATION_FALLBACK_SECONDS:
        return False

    try:
        commands = json.loads(interaction_row['commands'] or '[]')
    except Exception:
        commands = []
    text = str(getattr(message, 'text', '') or '').strip()
    if text and text in {str(cmd).strip() for cmd in commands}:
        return True

    return False

def mark_manual_interaction_intervention(db, interaction_row, message_date):
    if not interaction_row or 'id' not in interaction_row.keys():
        return
    seen_at = telegram_message_date_to_local_naive(message_date).strftime('%Y-%m-%d %H:%M:%S')
    db.execute(
        "UPDATE interaction_bots SET manual_intervention_at = ? WHERE id = ?",
        (seen_at, interaction_row['id'])
    )

def safe_int(value, default=0):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default

def datetime_sort_value(value):
    parsed = parse_stored_datetime(value)
    if not parsed:
        return 0
    return parsed.timestamp()

def get_interaction_display_windows(db, user_id):
    rows = db.execute("""
        SELECT id, resolved_chat_id, last_run, last_response_at
        FROM interaction_bots
        WHERE user_id = ?
          AND status = 'active'
          AND resolved_chat_id IS NOT NULL
          AND last_run IS NOT NULL
    """, (user_id,)).fetchall()
    windows = []
    for row in rows:
        start = parse_stored_datetime(row['last_run'])
        if not start:
            continue
        last_response = parse_stored_datetime(row['last_response_at'] if 'last_response_at' in row.keys() else None)
        end = (last_response + timedelta(seconds=INTERACTION_RESPONSE_IDLE_SECONDS)) if last_response else (start + timedelta(seconds=INTERACTION_RESPONSE_START_WINDOW_SECONDS))
        windows.append({
            'id': row['id'],
            'chat_id': str(row['resolved_chat_id']),
            'start': start,
            'end': end,
        })
    return windows

def find_interaction_window_for_row(row, windows):
    chat_id = str(row.get('chat_id') or '')
    msg_dt = parse_stored_datetime(row.get('date'))
    if not chat_id or not msg_dt:
        return None
    for window in windows:
        if window['chat_id'] == chat_id and window['start'] - timedelta(seconds=5) <= msg_dt <= window['end']:
            return window
    return None

def get_interaction_bot_chat_ids(db, user_id):
    rows = db.execute(
        """
        SELECT resolved_chat_id
        FROM interaction_bots
        WHERE user_id = ?
          AND status = 'active'
          AND resolved_chat_id IS NOT NULL
        """,
        (user_id,),
    ).fetchall()
    return {str(row['resolved_chat_id']) for row in rows if row['resolved_chat_id'] is not None}

def build_row_sort_identity(row):
    return (
        str(row.get('chat_id') or ''),
        safe_int(row.get('telegram_message_id'), safe_int(row.get('message_id'), safe_int(row.get('id'), 0))),
        safe_int(row.get('id'), safe_int(row.get('message_id'), 0)),
        safe_int(row.get('line_index'), -1),
    )

def build_interaction_burst_sort_keys(rows, db, user_id):
    bot_chat_ids = get_interaction_bot_chat_ids(db, user_id)
    if not bot_chat_ids:
        return {}

    rows_by_chat = {}
    for row in rows:
        chat_id = str(row.get('chat_id') or '')
        if chat_id not in bot_chat_ids:
            continue
        msg_dt = parse_stored_datetime(row.get('date'))
        if not msg_dt:
            continue
        rows_by_chat.setdefault(chat_id, []).append((msg_dt, row))

    sort_keys = {}
    for chat_rows in rows_by_chat.values():
        chat_rows.sort(
            key=lambda item: (
                item[0],
                safe_int(item[1].get('telegram_message_id'), safe_int(item[1].get('message_id'), safe_int(item[1].get('id'), 0))),
                safe_int(item[1].get('id'), safe_int(item[1].get('message_id'), 0)),
            )
        )

        burst = []
        previous_dt = None

        def flush_burst():
            if not burst:
                return
            burst_end = max(item[0] for item in burst)
            burst_order = -burst_end.timestamp()
            for _, burst_row in burst:
                telegram_order = safe_int(
                    burst_row.get('telegram_message_id'),
                    safe_int(burst_row.get('message_id'), safe_int(burst_row.get('id'), 0)),
                )
                line_order = safe_int(burst_row.get('line_index'), -1)
                row_id = safe_int(burst_row.get('id'), safe_int(burst_row.get('message_id'), 0))
                sort_keys[build_row_sort_identity(burst_row)] = (
                    burst_order,
                    0,
                    telegram_order,
                    line_order,
                    row_id,
                )

        for msg_dt, row in chat_rows:
            if previous_dt and (msg_dt - previous_dt).total_seconds() > INTERACTION_RESPONSE_BURST_GAP_SECONDS:
                flush_burst()
                burst = []
            burst.append((msg_dt, row))
            previous_dt = msg_dt
        flush_burst()

    return sort_keys

def sort_message_rows_preserving_bot_order(rows, db, user_id):
    windows = get_interaction_display_windows(db, user_id)
    interaction_burst_sort_keys = build_interaction_burst_sort_keys(rows, db, user_id)

    def sort_key(row):
        window = find_interaction_window_for_row(row, windows)
        telegram_order = safe_int(row.get('telegram_message_id'), safe_int(row.get('message_id'), safe_int(row.get('id'), 0)))
        line_order = safe_int(row.get('line_index'), -1)
        row_id = safe_int(row.get('id'), safe_int(row.get('message_id'), 0))
        burst_key = interaction_burst_sort_keys.get(build_row_sort_identity(row))
        if burst_key:
            return burst_key
        if window:
            return (-window['start'].timestamp(), 0, telegram_order, line_order, row_id)
        return (-datetime_sort_value(row.get('date')), 1, -telegram_order, line_order, -row_id)

    return sorted(rows, key=sort_key)

def should_cleanup_interaction_response_snapshot(interaction_row, user_id, chat_id):
    if not interaction_row or 'id' not in interaction_row.keys() or 'last_run' not in interaction_row.keys():
        return False
    last_run = str(interaction_row['last_run'] or '').split('.')[0]
    if not last_run:
        return False

    cleanup_key = (int(user_id), str(chat_id), int(interaction_row['id']), last_run)
    with interaction_response_cleanup_lock:
        if cleanup_key in interaction_response_cleanup_runs:
            return False
        if len(interaction_response_cleanup_runs) >= INTERACTION_CLEANUP_RUN_MAX_ENTRIES:
            interaction_response_cleanup_runs.clear()
        interaction_response_cleanup_runs.add(cleanup_key)
        return True
# ---------- Глобальные структуры ----------
background_tasks = {}      # session_id -> thread
user_clients = {}          # session_id -> (thread, client, user_id)
qr_sessions = {}           # session_id -> данные для QR
listener_locks = {}        # блокировки для каждого session_id
user_loops = {}            # session_id -> asyncio event loop


import asyncio
import logging

# --- ПОДАВЛЕНИЕ СИСТЕМНЫХ ОШИБОК TELETHON ПРИ ЗАКРЫТИИ ПОТОКОВ ---
def custom_exception_handler(loop, context):
    exception = context.get("exception")
    msg_str = str(exception) if exception else str(context.get("message", ""))
    
    # 1. Игнорируем ошибки закрытого цикла
    if isinstance(exception, (RuntimeError, GeneratorExit)) or "Task was destroyed" in msg_str or "Event loop is closed" in msg_str:
        return
    
    # 2. Игнорируем внутренний баг Telethon при отмене задач (AttributeError)
    if isinstance(exception, AttributeError) and "'NoneType' object has no attribute" in msg_str and ("recv" in msg_str or "disconnect" in msg_str):
        return
        
    # Для остальных, реальных ошибок используем стандартный обработчик
    loop.default_exception_handler(context)

# Применяем этот обработчик ко всем новым потокам
# (Вам нужно добавить `loop.set_exception_handler(custom_exception_handler)` 
# в функцию, где вы создаете новый цикл `loop = asyncio.new_event_loop()`)

import functools
from flask import session, redirect, url_for
from datetime import datetime, timedelta, timezone

def build_expired_session_response():
    session.clear()
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Сеанс завершен'}), 401
    return redirect(url_for('login_page'))

def get_session_user_record():
    if 'user_id' not in session:
        return None
    db = get_db()
    return db.execute(
        "SELECT id, role, session_token FROM users WHERE id = ?",
        (session['user_id'],)
    ).fetchone()

def has_invalid_session_token(user):
    if not user:
        return True
    if 'session_token' in user.keys() and user['session_token']:
        return session.get('session_token') != user['session_token']
    return False

def login_required(f):
    @functools.wraps(f)
    def wrapped(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login_page'))
            
        user = get_session_user_record()
        if has_invalid_session_token(user):
            return build_expired_session_response()
        
        return f(*args, **kwargs)
    return wrapped

def ensure_event_loop():
    try:
        return asyncio.get_running_loop()
    except RuntimeError:
        loop = asyncio.new_event_loop()
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        return loop

# ---------- Работа с БД ----------
def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        # 1. Если база занята, поток ждет вместо мгновенного падения.
        db = g._database = sqlite3.connect(
            app.config['DATABASE'],
            timeout=SQLITE_BUSY_TIMEOUT_MS / 1000,
        )
        db.execute(f"PRAGMA busy_timeout={SQLITE_BUSY_TIMEOUT_MS};")
        
        # 2. Включаем режим WAL для параллельной работы
        db.execute("PRAGMA journal_mode=WAL;")
        
        # 3. Синхронизация (опционально для ускорения записи)
        db.execute("PRAGMA synchronous=NORMAL;")
        
        db.row_factory = sqlite3.Row
    return db

def run_db_write_transaction(db, writer, retries=5, retry_delay=0.5, use_lock=True):
    for attempt in range(retries):
        try:
            if use_lock:
                with sqlite_write_lock:
                    result = writer()
                    db.commit()
                    return result
            result = writer()
            db.commit()
            return result
        except sqlite3.OperationalError as e:
            if 'locked' not in str(e).lower():
                raise
            try:
                db.rollback()
            except Exception:
                pass
            if attempt + 1 >= retries:
                raise
            time.sleep(retry_delay * (attempt + 1))

def enqueue_auto_confirm_source_message(user_id, message_db_id, source_label=""):
    if not message_db_id:
        return

    def worker():
        try:
            with app.app_context():
                db = get_db()
                auto_confirm_result = run_db_write_transaction(
                    db,
                    lambda: auto_confirm_source_message_matches(
                        db,
                        user_id,
                        message_db_id,
                        max_seconds=AUTO_CONFIRM_MESSAGE_MAX_SECONDS,
                    ),
                    use_lock=True,
                )
                if auto_confirm_result and auto_confirm_result["stats"].get("source_enabled"):
                    logger.info(
                        "Автопривязка источника %s: %s",
                        source_label or message_db_id,
                        auto_confirm_result["stats"],
                    )
                notify_clients("products")
        except Exception:
            logger.exception(
                "Фоновая автопривязка source=%s message_db_id=%s упала: %s",
                source_label,
                message_db_id,
                "см. traceback",
            )

    threading.Thread(target=worker, daemon=True).start()

def get_table_sql(db, table_name):
    row = db.execute(
        "SELECT sql FROM sqlite_master WHERE type = 'table' AND name = ?",
        (table_name,),
    ).fetchone()
    return (row[0] if row else '') or ''


def table_exists(db, table_name):
    return bool(get_table_sql(db, table_name))


def row_has_column(row, column_name):
    return row is not None and column_name in row.keys()


def normalize_telegram_ref(value):
    raw = str(value or '').strip().lower()
    if not raw:
        return ''
    raw = raw.split('?', 1)[0].strip()
    raw = raw.rstrip('/').strip()
    match = re.search(r'(?:t\.me|telegram\.me)/([^/\s)]+)', raw)
    if match:
        raw = match.group(1)
    raw = re.sub(r'^(?:https?://)?(?:www\.)?', '', raw)
    raw = raw.lstrip('@').strip()
    return raw


def get_interaction_bot_tracked_chat_ids(db, user_id, bot_rows):
    tracked_ids = set()
    bot_rows = [row for row in bot_rows if row]
    if not bot_rows:
        return tracked_ids

    for bot in bot_rows:
        if row_has_column(bot, 'tracked_chat_id') and bot['tracked_chat_id']:
            tracked_ids.add(int(bot['tracked_chat_id']))
        if row_has_column(bot, 'resolved_chat_id') and bot['resolved_chat_id']:
            rows = db.execute(
                "SELECT id FROM tracked_chats WHERE user_id=? AND chat_id=?",
                (user_id, bot['resolved_chat_id'])
            ).fetchall()
            tracked_ids.update(int(row['id']) for row in rows)

    tracked_rows = db.execute(
        "SELECT id, chat_id, chat_title, custom_name FROM tracked_chats WHERE user_id=?",
        (user_id,)
    ).fetchall()
    for bot in bot_rows:
        bot_refs = {
            normalize_telegram_ref(bot['bot_username'] if 'bot_username' in bot.keys() else ''),
            normalize_telegram_ref(bot['custom_name'] if 'custom_name' in bot.keys() else ''),
        }
        bot_refs.discard('')
        if not bot_refs:
            continue
        for chat in tracked_rows:
            chat_refs = {
                normalize_telegram_ref(chat['chat_id']),
                normalize_telegram_ref(chat['chat_title']),
                normalize_telegram_ref(chat['custom_name']),
            }
            chat_refs.discard('')
            if bot_refs.intersection(chat_refs):
                tracked_ids.add(int(chat['id']))
    return tracked_ids


def delete_tracked_chats_by_ids(db, user_id, tracked_ids):
    deleted = 0
    for tracked_id in sorted({int(item) for item in tracked_ids if item}):
        row = db.execute(
            "SELECT chat_id FROM tracked_chats WHERE id=? AND user_id=?",
            (tracked_id, user_id)
        ).fetchone()
        if not row:
            continue
        real_chat_id = row['chat_id']
        db.execute(
            "DELETE FROM product_messages WHERE message_id IN (SELECT id FROM messages WHERE chat_id=? AND user_id=?)",
            (real_chat_id, user_id)
        )
        db.execute("DELETE FROM messages WHERE chat_id=? AND user_id=?", (real_chat_id, user_id))
        db.execute("DELETE FROM excel_configs WHERE chat_id=? AND user_id=?", (real_chat_id, user_id))
        db.execute("DELETE FROM excel_missing_sheets WHERE chat_id=? AND user_id=?", (real_chat_id, user_id))
        db.execute("DELETE FROM tracked_chats WHERE id=? AND user_id=?", (tracked_id, user_id))
        deleted += 1
    return deleted


def delete_all_tracked_chats_for_user(db, user_id):
    rows = db.execute("SELECT id FROM tracked_chats WHERE user_id=?", (user_id,)).fetchall()
    return delete_tracked_chats_by_ids(db, user_id, [row['id'] for row in rows])


def get_next_product_sort_index(db, user_id):
    row = db.execute(
        "SELECT COALESCE(MAX(sort_index), 0) + 1 FROM products WHERE user_id = ?",
        (user_id,),
    ).fetchone()
    return int(row[0] if row and row[0] is not None else 1)


def build_product_sort_index(base_index, product_name=None):
    try:
        normalized_base = int(base_index)
    except (TypeError, ValueError):
        return None

    raw_name = str(product_name or '').lower()
    if 'esim' in raw_name and '1sim' not in raw_name and '2sim' not in raw_name:
        variant_rank = 0
    elif '1sim' in raw_name or 'sim+esim' in raw_name:
        variant_rank = 1
    elif '2sim' in raw_name:
        variant_rank = 2
    else:
        variant_rank = 3
    return normalized_base * 10 + variant_rank


def update_product_sort_index(db, user_id, product_id, sort_index):
    if sort_index is None:
        return
    try:
        normalized_sort_index = int(sort_index)
    except (TypeError, ValueError):
        return
    db.execute(
        "UPDATE products SET sort_index = ? WHERE id = ? AND user_id = ?",
        (normalized_sort_index, product_id, user_id),
    )


def maybe_sync_product_sort_from_message(db, user_id, product_id, message_id, line_index):
    if line_index is None or int(line_index) < 0:
        return
    product = db.execute(
        "SELECT name FROM products WHERE id = ? AND user_id = ?",
        (product_id, user_id),
    ).fetchone()
    msg = db.execute(
        "SELECT chat_id, type FROM messages WHERE id = ? AND user_id = ?",
        (message_id, user_id),
    ).fetchone()
    if not msg or not product:
        return

    msg_type = str(msg['type'] or '')
    chat_id = str(msg['chat_id'] or '')
    if msg_type.startswith('excel') or chat_id.startswith('api_src_'):
        update_product_sort_index(
            db,
            user_id,
            product_id,
            build_product_sort_index(int(line_index), product['name']),
        )


INTERACTION_RESPONSE_BACKFILL_DELAYS_SECONDS = (4, 15, 45, 90, 150, 240)
INTERACTION_RESPONSE_BACKFILL_LIMIT = 100
INTERACTION_COMMAND_SEND_TIMEOUT_SECONDS = 30
INTERACTION_BACKFILL_ATTEMPT_TIMEOUT_SECONDS = 35


async def backfill_recent_interaction_bot_responses(client, entity, user_id, userbot_id, interaction_id, bot_username, since_dt, after_message_id=None):
    """Save all recent bot response parts first; product auto-confirm runs only after rows are durable."""
    candidates = []
    iter_kwargs = {'limit': INTERACTION_RESPONSE_BACKFILL_LIMIT}
    if after_message_id:
        try:
            iter_kwargs['min_id'] = int(after_message_id)
        except (TypeError, ValueError):
            pass
    async for message in client.iter_messages(entity, **iter_kwargs):
        if getattr(message, 'out', False):
            continue
        if not message.text and not message.document:
            continue
        msg_dt = telegram_message_date_to_local_naive(message.date)
        if 'min_id' not in iter_kwargs and msg_dt + timedelta(seconds=5) < since_dt:
            break
        candidates.append(message)

    prepared = []
    for message in reversed(candidates):
        try:
            chat = await message.get_chat()
            sender = await message.get_sender()
            chat_id = getattr(chat, 'id', getattr(entity, 'id', None))
            if chat_id is None:
                continue

            first = getattr(chat, 'first_name', '') or ''
            last = getattr(chat, 'last_name', '') or ''
            chat_title = f"{first} {last}".strip() or getattr(chat, 'title', None) or str(chat_id)
            msg_type = 'private' if getattr(chat, 'bot', False) else 'channel'
            sender_name = getattr(sender, 'username', '') if sender else ''
            if not sender_name:
                sender_name = getattr(sender, 'first_name', 'Бот') if sender else 'Бот'
            bot_un_clean = normalize_telegram_ref(sender_name or bot_username)

            parsed_text = await parse_excel_message(client, message, chat_id, user_id)
            if not parsed_text:
                logger.warning(
                    "Backfill пропустил пустой ответ бота %s message_id=%s",
                    bot_username,
                    getattr(message, 'id', None),
                )
                continue
            prepared.append({
                'message_id': message.id,
                'message_date': message.date,
                'chat_id': chat_id,
                'chat_title': chat_title,
                'msg_type': msg_type,
                'sender_name': sender_name,
                'bot_un_clean': bot_un_clean,
                'text': parsed_text,
                'date': telegram_message_date_to_local_naive(message.date).strftime("%Y-%m-%d %H:%M:%S"),
                'reply_to_msg_id': getattr(message, 'reply_to_msg_id', None),
            })
        except Exception as e:
            logger.warning(
                "Backfill не смог подготовить ответ бота %s message_id=%s: %s",
                bot_username,
                getattr(message, 'id', None),
                e,
            )

    if not prepared:
        return 0

    auto_confirm_ids = []
    saved = 0
    with app.app_context():
        db = get_db()
        interaction_row = db.execute(
            """
            SELECT id, userbot_id, bot_username, commands, last_run, interval_minutes,
                   last_command_message_id, last_response_at, manual_intervention_at,
                   custom_name,
                   tracked_chat_id, resolved_chat_id
            FROM interaction_bots
            WHERE id = ? AND user_id = ? AND userbot_id = ? AND status = 'active'
            """,
            (interaction_id, user_id, userbot_id),
        ).fetchone()
        if not interaction_row:
            return 0

        def writer():
            nonlocal saved
            persisted_ids = []
            snapshot_has_price_lines = any(
                telegram_message_date_to_local_naive(item['message_date']) + timedelta(seconds=5) >= since_dt
                and source_text_has_supplier_price_lines(item['text'])
                for item in prepared
            )
            snapshot_cleanup_done = False
            for item in prepared:
                msg_dt = telegram_message_date_to_local_naive(item['message_date'])
                if msg_dt + timedelta(seconds=5) < since_dt:
                    continue
                tracked_row = ensure_tracked_source_row(
                    db,
                    user_id,
                    item['chat_id'],
                    item['chat_title'],
                    interaction_row['custom_name'] if 'custom_name' in interaction_row.keys() else (item['bot_un_clean'] or bot_username),
                )
                if not tracked_row:
                    logger.warning("Backfill не смог создать источник для бота %s chat_id=%s", bot_username, item['chat_id'])
                    continue
                if not interaction_row['tracked_chat_id'] or not interaction_row['resolved_chat_id']:
                    db.execute(
                        """
                        UPDATE interaction_bots
                        SET tracked_chat_id = CASE
                                WHEN tracked_chat_id IS NULL OR tracked_chat_id = '' THEN ?
                                ELSE tracked_chat_id
                            END,
                            resolved_chat_id = CASE
                                WHEN resolved_chat_id IS NULL OR resolved_chat_id = '' THEN ?
                                ELSE resolved_chat_id
                            END
                        WHERE id = ? AND user_id = ? AND userbot_id = ?
                        """,
                        (tracked_row['id'], item['chat_id'], interaction_row['id'], user_id, userbot_id),
                    )
                if (
                    snapshot_has_price_lines
                    and not snapshot_cleanup_done
                    and should_cleanup_interaction_response_snapshot(interaction_row, user_id, item['chat_id'])
                ):
                    cleanup_source_before_new_snapshot(db, user_id, item['chat_id'])
                    snapshot_cleanup_done = True
                item_has_price_lines = source_text_has_supplier_price_lines(item['text'])
                insert_cursor = db.execute(
                    """
                    INSERT OR IGNORE INTO messages
                    (user_id, telegram_message_id, type, text, date, chat_id, chat_title, sender_name, is_delayed, is_blocked)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, 0, 0)
                    """,
                    (
                        user_id,
                        item['message_id'],
                        item['msg_type'],
                        item['text'],
                        item['date'],
                        item['chat_id'],
                        item['chat_title'],
                        item['sender_name'],
                    ),
                )
                inserted = getattr(insert_cursor, 'rowcount', 0) > 0
                existing = db.execute(
                    """
                    SELECT id, text, date, is_blocked, is_delayed
                    FROM messages
                    WHERE telegram_message_id = ? AND chat_id = ? AND user_id = ?
                    """,
                    (item['message_id'], item['chat_id'], user_id),
                ).fetchone()
                if not existing:
                    continue
                already_current = (not inserted) and (
                    str(existing['text'] or '') == str(item['text'] or '')
                    and str(existing['date'] or '') == str(item['date'] or '')
                    and int(existing['is_blocked'] or 0) == 0
                    and int(existing['is_delayed'] or 0) == 0
                )
                if already_current:
                    has_active_bindings = db.execute(
                        """
                        SELECT 1
                        FROM product_messages
                        WHERE message_id = ?
                          AND status = 'confirmed'
                          AND COALESCE(is_actual, 1) = 1
                        LIMIT 1
                        """,
                        (existing['id'],),
                    ).fetchone()
                    if not has_active_bindings and source_text_has_supplier_price_lines(item['text']):
                        persisted_ids.append(existing['id'])
                    if not item_has_price_lines:
                        mark_interaction_bot_response_seen(db, interaction_row, item['message_date'])
                    continue
                db.execute(
                    """
                    UPDATE messages
                    SET text = ?, date = ?, chat_title = ?, sender_name = ?, type = ?, is_blocked = 0, is_delayed = 0
                    WHERE id = ?
                    """,
                    (
                        item['text'],
                        item['date'],
                        item['chat_title'],
                        item['sender_name'],
                        item['msg_type'],
                        existing['id'],
                    ),
                )
                mark_interaction_bot_response_seen(db, interaction_row, item['message_date'])
                if item_has_price_lines:
                    persisted_ids.append(existing['id'])
                saved += 1
            return persisted_ids

        auto_confirm_ids = run_db_write_transaction(db, writer, use_lock=True) or []

        if auto_confirm_ids:
            notify_clients()

        for message_db_id in auto_confirm_ids:
            enqueue_auto_confirm_source_message(user_id, message_db_id, f"Backfill {bot_username}")

    return saved


async def backfill_interaction_bot_responses_after_send(client, entity, user_id, userbot_id, interaction_id, bot_username, since_dt, after_message_id=None):
    total_saved = 0
    started = time.monotonic()
    for delay in INTERACTION_RESPONSE_BACKFILL_DELAYS_SECONDS:
        await asyncio.sleep(max(0, delay - (time.monotonic() - started)))
        total_saved += await backfill_recent_interaction_bot_responses(
            client,
            entity,
            user_id,
            userbot_id,
            interaction_id,
            bot_username,
            since_dt,
            after_message_id=after_message_id,
        )
    return total_saved


async def send_interaction_bot_command_and_backfill(client, user_id, userbot_id, interaction_id, bot_username, cmd, since_dt, resolved_chat_id=None):
    try:
        target = bot_username
        if resolved_chat_id:
            try:
                target = int(resolved_chat_id)
            except (TypeError, ValueError):
                target = bot_username

        async def send_command():
            entity = await client.get_entity(target)
            sent = await client.send_message(entity, cmd)
            return entity, sent

        entity, sent_message = await asyncio.wait_for(
            send_command(),
            timeout=INTERACTION_COMMAND_SEND_TIMEOUT_SECONDS,
        )
        sent_message_id = getattr(sent_message, 'id', None)
        entity_chat_id = getattr(entity, 'id', resolved_chat_id)
        first = getattr(entity, 'first_name', '') or ''
        last = getattr(entity, 'last_name', '') or ''
        entity_title = (
            f"{first} {last}".strip()
            or getattr(entity, 'title', None)
            or getattr(entity, 'username', None)
            or str(entity_chat_id or bot_username)
        )

        if sent_message_id:
            with app.app_context():
                db = get_db()
                def remember_command_and_source():
                    tracked_row = None
                    if entity_chat_id is not None:
                        tracked_row = ensure_tracked_source_row(db, user_id, entity_chat_id, entity_title, bot_username)
                    db.execute(
                        """
                        UPDATE interaction_bots
                        SET resolved_chat_id = CASE
                                WHEN ? IS NOT NULL THEN ?
                                ELSE resolved_chat_id
                            END,
                            tracked_chat_id = CASE
                                WHEN ? IS NOT NULL THEN ?
                                ELSE tracked_chat_id
                            END,
                            last_command_message_id = ?,
                            manual_intervention_at = NULL
                        WHERE id = ? AND user_id = ? AND userbot_id = ?
                        """,
                        (
                            entity_chat_id,
                            entity_chat_id,
                            tracked_row['id'] if tracked_row else None,
                            tracked_row['id'] if tracked_row else None,
                            sent_message_id,
                            interaction_id,
                            user_id,
                            userbot_id,
                        ),
                    )

                run_db_write_transaction(
                    db,
                    remember_command_and_source,
                    use_lock=True,
                )

        logger.info("✅ Успешно отправлена команда '%s' боту %s", cmd, bot_username)
        backfilled = await backfill_interaction_bot_responses_after_send(
            client,
            entity,
            user_id,
            userbot_id,
            interaction_id,
            bot_username,
            since_dt,
            after_message_id=sent_message_id,
        )
        if backfilled:
            logger.info(
                "✅ Backfill сохранил %s ответ(ов) бота %s после команды '%s'",
                backfilled,
                bot_username,
                cmd,
            )
    except Exception as e:
        logger.error("❌ Ошибка отправки/backfill команды '%s' боту %s: %s", cmd, bot_username, e)


async def recover_interaction_bot_responses_once(
    client,
    user_id,
    userbot_id,
    interaction_id,
    bot_username,
    since_dt,
    last_command_message_id,
    resolved_chat_id=None,
):
    if not since_dt or not last_command_message_id:
        return 0
    target = bot_username
    if resolved_chat_id:
        try:
            target = int(resolved_chat_id)
        except (TypeError, ValueError):
            target = bot_username
    entity = await client.get_entity(target)
    return await asyncio.wait_for(
        backfill_recent_interaction_bot_responses(
            client,
            entity,
            user_id,
            userbot_id,
            interaction_id,
            bot_username,
            since_dt,
            after_message_id=last_command_message_id,
        ),
        timeout=INTERACTION_BACKFILL_ATTEMPT_TIMEOUT_SECONDS,
    )


def bot_interaction_scheduler():
    """Фоновый поток для автоматической отправки команд ботам по расписанию"""
    first_run = True
    while True:
        time.sleep(5 if first_run else 60)
        first_run = False
        try:
            with app.app_context():
                db = get_db()
                active_interactions = db.execute("SELECT * FROM interaction_bots WHERE status = 'active'").fetchall()
                from datetime import timedelta
                now = datetime.now(timezone.utc).replace(tzinfo=None) + timedelta(hours=3)
                
                for interaction in active_interactions:
                    # Оборачиваем КАЖДОГО бота в свой try..except, чтобы ошибка в одном не ломала других
                    try:
                        userbot_id = interaction['userbot_id']
                        
                        if not is_parsing_allowed(userbot_id):
                            continue 
                            
                        last_run = interaction['last_run']
                        last_run_dt = None
                        if last_run:
                            try:
                                last_run_dt = datetime.strptime(last_run, '%Y-%m-%d %H:%M:%S.%f')
                            except ValueError:
                                last_run_dt = datetime.strptime(last_run, '%Y-%m-%d %H:%M:%S')
                        
                        if userbot_id in user_clients:
                            _, client, _ = user_clients[userbot_id]
                            loop = user_loops.get(userbot_id)
                            
                            if loop and loop.is_running() and client.is_connected():
                                last_command_message_id = interaction['last_command_message_id'] if 'last_command_message_id' in interaction.keys() else None
                                if last_run_dt and last_command_message_id:
                                    age_seconds = (now - last_run_dt).total_seconds()
                                    recovery_window_seconds = max(
                                        INTERACTION_RESPONSE_IDLE_SECONDS,
                                        safe_int(interaction['interval_minutes'], 5) * 60,
                                    )
                                    if 0 <= age_seconds <= recovery_window_seconds:
                                        recover_future = asyncio.run_coroutine_threadsafe(
                                            recover_interaction_bot_responses_once(
                                                client,
                                                interaction['user_id'],
                                                userbot_id,
                                                interaction['id'],
                                                interaction['bot_username'],
                                                last_run_dt,
                                                last_command_message_id,
                                                interaction['resolved_chat_id'] if 'resolved_chat_id' in interaction.keys() else None,
                                            ),
                                            loop,
                                        )
                                        try:
                                            recovered = recover_future.result(
                                                timeout=INTERACTION_BACKFILL_ATTEMPT_TIMEOUT_SECONDS + 5
                                            )
                                            if recovered:
                                                logger.info(
                                                    "✅ Recovery backfill сохранил %s ответ(ов) бота %s",
                                                    recovered,
                                                    interaction['bot_username'],
                                                )
                                        except Exception as recover_e:
                                            recover_future.cancel()
                                            logger.warning(
                                                "Recovery backfill для бота %s не завершился: %s",
                                                interaction['bot_username'],
                                                recover_e,
                                            )

                                if last_run_dt:
                                    diff_minutes = (now - last_run_dt).total_seconds() / 60
                                    if diff_minutes < interaction['interval_minutes']:
                                        continue  # Время еще не пришло

                                commands = json.loads(interaction['commands'])
                                bot_username = interaction['bot_username']
                                
                                logger.info(f"🔄 Запуск команд для бота {bot_username}...")
                                run_db_write_transaction(
                                    db,
                                    lambda: db.execute(
                                        """
                                        UPDATE interaction_bots
                                        SET last_run = ?,
                                            last_command_message_id = NULL,
                                            last_response_at = NULL,
                                            manual_intervention_at = NULL
                                        WHERE id = ?
                                        """,
                                        (now, interaction['id'])
                                    ),
                                    use_lock=True,
                                )
                                
                                for cmd in commands:
                                    asyncio.run_coroutine_threadsafe(
                                        send_interaction_bot_command_and_backfill(
                                            client,
                                            interaction['user_id'],
                                            userbot_id,
                                            interaction['id'],
                                            bot_username,
                                            str(cmd),
                                            now,
                                            interaction['resolved_chat_id'] if 'resolved_chat_id' in interaction.keys() else None,
                                        ),
                                        loop,
                                    )
                                    time.sleep(0.2)
                                
                                if 'notify_clients' in globals():
                                    notify_clients()
                                    
                    except Exception as bot_e:
                        logger.error(f"Сбой в логике расписания для бота {interaction['bot_username']}: {bot_e}")
                             
        except Exception as e:
            logger.error(f"Глобальная ошибка в планировщике ботов: {e}")





# --- Маршруты для Товаров ---
@app.before_request
def check_session_token():
    # Пропускаем проверку для страниц логина и логаута
    if request.endpoint in ['login_page', 'static', 'logout']:
        return
        
    if 'user_id' in session:
        try:
            user = get_session_user_record()
            if has_invalid_session_token(user):
                return build_expired_session_response()
        except Exception:
            pass

import google.generativeai as genai
import json
from flask import request, jsonify

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY") or ""
if GEMINI_API_KEY:
    genai.configure(api_key=GEMINI_API_KEY)

AUTOFILL_RESPONSE_KEYS = [
    "brand",
    "country_sim",
    "weight",
    "color",
    "storage",
    "ram",
    "display",
    "processor",
    "camera",
    "front_camera",
    "video",
    "connectivity",
    "battery",
    "os",
    "biometrics",
    "charging",
    "warranty",
    "model_no",
    "description",
]

AUTOFILL_REQUIRED_NOTE = "Версия без RuStore"
AUTOFILL_APPLE_NOTE = "Оригинальная продукция Apple"
AUTOFILL_MODELS = ["gemini-2.5-flash", "gemini-2.5-flash-lite"]
AUTOFILL_SUPPLEMENT_MIN_SCORE = 1.2
DEFAULT_IPHONE_OS_VERSION = "iOS 18"
STORE_WARRANTY_DEFAULT = "Гарантия от магазина 1 год"
GENERIC_BRAND_LABELS = {
    "card",
    "cm",
    "coca",
    "crossbody",
    "digital",
    "finewoven",
    "iphone",
    "magic",
    "nsw",
    "nsw2",
    "ps",
    "ps4",
    "ps5",
    "roll",
    "russian",
    "smart",
    "steamdeck",
    "switch",
    "tab",
    "type-c",
    "usb-c",
    "utility",
    "аксессуары",
    "аксессуары для",
    "гаджеты",
    "гейминг",
    "геймпады для",
    "детские умные часы",
    "игры и подписки для",
    "наушники и колонки",
    "портативные игровые приставки",
    "планшеты",
    "продукция dyson",
    "смартфоны",
    "умные часы",
    "умные часы и фитнес-браслеты",
}
BRAND_COUNTRY_FALLBACKS = {
    "A4TECH": "Китай",
    "Anker": "Китай",
    "Aukey": "Китай",
    "Awei": "Китай",
    "Asus": "Тайвань",
    "Blackmagic": "Сингапур",
    "Baseus": "Китай",
    "Belkin": "США",
    "Borofone": "Китай",
    "Canon": "Япония",
    "Canyon": "Китай",
    "DJI": "Китай",
    "Dyson": "Малайзия",
    "Elago": "Южная Корея",
    "ELARI": "Китай",
    "Fujifilm": "Япония",
    "Garmin": "Тайвань",
    "Google": "Вьетнам",
    "GoPro": "Китай",
    "HOCO": "Китай",
    "Honor": "Китай",
    "HP": "Китай",
    "Hollyland": "Китай",
    "Huawei": "Китай",
    "Inkax": "Китай",
    "Kodak": "Китай",
    "Lenovo": "Китай",
    "Ldnio": "Китай",
    "Logitech": "Китай",
    "Meta": "Китай",
    "MSI": "Тайвань",
    "Native Union": "Гонконг",
    "Nacon": "Китай",
    "Nintendo": "Япония",
    "Nikon": "Таиланд",
    "Nothing": "Китай",
    "Nubia": "Китай",
    "Olympus": "Япония",
    "OnePlus": "Китай",
    "Panasonic": "Малайзия",
    "Philips": "Нидерланды",
    "PlayStation": "Япония",
    "PowerA": "США",
    "Pop Mart": "Китай",
    "reMarkable": "Китай",
    "Ritmix": "Китай",
    "Rode": "Австралия",
    "Razer": "Китай",
    "Samsung": "Вьетнам",
    "SanDisk": "США",
    "Samyang": "Южная Корея",
    "Satechi": "США",
    "Sigma": "Япония",
    "Sony": "Китай",
    "Tamron": "Япония",
    "Tecno": "Китай",
    "Tokina": "Япония",
    "Transcend": "Тайвань",
    "Syntech": "Китай",
    "UAG": "США",
    "Valve": "США",
    "Victrix": "США",
    "WIWU": "Китай",
    "Whoop": "США",
    "Xbox": "США",
    "Xiaomi": "Китай",
    "Яндекс": "Россия",
}
FOLDER_NAME_TRAILING_FILLER_RE = re.compile(r"\s+для\s*$", re.IGNORECASE)
BRAND_KEYWORDS = (
    ("Anker", ("anker",)),
    ("Aukey", ("aukey",)),
    ("Awei", ("awei",)),
    ("Asus", ("asus", "rog phone", "zenfone")),
    ("Blackmagic", ("blackmagic",)),
    ("Canon", ("canon", "eos", "powershot", "ixus")),
    ("DJI", ("dji", "osmo", "mavic", "avata")),
    ("Fujifilm", ("fujifilm", "fuji", "instax")),
    ("Garmin", ("garmin", "fenix", "forerunner", "venu", "epix", "tactix", "etrex")),
    ("Samsung", ("galaxy", "samsung", "z fold", "z flip")),
    ("GoPro", ("gopro", "go pro", "hero ", "hero-", "hero(", "hero)")),
    ("Hollyland", ("hollyland", "lark ")),
    ("Kodak", ("kodak", "pixpro")),
    ("Nikon", ("nikon", "nikkor")),
    ("Olympus", ("olympus", "om system", "om-d", "pen e-p")),
    ("Xiaomi", ("xiaomi", "redmi", "poco")),
    ("Google", ("pixel", "google")),
    ("Honor", ("honor",)),
    ("Realme", ("realme",)),
    ("OnePlus", ("oneplus", "one plus")),
    ("Nothing", ("nothing phone", "cmf phone", "cmf by nothing")),
    ("Oppo", ("oppo",)),
    ("Vivo", ("vivo",)),
    ("Huawei", ("huawei",)),
    ("Яндекс", ("яндекс", "yandex")),
    ("Motorola", ("motorola", " moto ")),
    ("Sony", ("sony", "xperia")),
    ("Infinix", ("infinix",)),
    ("Tecno", ("tecno",)),
    ("Valve", ("steam deck", "valve")),
    ("Lenovo", ("lenovo legion", "legion go")),
    ("MSI", ("msi claw",)),
    ("Meta", ("meta quest", "oculus quest")),
    ("Nacon", ("nacon", "revolution 5 pro")),
    ("Victrix", ("victrix",)),
    ("Razer", ("razer",)),
    ("Backbone", ("backbone",)),
    ("Pop Mart", ("pop mart", "labubu", "the monsters")),
    ("Syntech", ("syntech",)),
    ("Nintendo", ("nintendo switch", "switch oled", "switch lite")),
    ("PlayStation", ("playstation", " ps5", " ps4")),
    ("Xbox", ("xbox",)),
    ("PowerA", ("powera",)),
    ("Dyson", ("dyson",)),
    ("Insta360", ("insta360", "insta 360")),
    ("Panasonic", ("panasonic", "lumix")),
    ("Rode", ("rode",)),
    ("SanDisk", ("sandisk",)),
    ("Samyang", ("samyang",)),
    ("Sigma", ("sigma",)),
    ("Kingston", ("kingston",)),
    ("Philips", ("philips",)),
    ("HP", ("hp ", "hp elite", "hp pro")),
    ("Belkin", ("belkin",)),
    ("Tamron", ("tamron",)),
    ("Tokina", ("tokina",)),
    ("Transcend", ("transcend",)),
    ("Baseus", ("baseus",)),
    ("Borofone", ("borofone",)),
    ("uBear", ("ubear",)),
    ("UNIQ", ("uniq",)),
    ("Breaking", ("breaking",)),
    ("Remax", ("remax",)),
    ("Magssory", ("magssory",)),
    ("Netac", ("netac",)),
    ("OltraMax", ("oltramax",)),
    ("TV-COM", ("tv-com",)),
    ("ADATA", ("adata",)),
    ("Silicon Power", ("silicon power",)),
    ("WIWU", ("wiwu",)),
    ("UAG", ("uag", "urban armor gear")),
    ("Elago", ("elago",)),
    ("Satechi", ("satechi",)),
    ("Native Union", ("native union",)),
    ("Ldnio", ("ldnio",)),
    ("Inkax", ("inkax",)),
    ("Anker", ("anker",)),
    ("Awei", ("awei",)),
    ("Aukey", ("aukey",)),
    ("Prime Line", ("prime line",)),
    ("Pitaka", ("pitaka",)),
    ("Deppa", ("deppa",)),
    ("Defender", ("defender",)),
    ("DEXP", ("dexp",)),
    ("Oxion", ("oxion",)),
    ("PNY", ("pny",)),
    ("HOCO", ("hoco",)),
    ("Kodak", ("kodak",)),
    ("JBL", ("jbl",)),
    ("Logitech", ("logitech",)),
    ("Canyon", ("canyon",)),
    ("Ritmix", ("ritmix",)),
    ("A4TECH", ("a4tech",)),
    ("reMarkable", ("remarkable",)),
    ("Whoop", ("whoop",)),
    ("Nubia", ("nubia", "redmagic")),
)
ACCESSORY_COMPATIBILITY_MARKERS = (
    "airpods",
    "apple watch",
    "iphone",
    "ipad",
    "для apple",
    "for apple",
    "для samsung",
    "for samsung",
    "mag safe",
    "magsafe",
)
DESCRIPTION_TEMPLATE_MARKERS = (
    "это понятная и аккуратно оформленная карточка для магазина",
    "карточка, собранная из актуального каталога cmstore",
    "товар оформлен для каталога магазина",
    "ключевые характеристики:",
    "исполнение:",
    "читать полностью",
    "<div",
)
APPLE_DESCRIPTION_NOTE_KINDS = {"smartphone", "tablet", "laptop", "watch"}
ANDROID_BRANDS = {
    "Asus",
    "Google",
    "Honor",
    "Huawei",
    "Motorola",
    "Nothing",
    "OnePlus",
    "Oppo",
    "Realme",
    "Samsung",
    "Sony",
    "Tecno",
    "Vivo",
    "Xiaomi",
}
IPHONE_MODEL_WEIGHT_MAP = {
    "11": "0.194",
    "12": "0.164",
    "13 Mini": "0.141",
    "13": "0.174",
    "14": "0.172",
    "14 Plus": "0.203",
    "15": "0.171",
    "15 Plus": "0.201",
    "15 Pro": "0.187",
    "15 Pro Max": "0.221",
    "16e": "0.167",
    "16": "0.170",
    "16 Plus": "0.199",
    "16 Pro": "0.199",
    "16 Pro Max": "0.227",
    "17e": "0.180",
    "17": "0.190",
    "17 Air": "0.170",
    "17 Pro": "0.200",
    "17 Pro Max": "0.240",
    "SE 2022": "0.144",
}
IPHONE_MODEL_PATTERNS = [
    (r"^17\s+pro\s+max\b", "17 Pro Max"),
    (r"^17\s+pro\b", "17 Pro"),
    (r"^17\s+air\b", "17 Air"),
    (r"^17e\b", "17e"),
    (r"^17\b", "17"),
    (r"^16\s+pro\s+max\b", "16 Pro Max"),
    (r"^16\s+pro\b", "16 Pro"),
    (r"^16\s+plus\b", "16 Plus"),
    (r"^16e\b", "16e"),
    (r"^16\b", "16"),
    (r"^15\s+pro\s+max\b", "15 Pro Max"),
    (r"^15\s+pro\b", "15 Pro"),
    (r"^15\s+plus\b", "15 Plus"),
    (r"^15\b", "15"),
    (r"^14\s+plus\b", "14 Plus"),
    (r"^14\b", "14"),
    (r"^13\s+mini\b", "13 Mini"),
    (r"^13\b", "13"),
    (r"^12\b", "12"),
    (r"^11\b", "11"),
    (r"^se(?:\s+2022)?\b", "SE 2022"),
]
APPLE_PART_NUMBER_PATTERN = re.compile(r"\bM(?=[A-Z0-9]{4,6}(?:/[A-Z])?\b)(?=[A-Z0-9]*\d)[A-Z0-9]{4,6}(?:/[A-Z])?\b", re.IGNORECASE)
APPLE_FAMILY_PREFIX_PATTERN = re.compile(
    r"^(macbook|ipad|airpods|apple watch|imac|mac mini|mac studio|vision pro)\b",
    re.IGNORECASE,
)
SIM_INLINE_VARIANT_PATTERN = re.compile(r"\s*\((?:[^)]*\b(?:e\s*sim|sim)\b[^)]*)\)", re.IGNORECASE)
SIM_DESCRIPTION_PATTERN = re.compile(
    r"(?:\b(?:e\s*sim|1\s*sim|2\s*sim|sim\+e\s*sim)\b|sim-карт|sim карт|физическ\w+\s+sim|слот\w*\s+для\s+sim)",
    re.IGNORECASE,
)
AUTOFILL_SPEC_KEYS = [
    "display",
    "processor",
    "camera",
    "front_camera",
    "video",
    "connectivity",
    "battery",
    "os",
    "biometrics",
    "charging",
]
AUTOFILL_CACHE_MAX_AGE_DAYS = 30
AI_REPAIR_JOB_BATCH_SIZE = 10
VARIANT_FOLDER_NAMES = {"esim", "sim+esim", "2sim"}
STORAGE_FOLDER_NAME_RE = re.compile(r"^\d+(?:\.\d+)?(?:gb|tb)$", re.IGNORECASE)
IPHONE_CATALOG_PATH = ("Электроника", "Смартфоны", "iPhone")
IPHONE_CASES_CATALOG_PATH = ("Электроника", "Аксессуары", "Чехлы", "Чехлы для телефонов", "Чехлы для iPhone")
IPHONE_SCREEN_PROTECTOR_CATALOG_PATH = (
    "Электроника",
    "Аксессуары",
    "Защитные стекла и пленки",
    "Для смартфонов",
    "Защитные стёкла и плёнки для iPhone",
)
DATA_CABLES_CATALOG_PATH = ("Электроника", "Аксессуары", "Зарядные устройства", "Дата-кабели")
NETWORK_CHARGERS_CATALOG_PATH = ("Электроника", "Аксессуары", "Зарядные устройства", "Сетевые зарядные устройства")
WATCH_CHARGERS_CATALOG_PATH = (
    "Электроника",
    "Аксессуары",
    "Зарядные устройства",
    "Зарядные устройства для умных часов и фитнес-браслетов",
)
YANDEX_STATION_CATALOG_PATH = ("Электроника", "Наушники и колонки", "Умные колонки", "Яндекс станция")
JBL_SPEAKER_CATALOG_PATH = (
    "Электроника",
    "Наушники и колонки",
    "Портативные колонки и акустические системы",
    "JBL",
)
PLAYSTATION_SUBSCRIPTIONS_CATALOG_PATH = ("Электроника", "Гейминг", "PlayStation", "Игры и подписки для")
XBOX_SUBSCRIPTIONS_CATALOG_PATH = ("Электроника", "Гейминг", "Xbox", "Игры и подписки для")
STEAM_DECK_CATALOG_PATH = ("Электроника", "Гейминг", "Портативные игровые приставки", "Steam Deck")
GAMING_CATALOG_PATH = ("Электроника", "Гейминг")
PLAYSTATION_CATALOG_PATH = (*GAMING_CATALOG_PATH, "PlayStation")
PLAYSTATION_CONSOLES_CATALOG_PATH = (*PLAYSTATION_CATALOG_PATH, "Приставки")
PLAYSTATION_GAMEPADS_CATALOG_PATH = (*PLAYSTATION_CATALOG_PATH, "Геймпады")
PLAYSTATION_ACCESSORIES_CATALOG_PATH = (*PLAYSTATION_CATALOG_PATH, "Аксессуары")
XBOX_CATALOG_PATH = (*GAMING_CATALOG_PATH, "Xbox")
XBOX_CONSOLES_CATALOG_PATH = (*XBOX_CATALOG_PATH, "Приставки")
XBOX_GAMEPADS_CATALOG_PATH = (*XBOX_CATALOG_PATH, "Геймпады для")
GAMING_ACCESSORIES_CATALOG_PATH = (*GAMING_CATALOG_PATH, "Игровые аксессуары")
PORTABLE_GAMING_CATALOG_PATH = (*GAMING_CATALOG_PATH, "Портативные игровые приставки")
NINTENDO_CATALOG_PATH = (*PORTABLE_GAMING_CATALOG_PATH, "Nintendo")
NINTENDO_CONSOLES_CATALOG_PATH = (*NINTENDO_CATALOG_PATH, "Приставки Nintendo Switch")
NINTENDO_GAMEPADS_CATALOG_PATH = (*NINTENDO_CATALOG_PATH, "Геймпады для Nintendo")
NINTENDO_ACCESSORIES_CATALOG_PATH = (*GAMING_ACCESSORIES_CATALOG_PATH, "Аксессуары для Nintendo")
STEAM_DECK_ACCESSORIES_CATALOG_PATH = (*GAMING_ACCESSORIES_CATALOG_PATH, "Аксессуары для Steam Deck")
OCULUS_CATALOG_PATH = (*GAMING_CATALOG_PATH, "Системы VR", "Oculus")
CONSTRUCTORS_CATALOG_PATH = ("Электроника", "Гаджеты", "Хобби и отдых", "Конструкторы")
COMPUTER_ACCESSORIES_CATALOG_PATH = ("Электроника", "Аксессуары", "Аксессуары для компьютера")
AIRPODS_CATALOG_PATH = ("Электроника", "Наушники и колонки", "Наушники", "AirPods")
IPHONE_MODEL_FOLDER_PATTERNS = (
    (r"\b17\s+pro\s+max\b", "iPhone 17 Pro Max"),
    (r"\b17\s+pro\b", "iPhone 17 Pro"),
    (r"\b17\s+air\b", "iPhone 17 Air"),
    (r"\b17\s*[eе]\b", "iPhone 17e"),
    (r"\b17\b", "iPhone 17"),
    (r"\b16\s+pro\s+max\b", "iPhone 16 Pro Max"),
    (r"\b16\s+pro\b", "iPhone 16 Pro"),
    (r"\b16\s+plus\b", "iPhone 16 Plus"),
    (r"\b16\s*[eе]\b", "iPhone 16e"),
    (r"\b16\b", "iPhone 16"),
    (r"\b15\s+pro\s+max\b", "iPhone 15 Pro Max"),
    (r"\b15\s+pro\b", "iPhone 15 Pro"),
    (r"\b15\s+plus\b", "iPhone 15 Plus"),
    (r"\b15\b", "iPhone 15"),
    (r"\b14\s+pro\s+max\b", "iPhone 14 Pro Max"),
    (r"\b14\s+pro\b", "iPhone 14 Pro"),
    (r"\b14\s+plus\b", "iPhone 14 Plus"),
    (r"\b14\b", "iPhone 14"),
    (r"\b13\s+mini\b", "iPhone 13 Mini"),
    (r"\b13\b", "iPhone 13"),
    (r"\b12\b", "iPhone 12"),
    (r"\b11\b", "iPhone 11"),
    (r"\bse(?:\s+2022)?\b", "iPhone SE 2022"),
)
IPHONE_17_PRO_ESIM_FLAGS = {"🇧🇭", "🇨🇦", "🇬🇺", "🇯🇵", "🇰🇼", "🇲🇽", "🇴🇲", "🇶🇦", "🇸🇦", "🇦🇪", "🇺🇸", "🇻🇮"}
IPHONE_17_PRO_GLOBAL_FLAGS = {"🇭🇰", "🇲🇴", "🇫🇷", "🇩🇪", "🇬🇧", "🇮🇹", "🇪🇸", "🇸🇪", "🇳🇱", "🇵🇱", "🇦🇺", "🇳🇿", "🇰🇷", "🇸🇬", "🇪🇺", "🇮🇳"}
IPHONE_17_PRO_2SIM_FLAGS = {"🇨🇳"}
IPHONE_17E_GLOBAL_FLAGS = IPHONE_17_PRO_GLOBAL_FLAGS | IPHONE_17_PRO_2SIM_FLAGS | {"🇲🇾"}
IPHONE_14_16_ESIM_FLAGS = {"🇺🇸"}
IPHONE_14_16_2SIM_FLAGS = {"🇭🇰", "🇲🇴", "🇨🇳"}
IPHONE_14_16_GLOBAL_FLAGS = {
    "🇦🇪", "🇯🇵", "🇫🇷", "🇩🇪", "🇬🇧", "🇮🇹", "🇪🇸", "🇸🇪",
    "🇳🇱", "🇵🇱", "🇦🇺", "🇳🇿", "🇰🇷", "🇸🇬", "🇪🇺", "🇮🇳",
}
IPHONE_REGION_CODE_GROUPS = {
    "esim_17": {
        "AA", "AE", "AH", "C", "CA", "CL", "E", "J", "LL",
    },
    "global": {
        "A", "AB", "B", "BT", "DN", "FD", "HN", "IN", "IP", "KS", "PO", "PP",
        "T", "TU", "X", "Y", "ZA", "ZD", "ZP", "EU",
    },
    "china": {"CH", "CN", "ZP"},
}
IPHONE_17_PRO_ESIM_MARKERS = (
    "бахрейн", "bahrain", "канада", "canada", "гуам", "guam", "япония", "japan",
    "кувейт", "kuwait", "мексика", "mexico", "оман", "oman", "катар", "qatar",
    "саудовская аравия", "саудовская", "saudi", "оаэ", "эмираты", "арабы",
    "uae", "united arab emirates", "сша", "америка", "usa", "united states",
    "us version", "virgin islands", "виргинские",
)
IPHONE_17_PRO_GLOBAL_MARKERS = (
    "гонконг", "hong kong", "hongkong", "hk", "франция", "france", "германия",
    "germany", "великобритания", "британия", "uk", "италия", "italy", "испания",
    "spain", "швеция", "sweden", "нидерланды", "netherlands", "польша", "poland",
    "австралия", "australia", "новая зеландия", "new zealand", "южная корея",
    "корея", "south korea", "singapore", "сингапур", "европа", "europe", "eu",
    "индия", "india", "макао", "macau", "macao", "глобальная", "global", "global version",
)
IPHONE_2SIM_COUNTRY_MARKERS = (
    "китай", "материковый китай", "china mainland", "mainland china", "cn",
)
IPHONE_17E_GLOBAL_MARKERS = (
    *IPHONE_17_PRO_GLOBAL_MARKERS,
    *IPHONE_2SIM_COUNTRY_MARKERS,
    "малайзия",
    "malaysia",
)
IPHONE_14_16_2SIM_MARKERS = (
    *IPHONE_2SIM_COUNTRY_MARKERS,
    "гонконг", "hong kong", "hongkong", "hk", "макао", "macau", "macao",
)
IPHONE_14_16_GLOBAL_MARKERS = (
    "оаэ", "эмираты", "арабы", "uae", "united arab emirates", "global",
    "глобальная", "япония", "japan", "sim esim", "1sim esim",
)
IPHONE_REGION_CODE_RE = re.compile(r"\b(M[A-Z0-9]{4})([A-Z]{1,3})/A\b", re.IGNORECASE)
IPHONE_SIM_VARIANT_LABELS = {
    "esim": "eSim Only",
    "sim+esim": "Глобальная версия",
    "2sim": "Китай",
}
BRAND_FALLBACK_TOKEN_STOPWORDS = {
    "cm",
    "wi-fi",
    "usb",
    "vga",
    "hdmi",
    "displayport",
    "type-c",
    "lightning",
    "digital",
    "silicone",
    "universal",
    "mini",
}
ACCESSORY_COUNTRY_FALLBACK_MARKERS = (
    "заряд",
    "charger",
    "адаптер",
    "кабель",
    "cable",
    "чех",
    "case",
    "cover",
    "держатель",
    "airpods",
    "науш",
    "гарнитур",
    "колонк",
    "speaker",
    "power bank",
    "powerbank",
    "флеш",
    "карта памяти",
    "ssd",
    "airtag",
    "hub",
    "dock",
    "док-станц",
    "ремеш",
    "strap",
)
BRAND_FALLBACK_PREFIX_RE = re.compile(
    r"^(?:автомобильное зарядное устройство|сетевое зарядное устройство|беспроводное зарядное устройство|"
    r"автомобильный держатель(?: магнитный)?|адаптер-разветвитель для компьютера|адаптер-переходник|"
    r"внешний аккумулятор|дата-кабель|аудиокабель|кабель|чехол(?:-книжка)?|usb[- ]флешка|usb[- ]фонарик|"
    r"детские умные часы|графический планшет|мышь беспроводная|стилус|ремешок на запястье для смартфона|"
    r"батарейка|патч-корд|док-станция|кардхолдер)\s+",
    re.IGNORECASE,
)


def brand_marker_in_text(text, marker):
    haystack = str(text or "").lower()
    needle = str(marker or "").lower()
    if not haystack or not needle:
        return False
    if re.fullmatch(r"[a-z0-9]+", needle):
        return bool(re.search(rf"(?<![a-z0-9]){re.escape(needle)}(?![a-z0-9])", haystack))
    return needle in haystack


def detect_brand_fallback_candidate(product_name):
    original = str(product_name or "").strip()
    normalized = BRAND_FALLBACK_PREFIX_RE.sub("", original, count=1).strip()
    if not normalized:
        normalized = original
    for token in re.findall(r"[A-Za-z][A-Za-z0-9+\-]{1,}", normalized):
        lowered = token.lower()
        if lowered in BRAND_FALLBACK_TOKEN_STOPWORDS:
            continue
        return token
    return ""


def detect_brand_from_name(product_name):
    explicit_brand = explicit_brand_from_product_name(product_name)
    if explicit_brand:
        return explicit_brand
    name = str(product_name or "").lower()
    if APPLE_PART_NUMBER_PATTERN.search(str(product_name or "")):
        return "Apple"
    for brand, markers in BRAND_KEYWORDS:
        if brand == "Apple":
            continue
        if any(brand_marker_in_text(name, marker) for marker in markers) and any(marker in name for marker in ACCESSORY_COMPATIBILITY_MARKERS):
            return brand
    apple_markers = (
        "iphone", "ipad", "macbook", "apple watch", "airpods", "imac", "mac mini", "vision pro"
    )
    if any(marker in name for marker in apple_markers):
        return "Apple"
    if looks_like_iphone_shorthand(product_name):
        return "Apple"
    for brand, markers in BRAND_KEYWORDS:
        if any(brand_marker_in_text(name, marker) for marker in markers):
            return brand
    fallback_brand = sanitize_brand_value(detect_brand_fallback_candidate(product_name))
    if fallback_brand:
        return fallback_brand
    return ""


def normalize_folder_display_name(folder_name):
    display_name = re.sub(r"\s+", " ", str(folder_name or "").strip())
    if not display_name:
        return ""
    return FOLDER_NAME_TRAILING_FILLER_RE.sub("", display_name).strip()


def is_generic_brand_value(value):
    normalized = re.sub(r"\s+", " ", str(value or "").strip()).casefold()
    if not normalized:
        return True
    compact = re.sub(r"[\s_\-]+", "", normalized)
    if normalized in GENERIC_BRAND_LABELS or compact in GENERIC_BRAND_LABELS or normalized.startswith("для "):
        return True
    if re.fullmatch(r"\d+(?:\s*(?:w|mm|gb|tb|гб|тб))?", normalized):
        return True
    if re.fullmatch(r"s\d{2}\+?", normalized):
        return True
    if re.fullmatch(r"(?:ps|ps4|ps5|nsw|nsw2|usb\s*c|type\s*c)", normalized):
        return True
    return False


def sanitize_brand_value(value):
    brand = re.sub(r"\s+", " ", str(value or "").strip())
    return "" if is_generic_brand_value(brand) else brand


PLATFORM_COMPATIBILITY_BRANDS = {
    "PlayStation",
    "PS",
    "PS4",
    "PS5",
    "Nintendo Switch",
    "NSW",
    "NSW2",
    "Switch",
    "Steamdeck",
}


def explicit_brand_from_product_name(product_name, article=""):
    raw = re.sub(r"\s+", " ", str(product_name or "").replace("\xa0", " ").strip())
    lower = raw.lower()
    if not raw:
        return ""
    if APPLE_PART_NUMBER_PATTERN.search(str(article or "")) or APPLE_PART_NUMBER_PATTERN.search(raw):
        return "Apple"

    explicit_patterns = (
        ("Nacon", r"\bnacon\b|\brevolution\s*5\s*pro\b"),
        ("Victrix", r"\bvictrix\b"),
        ("Razer", r"\brazer\b"),
        ("Logitech", r"\blogitech\b"),
        ("Backbone", r"\bbackbone\b"),
        ("MSI", r"\bmsi\s+claw\b"),
        ("Asus", r"\b(?:asus\s+)?rog\s+ally\b|\bxbox\s+ally\b"),
        ("Meta", r"\bmeta\s+(?:quest\s*)?\d+\b|\boculus\s+quest\b"),
        ("Valve", r"\bsteam\s*deck\b|\bsteamdeck\b"),
        ("Samsung", r"\bgalaxy\s+tab\b|\btab\s+a8\s+book\s+cover\b|\bgalaxy\s+s\d{1,2}\b|\bs\d{2}(?:\s+ultra|\+)?\s+(?:smart|clear|silicone|wallet|case|gadget)\b"),
        ("Insta360", r"\binsta\s*360\b|\binsta360\b|\butility\s+frame\s+for\s+x4\b"),
        ("Garmin", r"\brussian\s+map\s+sd\s+card\b"),
        ("Pop Mart", r"\bpop\s*mart\b|\blabubu\b|\bthe\s+monsters\b"),
        ("Syntech", r"\bsyntech\b"),
        ("Dyson", r"\broll\s+brush\s+\+\s+brush\b.*\bblack\s+rose\b"),
        ("Sony", r"\bsony\b|\bwh[-\s]?(?:1000)?xm\d+\b|\bdual[\s-]*sen[cs]e\b|\bdualsense\b|\bps\s*portal\b|\bpulse\s+elite\b|\binzone\b|\bplaystation\s+(?:portal|5|4)\b"),
        ("Nintendo", r"\bnintendo\b|\bswitch\s+(?:lite|oled|2)\b|\bnsw2?\b|\bjoy[-\s]?con\b"),
        ("Xbox", r"\bxbox\b"),
    )
    for brand, pattern in explicit_patterns:
        if re.search(pattern, lower, re.IGNORECASE):
            return brand

    apple_accessory_patterns = (
        r"\bmagic\s+(?:mouse|trackpad|keyboard)\b",
        r"\b(?:homepod|airtag|studio\s+display|smart\s+folio|crossbody\s+strap|iphone\s+air\s+bumper|digital\s+av\s+adapter)\b",
        r"\bfinewoven\b",
        r"\b(?:clear|silicone|silicon)\s+case\b.*\bmagsafe\b",
        r"\b\d{2}\s+pro(?:\s+max)?\s+(?:clear|silicone|silicon)\s+case\b",
        r"\busb[-\s]?c\s+to\s+lightning\s+adapter\b",
        r"\b(?:usb[-\s]?c\s+to\s+magsafe|magsafe\s+3)\b",
        r"\b(?:woven\s+cable|power\s+adapter|watch\s+charger)\b.*\bM[A-Z0-9]{4,6}\b",
        r"\b(?:\d{2}\s*mm)\b.*\b(?:modern\s+buckle|magnetic\s+link|milanese|link\s+bracelet|sport\s+band)\b",
        r"\biphone\s+\d{2}\b",
        r"\bimac\b",
    )
    if any(re.search(pattern, lower, re.IGNORECASE) for pattern in apple_accessory_patterns):
        return "Apple"

    if re.search(r"\bps[45]\s+", lower):
        return "PlayStation"
    if re.search(r"\bnsw2?\s+", lower):
        return "Nintendo"
    return ""


def resolve_product_brand_value(product_name, article="", candidate_brand=""):
    candidate = sanitize_brand_value(candidate_brand)
    explicit_brand = explicit_brand_from_product_name(product_name, article)
    if explicit_brand:
        return explicit_brand
    detected_brand = sanitize_brand_value(detect_brand_from_name(product_name))
    if detected_brand and (
        not candidate
        or candidate in PLATFORM_COMPATIBILITY_BRANDS
        or is_generic_brand_value(candidate)
    ):
        return detected_brand
    return candidate or detected_brand


def infer_brand_country_fallback(brand):
    return BRAND_COUNTRY_FALLBACKS.get(str(brand or "").strip(), "")


def infer_accessory_country_fallback(product_name, brand=""):
    brand_country = infer_brand_country_fallback(brand)
    if brand_country:
        return brand_country
    normalized = str(product_name or "").lower()
    if any(marker in normalized for marker in ACCESSORY_COUNTRY_FALLBACK_MARKERS):
        return "Китай"
    return ""


def infer_handheld_processor(product_name):
    lower = str(product_name or "").lower()
    if "z2 extreme" in lower:
        return "AMD Ryzen Z2 Extreme"
    if "z1 extreme" in lower:
        return "AMD Ryzen Z1 Extreme"
    if "z2 go" in lower:
        return "AMD Ryzen Z2 Go"
    return ""


def build_known_product_description(product_name, kind_title, highlights, extra_line=""):
    safe_name = re.sub(r"\s+", " ", str(product_name or "").strip())
    lines = [f"{safe_name} — {kind_title}, который хорошо подходит для повседневного использования."]
    if extra_line:
        lines.append(str(extra_line).strip().rstrip(".") + ".")
    if highlights:
        lines.append("Среди основных особенностей — " + ", ".join(highlights[:4]) + ".")
    return "\n\n".join(lines)


def supplement_autofill_payload_from_known_products(product_name, article="", payload=None):
    next_payload = dict(payload or {})
    raw_name = re.sub(r"\s+", " ", str(product_name or "").strip())
    lower = raw_name.lower()
    inferred_ram, inferred_storage = extract_ram_storage_from_name(raw_name)

    def fill_if_missing(key, value):
        value = str(value or "").strip()
        if value and not str(next_payload.get(key) or "").strip():
            next_payload[key] = value

    def fill_many(values):
        for key, value in (values or {}).items():
            fill_if_missing(key, value)

    if lower.startswith("игра для nintendo switch"):
        game_title = re.sub(r"^игра\s+для\s+nintendo\s+switch(?:\s+2)?\s*", "", raw_name, flags=re.IGNORECASE).strip()
        fill_many({
            "brand": "Nintendo",
            "country_sim": "Япония",
            "display": "Зависит от консоли",
            "processor": "Зависит от консоли",
            "camera": "Не применяется",
            "front_camera": "Не применяется",
            "video": "Зависит от режима вывода консоли",
            "connectivity": "Совместимость с Nintendo Switch",
            "battery": "Зависит от консоли",
            "os": "Nintendo Switch system software",
            "biometrics": "Не применяется",
            "charging": "Не применяется",
            "warranty": STORE_WARRANTY_DEFAULT,
        })
        if game_title and not str(next_payload.get("description") or "").strip():
            next_payload["description"] = build_known_product_description(
                raw_name,
                "игровое издание для Nintendo Switch",
                [
                    f"Платформа: {'Nintendo Switch 2' if 'switch 2' in lower else 'Nintendo Switch'}",
                    f"Название: {game_title}",
                    "Формат: игровой картридж или цифровая версия",
                ],
            )
        return next_payload

    if "nintendo switch" in lower and any(token in lower for token in ("joy-con", "controller", "powera", "wheel", "геймпад", "руль")):
        accessory_brand = "PowerA" if "powera" in lower else "Nintendo"
        fill_many({
            "brand": accessory_brand,
            "country_sim": infer_brand_country_fallback(accessory_brand),
            "display": "Не применяется",
            "processor": "Не применяется",
            "camera": "Не применяется",
            "front_camera": "Не применяется",
            "video": "Не применяется",
            "connectivity": "Совместимость с Nintendo Switch",
            "battery": "Зависит от модели аксессуара",
            "os": "Не применяется",
            "biometrics": "Не применяется",
            "charging": "Зависит от модели аксессуара",
            "warranty": STORE_WARRANTY_DEFAULT,
        })
        if not str(next_payload.get("description") or "").strip():
            next_payload["description"] = build_known_product_description(
                raw_name,
                "игровой аксессуар",
                [
                    f"Бренд: {accessory_brand}",
                    f"Совместимость: {'Nintendo Switch 2' if 'switch 2' in lower else 'Nintendo Switch'}",
                ],
            )
        return next_payload

    if "steam deck oled" in lower:
        fill_many({
            "brand": "Valve",
            "country_sim": infer_brand_country_fallback("Valve"),
            "weight": "0.640",
            "storage": inferred_storage or "512 GB",
            "ram": "16 GB",
            "display": '7.4" OLED HDR, 90 Гц',
            "processor": "AMD APU Zen 2 / RDNA 2",
            "camera": "Нет",
            "front_camera": "Нет",
            "video": "1280x800, HDR, 90 Гц",
            "connectivity": "Wi-Fi 6E, Bluetooth 5.3, USB-C",
            "battery": "50 Wh",
            "os": "SteamOS 3",
            "biometrics": "Нет",
            "charging": "USB-C PD 45 W",
            "warranty": STORE_WARRANTY_DEFAULT,
        })
        if not str(next_payload.get("description") or "").strip():
            next_payload["description"] = build_known_product_description(
                raw_name,
                "портативная игровая консоль",
                ["Дисплей: 7.4\" OLED HDR", "ОЗУ: 16 GB", f"Память: {inferred_storage or '512 GB'}", "ОС: SteamOS 3"],
            )
        return next_payload

    if "steam deck" in lower:
        fill_many({
            "brand": "Valve",
            "country_sim": infer_brand_country_fallback("Valve"),
            "weight": "0.669",
            "storage": inferred_storage or "256 GB",
            "ram": "16 GB",
            "display": '7.0" IPS, 60 Гц',
            "processor": "AMD APU Zen 2 / RDNA 2",
            "camera": "Нет",
            "front_camera": "Нет",
            "video": "1280x800, 60 Гц",
            "connectivity": "Wi-Fi 5, Bluetooth 5.0, USB-C",
            "battery": "40 Wh",
            "os": "SteamOS 3",
            "biometrics": "Нет",
            "charging": "USB-C PD 45 W",
            "warranty": STORE_WARRANTY_DEFAULT,
        })
        if not str(next_payload.get("description") or "").strip():
            next_payload["description"] = build_known_product_description(
                raw_name,
                "портативная игровая консоль",
                ["Дисплей: 7.0\" IPS", "ОЗУ: 16 GB", f"Память: {inferred_storage or '256 GB'}", "ОС: SteamOS 3"],
            )
        return next_payload

    if "nintendo switch 2" in lower:
        fill_many({
            "brand": "Nintendo",
            "country_sim": "Япония",
            "weight": "0.534",
            "storage": inferred_storage or "256 GB",
            "display": '7.9" LCD, 120 Гц',
            "processor": "NVIDIA custom processor",
            "camera": "Нет",
            "front_camera": "Нет",
            "video": "До 4K в ТВ-режиме / до 1080p в портативном режиме",
            "connectivity": "Wi-Fi 6, Bluetooth, USB-C, NFC",
            "battery": "5220 мАч",
            "os": "Nintendo Switch system software",
            "biometrics": "Нет",
            "charging": "USB-C",
            "warranty": STORE_WARRANTY_DEFAULT,
        })
        if not str(next_payload.get("description") or "").strip():
            next_payload["description"] = build_known_product_description(
                raw_name,
                "портативная игровая консоль",
                ["Дисплей: 7.9\" LCD 120 Гц", f"Память: {inferred_storage or '256 GB'}", "Вывод изображения: до 4K в док-режиме"],
            )
        return next_payload

    if "nintendo switch oled" in lower:
        fill_many({
            "brand": "Nintendo",
            "country_sim": "Япония",
            "weight": "0.420",
            "storage": inferred_storage or "64 GB",
            "display": '7.0" OLED',
            "processor": "NVIDIA custom processor",
            "camera": "Нет",
            "front_camera": "Нет",
            "video": "До 1080p в ТВ-режиме / 720p в портативном режиме",
            "connectivity": "Wi-Fi, Bluetooth 4.1, USB-C, NFC",
            "battery": "До 9 часов работы",
            "os": "Nintendo Switch system software",
            "biometrics": "Нет",
            "charging": "USB-C",
            "warranty": STORE_WARRANTY_DEFAULT,
        })
        if not str(next_payload.get("description") or "").strip():
            next_payload["description"] = build_known_product_description(
                raw_name,
                "портативная игровая консоль",
                ["Дисплей: 7.0\" OLED", f"Память: {inferred_storage or '64 GB'}", "Вывод изображения: до Full HD в ТВ-режиме"],
            )
        return next_payload

    if "nintendo switch lite" in lower:
        fill_many({
            "brand": "Nintendo",
            "country_sim": "Япония",
            "weight": "0.275",
            "storage": inferred_storage or "32 GB",
            "display": '5.5" LCD',
            "processor": "NVIDIA custom processor",
            "camera": "Нет",
            "front_camera": "Нет",
            "video": "До 720p на встроенном экране",
            "connectivity": "Wi-Fi, Bluetooth, USB-C",
            "battery": "До 7 часов работы",
            "os": "Nintendo Switch system software",
            "biometrics": "Нет",
            "charging": "USB-C",
            "warranty": STORE_WARRANTY_DEFAULT,
        })
        if not str(next_payload.get("description") or "").strip():
            next_payload["description"] = build_known_product_description(
                raw_name,
                "портативная игровая консоль",
                ["Дисплей: 5.5\" LCD", f"Память: {inferred_storage or '32 GB'}", "Формат: компактная портативная версия"],
            )
        return next_payload

    if re.search(r"\bnintendo switch\b", lower):
        fill_many({
            "brand": "Nintendo",
            "country_sim": "Япония",
            "weight": "0.398",
            "storage": inferred_storage or "32 GB",
            "display": '6.2" LCD',
            "processor": "NVIDIA custom processor",
            "camera": "Нет",
            "front_camera": "Нет",
            "video": "До 1080p в ТВ-режиме / 720p в портативном режиме",
            "connectivity": "Wi-Fi, Bluetooth 4.1, USB-C, NFC",
            "battery": "До 9 часов работы",
            "os": "Nintendo Switch system software",
            "biometrics": "Нет",
            "charging": "USB-C",
            "warranty": STORE_WARRANTY_DEFAULT,
        })
        if not str(next_payload.get("description") or "").strip():
            next_payload["description"] = build_known_product_description(
                raw_name,
                "портативная игровая консоль",
                ["Дисплей: 6.2\" LCD", f"Память: {inferred_storage or '32 GB'}", "Вывод изображения: до Full HD в ТВ-режиме"],
            )
        return next_payload

    if any(marker in lower for marker in ("rog ally", "legion go", "xbox ally")):
        handheld_brand = sanitize_brand_value(next_payload.get("brand")) or detect_brand_from_inputs(raw_name, article)
        processor = infer_handheld_processor(raw_name)
        fill_many({
            "brand": handheld_brand,
            "country_sim": infer_brand_country_fallback(handheld_brand),
            "storage": inferred_storage,
            "ram": inferred_ram,
            "processor": processor,
            "camera": "Нет",
            "front_camera": "Нет",
            "os": "SteamOS" if "steamos" in lower else "Windows 11",
            "biometrics": "Нет",
            "charging": "USB-C PD",
            "warranty": STORE_WARRANTY_DEFAULT,
        })
        if not str(next_payload.get("description") or "").strip():
            highlights = []
            if processor:
                highlights.append(f"Процессор: {processor}")
            if inferred_ram:
                highlights.append(f"ОЗУ: {inferred_ram}")
            if inferred_storage:
                highlights.append(f"Память: {inferred_storage}")
            next_payload["description"] = build_known_product_description(
                raw_name,
                "портативная игровая консоль",
                highlights,
                "Формат: компактная игровая система для портативного использования.",
            )
        return next_payload

    return next_payload


def has_apple_part_number(product_name):
    raw = str(product_name or "")
    return bool(APPLE_PART_NUMBER_PATTERN.search(raw) or IPHONE_REGION_CODE_RE.search(raw))


def extract_apple_model_no(value):
    region_match = IPHONE_REGION_CODE_RE.search(str(value or ""))
    if region_match:
        return region_match.group(1).upper()
    match = APPLE_PART_NUMBER_PATTERN.search(str(value or ""))
    if not match:
        return ""
    return match.group(0).split("/")[0].upper()


def extract_apple_model_tokens(value):
    tokens = set()
    raw = str(value or "")
    for match in IPHONE_REGION_CODE_RE.finditer(raw):
        tokens.add(match.group(1).upper())
    for match in APPLE_PART_NUMBER_PATTERN.finditer(raw):
        token = match.group(0).split("/")[0].upper()
        if not token:
            continue
        tokens.add(token)
        if len(token) > 5 and re.fullmatch(r"[A-Z]{1,3}", token[5:]):
            tokens.add(token[:5])
    return tokens


GENERIC_MODEL_TOKEN_PATTERN = re.compile(
    r"(?<![A-Z0-9])(?=[A-Z0-9][A-Z0-9\-/]{4,19}(?![A-Z0-9]))"
    r"(?=[A-Z0-9\-/]*[A-Z])(?=[A-Z0-9\-/]*\d)[A-Z0-9][A-Z0-9\-/]{4,19}(?![A-Z0-9])",
    re.IGNORECASE,
)
SHORT_MODEL_TOKEN_PATTERN = re.compile(
    r"(?<![A-Z0-9])([A-Z]{1,3})[-\s]?(\d{2,4}[A-Z]?)(?![A-Z0-9])",
    re.IGNORECASE,
)
SHORT_MODEL_TOKEN_RE = re.compile(
    r"(?:HT|HS|HD|SV|RB|PH|DS)\d{2,3}[A-Z]?|[RXL]\d{3,4}[A-Z]?",
    re.IGNORECASE,
)
SAMSUNG_MODEL_TOKEN_RE = re.compile(r"^SM([A-Z]\d{3,4}[A-Z]?)$", re.IGNORECASE)
GENERIC_MODEL_TOKEN_STOPWORDS = {
    "IPHONE", "MACBOOK", "APPLE", "SILVER", "BLACK", "WHITE", "BLUE", "GREEN", "PINK",
    "ORANGE", "YELLOW", "PURPLE", "MIDNIGHT", "STARLIGHT", "NATURAL", "DESERT",
    "SPACE", "GRAY", "GREY", "GOLD", "WIFI", "CELLULAR", "GLOBAL", "VERSION",
}


def normalize_catalog_model_token(value):
    token = re.sub(r"[^A-Z0-9]", "", str(value or "").upper())
    if len(token) > 20:
        return ""
    if len(token) < 5 and not SHORT_MODEL_TOKEN_RE.fullmatch(token):
        return ""
    if token in GENERIC_MODEL_TOKEN_STOPWORDS:
        return ""
    if not re.search(r"[A-Z]", token) or not re.search(r"\d", token):
        return ""
    if re.fullmatch(r"\d+(?:GB|TB|ГБ|ТБ)", token):
        return ""
    return token


def add_catalog_model_token_variants(tokens, token):
    token = normalize_catalog_model_token(token)
    if not token:
        return
    tokens.add(token)
    samsung_match = SAMSUNG_MODEL_TOKEN_RE.fullmatch(token)
    if samsung_match:
        suffix = samsung_match.group(1).upper()
        tokens.add(suffix)
        if re.fullmatch(r"[A-Z]\d{3,4}[A-Z]", suffix):
            tokens.add(suffix[:-1])
    if token.startswith("M") and len(token) > 5 and re.fullmatch(r"[A-Z]{1,3}", token[5:]):
        tokens.add(token[:5])


def extract_catalog_model_tokens(*values):
    tokens = set()
    for value in values:
        raw = str(value or "")
        tokens.update(extract_apple_model_tokens(raw))
        for match in GENERIC_MODEL_TOKEN_PATTERN.finditer(raw):
            add_catalog_model_token_variants(tokens, match.group(0))
        for match in SHORT_MODEL_TOKEN_PATTERN.finditer(raw):
            candidate = f"{match.group(1)}{match.group(2)}"
            if SHORT_MODEL_TOKEN_RE.fullmatch(candidate):
                add_catalog_model_token_variants(tokens, candidate)
    return tokens


def extract_exact_catalog_article_tokens(*values):
    tokens = set()
    for value in values:
        raw = str(value or "").replace("\xa0", " ")
        raw = re.sub(
            r"[-–—]\s*\d{2,7}(?:[.,]\d{1,2})?\s*(?:₽|руб\.?|р\.?|p\.?|rub|rur|usd|\$|€|к|k)?\s*$",
            " ",
            raw,
            flags=re.IGNORECASE,
        )
        raw = re.sub(r"\b(?:USB|TYPE)[-\s]*C(?=[A-Z0-9]{4,})", " ", raw, flags=re.IGNORECASE)
        for match in IPHONE_REGION_CODE_RE.finditer(raw):
            token = normalize_catalog_model_token(match.group(1))
            if token:
                tokens.add(token)
        for match in APPLE_PART_NUMBER_PATTERN.finditer(raw):
            token = normalize_catalog_model_token(match.group(0))
            if token:
                tokens.add(token)
        for match in GENERIC_MODEL_TOKEN_PATTERN.finditer(raw):
            token = normalize_catalog_model_token(match.group(0))
            if token:
                tokens.add(token)
        for match in SHORT_MODEL_TOKEN_PATTERN.finditer(raw):
            candidate = f"{match.group(1)}{match.group(2)}"
            if SHORT_MODEL_TOKEN_RE.fullmatch(candidate):
                token = normalize_catalog_model_token(candidate)
                if token:
                    tokens.add(token)
    spec_like_re = re.compile(
        r"^(?:\d+(?:GB|TB|MM|MP|CPU|GPU|MAH|HZ|W|WH|NM|CM|IN|CORE|CORES))+$",
        re.IGNORECASE,
    )
    filtered = {token for token in tokens if not spec_like_re.fullmatch(token)}
    for token in list(tokens):
        if token not in filtered:
            continue
        if any(other != token and other in filtered and other.startswith(token) for other in tokens):
            filtered.discard(token)
    return filtered


def apple_model_no_in_text(model_no, text):
    model_tokens = extract_catalog_model_tokens(model_no) or {normalize_model_no(model_no, "Apple")}
    return bool({token for token in model_tokens if token} & extract_catalog_model_tokens(text))


def extract_apple_region_code(value):
    match = IPHONE_REGION_CODE_RE.search(str(value or ""))
    if not match:
        return ""
    return match.group(2).upper()


def iphone_generation_number(product_name):
    family = detect_iphone_model_family(product_name)
    match = re.match(r"^(\d+)", family or "")
    if not match:
        return None
    try:
        return int(match.group(1))
    except (TypeError, ValueError):
        return None


def contains_any_marker(text, markers):
    normalized = normalize_product_search_text(text)
    haystack = f" {normalized} "
    for marker in markers:
        normalized_marker = normalize_product_search_text(marker)
        if normalized_marker and f" {normalized_marker} " in haystack:
            return True
    return False


def normalize_iphone_sim_variant_key(value):
    text = str(value or "")
    compact = re.sub(r"[\s_\-()]+", "", text.lower())
    normalized = normalize_product_search_text(text)

    if (
        "1sim+esim" in compact
        or "sim+esim" in compact
        or "nanosim+esim" in compact
        or "1simesim" in compact
        or (("1sim" in compact or re.search(r"\b(?:1|one)\s*(?:sim|сим)\b", normalized)) and "2sim" not in compact and "esim" not in compact)
        or re.search(r"\b(?:1|one)\s*(?:sim|сим)\s*(?:and|и|plus|плюс)?\s*e\s*sim\b", normalized)
        or re.search(r"\bnano\s*sim\s*(?:and|и|plus|плюс)?\s*e\s*sim\b", normalized)
        or re.search(r"\bphysical\s*sim\s*(?:and|и|plus|плюс)?\s*e\s*sim\b", normalized)
        or "nano sim esim" in normalized
        or "physical sim esim" in normalized
        or "физическая sim esim" in normalized
    ):
        return "sim+esim"
    if (
        "2sim" in compact
        or re.search(r"\b2\s*(?:sim|сим)\b", normalized)
        or (re.search(r"\bdual\s*nano\s*sim\b", normalized) and "esim" not in compact)
        or (re.search(r"\b(?:two|dual)\s+(?:physical\s+)?nano\s*sim(?:\s*cards?)?\b", normalized) and "esim" not in compact)
        or (re.search(r"\b(?:two|dual)\s+physical\s+sim(?:\s*cards?)?\b", normalized) and "esim" not in compact)
        or (re.search(r"\btwo\s+sim(?:\s*cards?)?\b", normalized) and "esim" not in compact)
        or "две физ" in normalized
        or "два физических" in normalized
    ):
        return "2sim"
    if (
        "esimonly" in compact
        or re.search(r"\b(?:e\s*sim|esim)\b", normalized)
        or "только esim" in normalized
        or "без физ" in normalized
        or "без слота" in normalized
    ):
        return "esim"
    return ""


def iphone_region_code_variant_key(product_name, generation):
    region_code = extract_apple_region_code(product_name)
    if not region_code:
        return ""
    if region_code in IPHONE_REGION_CODE_GROUPS["china"]:
        if generation and generation >= 17 and region_code == "ZP":
            return "sim+esim"
        return "2sim"
    if generation and generation >= 17 and region_code in IPHONE_REGION_CODE_GROUPS["esim_17"]:
        return "esim"
    if generation and 14 <= generation <= 16 and region_code == "LL":
        return "esim"
    if region_code in IPHONE_REGION_CODE_GROUPS["global"]:
        return "sim+esim"
    return ""


def infer_iphone_variant_key(product_name):
    name = str(product_name or "")
    explicit_variant = normalize_iphone_sim_variant_key(name)
    if explicit_variant:
        return explicit_variant

    generation = iphone_generation_number(name)
    if not generation:
        return ""
    family = detect_iphone_model_family(name)

    if family == "17 Air":
        return "esim"

    region_code_variant = iphone_region_code_variant_key(name, generation)
    if region_code_variant:
        return region_code_variant

    flags = set(extract_country_flags(name))
    if family == "17e":
        if flags & IPHONE_17E_GLOBAL_FLAGS or contains_any_marker(name, IPHONE_17E_GLOBAL_MARKERS):
            return "sim+esim"
        if flags & IPHONE_17_PRO_ESIM_FLAGS or contains_any_marker(name, IPHONE_17_PRO_ESIM_MARKERS):
            return "esim"
        return ""

    if generation >= 17 and family in {"17", "17 Pro", "17 Pro Max"}:
        if flags & IPHONE_17_PRO_2SIM_FLAGS or contains_any_marker(name, IPHONE_2SIM_COUNTRY_MARKERS):
            return "2sim"
        if flags & IPHONE_17_PRO_ESIM_FLAGS or contains_any_marker(name, IPHONE_17_PRO_ESIM_MARKERS):
            return "esim"
        if flags & IPHONE_17_PRO_GLOBAL_FLAGS or contains_any_marker(name, IPHONE_17_PRO_GLOBAL_MARKERS):
            return "sim+esim"
        return ""

    if 14 <= generation <= 16:
        if flags & IPHONE_14_16_2SIM_FLAGS or contains_any_marker(name, IPHONE_14_16_2SIM_MARKERS):
            return "2sim"
        if flags & IPHONE_14_16_ESIM_FLAGS or contains_any_marker(name, ("сша", "америка", "usa", "united states", "us version")):
            return "esim"
        if flags & IPHONE_14_16_GLOBAL_FLAGS or contains_any_marker(name, (*IPHONE_14_16_GLOBAL_MARKERS, *IPHONE_17_PRO_GLOBAL_MARKERS)):
            return "sim+esim"

    return ""


def iphone_variant_label_from_key(variant_key):
    return IPHONE_SIM_VARIANT_LABELS.get(str(variant_key or "").strip().lower(), "")


def is_iphone_variant_sensitive(product_name):
    generation = iphone_generation_number(product_name)
    return bool(generation and generation >= 14)


def product_message_variants_compatible(product_name, product_country, line_text, product_condition_text=""):
    if not product_condition_markers_compatible(product_name, line_text, product_condition_text=product_condition_text):
        return False

    if not is_apple_iphone_product(product_name, "Apple"):
        return True

    product_family = detect_iphone_model_family(product_name)
    line_family = detect_iphone_model_family(strip_supplier_line_noise(line_text))
    if product_family and line_family and product_family != line_family:
        return False

    if not is_iphone_variant_sensitive(product_name):
        return True

    product_key = infer_iphone_variant_key(f"{product_name} {product_country or ''}")
    line_key = infer_iphone_variant_key(line_text)
    if not product_key:
        return True
    return line_key == product_key


def detect_brand_from_inputs(product_name, article=""):
    if has_apple_part_number(article) or has_apple_part_number(product_name):
        return "Apple"
    return detect_brand_from_name(product_name)


def looks_like_iphone_shorthand(product_name):
    normalized = str(product_name or "").strip().lower()
    if re.search(r"\b(?:iphone|айфон)\s+(?:se(?:\s+2022)?|1[1-9][a-zа-я]?|2[0-9][a-zа-я]?)\b", normalized):
        return True
    if re.match(r"^(?:📱\s*)?se(?:\s+2022)?\b", normalized):
        return True
    return bool(re.match(r"^(?:📱\s*)?(?:\d{2}|\d{2}[a-z]?)(?:\s+(?:pro max|pro|plus|mini|air|e))?\b", normalized))


def detect_iphone_model_family(product_name):
    normalized = str(product_name or "").strip()
    normalized = SIM_INLINE_VARIANT_PATTERN.sub("", normalized)
    normalized = normalized.lower().replace("ё", "е")
    normalized = re.sub(r"\b(?:смартфон|apple|iphone|айфон)\b", " ", normalized)
    normalized = re.sub(r"^[^a-zа-я0-9]+", "", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    for pattern, family in IPHONE_MODEL_PATTERNS:
        if re.search(pattern, normalized):
            return family
    return ""


def is_apple_iphone_product(product_name, brand=""):
    normalized_brand = str(brand or "").strip()
    if normalized_brand and normalized_brand != "Apple":
        return False
    if not looks_like_iphone_shorthand(product_name):
        return False
    return normalized_brand == "Apple" or detect_brand_from_name(product_name) == "Apple"


def infer_iphone_accessory_catalog_path(product_name, brand=""):
    normalized_name = re.sub(r"\s+", " ", str(product_name or "").replace("ё", "е").lower()).strip()
    if not normalized_name:
        return None
    if "iphone" not in normalized_name and "айфон" not in normalized_name:
        return None

    case_markers = (
        "чех",
        "клип-кейс",
        "наклад",
        "бампер",
        "case",
        "cover",
        "wallet",
        "folio",
    )
    protector_markers = (
        "стекл",
        "пленк",
        "защит",
        "glass",
        "protector",
        "privacy",
        "camera lens",
        "линз",
        "screen guard",
    )

    if any(marker in normalized_name for marker in case_markers):
        return IPHONE_CASES_CATALOG_PATH
    if any(marker in normalized_name for marker in protector_markers):
        return IPHONE_SCREEN_PROTECTOR_CATALOG_PATH
    return None


def infer_iphone_weight_from_name(product_name):
    family = detect_iphone_model_family(product_name)
    return normalize_autofill_weight(IPHONE_MODEL_WEIGHT_MAP.get(family, ""))


def normalize_iphone_os_value(value, product_name, brand=""):
    raw = str(value or "").strip()
    if not is_apple_iphone_product(product_name, brand):
        return raw
    if not raw or raw.lower() == "ios":
        return DEFAULT_IPHONE_OS_VERSION
    return raw


def detect_condition_marker(product_name):
    markers = detect_condition_markers(product_name)
    if "ASIS" in markers:
        return "ASIS"
    if "CPO" in markers:
        return "CPO"
    if "DEFECT" in markers:
        return "DEFECT"
    if "REFURB" in markers:
        return "REFURB"
    return ""


CONDITION_NEGATION_RE = re.compile(
    r"\b(?:без|нет|no|without)\s+"
    r"(?:дефект\w*|брака|брак\w*|царап\w*|скол\w*|трещин\w*|"
    r"defects?|defective|damage|damages)\b",
    re.IGNORECASE,
)


def normalize_condition_text(value):
    text = normalize_match_text(value)
    return CONDITION_NEGATION_RE.sub(" ", text)


PRODUCT_CONDITION_MARKER_PATTERNS = (
    ("ASIS", r"\b(?:as[\s-]*is|asis)\b"),
    ("CPO", r"\b(?:cpo|сро)\b"),
    ("REFURB", r"\b(?:refurb(?:ished)?|pre[\s-]*owned|open[\s-]*box|витрин\w*)\b"),
    ("DEFECT", (
        r"\b(?:"
        r"у\s*ц\s*е\s*н\s*к\w*|уценк\w*|markdown|clearance|"
        r"дефект\w*|defect(?:ive)?|брак\w*|"
        r"предактив\w*|актив\b|активк\w*|активирован\w*|активаци\w*|"
        r"мят\w*\s+короб\w*|мят\w*\s+уг\w*|"
        r"отсутств\w+\s+(?:кабел\w*|насад\w*|комплект\w*|короб\w*)|"
        r"неисправ\w*|нерабоч\w*"
        r")\b"
    )),
)


def detect_condition_markers(value):
    normalized = normalize_condition_text(value)
    markers = set()
    for marker, pattern in PRODUCT_CONDITION_MARKER_PATTERNS:
        if re.search(pattern, normalized, flags=re.IGNORECASE):
            markers.add(marker)
    return markers


PRODUCT_CONDITION_FIELDS = (
    "name",
    "warranty",
    "description",
    "description_html",
    "specs",
)


def build_product_condition_text(product, extra_text=""):
    if isinstance(product, str):
        values = [product, extra_text]
    else:
        values = [row_value(product, key, "") for key in PRODUCT_CONDITION_FIELDS]
        if extra_text:
            values.append(extra_text)
    return " ".join(str(value or "") for value in values if str(value or "").strip())


def product_condition_markers_compatible(product_name, line_text, product_condition_text=""):
    product_markers = detect_condition_markers(build_product_condition_text(product_name, product_condition_text))
    line_markers = detect_condition_markers(line_text)
    return product_markers == line_markers


def extract_country_flags(product_name):
    text = str(product_name or "")
    known_flags = (
        IPHONE_17_PRO_ESIM_FLAGS
        | IPHONE_17_PRO_GLOBAL_FLAGS
        | IPHONE_17_PRO_2SIM_FLAGS
        | IPHONE_14_16_ESIM_FLAGS
        | IPHONE_14_16_2SIM_FLAGS
        | {"🇮🇳", "🇷🇺", "🇲🇾", "🇮🇩", "🇰🇿", "🇨🇦", "🇯🇵", "🇦🇪", "🇪🇺", "🇬🇧", "🇦🇺", "🇰🇷"}
    )
    return [flag for flag in known_flags if flag in text]


def infer_iphone_variant_from_flags(product_name):
    return iphone_variant_label_from_key(infer_iphone_variant_key(product_name))


def detect_variant_folder_name(product_name):
    return normalize_iphone_sim_variant_key(product_name)


def detect_country_variant(product_name):
    return iphone_variant_label_from_key(infer_iphone_variant_key(product_name))


def normalize_folder_id(value):
    if value in (None, "", "null", "None"):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def is_variant_folder_name(folder_name):
    return str(folder_name or "").strip().lower() in VARIANT_FOLDER_NAMES


def is_storage_folder_name(folder_name):
    normalized_name = re.sub(r"\s+", "", str(folder_name or "").strip())
    return bool(STORAGE_FOLDER_NAME_RE.fullmatch(normalized_name))


def get_user_folder(db, user_id, folder_id):
    normalized_folder_id = normalize_folder_id(folder_id)
    if not normalized_folder_id:
        return None
    return db.execute(
        "SELECT id, name, parent_id FROM folders WHERE id = ? AND user_id = ?",
        (normalized_folder_id, user_id),
    ).fetchone()


def normalize_folder_lookup_name(folder_name):
    return normalize_folder_display_name(folder_name).casefold()


def find_folder_by_name_and_parent(db, user_id, folder_name, parent_id):
    normalized_name = normalize_folder_lookup_name(folder_name)
    if not normalized_name:
        return None
    normalized_parent_id = normalize_folder_id(parent_id)
    if normalized_parent_id is None:
        rows = db.execute(
            "SELECT id, name, parent_id FROM folders WHERE user_id = ? AND parent_id IS NULL ORDER BY id",
            (user_id,),
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT id, name, parent_id FROM folders WHERE user_id = ? AND parent_id = ? ORDER BY id",
            (user_id, normalized_parent_id),
        ).fetchall()
    for row in rows:
        if normalize_folder_lookup_name(row["name"]) == normalized_name:
            return row
    return None


def ensure_named_folder(db, user_id, folder_name, parent_id):
    display_name = normalize_folder_display_name(folder_name)
    normalized_name = normalize_folder_lookup_name(display_name)
    if not normalized_name:
        return normalize_folder_id(parent_id)
    existing = find_folder_by_name_and_parent(db, user_id, display_name, parent_id)
    if existing:
        return int(existing["id"])
    cursor = db.execute(
        "INSERT INTO folders (user_id, name, parent_id) VALUES (?, ?, ?)",
        (user_id, display_name, normalize_folder_id(parent_id)),
    )
    return int(cursor.lastrowid)


def get_folder_path_names(db, user_id, folder_id):
    normalized_folder_id = normalize_folder_id(folder_id)
    visited = set()
    path = []
    while normalized_folder_id and normalized_folder_id not in visited:
        visited.add(normalized_folder_id)
        folder = get_user_folder(db, user_id, normalized_folder_id)
        if not folder:
            break
        path.append(normalize_folder_display_name(folder["name"]))
        normalized_folder_id = normalize_folder_id(folder["parent_id"])
    return list(reversed(path))


def folder_path_startswith(db, user_id, folder_id, expected_path):
    actual_path = [part.lower() for part in get_folder_path_names(db, user_id, folder_id)]
    expected = [str(part or "").strip().lower() for part in expected_path if str(part or "").strip()]
    if len(actual_path) < len(expected):
        return False
    return actual_path[:len(expected)] == expected


def detect_storage_folder_name(product_name="", storage_value=""):
    normalized_storage = normalize_capacity(storage_value, allow_tb=True)
    if normalized_storage:
        return normalized_storage.replace(" ", "")

    text = str(product_name or "")
    explicit_match = re.search(r"\b(\d+(?:\.\d+)?)\s*(TB|GB)\b", text, re.IGNORECASE)
    if explicit_match:
        return f"{explicit_match.group(1)}{explicit_match.group(2).upper()}"

    bare_match = re.search(r"\b(64|128|256|512|1024|2048)\b", text)
    if not bare_match:
        return ""

    amount = bare_match.group(1)
    if amount == "1024":
        return "1TB"
    if amount == "2048":
        return "2TB"
    return f"{amount}GB"


def normalize_folder_path_parts(parts):
    normalized_parts = []
    for part in parts or ():
        display_name = normalize_folder_display_name(part)
        if display_name:
            normalized_parts.append(display_name)
    return tuple(normalized_parts)


def folder_path_equals_parts(actual_path, expected_path):
    actual = [part.casefold() for part in normalize_folder_path_parts(actual_path)]
    expected = [part.casefold() for part in normalize_folder_path_parts(expected_path)]
    return actual == expected


def infer_airpods_model_folder_name(product_name):
    normalized_name = re.sub(r"\s+", " ", str(product_name or "").strip()).lower()
    if not normalized_name:
        return None
    if "airpods max 2" in normalized_name:
        return "AirPods Max 2"
    if "airpods max" in normalized_name:
        return "AirPods Max"
    if "airpods pro 3" in normalized_name:
        return "AirPods Pro 3"
    if "airpods pro 2" in normalized_name or ("airpods pro" in normalized_name and "type-c" in normalized_name):
        return "AirPods Pro 2 (Type-C)"
    if "airpods 4" in normalized_name:
        if any(marker in normalized_name for marker in ("anc", "шумоподав", "noise cancel")):
            return "AirPods 4 с активным шумоподавлением"
        return "AirPods 4"
    if "airpods 3" in normalized_name:
        return "AirPods 3"
    if "airpods 2" in normalized_name:
        return "AirPods 2"
    return None


def infer_gaming_catalog_folder_path(product_name):
    normalized_name = re.sub(r"\s+", " ", str(product_name or "").strip())
    lower_name = normalize_match_text(normalized_name)
    if not lower_name:
        return None

    if "rog ally" in lower_name or "xbox ally" in lower_name:
        return (*PORTABLE_GAMING_CATALOG_PATH, "Asus ROG Ally")
    if "legion go" in lower_name:
        return (*PORTABLE_GAMING_CATALOG_PATH, "Lenovo Legion Go")
    if "msi claw" in lower_name:
        return (*PORTABLE_GAMING_CATALOG_PATH, "MSI Claw")

    if "steamdeck" in lower_name or "steam deck" in lower_name:
        if any(marker in lower_name for marker in ("dock", "док", "charging", "charge", "заряд")):
            return STEAM_DECK_ACCESSORIES_CATALOG_PATH
        return STEAM_DECK_CATALOG_PATH

    if re.search(r"\b(?:meta|oculus)(?:\s+quest)?\s*3\b", lower_name):
        return OCULUS_CATALOG_PATH

    has_switch = bool(re.search(r"\b(?:nintendo\s+)?switch\b|\bnsw2?\b", lower_name))
    if has_switch or "joy con" in lower_name or "joy-con" in lower_name:
        if any(marker in lower_name for marker in ("pro controller", "joy con", "joy-con", "геймпад", "контроллер")):
            return NINTENDO_GAMEPADS_CATALOG_PATH
        if any(marker in lower_name for marker in (
            "case", "bag", "camera", "micro sd", "sd express", "screen protector",
            "protective", "защ", "чех", "стекло", "карта памяти",
        )):
            return NINTENDO_ACCESSORIES_CATALOG_PATH
        if any(marker in lower_name for marker in ("switch lite", "switch oled", "nintendo switch", "nsw oled")):
            return NINTENDO_CONSOLES_CATALOG_PATH

    if re.search(r"\bps\s*portal\b|\bplaystation\s+portal\b", lower_name):
        return PLAYSTATION_CONSOLES_CATALOG_PATH
    if re.search(r"\bps\s*5\b|\bplaystation\s*5\b|\bcfi[\s-]?\d{2,4}\s*[ab]\b", lower_name):
        if any(marker in lower_name for marker in ("disc drive", "дисковод", "panel", "панел", "cover", "наклад")):
            return PLAYSTATION_ACCESSORIES_CATALOG_PATH
        return PLAYSTATION_CONSOLES_CATALOG_PATH
    if re.search(r"\bdual[\s-]*sen[cs]e\b|\bdualsense\b", lower_name):
        if any(marker in lower_name for marker in ("charging", "charge", "dock", "station", "заряд", "станц")):
            return PLAYSTATION_ACCESSORIES_CATALOG_PATH
        return PLAYSTATION_GAMEPADS_CATALOG_PATH
    if any(marker in lower_name for marker in ("nacon revolution", "victrix pro", "razer wolverine")):
        return PLAYSTATION_GAMEPADS_CATALOG_PATH
    if any(marker in lower_name for marker in ("pulse elite", "inzone", "razer kaira", "logitech shifter", "logitech g29", "logitech g923")):
        return PLAYSTATION_ACCESSORIES_CATALOG_PATH

    if "xbox" in lower_name:
        if re.search(r"\bxbox\s+(?:series\s+)?[xs]\b", lower_name) and not any(marker in lower_name for marker in (
            "controller", "геймпад", "elite", "deep pink", "shock blue", "velocity",
            "pulse", "lunar", "nocturnal", "cipher", "breaker", "camo",
        )):
            return XBOX_CONSOLES_CATALOG_PATH
        if any(marker in lower_name for marker in (
            "controller", "геймпад", "elite", "deep pink", "shock blue", "velocity",
            "pulse", "lunar", "nocturnal", "cipher", "breaker", "camo", "doom dark ages",
        )):
            return XBOX_GAMEPADS_CATALOG_PATH
        if re.search(r"\bxbox\s+[xs]\b", lower_name):
            return XBOX_CONSOLES_CATALOG_PATH

    return None


def infer_specific_catalog_folder_path(product_name, current_path, storage_value=""):
    normalized_name = re.sub(r"\s+", " ", str(product_name or "").strip())
    lower_name = normalized_name.lower()
    normalized_path = normalize_folder_path_parts(current_path)
    if not normalized_name or not normalized_path:
        return None

    if folder_path_equals_parts(normalized_path, DATA_CABLES_CATALOG_PATH):
        if "apple watch" in lower_name:
            return (*WATCH_CHARGERS_CATALOG_PATH, "Watch")
        if "galaxy watch" in lower_name or "samsung watch" in lower_name:
            return (*WATCH_CHARGERS_CATALOG_PATH, "Samsung")
        if any(marker in lower_name for marker in ("xiaomi watch", "redmi watch", "mi band")):
            return (*WATCH_CHARGERS_CATALOG_PATH, "Xiaomi")
        if "magsafe 3" in lower_name or ("mag safe 3" in lower_name) or ("macbook" in lower_name and "magsafe" in lower_name):
            return (*NETWORK_CHARGERS_CATALOG_PATH, "СЗУ для ноутбука")
        if "lightning" in lower_name:
            return (*DATA_CABLES_CATALOG_PATH, "Дата-кабели Lightning")
        if "micro" in lower_name:
            return (*DATA_CABLES_CATALOG_PATH, "Дата-кабели Micro")
        if any(marker in lower_name for marker in ("type-c", "usb-c", "thunderbolt")):
            return (*DATA_CABLES_CATALOG_PATH, "Дата-кабели Type-C")
        return (*DATA_CABLES_CATALOG_PATH, "Прочие дата-кабели")

    if folder_path_equals_parts(normalized_path, NETWORK_CHARGERS_CATALOG_PATH):
        watch_target = None
        has_apple_watch = "apple watch" in lower_name
        has_samsung_watch = "galaxy watch" in lower_name or "samsung watch" in lower_name
        has_xiaomi_watch = any(marker in lower_name for marker in ("xiaomi watch", "redmi watch", "mi band"))
        watch_target_count = sum((has_apple_watch, has_samsung_watch, has_xiaomi_watch))
        if watch_target_count > 1:
            watch_target = "Универсальные"
        elif has_apple_watch:
            watch_target = "Watch"
        elif has_samsung_watch:
            watch_target = "Samsung"
        elif has_xiaomi_watch:
            watch_target = "Xiaomi"
        if watch_target:
            return (*WATCH_CHARGERS_CATALOG_PATH, watch_target)

        if any(marker in lower_name for marker in ("macbook", "magsafe", "mag safe", "ноутбук")):
            return (*NETWORK_CHARGERS_CATALOG_PATH, "СЗУ для ноутбука")

        has_bundle_marker = any(marker in lower_name for marker in ("с кабелем", "+ кабель", "with cable"))
        if has_bundle_marker and "lightning" in lower_name:
            return (*NETWORK_CHARGERS_CATALOG_PATH, "СЗУ + Lightning")
        if has_bundle_marker and "micro" in lower_name:
            return (*NETWORK_CHARGERS_CATALOG_PATH, "СЗУ + Micro")
        if has_bundle_marker and any(marker in lower_name for marker in ("type-c", "usb-c")):
            return (*NETWORK_CHARGERS_CATALOG_PATH, "СЗУ + Type-C")
        return (*NETWORK_CHARGERS_CATALOG_PATH, "СЗУ блок")

    if folder_path_equals_parts(normalized_path, WATCH_CHARGERS_CATALOG_PATH):
        has_apple_watch = "apple watch" in lower_name
        has_samsung_watch = "galaxy watch" in lower_name or "samsung watch" in lower_name
        has_xiaomi_watch = any(marker in lower_name for marker in ("xiaomi watch", "redmi watch", "mi band"))
        target_count = sum((has_apple_watch, has_samsung_watch, has_xiaomi_watch))
        if target_count > 1:
            return (*WATCH_CHARGERS_CATALOG_PATH, "Универсальные")
        if has_apple_watch:
            return (*WATCH_CHARGERS_CATALOG_PATH, "Watch")
        if has_samsung_watch:
            return (*WATCH_CHARGERS_CATALOG_PATH, "Samsung")
        if has_xiaomi_watch:
            return (*WATCH_CHARGERS_CATALOG_PATH, "Xiaomi")
        return (*WATCH_CHARGERS_CATALOG_PATH, "Универсальные")

    if folder_path_equals_parts(normalized_path, YANDEX_STATION_CATALOG_PATH):
        if "мини 3 pro" in lower_name or "mini 3 pro" in lower_name:
            return (*YANDEX_STATION_CATALOG_PATH, "Станция Мини 3 Pro")
        if "мини 3" in lower_name or "mini 3" in lower_name:
            return (*YANDEX_STATION_CATALOG_PATH, "Станция Мини 3")
        if re.search(r"\bстанция\s+3\b", lower_name) and "мини" not in lower_name:
            return (*YANDEX_STATION_CATALOG_PATH, "Станция 3")

    if folder_path_equals_parts(normalized_path, JBL_SPEAKER_CATALOG_PATH):
        if "partybox" in lower_name:
            return (*JBL_SPEAKER_CATALOG_PATH, "PartyBox")

    if folder_path_equals_parts(normalized_path, PLAYSTATION_SUBSCRIPTIONS_CATALOG_PATH):
        if any(marker in lower_name for marker in ("подпис", "plus", "store", "учетной записи", "учётной записи", "аккаунт", "пополнен")):
            return (*PLAYSTATION_SUBSCRIPTIONS_CATALOG_PATH, "Подписки")
        return (*PLAYSTATION_SUBSCRIPTIONS_CATALOG_PATH, "Игры")

    if folder_path_equals_parts(normalized_path, XBOX_SUBSCRIPTIONS_CATALOG_PATH):
        if any(marker in lower_name for marker in ("подпис", "game pass", "core", "ultimate", "live", "gold", "пополнен")):
            return (*XBOX_SUBSCRIPTIONS_CATALOG_PATH, "Подписки")
        return (*XBOX_SUBSCRIPTIONS_CATALOG_PATH, "Игры")

    if folder_path_equals_parts(normalized_path, STEAM_DECK_CATALOG_PATH):
        target_storage_folder = detect_storage_folder_name(normalized_name, storage_value)
        if target_storage_folder:
            return (*STEAM_DECK_CATALOG_PATH, target_storage_folder.replace(" ", ""))

    if folder_path_equals_parts(normalized_path, CONSTRUCTORS_CATALOG_PATH):
        if "franzis" in lower_name:
            return (*CONSTRUCTORS_CATALOG_PATH, "Franzis")
        if "lego" in lower_name or re.search(r"\btechnic\b", lower_name):
            return (*CONSTRUCTORS_CATALOG_PATH, "LEGO")
        if any(marker in lower_name for marker in ("xiaomi", "onebot", "dawn of jupiter")):
            return (*CONSTRUCTORS_CATALOG_PATH, "Xiaomi")

    if folder_path_equals_parts(normalized_path, COMPUTER_ACCESSORIES_CATALOG_PATH):
        if any(marker in lower_name for marker in ("чех", "case", "storage box", "organizer")):
            return (*COMPUTER_ACCESSORIES_CATALOG_PATH, "Чехлы и органайзеры")
        if any(marker in lower_name for marker in ("клавиат", "keyboard")):
            return (*COMPUTER_ACCESSORIES_CATALOG_PATH, "Клавиатуры")
        if any(marker in lower_name for marker in ("мыш", "mouse", "трекпад", "trackpad")):
            return (*COMPUTER_ACCESSORIES_CATALOG_PATH, "Мышки и трекпады")
        if any(marker in lower_name for marker in ("веб-кам", "webcam")):
            return (*COMPUTER_ACCESSORIES_CATALOG_PATH, "Веб-камеры")
        if any(marker in lower_name for marker in ("router", "роутер", "wi-fi", "wifi")):
            return (*COMPUTER_ACCESSORIES_CATALOG_PATH, "Роутеры и адаптеры Wi-Fi")
        if any(marker in lower_name for marker in ("hub", "адаптер", "adapter", "dock", "док-станц")):
            return (*COMPUTER_ACCESSORIES_CATALOG_PATH, "Разветвители и адаптеры")
        if "графический планшет" in lower_name:
            return (*COMPUTER_ACCESSORIES_CATALOG_PATH, "Графические планшеты")

    if folder_path_equals_parts(normalized_path, AIRPODS_CATALOG_PATH):
        if any(marker in lower_name for marker in ("подстав", "holder", "stand")):
            return ("Электроника", "Аксессуары", "Прочие аксессуары", "Подставки")
        model_folder_name = infer_airpods_model_folder_name(normalized_name)
        if model_folder_name:
            return (*AIRPODS_CATALOG_PATH, model_folder_name)

    return None


def infer_iphone_model_folder_name(product_name):
    normalized_name = re.sub(r"\s+", " ", str(product_name or "").replace("ё", "е").lower()).strip()
    if not normalized_name:
        return None
    for pattern, folder_name in IPHONE_MODEL_FOLDER_PATTERNS:
        if re.search(pattern, normalized_name):
            return folder_name
    return None


def ensure_folder_path(db, user_id, folder_names):
    parent_id = None
    for folder_name in folder_names:
        parent_id = ensure_named_folder(db, user_id, folder_name, parent_id)
    return parent_id


def find_or_create_default_model_folder(db, user_id, product_name, brand=""):
    if infer_iphone_accessory_catalog_path(product_name, brand):
        return None
    if not is_apple_iphone_product(product_name, brand):
        return None
    iphone_model_folder_name = infer_iphone_model_folder_name(product_name)
    if iphone_model_folder_name:
        return ensure_folder_path(db, user_id, (*IPHONE_CATALOG_PATH, iphone_model_folder_name))
    return None


def resolve_product_folder_assignment(db, user_id, product_name, folder_id=None, storage_value="", brand=""):
    normalized_folder_id = normalize_folder_id(folder_id)
    gaming_catalog_path = infer_gaming_catalog_folder_path(product_name)
    if gaming_catalog_path:
        return ensure_folder_path(db, user_id, gaming_catalog_path)
    accessory_catalog_path = infer_iphone_accessory_catalog_path(product_name, brand)
    if accessory_catalog_path:
        return ensure_folder_path(db, user_id, accessory_catalog_path)
    target_variant_folder = detect_variant_folder_name(product_name)
    target_storage_folder = detect_storage_folder_name(product_name, storage_value)
    inferred_model_folder_id = find_or_create_default_model_folder(db, user_id, product_name, brand)

    current_folder = None
    if normalized_folder_id:
        if inferred_model_folder_id and not folder_path_startswith(db, user_id, normalized_folder_id, IPHONE_CATALOG_PATH):
            current_folder = None
        else:
            current_folder = get_user_folder(db, user_id, normalized_folder_id)

    model_folder_id = None
    variant_folder_id = None
    if current_folder:
        current_name = str(current_folder["name"] or "").strip()
        current_name_lower = current_name.lower()
        current_parent_id = normalize_folder_id(current_folder["parent_id"])
        current_path = get_folder_path_names(db, user_id, current_folder["id"])

        if is_storage_folder_name(current_name):
            parent_variant_folder = get_user_folder(db, user_id, current_parent_id)
            if parent_variant_folder and is_variant_folder_name(parent_variant_folder["name"]):
                model_folder_id = normalize_folder_id(parent_variant_folder["parent_id"])
                if not target_variant_folder or str(parent_variant_folder["name"] or "").strip().lower() == target_variant_folder:
                    variant_folder_id = int(parent_variant_folder["id"])
                    if not target_storage_folder or current_name_lower == target_storage_folder.lower():
                        return int(current_folder["id"])
            else:
                model_folder_id = current_parent_id
                if not target_variant_folder and (not target_storage_folder or current_name_lower == target_storage_folder.lower()):
                    return int(current_folder["id"])
        elif is_variant_folder_name(current_name):
            model_folder_id = current_parent_id
            if not target_variant_folder or current_name_lower == target_variant_folder:
                variant_folder_id = int(current_folder["id"])
        elif inferred_model_folder_id and [part.lower() for part in current_path[:len(IPHONE_CATALOG_PATH)]] == [part.lower() for part in IPHONE_CATALOG_PATH]:
            if len(current_path) >= len(IPHONE_CATALOG_PATH) + 1:
                model_folder_id = int(current_folder["id"])
        else:
            model_folder_id = int(current_folder["id"])

    if inferred_model_folder_id and not folder_path_startswith(db, user_id, model_folder_id, IPHONE_CATALOG_PATH):
        model_folder_id = inferred_model_folder_id
    if model_folder_id is None:
        model_folder_id = inferred_model_folder_id

    iphone_folder_context = folder_path_startswith(db, user_id, model_folder_id, IPHONE_CATALOG_PATH)

    if target_variant_folder and iphone_folder_context:
        if variant_folder_id is None:
            if model_folder_id is None:
                return normalized_folder_id
            variant_folder_id = ensure_named_folder(db, user_id, target_variant_folder, model_folder_id)
        if target_storage_folder:
            return ensure_named_folder(db, user_id, target_storage_folder, variant_folder_id)
        return variant_folder_id

    if target_storage_folder and model_folder_id is not None and iphone_folder_context:
        resolved_folder_id = ensure_named_folder(db, user_id, target_storage_folder, model_folder_id)
    else:
        resolved_folder_id = model_folder_id if model_folder_id is not None else normalized_folder_id

    refined_current_path = get_folder_path_names(db, user_id, resolved_folder_id) if resolved_folder_id else []
    refined_catalog_path = infer_specific_catalog_folder_path(product_name, refined_current_path, storage_value)
    if refined_catalog_path:
        return ensure_folder_path(db, user_id, refined_catalog_path)
    return resolved_folder_id


def normalize_autofill_weight(value):
    raw = str(value or "").strip().replace(",", ".")
    if not raw:
        return ""
    match = re.search(r"(\d+(?:\.\d+)?)", raw)
    if not match:
        return ""
    try:
        numeric = float(match.group(1))
    except ValueError:
        return ""
    lowered = raw.lower()
    if "кг" in lowered or "kg" in lowered:
        grams = numeric * 1000.0
    elif "г" in lowered:
        grams = numeric
    else:
        grams = numeric * 1000.0 if numeric < 10 else numeric
    if grams <= 0 or grams >= 100000:
        return ""
    return str(int(round(grams)))


def normalize_capacity(value, allow_tb=False):
    raw = str(value or "").strip().upper().replace(" ", "")
    if not raw:
        return ""
    match = re.fullmatch(r"(\d+(?:\.\d+)?)(GB|TB)", raw)
    if not match:
        return ""
    amount, unit = match.groups()
    if unit == "TB" and not allow_tb:
        return ""
    return f"{amount} {unit}"


def capacity_to_gb(value):
    normalized = normalize_capacity(value, allow_tb=True)
    if not normalized:
        return 0.0
    amount_str, unit = normalized.split()
    amount = float(amount_str)
    return amount * 1024.0 if unit == "TB" else amount


def extract_ram_storage_from_name(product_name):
    raw_name = str(product_name or "").replace("ГБ", "GB").replace("гб", "GB").replace("ТБ", "TB").replace("тб", "TB")
    pair_patterns = (
        re.compile(r"\b(\d{1,3})\s*/\s*(\d{2,4})\s*(GB|TB)\b", re.IGNORECASE),
        re.compile(r"\b(\d{1,3})\s*(GB)\s*/\s*(\d{2,4})\s*(GB|TB)\b", re.IGNORECASE),
    )
    for pattern in pair_patterns:
        match = pattern.search(raw_name)
        if not match:
            continue
        groups = match.groups()
        if len(groups) == 3:
            ram_amount, storage_amount, storage_unit = groups
            ram_value = normalize_capacity(f"{ram_amount}GB", allow_tb=False)
            storage_value = normalize_capacity(f"{storage_amount}{storage_unit}", allow_tb=True)
        else:
            ram_amount, ram_unit, storage_amount, storage_unit = groups
            ram_value = normalize_capacity(f"{ram_amount}{ram_unit}", allow_tb=False)
            storage_value = normalize_capacity(f"{storage_amount}{storage_unit}", allow_tb=True)
        if ram_value or storage_value:
            return ram_value, storage_value

    capacities = []
    for amount, unit in re.findall(r"\b(\d+(?:\.\d+)?)\s*(GB|TB)\b", raw_name, flags=re.IGNORECASE):
        normalized = normalize_capacity(f"{amount}{unit}", allow_tb=True)
        if normalized:
            capacities.append(normalized)
    unique_capacities = []
    seen = set()
    for capacity in capacities:
        key = capacity.lower()
        if key in seen:
            continue
        seen.add(key)
        unique_capacities.append(capacity)
    if not unique_capacities:
        return "", ""
    if len(unique_capacities) == 1:
        return "", unique_capacities[0]
    ordered = sorted(unique_capacities, key=capacity_to_gb)
    ram_candidate = ordered[0] if capacity_to_gb(ordered[0]) <= 64 else ""
    storage_candidate = ordered[-1]
    if ram_candidate and ram_candidate.lower() == storage_candidate.lower():
        ram_candidate = ""
    return ram_candidate, storage_candidate


def infer_basic_os(product_name, brand=""):
    lower = str(product_name or "").lower()
    normalized_brand = str(brand or "").strip()
    if normalized_brand == "Apple":
        if "ipad" in lower:
            return "iPadOS"
        if "watch" in lower:
            return "watchOS"
        if any(marker in lower for marker in ("macbook", "imac", "mac mini", "mac studio", "mac pro")):
            return "macOS"
        if "vision pro" in lower:
            return "visionOS"
        if any(marker in lower for marker in ("iphone",)):
            return DEFAULT_IPHONE_OS_VERSION
        return ""
    if normalized_brand in ANDROID_BRANDS:
        return "Android"
    if "galaxy watch" in lower or "wear os" in lower:
        return "Wear OS"
    return ""


def normalize_model_no(value, brand):
    model_no = str(value or "").strip().upper()
    if not model_no:
        return ""
    if brand == "Apple":
        return model_no if re.fullmatch(r"M[A-Z0-9]{4,6}", model_no) else ""
    if len(model_no) > 32:
        return ""
    if not re.fullmatch(r"[A-Z0-9][A-Z0-9\-/ ]{1,31}", model_no):
        return ""
    if not re.search(r"\d", model_no):
        return ""
    return model_no


def strip_description_notes(description):
    raw_text = str(description or "").replace("\r", "")
    raw_text = re.sub(r"<[^>]+>", " ", raw_text)
    raw_text = raw_text.replace("&nbsp;", " ").replace("&amp;", "&").replace("&quot;", '"')
    raw_text = re.sub(r"читать полностью.*", "", raw_text, flags=re.IGNORECASE)
    for note in (AUTOFILL_REQUIRED_NOTE, AUTOFILL_APPLE_NOTE):
        raw_text = re.sub(rf"\s*{re.escape(note)}\s*", "\n", raw_text, flags=re.IGNORECASE)
    raw_text = re.sub(r"\s*без\s+rustore\s*", "\n", raw_text, flags=re.IGNORECASE)
    lines = [line.rstrip() for line in raw_text.split("\n")]
    banned_lines = {
        AUTOFILL_REQUIRED_NOTE.lower(),
        AUTOFILL_APPLE_NOTE.lower(),
    }
    cleaned_lines = []
    for line in lines:
        normalized_line = line.strip().lower()
        if not normalized_line:
            cleaned_lines.append(line)
            continue
        if normalized_line in banned_lines:
            continue
        if any(marker in normalized_line for marker in DESCRIPTION_TEMPLATE_MARKERS):
            continue
        if normalized_line.startswith("преимущества модели:"):
            continue
        if normalized_line.startswith("исполнение:") or normalized_line.startswith("ключевые характеристики:"):
            continue
        if "rustore" in normalized_line:
            continue
        if normalized_line.startswith("недостатки:"):
            continue
        if normalized_line == "◊ юридическая информация" or normalized_line.startswith("юридическая информация"):
            continue
        if "apple.com/" in normalized_line:
            continue
        cleaned_lines.append(line)
    return "\n".join(cleaned_lines).strip()


def remove_sim_mentions_from_description(description):
    text = SIM_INLINE_VARIANT_PATTERN.sub("", str(description or "").replace("\r", ""))
    parts = [part.strip() for part in re.split(r"\n\s*\n", text) if part.strip()]
    cleaned_parts = []
    for part in parts:
        if SIM_DESCRIPTION_PATTERN.search(part):
            continue
        cleaned_parts.append(part)
    return "\n\n".join(cleaned_parts).strip()


def infer_description_kind_from_name(product_name):
    lower = str(product_name or "").lower()
    if not lower:
        return ""
    if any(marker in lower for marker in ("чех", "case", "cover")):
        return "case"
    if any(marker in lower for marker in ("ремеш", "strap", "браслет для")):
        return "strap"
    if any(marker in lower for marker in ("стекл", "пленк", "protector", "glass")):
        return "protection"
    if any(marker in lower for marker in ("держател", "mount", "stand", "подстав")):
        return "mount"
    if any(marker in lower for marker in ("кабель", "cable")):
        return "cable"
    if any(marker in lower for marker in ("заряд", "charger", "адаптер")):
        return "charger"
    if any(marker in lower for marker in ("airpods", "науш", "гарнитур", "buds")):
        return "headphones"
    if any(marker in lower for marker in ("колонк", "speaker", "homepod", "станция")):
        return "speaker"
    if any(marker in lower for marker in ("watch", "часы", "fit")):
        return "watch"
    if any(marker in lower for marker in ("ipad", "планш")):
        return "tablet"
    if any(marker in lower for marker in ("macbook", "ноут", "laptop")):
        return "laptop"
    if looks_like_iphone_shorthand(product_name) or any(marker in lower for marker in ("iphone", "смартфон", "galaxy", "pixel", "redmi note", "nothing phone")):
        return "smartphone"
    return ""


def should_append_apple_notes(brand, product_name="", kind=""):
    if str(brand or "").strip() != "Apple":
        return False
    effective_kind = str(kind or "").strip().lower() or infer_description_kind_from_name(product_name)
    if effective_kind in APPLE_DESCRIPTION_NOTE_KINDS:
        return True
    lower = str(product_name or "").lower()
    return any(marker in lower for marker in ("iphone", "ipad", "macbook", "apple watch", "imac", "mac mini", "vision pro"))


def finalize_description(description, brand, product_name="", kind=""):
    base_description = remove_sim_mentions_from_description(strip_description_notes(description))
    parts = []
    seen_parts = set()
    for part in re.split(r"\n\s*\n", base_description):
        cleaned_part = re.sub(r"\s+", " ", str(part or "").strip())
        normalized_part = cleaned_part.lower()
        if not cleaned_part or normalized_part in seen_parts:
            continue
        seen_parts.add(normalized_part)
        parts.append(cleaned_part)
    return "\n\n".join(parts[:4]).strip()


def build_synonyms(product_name, brand, article=""):
    original = re.sub(r"\s+", " ", str(product_name or "").strip())
    if not original:
        return []

    def append_variant(variants_list, value):
        cleaned = re.sub(r"\s+", " ", str(value or "").strip())
        if cleaned:
            variants_list.append(cleaned)

    variants = []
    variant_match = re.search(r"\(([^)]*(?:e\s*sim|sim)[^)]*)\)", original, re.IGNORECASE)
    variant_suffix = variant_match.group(1).strip() if variant_match else ""
    variant_suffix_lower = variant_suffix.lower()
    no_variant = (
        original[:variant_match.start()] + " " + original[variant_match.end():]
        if variant_match else original
    )
    no_variant = re.sub(r"\s+", " ", no_variant).strip()
    storage_case_normalized = re.sub(
        r"\b(\d+(?:\.\d+)?)\s*(gb|tb)\b",
        lambda match: f"{match.group(1)}{match.group(2).upper()}",
        no_variant,
        flags=re.IGNORECASE,
    )
    storage_case_normalized = re.sub(r"\s+", " ", storage_case_normalized).strip()
    normalized_storage = re.sub(
        r"(?<!\d)(\d{3,4})(?!\s*(?:GB|TB)\b)(?=\s+[A-Za-z])",
        r"\1GB",
        storage_case_normalized,
        flags=re.IGNORECASE,
    )
    normalized_storage = re.sub(r"\s+", " ", normalized_storage).strip()
    base_name = normalized_storage or storage_case_normalized or no_variant
    article_model_no = extract_apple_model_no(article) or extract_apple_model_no(product_name)
    strict_iphone_variant = bool(
        brand == "Apple"
        and is_iphone_variant_sensitive(original)
        and infer_iphone_variant_key(original)
    )

    if no_variant and no_variant.lower() != original.lower() and not strict_iphone_variant:
        append_variant(variants, no_variant)
    if storage_case_normalized and storage_case_normalized.lower() not in {original.lower(), no_variant.lower()} and not strict_iphone_variant:
        append_variant(variants, storage_case_normalized)
    if no_variant and variant_suffix:
        append_variant(variants, f"{no_variant} {variant_suffix}")
        if variant_suffix_lower == "1sim+esim":
            append_variant(variants, f"{no_variant} 1Sim")
        elif variant_suffix_lower == "esim":
            append_variant(variants, f"{no_variant} eSim")
        elif variant_suffix_lower == "2sim":
            append_variant(variants, f"{no_variant} 2Sim")
    if storage_case_normalized and variant_suffix:
        append_variant(variants, f"{storage_case_normalized} {variant_suffix}")
        if variant_suffix_lower == "1sim+esim":
            append_variant(variants, f"{storage_case_normalized} 1Sim")
        elif variant_suffix_lower == "esim":
            append_variant(variants, f"{storage_case_normalized} eSim")
        elif variant_suffix_lower == "2sim":
            append_variant(variants, f"{storage_case_normalized} 2Sim")
    if normalized_storage and normalized_storage.lower() not in {original.lower(), no_variant.lower()} and not strict_iphone_variant:
        append_variant(variants, normalized_storage)
        if variant_suffix:
            append_variant(variants, f"{normalized_storage} {variant_suffix}")
            if variant_suffix_lower == "1sim+esim":
                append_variant(variants, f"{normalized_storage} 1Sim")
            elif variant_suffix_lower == "esim":
                append_variant(variants, f"{normalized_storage} eSim")
            elif variant_suffix_lower == "2sim":
                append_variant(variants, f"{normalized_storage} 2Sim")
    if (
        brand == "Apple"
        and normalized_storage
        and looks_like_iphone_shorthand(original)
        and not normalized_storage.lower().startswith("iphone ")
        and not has_apple_part_number(original)
    ):
        if not strict_iphone_variant:
            append_variant(variants, f"iPhone {normalized_storage}")
        if variant_suffix:
            append_variant(variants, f"iPhone {normalized_storage} {variant_suffix}")
            if variant_suffix_lower == "1sim+esim":
                append_variant(variants, f"iPhone {normalized_storage} 1Sim")
            elif variant_suffix_lower == "esim":
                append_variant(variants, f"iPhone {normalized_storage} eSim")
            elif variant_suffix_lower == "2sim":
                append_variant(variants, f"iPhone {normalized_storage} 2Sim")
    if brand == "Apple" and base_name and APPLE_FAMILY_PREFIX_PATTERN.match(base_name):
        if not base_name.lower().startswith("apple "):
            if not strict_iphone_variant:
                append_variant(variants, f"Apple {base_name}")
            if variant_suffix:
                append_variant(variants, f"Apple {base_name} {variant_suffix}")
        if article_model_no:
            append_variant(variants, f"{base_name} {article_model_no}")
            if not base_name.lower().startswith("apple "):
                append_variant(variants, f"Apple {base_name} {article_model_no}")

    unique_variants = []
    seen = {original.lower()}
    for variant in variants:
        key = variant.lower()
        if not variant or key in seen:
            continue
        seen.add(key)
        unique_variants.append(variant)
        if len(unique_variants) >= 12:
            break
    return unique_variants


def build_offline_autofill_payload(product_name, article=""):
    brand = resolve_product_brand_value(product_name, article, detect_brand_from_inputs(product_name, article))
    ram_value, storage_value = extract_ram_storage_from_name(product_name)
    apple_model_no = extract_apple_model_no(article) or extract_apple_model_no(product_name)
    payload = {key: "" for key in AUTOFILL_RESPONSE_KEYS}
    payload["brand"] = brand
    payload["country_sim"] = detect_country_variant(product_name) or infer_accessory_country_fallback(product_name, brand)
    payload["weight"] = infer_iphone_weight_from_name(product_name) if is_apple_iphone_product(product_name, brand) else ""
    payload["storage"] = storage_value
    payload["ram"] = ram_value
    payload["model_no"] = apple_model_no if brand == "Apple" else ""
    payload["os"] = infer_basic_os(product_name, brand)
    payload["warranty"] = STORE_WARRANTY_DEFAULT
    payload["description"] = ""
    payload = supplement_autofill_payload_from_known_products(product_name, article, payload)
    payload = supplement_autofill_payload_from_catalog(product_name, article, payload)
    return normalize_autofill_payload(product_name, payload, article)


def build_autofill_prompt(product_name, article=""):
    article_line = f'Артикул / part number: "{article}"\n' if article else ""
    return f"""
Ты заполняешь карточку товара для магазина в русском стиле.
Товар: "{product_name}"
{article_line}Данные должны быть актуальны на 2026 год и позже, включая текущее состояние на момент запроса.
Если можешь использовать онлайн-поиск, обязательно опирайся на него. Если в точности не уверен, оставляй поле пустым.
Если артикул передан, сначала выполни поиск именно по точному артикулу в кавычках и только потом сверяй название товара.

СТИЛЬ КАРТОЧКИ:
1. Описание только на русском языке.
2. Тон строго под российский интернет-магазин: деловой, понятный, без англоязычного маркетингового мусора.
3. 3 коротких абзаца без списков и markdown.
4. Первое предложение начинай с точного названия модели в стиле: "iPhone 17 Pro — ...", "Samsung Galaxy S25 — ...".
5. Не добавляй в описание служебные пометки вроде "Версия без RuStore", "Без RuStore", "Оригинальная продукция Apple".
6. Не перечисляй характеристики списком внутри описания и не дублируй поля карточки дословно.

ПРАВИЛА:
- Если название вида "17", "17 Pro", "17 Pro Max", "17 Air", "17e" без бренда, считай что это iPhone соответствующей модели.
- Если в названии есть ASIS, это б/у товар. Не описывай его как новый, запечатанный или заводской.
- Если в названии есть CPO, это Certified Pre-Owned. Тоже не описывай как абсолютно новый товар.
- Флаги стран вроде 🇮🇳 🇦🇪 🇭🇰 🇯🇵 — это регион поставки/происхождения и подсказка для поиска, но НЕ значение поля country_sim.
- Если в названии есть Apple part number вроде "MQL03", "MU773", "MC7X4", ты ОБЯЗАН определить устройство именно по part number через онлайн-поиск. Никогда не трактуй такой код как номер iPhone.
- Если артикул передан отдельно, он важнее названия для определения точной модели и конфигурации.
- При конфликте между названием и артикулом доверяй артикулу, а не догадке по названию.
- Для Apple part number сначала ищи точное совпадение part number, затем сверяй цвет, память и тип устройства. Если тип устройства не совпал, не заполняй догадкой.
- Если запрос невозможно подтвердить онлайн, лучше верни пустые поля, чем неправильную модель.
- Поле brand заполняй только реальным брендом/производителем товара из названия или артикула. Не используй прайсовый заголовок, секцию или совместимость как бренд: PS5/PS4/PlayStation, Xbox, Nintendo Switch/NSW, USB-C, Magic, 16, 42mm, 60W и похожие слова не являются брендом товара сами по себе.
- Если это Apple, постарайся заполнить максимум подтверждаемых полей в стиле карточек магазина: display, processor, camera, front_camera, connectivity, os, biometrics, charging.
- Не выдумывай артикулы, вес, камеры, RAM, частоты, батарею и SIM-версию.
- Если поле не удалось подтвердить, верни "".
- Цвет только на английском, как в названии.
- storage и ram возвращай только с пробелом: "256 GB", "8 GB", "1 TB".
- Гарантия всегда "12 месяцев".
- country_sim только из этого набора, если можно определить по названию:
  eSIM -> "eSim Only"
  1SIM+eSIM / SIM+eSIM -> "Глобальная версия"
  2SIM -> "Китай"
- model_no для Apple только если это реальный код из 5 символов, начинающийся на M.

Верни только JSON без пояснений.
{{
"brand": "",
"country_sim": "",
"weight": "",
"color": "",
"storage": "",
"ram": "",
"display": "",
"processor": "",
"camera": "",
"front_camera": "",
"video": "",
"connectivity": "",
"battery": "",
"os": "",
"biometrics": "",
"charging": "",
"warranty": "12 месяцев",
"model_no": "",
"description": ""
}}
    """


def call_gemini_online(prompt):
    if not google_genai_sdk or not google_genai_types:
        raise RuntimeError("Модуль google-genai не установлен, онлайн-поиск Gemini недоступен")
    if not GEMINI_API_KEY:
        raise RuntimeError("GEMINI_API_KEY не задан")
    client = google_genai_sdk.Client(api_key=GEMINI_API_KEY)
    last_error = None
    for model_name in AUTOFILL_MODELS:
        try:
            response = client.models.generate_content(
                model=model_name,
                contents=prompt,
                config=google_genai_types.GenerateContentConfig(
                    temperature=0.1,
                    tools=[google_genai_types.Tool(google_search=google_genai_types.GoogleSearch())],
                ),
            )
            response_text = getattr(response, "text", None)
            if response_text and response_text.strip():
                return response_text.strip()
            raise RuntimeError(f"{model_name} вернул пустой ответ после онлайн-поиска")
        except Exception as e:
            last_error = e
            logger.warning(f"Gemini online grounding недоступен на {model_name}: {e}")
    raise RuntimeError(
        f"Онлайн-поиск Gemini сейчас недоступен. Автозаполнение остановлено, чтобы не подставлять устаревшие или выдуманные данные. Последняя ошибка: {last_error}"
    )


def parse_autofill_json(result_text):
    cleaned = str(result_text or "").replace("```json", "").replace("```", "").strip()
    if not cleaned:
        raise RuntimeError("Gemini вернул пустой ответ")
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        start = cleaned.find("{")
        end = cleaned.rfind("}")
        if start != -1 and end > start:
            return json.loads(cleaned[start:end + 1])
        raise


def supplement_autofill_payload_from_catalog(product_name, article="", payload=None):
    next_payload = dict(payload or {})
    try:
        from scripts import backfill_cmstore_product_details as bf
        from scripts import fill_catalog_from_cmstore as fc
    except Exception as e:
        logger.warning(f"Не удалось подключить каталоговый добор ТХ: {e}")
        return next_payload

    source_url = str(next_payload.get("source_url") or "").strip()
    if not source_url:
        try:
            candidate, score = bf.choose_candidate(product_name, timeout=20)
        except Exception as e:
            logger.warning(f"Поиск карточки-донора для '{product_name}' не удался: {e}")
            return next_payload
        if not candidate or score < AUTOFILL_SUPPLEMENT_MIN_SCORE:
            return next_payload
        source_url = str(candidate.get("url") or "").strip()
        if source_url:
            next_payload["source_url"] = source_url

    if not source_url:
        return next_payload

    try:
        details, page_description, _ = fc.get_page_payload(source_url, {})
    except Exception as e:
        logger.warning(f"Не удалось получить детали товара из '{source_url}': {e}")
        return next_payload

    detail_specs = bf.build_spec_updates(details)

    def fill_if_missing(key, value):
        value = str(value or "").strip()
        if value and not str(next_payload.get(key) or "").strip():
            next_payload[key] = value

    fill_if_missing("brand", detect_brand_from_inputs(product_name, article))
    fill_if_missing("weight", normalize_autofill_weight(bf.find_detail(details, *fc.WEIGHT_DETAIL_LABELS)))
    fill_if_missing("model_no", bf.find_detail(details, *fc.MODEL_DETAIL_LABELS))
    fill_if_missing("color", fc.detect_color(
        {
            "title": product_name,
            "brand": next_payload.get("brand") or detect_brand_from_inputs(product_name, article),
            "shorty_map": {},
        },
        details,
    ))

    for key in AUTOFILL_SPEC_KEYS:
        fill_if_missing(key, detail_specs.get(key))

    if page_description and not str(next_payload.get("description") or "").strip():
        next_payload["description"] = finalize_description(
            page_description,
            next_payload.get("brand"),
            product_name=product_name,
        )

    return next_payload

SPEC_DUPLICATE_LABEL_TARGETS = {
    "дисплей": ("display",),
    "процессор": ("processor",),
    "камера": ("camera",),
    "фронтальная камера": ("front_camera",),
    "видео": ("video",),
    "связь": ("connectivity",),
    "подключение": ("connectivity",),
    "интерфейс": ("connectivity", "charging"),
    "батарея": ("battery",),
    "ос": ("os",),
    "операционная система": ("os",),
    "биометрия": ("biometrics",),
    "зарядка": ("charging",),
    "память": ("storage",),
    "встроенная память": ("storage",),
    "объем встроенной памяти": ("storage",),
    "озу": ("ram",),
    "ram": ("ram",),
    "оперативная память": ("ram",),
    "объем оперативной памяти": ("ram",),
    "цвет": ("color",),
    "бренд": ("brand",),
    "производитель": ("brand",),
    "страна": ("country",),
    "страна производства": ("country",),
    "страна-производитель": ("country",),
    "вес": ("weight",),
    "модель": ("model_number",),
    "код производителя": ("model_number",),
    "артикул": ("model_number",),
    "гарантия": ("warranty",),
}
SPEC_ALWAYS_REDUNDANT_LABELS = {
    "тип",
    "тип товара",
    "категория",
}


def get_product_field_value(product, key):
    if product is None:
        return ""
    if isinstance(product, dict):
        return product.get(key, "")
    try:
        return product[key]
    except Exception:
        return ""


def normalize_spec_label_for_compare(value):
    text = str(value or "").strip().rstrip(":").lower().replace("ё", "е")
    text = text.replace("№", "").replace("(color)", "")
    return re.sub(r"\s+", " ", text).strip()


def normalize_spec_value_for_compare(value):
    text = str(value or "").strip().lower().replace("ё", "е")
    replacements = {
        "wi fi": "wi-fi",
        "type c": "type-c",
        "usb c": "usb-c",
        "гб": "gb",
        "тб": "tb",
        "мпикс": "mp",
        "мач": "mah",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    return re.sub(r"[^a-zа-я0-9]+", "", text)


def normalize_weight_for_compare(value):
    return normalize_autofill_weight(value)


def are_duplicate_spec_values(label, left, right):
    if not str(left or "").strip() or not str(right or "").strip():
        return False
    normalized_label = normalize_spec_label_for_compare(label)
    if normalized_label == "вес":
        return normalize_weight_for_compare(left) == normalize_weight_for_compare(right)
    return normalize_spec_value_for_compare(left) == normalize_spec_value_for_compare(right)


def sanitize_specs_payload(raw_value, product=None):
    if isinstance(raw_value, dict):
        parsed = {key: str(value).strip() for key, value in raw_value.items() if str(value).strip()}
    else:
        parsed = parse_specs_payload(raw_value)
    if not parsed:
        return {}

    compare_values = {}
    for key in AUTOFILL_SPEC_KEYS:
        value = str(parsed.get(key) or "").strip()
        if value:
            compare_values[key] = value
    for key in ("brand", "color", "country", "weight", "model_number", "storage", "ram", "warranty"):
        value = str(get_product_field_value(product, key) or "").strip()
        if value:
            compare_values[key] = value

    cleaned = {}
    for key, value in parsed.items():
        clean_key = str(key or "").strip()
        clean_value = str(value or "").strip()
        if not clean_key or not clean_value:
            continue
        if clean_key in AUTOFILL_SPEC_KEYS:
            cleaned[clean_key] = clean_value
            continue
        normalized_label = normalize_spec_label_for_compare(clean_key)
        if normalized_label in SPEC_ALWAYS_REDUNDANT_LABELS:
            continue
        duplicate_targets = SPEC_DUPLICATE_LABEL_TARGETS.get(normalized_label, ())
        if any(are_duplicate_spec_values(normalized_label, clean_value, compare_values.get(target)) for target in duplicate_targets):
            continue
        cleaned[clean_key] = clean_value
    return cleaned


def parse_specs_payload(raw_value, product=None):
    if isinstance(raw_value, dict):
        parsed = {key: str(value).strip() for key, value in raw_value.items() if str(value).strip()}
        return sanitize_specs_payload(parsed, product)
    if not raw_value:
        return {}
    try:
        parsed = json.loads(raw_value)
    except (TypeError, ValueError, json.JSONDecodeError):
        return {}
    if not isinstance(parsed, dict):
        return {}
    cleaned = {key: str(value).strip() for key, value in parsed.items() if str(value).strip()}
    return sanitize_specs_payload(cleaned, product)


def build_specs_payload_from_autofill(existing_specs, payload):
    merged = dict(existing_specs or {})
    for key in AUTOFILL_SPEC_KEYS:
        value = str(payload.get(key, "") or "").strip()
        if value:
            merged[key] = value
    return sanitize_specs_payload(merged, payload)


def merge_synonyms(existing_synonyms, incoming_synonyms):
    merged = []
    seen = set()
    for source in (incoming_synonyms or [], str(existing_synonyms or "").split(",")):
        if isinstance(source, str):
            candidates = [source]
        else:
            candidates = list(source or [])
        for candidate in candidates:
            value = str(candidate or "").strip()
            key = value.lower()
            if not value or key in seen:
                continue
            seen.add(key)
            merged.append(value)
    return ", ".join(merged)


def build_autofill_cache_key(product_name, article=""):
    raw = f"{str(product_name or '').strip().lower()}|{str(article or '').strip().lower()}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def get_cached_autofill_payload(db, product_name, article=""):
    cache_key = build_autofill_cache_key(product_name, article)
    max_age = f"-{int(AUTOFILL_CACHE_MAX_AGE_DAYS)} days"
    row = db.execute(
        """
        SELECT payload
        FROM ai_autofill_cache
        WHERE cache_key = ?
          AND updated_at >= datetime('now', ?)
        ORDER BY updated_at DESC
        LIMIT 1
        """,
        (cache_key, max_age),
    ).fetchone()
    if not row or not row["payload"]:
        return None
    try:
        payload = json.loads(row["payload"])
    except (TypeError, ValueError, json.JSONDecodeError):
        return None
    if not isinstance(payload, dict):
        return None
    payload["synonyms"] = payload.get("synonyms") or []
    return payload


def save_autofill_cache(db, user_id, product_name, article, payload):
    cache_key = build_autofill_cache_key(product_name, article)
    serialized_payload = json.dumps(payload, ensure_ascii=False)
    existing = db.execute(
        "SELECT id FROM ai_autofill_cache WHERE cache_key = ?",
        (cache_key,),
    ).fetchone()
    if existing:
        db.execute(
            """
            UPDATE ai_autofill_cache
            SET user_id = ?, product_name = ?, article = ?, payload = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (user_id, product_name, article, serialized_payload, existing["id"]),
        )
    else:
        db.execute(
            """
            INSERT INTO ai_autofill_cache (cache_key, user_id, product_name, article, payload)
            VALUES (?, ?, ?, ?, ?)
            """,
            (cache_key, user_id, product_name, article, serialized_payload),
        )


def payload_improves_product(product_row, payload):
    product = dict(product_row or {})
    comparisons = (
        ("brand", sanitize_brand_value(product.get("brand")), sanitize_brand_value(payload.get("brand"))),
        ("country", str(product.get("country") or "").strip(), str(payload.get("country_sim") or "").strip()),
        ("weight", str(product.get("weight") or "").strip(), str(payload.get("weight") or "").strip()),
        ("model_number", normalize_model_no(product.get("model_number"), sanitize_brand_value(product.get("brand"))), normalize_model_no(payload.get("model_no"), sanitize_brand_value(payload.get("brand")))),
        ("color", str(product.get("color") or "").strip(), str(payload.get("color") or "").strip()),
        ("storage", str(product.get("storage") or "").strip(), str(payload.get("storage") or "").strip()),
        ("ram", str(product.get("ram") or "").strip(), str(payload.get("ram") or "").strip()),
        ("warranty", str(product.get("warranty") or "").strip(), str(payload.get("warranty") or "").strip()),
        ("description", strip_description_notes(product.get("description")), strip_description_notes(payload.get("description"))),
    )
    for _, current_value, next_value in comparisons:
        if next_value and next_value != current_value:
            return True

    existing_specs = parse_specs_payload(product.get("specs"))
    for key in AUTOFILL_SPEC_KEYS:
        next_value = str(payload.get(key) or "").strip()
        current_value = str(existing_specs.get(key) or "").strip()
        if next_value and next_value != current_value:
            return True

    existing_synonyms = {str(item or "").strip().lower() for item in str(product.get("synonyms") or "").split(",") if str(item or "").strip()}
    for synonym in payload.get("synonyms", []):
        normalized = str(synonym or "").strip().lower()
        if normalized and normalized not in existing_synonyms:
            return True
    return False


def upsert_product_from_autofill(db, user_id, product_row, payload, preferred_folder_id=None):
    product = dict(product_row)
    resolved_brand_hint = resolve_product_brand_value(
        product.get("name", ""),
        payload.get("model_no") or product.get("model_number") or "",
        payload.get("brand") or product.get("brand"),
    )
    resolved_folder_id = resolve_product_folder_assignment(
        db,
        user_id,
        product.get("name", ""),
        preferred_folder_id if preferred_folder_id is not None else product.get("folder_id"),
        payload.get("storage") or product.get("storage"),
        resolved_brand_hint,
    )
    merged_specs = build_specs_payload_from_autofill(parse_specs_payload(product.get("specs")), payload)
    merged_synonyms = merge_synonyms(product.get("synonyms"), payload.get("synonyms", []))
    resolved_brand = resolved_brand_hint
    resolved_country = (
        payload.get("country_sim")
        or product.get("country")
        or infer_brand_country_fallback(resolved_brand)
        or infer_accessory_country_fallback(product.get("name", ""), resolved_brand)
    )
    resolved_model_no = normalize_model_no(payload.get("model_no"), resolved_brand)
    if not resolved_model_no:
        resolved_model_no = normalize_model_no(product.get("model_number"), sanitize_brand_value(product.get("brand")))
    resolved_warranty = str(payload.get("warranty") or product.get("warranty") or STORE_WARRANTY_DEFAULT).strip()
    if resolved_warranty in {"Гарантия CMstore", "Гарантия от Магазина", "Гарантия от магазина", "Гарантия от магазина 1 год."}:
        resolved_warranty = STORE_WARRANTY_DEFAULT
    db.execute(
        """
        UPDATE products
        SET brand = ?, country = ?, weight = ?, model_number = ?, folder_id = ?,
            color = ?, storage = ?, ram = ?, warranty = ?, description = ?, specs = ?, synonyms = ?, source_url = ?
        WHERE id = ? AND user_id = ?
        """,
        (
            resolved_brand,
            resolved_country,
            payload.get("weight") or product.get("weight"),
            resolved_model_no,
            resolved_folder_id,
            payload.get("color") or product.get("color"),
            payload.get("storage") or product.get("storage"),
            payload.get("ram") or product.get("ram"),
            resolved_warranty,
            payload.get("description") or product.get("description"),
            json.dumps(merged_specs, ensure_ascii=False) if merged_specs else "",
            merged_synonyms,
            payload.get("source_url") or product.get("source_url"),
            product["id"],
            user_id,
        ),
    )
    return resolved_folder_id


def is_retryable_ai_error(error):
    text = str(error or "").lower()
    retry_markers = (
        "429",
        "503",
        "quota",
        "resource_exhausted",
        "rate limit",
        "high demand",
        "unavailable",
        "temporarily unavailable",
    )
    return any(marker in text for marker in retry_markers)


def compute_ai_retry_delay_seconds(attempts):
    if attempts <= 1:
        return 5 * 60
    if attempts == 2:
        return 15 * 60
    if attempts == 3:
        return 30 * 60
    return 60 * 60


def enqueue_ai_repair_job(db, user_id, product_id, product_name, article="", preferred_folder_id=None, last_error=""):
    existing = db.execute(
        """
        SELECT id, attempts
        FROM ai_repair_jobs
        WHERE user_id = ? AND product_id = ? AND status IN ('pending', 'processing', 'retry_wait')
        ORDER BY id DESC
        LIMIT 1
        """,
        (user_id, product_id),
    ).fetchone()
    if existing:
        db.execute(
            """
            UPDATE ai_repair_jobs
            SET product_name = ?, article = ?, preferred_folder_id = ?, status = 'retry_wait',
                last_error = ?, next_retry_at = CURRENT_TIMESTAMP, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (product_name, article, normalize_folder_id(preferred_folder_id), str(last_error or "")[:1000], existing["id"]),
        )
        return int(existing["id"])
    cursor = db.execute(
        """
        INSERT INTO ai_repair_jobs (
            user_id, product_id, product_name, article, preferred_folder_id,
            status, attempts, next_retry_at, last_error
        ) VALUES (?, ?, ?, ?, ?, 'retry_wait', 0, CURRENT_TIMESTAMP, ?)
        """,
        (
            user_id,
            product_id,
            product_name,
            article,
            normalize_folder_id(preferred_folder_id),
            str(last_error or "")[:1000],
        ),
    )
    return int(cursor.lastrowid)


def normalize_autofill_payload(product_name, parsed_data, article=""):
    payload = {key: str(parsed_data.get(key, "") or "").strip() for key in AUTOFILL_RESPONSE_KEYS}
    payload["source_url"] = str(parsed_data.get("source_url", "") or "").strip()
    detected_brand = sanitize_brand_value(detect_brand_from_inputs(product_name, article))
    payload["brand"] = resolve_product_brand_value(product_name, article, payload["brand"] or detected_brand)
    article_model_no = extract_apple_model_no(article) or extract_apple_model_no(product_name)
    if payload["brand"] == "Apple" and article_model_no:
        payload["model_no"] = article_model_no
    payload["country_sim"] = (
        detect_country_variant(product_name)
        or payload["country_sim"]
        or infer_brand_country_fallback(payload["brand"])
        or infer_accessory_country_fallback(product_name, payload["brand"])
    )
    payload["weight"] = normalize_autofill_weight(payload["weight"])
    if is_apple_iphone_product(product_name, payload["brand"]):
        payload["weight"] = infer_iphone_weight_from_name(product_name) or payload["weight"]
        payload["os"] = normalize_iphone_os_value(payload["os"], product_name, payload["brand"])
    elif not payload["os"]:
        payload["os"] = infer_basic_os(product_name, payload["brand"])
    inferred_ram, inferred_storage = extract_ram_storage_from_name(product_name)
    if inferred_storage and not payload["storage"]:
        payload["storage"] = inferred_storage
    if inferred_ram and not payload["ram"]:
        payload["ram"] = inferred_ram
    payload["storage"] = normalize_capacity(payload["storage"], allow_tb=True)
    payload["ram"] = normalize_capacity(payload["ram"], allow_tb=False)
    payload["model_no"] = normalize_model_no(payload["model_no"], payload["brand"])
    if payload["warranty"] in {"Гарантия CMstore", "Гарантия от Магазина", "Гарантия от магазина", "Гарантия от магазина 1 год."}:
        payload["warranty"] = STORE_WARRANTY_DEFAULT
    payload["warranty"] = payload["warranty"] or STORE_WARRANTY_DEFAULT
    payload["description"] = finalize_description(
        payload["description"],
        payload["brand"],
        product_name=product_name,
    )
    payload["synonyms"] = build_synonyms(product_name, payload["brand"], article)
    return payload


def get_or_generate_autofill_payload(db, user_id, product_name, article="", use_cache=True):
    if use_cache:
        cached_payload = get_cached_autofill_payload(db, product_name, article)
        if cached_payload:
            normalized_cached_payload = normalize_autofill_payload(product_name, cached_payload, article)
            if normalized_cached_payload != cached_payload:
                save_autofill_cache(db, user_id, product_name, article, normalized_cached_payload)
            return normalized_cached_payload, "cache"
    if not GEMINI_API_KEY:
        return build_offline_autofill_payload(product_name, article), "offline"
    result_text = call_gemini_online(build_autofill_prompt(product_name, article))
    payload = normalize_autofill_payload(product_name, parse_autofill_json(result_text), article)
    payload = normalize_autofill_payload(
        product_name,
        supplement_autofill_payload_from_known_products(product_name, article, payload),
        article,
    )
    payload = normalize_autofill_payload(
        product_name,
        supplement_autofill_payload_from_catalog(product_name, article, payload),
        article,
    )
    save_autofill_cache(db, user_id, product_name, article, payload)
    return payload, "live"


@app.route('/api/autofill', methods=['POST'])
def autofill():
    db = get_db()
    data = request.json or {}
    product_name = data.get('name', '').strip()
    article = data.get('article', '').strip()
    use_cache = not bool(data.get('force_online'))

    if not product_name:
        return jsonify({"error": "No product name"}), 400

    try:
        payload, source = get_or_generate_autofill_payload(db, 0, product_name, article, use_cache=use_cache)
        db.commit()
        return jsonify({**payload, "_source": source})
    except Exception as e:
        print("Ошибка Gemini:", str(e))
        return jsonify({"error": str(e)}), 503


@app.route('/api/products/<int:prod_id>/repair_with_ai', methods=['POST'])
@login_required
def repair_product_with_ai(prod_id):
    user_id = session['user_id']
    db = get_db()
    product_row = db.execute(
        "SELECT * FROM products WHERE id = ? AND user_id = ?",
        (prod_id, user_id),
    ).fetchone()
    if not product_row:
        return jsonify({"error": "Товар не найден"}), 404

    data = request.get_json(silent=True) or {}
    product = dict(product_row)
    product_name = str(data.get("name") or product.get("name") or "").strip()
    article = str(data.get("article") or product.get("model_number") or "").strip()
    preferred_folder_id = normalize_folder_id(data.get("folder_id"))

    if not product_name:
        return jsonify({"error": "У товара нет названия"}), 400

    try:
        payload, source = get_or_generate_autofill_payload(db, user_id, product_name, article, use_cache=True)
        if source == "cache" and not payload_improves_product(product_row, payload):
            payload, source = get_or_generate_autofill_payload(db, user_id, product_name, article, use_cache=False)
        if not payload_improves_product(product_row, payload):
            return jsonify({
                "success": False,
                "queued": False,
                "message": "Не удалось подтвердить новые данные для этой карточки. Нужен более точный источник или ручная проверка.",
            }), 200
        resolved_folder_id = upsert_product_from_autofill(
            db,
            user_id,
            product_row,
            payload,
            preferred_folder_id if preferred_folder_id is not None else product.get("folder_id"),
        )
        db.commit()
        notify_clients()
        variant_folder_name = detect_variant_folder_name(product_name)
        folder_message = f"Папка: {variant_folder_name}" if variant_folder_name else "Папка не менялась"
        return jsonify({
            "success": True,
            "queued": False,
            "folder_id": resolved_folder_id,
            "message": f"Карточка исправлена. {folder_message}. Источник: {'кэш' if source == 'cache' else 'онлайн'}",
        })
    except Exception as e:
        if is_retryable_ai_error(e):
            job_id = enqueue_ai_repair_job(
                db,
                user_id,
                prod_id,
                product_name,
                article,
                preferred_folder_id if preferred_folder_id is not None else product.get("folder_id"),
                last_error=e,
            )
            db.commit()
            notify_clients()
            return jsonify({
                "success": False,
                "queued": True,
                "job_id": job_id,
                "message": "Лимит или временная недоступность AI. Карточка поставлена в очередь и будет доисправлена автоматически.",
            }), 202
        logger.error(f"Ошибка AI-исправления товара {prod_id}: {e}")
        return jsonify({"error": str(e)}), 503



@app.route('/api/products/<int:prod_id>/messages')
@login_required
def get_product_messages(prod_id):
    user_id = session['user_id']
    db = get_db()
    truthy_values = {'1', 'true', 'yes'}
    confirmed_only = str(request.args.get('confirmed_only', '')).strip().lower() in truthy_values
    skip_maintenance = str(request.args.get('skip_maintenance', '')).strip().lower() in truthy_values or confirmed_only
    try:
        recent_limit = int(request.args.get('recent_limit') or 1000)
    except (TypeError, ValueError):
        recent_limit = 1000
    recent_limit = max(50, min(recent_limit, 1000))
    
    # 1. Получаем синонимы товара
    prod = db.execute(
        """
        SELECT name, country, synonyms, brand, model_number, color, storage
        FROM products
        WHERE id = ? AND user_id = ?
        """,
        (prod_id, user_id),
    ).fetchone()
    if not prod:
        return jsonify({'proposed': [], 'confirmed': [], 'supplier_summary': []})
    product_context = dict(prod)
    prepare_auto_confirm_product(product_context)
    product_name = prod['name'] or ""
    product_country = prod['country'] or ""
    product_synonyms = prod['synonyms'] or ""

    if AUTO_BINDING_MAINTENANCE_ENABLED and not skip_maintenance:
        should_notify_products = False
        with sqlite_write_lock:
            removed_duplicates = prune_duplicate_confirmed_bindings(db, product_id=prod_id)
            deactivated_bindings = deactivate_confirmed_bindings_from_inactive_messages(db, product_id=prod_id)
            removed_discount_bindings = remove_discount_section_confirmed_bindings(db, product_id=prod_id)
            sanitized_bindings = sanitize_confirmed_binding_records(db, product_id=prod_id, user_id=user_id)
            invalid_bindings = deactivate_semantically_invalid_confirmed_bindings(db, product_id=prod_id, user_id=user_id)
            refreshed_bindings = refresh_confirmed_bindings_from_latest_messages(db, product_id=prod_id, user_id=user_id)
            if removed_duplicates or deactivated_bindings or removed_discount_bindings or sanitized_bindings or invalid_bindings or refreshed_bindings:
                db.commit()
                should_notify_products = True
        if should_notify_products:
            notify_clients("products")
    
    # --- УМНАЯ ОЧИСТКА (превращает строку в набор слов) ---
    # --- УЛУЧШЕННАЯ ОЧИСТКА ---
    import re
    def clean_for_search(text):
        if not text: return set()
        
        # ДОБАВЛЯЕМ '*' (звездочку) в список удаляемых символов
        # Также добавили '.', если вдруг поставщик пишет "17.512"
        t = re.sub(r'(?i)\bwi\s*[- ]\s*fi\b', 'wifi', str(text or ""))
        t = re.sub(r'[()*.[\]{},;!|/+\-]', ' ', t)
        
        # Убираем лишние пробелы и неразрывные пробелы
        t = re.sub(r'\s+', ' ', t.replace('\xa0', ' ').replace('ё', 'е')).strip().lower()
        
        # Решаем проблему кириллицы/латиницы
        t = t.translate(str.maketrans('асеокрх', 'aceokpx'))
        
        # Возвращаем набор чистых слов, отделяя артикулы от флагов/эмодзи.
        words = set(re.findall(r'[a-zа-я0-9]+', t))
        words.update(token.lower() for token in extract_catalog_model_tokens(text))
        return words

    # Для предложенных привязок используем тот же набор сигналов, что и авто-confirm:
    # явные синонимы + сигнатуры из name/model, чтобы пустое поле synonyms не ломало матчинг.
    synonyms_as_word_sets = product_context.get('_synonym_sets') or []

    # --- ИСПРАВЛЕНИЕ: Получаем все последние сообщения (all_msgs) ---
    # Без этого блока была ошибка "all_msgs is not defined"
    all_msgs = []
    latest_fingerprints = {}
    if not confirmed_only:
        all_msgs_rows = db.execute("""
            SELECT m.id, m.text, m.date, m.chat_id, m.sender_name, m.chat_title, m.type, tc.custom_name
            FROM messages m
            LEFT JOIN tracked_chats tc ON m.chat_id = tc.chat_id AND m.user_id = tc.user_id
            WHERE m.user_id = ? 
              AND (m.is_blocked IS NULL OR m.is_blocked = 0)
              AND (m.is_delayed IS NULL OR m.is_delayed = 0)
            ORDER BY m.date DESC LIMIT ?
        """, (user_id, recent_limit)).fetchall()
        all_msgs = [dict(r) for r in all_msgs_rows]

        latest_fingerprints = {}
        for msg in all_msgs:
            if not msg['text']: continue
            lines = msg['text'].split('\n')
            in_discount_section = False
            for i, line in enumerate(lines):
                if not line.strip(): continue
                if line_is_discount_section_marker(line):
                    in_discount_section = True
                    continue
                if in_discount_section:
                    continue
                if line_has_discount_condition_marker(line):
                    continue
                fp = get_binding_fingerprint(
                    msg['chat_id'],
                    msg['sender_name'],
                    msg['text'],
                    i,
                    include_price=False,
                )
                if fp not in latest_fingerprints or msg['date'] > latest_fingerprints[fp]['date']:
                    price_val = extract_price_with_context(lines, i)

                    latest_fingerprints[fp] = {
                        'full_text': msg['text'],
                        'text': line,
                        'price': price_val,
                        'date': msg['date'],
                        'message_id': msg['id'],
                        'line_index': i
                    }

    # 2. Обработка ПОДТВЕРЖДЕННЫХ
    confirmed_rows = db.execute("""
        SELECT pm.id as pm_id, pm.group_id, pm.message_id, m.chat_title, m.chat_id, m.sender_name, 
               pm.extracted_price, pm.is_actual, m.text, m.date, pm.line_index, tc.custom_name, m.type
        FROM product_messages pm
        JOIN messages m ON pm.message_id = m.id
        LEFT JOIN tracked_chats tc ON m.chat_id = tc.chat_id AND m.user_id = tc.user_id
        WHERE pm.product_id = ? AND pm.status = 'confirmed'
    """, (prod_id,)).fetchall()
    
    confirmed_raw = []
    confirmed_fingerprints = set()
    has_updates = False
    delete_confirmed_ids = []

    for r in confirmed_rows:
        d = dict(r)
        if d['line_index'] != -1 and d['text']:
            lines = d['text'].split('\n')
            if 0 <= d['line_index'] < len(lines):
                orig_line = lines[d['line_index']]
                fp = get_binding_fingerprint(
                    d['chat_id'],
                    d['sender_name'],
                    d['text'],
                    d['line_index'],
                    d.get('extracted_price'),
                    include_price=True,
                )
                confirmed_fingerprints.add(fp)
                
                latest_key = get_binding_fingerprint(
                    d['chat_id'],
                    d['sender_name'],
                    d['text'],
                    d['line_index'],
                    include_price=False,
                )
                latest = latest_fingerprints.get(latest_key)
                if latest and latest['date'] > d['date']:
                    d['text'] = latest['full_text']
                    d['extracted_price'] = latest['price']
                    d['date'] = latest['date'] 
                    d['message_id'] = latest['message_id']
                    d['line_index'] = latest['line_index']
                    d['is_actual'] = 1
                    current_id = d.get('pm_id')
                    try:
                        db.execute("UPDATE product_messages SET message_id=?, line_index=?, extracted_price=?, is_actual=1 WHERE id=?",
                                   (latest['message_id'], latest['line_index'], latest['price'], current_id))
                        has_updates = True
                    except: pass
                else:
                    d['text'] = r['text']
                if d.get('extracted_price') in (None, ''):
                    recovered_price = extract_price_with_context(lines, d['line_index'])
                    if recovered_price is not None:
                        d['extracted_price'] = recovered_price
                        current_id = d.get('pm_id')
                        try:
                            db.execute("UPDATE product_messages SET extracted_price=? WHERE id=?", (recovered_price, current_id))
                            has_updates = True
                        except Exception:
                            pass
        source_line, resolved_price = resolve_binding_display_line_and_price(
            d.get('text'),
            d.get('line_index'),
            d.get('extracted_price'),
            require_price=False,
        )
        if not source_line:
            if d.get('pm_id'):
                delete_confirmed_ids.append(int(d['pm_id']))
                has_updates = True
            continue
        current_price = normalize_binding_price_value(d.get('extracted_price'))
        if resolved_price is not None and (current_price is None or abs(current_price - resolved_price) > 0.009):
            d['extracted_price'] = resolved_price
            current_id = d.get('pm_id')
            try:
                db.execute("UPDATE product_messages SET extracted_price=? WHERE id=?", (resolved_price, current_id))
                has_updates = True
            except Exception:
                pass
        identity_text = d.get('text')
        cleaned_line = strip_supplier_line_noise(source_line) or source_line.strip()
        d['match_line'] = cleaned_line
        d['text'] = cleaned_line
        d['_display_key'] = get_binding_fingerprint(
            d.get('chat_id'),
            d.get('sender_name'),
            identity_text,
            d.get('line_index'),
            d.get('extracted_price'),
            include_price=True,
        )
        confirmed_raw.append(d)

    if delete_confirmed_ids:
        for offset in range(0, len(delete_confirmed_ids), 500):
            chunk = delete_confirmed_ids[offset:offset + 500]
            placeholders = ",".join("?" for _ in chunk)
            db.execute(f"DELETE FROM product_messages WHERE id IN ({placeholders})", chunk)
        
    if has_updates:
        db.commit()
        notify_clients()
        
    grouped_confirmed = {}
    for d in confirmed_raw:
        g_id = f"group:{d['group_id']}" if d['group_id'] else f"fp:{d.get('_display_key') or d['pm_id']}"
        current = grouped_confirmed.get(g_id)
        if not current:
            grouped_confirmed[g_id] = d
            continue
        current_actual = 1 if int(current.get('is_actual') or 0) == 1 else 0
        next_actual = 1 if int(d.get('is_actual') or 0) == 1 else 0
        if (next_actual, str(d.get('date') or "")) > (current_actual, str(current.get('date') or "")):
            grouped_confirmed[g_id] = d
    
    confirmed = sorted(list(grouped_confirmed.values()), key=lambda x: x['date'], reverse=True)

    supplier_summary_map = {}
    for d in confirmed_raw:
        source_title = (d.get('custom_name') or d.get('chat_title') or d.get('sender_name') or 'Источник').strip()
        source_key = f"{d.get('chat_id') or ''}:{source_title.lower()}"
        row_date = str(d.get('date') or '')
        is_actual = 1 if int(d.get('is_actual') or 0) == 1 else 0
        line_text = strip_supplier_line_noise(d.get('match_line') or d.get('text') or '')
        raw_price = d.get('extracted_price')
        try:
            price_number = float(raw_price) if raw_price not in (None, '') else None
        except (TypeError, ValueError):
            price_number = None

        summary = supplier_summary_map.get(source_key)
        if not summary:
            summary = {
                'supplier_key': source_key,
                'chat_id': d.get('chat_id'),
                'chat_title': d.get('chat_title'),
                'sender_name': d.get('sender_name'),
                'custom_name': d.get('custom_name'),
                'type': d.get('type'),
                'confirmed_count': 0,
                'actual_count': 0,
                'best_actual_price': None,
                'lowest_seen_price': None,
                'latest_price': None,
                'latest_date': row_date,
                'latest_match_line': line_text,
                '_latest_rank': (is_actual, row_date),
            }
            supplier_summary_map[source_key] = summary

        summary['confirmed_count'] += 1
        if is_actual:
            summary['actual_count'] += 1

        if price_number is not None:
            if summary['lowest_seen_price'] is None or price_number < summary['lowest_seen_price']:
                summary['lowest_seen_price'] = price_number
            if is_actual and (summary['best_actual_price'] is None or price_number < summary['best_actual_price']):
                summary['best_actual_price'] = price_number

        if (is_actual, row_date) >= summary['_latest_rank']:
            summary['_latest_rank'] = (is_actual, row_date)
            summary['latest_date'] = row_date
            summary['latest_match_line'] = line_text
            summary['latest_price'] = price_number if price_number is not None else raw_price

    supplier_summary = list(supplier_summary_map.values())
    best_active_price = min(
        (item['best_actual_price'] for item in supplier_summary if item['best_actual_price'] is not None),
        default=None,
    )
    for item in supplier_summary:
        item['is_actual'] = 1 if item['actual_count'] > 0 else 0
        item['display_price'] = (
            item['best_actual_price']
            if item['best_actual_price'] is not None
            else item['latest_price']
        )
        item['is_best_price'] = bool(
            best_active_price is not None
            and item['best_actual_price'] is not None
            and float(item['best_actual_price']) == float(best_active_price)
        )
        item.pop('_latest_rank', None)

    supplier_summary.sort(
        key=lambda item: (
            0 if item['best_actual_price'] is not None else 1,
            float(item['best_actual_price']) if item['best_actual_price'] is not None else float(item['lowest_seen_price'] or 10**12),
            -int(item['actual_count'] or 0),
            str(item.get('latest_date') or ''),
        )
    )

    if confirmed_only:
        return jsonify({'proposed': [], 'confirmed': confirmed, 'supplier_summary': supplier_summary})

    # === НОВОЕ: Собираем все строки, которые уже привязаны к ДРУГИМИ товарам ===
    occupied_rows = db.execute(
        "SELECT message_id, line_index FROM product_messages WHERE status = 'confirmed' AND product_id != ?", 
        (prod_id,)
    ).fetchall()
    
    # Создаем удобный список занятых строк, чтобы быстро их проверять
    occupied = {(r['message_id'], r['line_index']) for r in occupied_rows}
    # ===========================================================================
    
    # 3. Обработка ПРЕДЛОЖЕННЫХ (склеивание дубликатов и поиск по словам)
    # 3. Обработка ПРЕДЛОЖЕННЫХ (склеивание дубликатов и поиск по словам)
    proposed_list = []
    seen_fps = set()
    
    for row in all_msgs:
        lines = row['text'].split('\n') if row['text'] else []
        in_discount_section = False
        for i, line in enumerate(lines):
            if not line.strip(): continue
            if line_is_discount_section_marker(line):
                in_discount_section = True
                continue
            if in_discount_section:
                continue
            if line_has_discount_condition_marker(line):
                continue
            
            # --- НОВОЕ СТРОГОЕ ПРАВИЛО ---
            # Если эта строка уже кем-то занята - мы просто игнорируем ее
            # и даже не показываем в списке "Предложенные"
            if (row['id'], i) in occupied or (row['id'], -1) in occupied:
                continue
            # -----------------------------

            fp = get_binding_fingerprint(
                row['chat_id'],
                row['sender_name'],
                row['text'],
                i,
                include_price=True,
            )
            
            if fp in seen_fps or fp in confirmed_fingerprints:
                continue
            # --- ЛОГИКА ЖЕСТКОГО ПОИСКА ПО СЛОВАМ ---
            # --- ЛОГИКА ЖЕСТКОГО ПОИСКА ПО СЛОВАМ ---
            line_words = clean_for_search(line)
            is_match = product_supplier_line_matches(
                product_context,
                line,
                line_words=line_words,
                synonym_sets=synonyms_as_word_sets,
            )
            
            if is_match:
                seen_fps.add(fp)
                
                # Ищем цену в текущей строке или на 1-2 строки ниже
                s_price = extract_price_with_context(lines, i)
                
                proposed_list.append({
                    'message_id': row['id'],
                    'line_index': i,
                    'chat_title': row['chat_title'],
                    'sender_name': row['sender_name'],
                    'custom_name': row.get('custom_name'),
                    'type': row['type'],
                    'match_line': strip_supplier_line_noise(line),
                    'suggested_price': s_price,
                    'date': row['date']
                })
    
    proposed_list.sort(key=lambda x: x['date'], reverse=True)
    return jsonify({'proposed': proposed_list, 'confirmed': confirmed, 'supplier_summary': supplier_summary})

import re

def parse_number(s):
    text = str(s or "").replace("\xa0", " ").strip()
    text = re.sub(r"[^\d,.\s]", "", text)
    compact = re.sub(r"\s+", "", text)
    if not re.search(r"\d", compact):
        return 0.0

    decimal_match = re.search(r"([,.])(\d{1,2})$", compact)
    if decimal_match:
        integer_part = re.sub(r"\D", "", compact[:decimal_match.start()])
        fraction_part = decimal_match.group(2)
        if integer_part:
            return float(f"{integer_part}.{fraction_part}")

    clean = re.sub(r"\D", "", compact)
    return float(clean) if clean else 0.0


def supplier_line_likely_low_price_accessory(line):
    normalized = normalize_match_text(line)
    if not normalized:
        return False
    return bool(re.search(
        r"\b(adapter|адаптер|переходник|charger|сзу|usb[\s-]*c|type[\s-]*c|"
        r"stand|подстав\w*|dock|док|station|станц\w*|macbook|euro\s*plug|евро\s*вилк)\b",
        normalized,
        re.IGNORECASE,
    ))


def extract_price_with_context(lines, index):
    if not lines:
        return None
    try:
        idx = int(index)
    except (TypeError, ValueError):
        idx = 0
    if idx < 0 or idx >= len(lines):
        return None
    price = extract_price(lines[idx])
    if price is not None:
        return price
    def extract_context_line_price(value):
        direct = extract_price(value)
        if direct is not None:
            return direct
        match = re.search(
            r"[-–—]\s*(\d{2,7}(?:[,.]\d{1,2})?)\s*(?:₽|руб\.?|р\.?|p\.?|rub|rur|usd|\$|€|к|k)?\s*$",
            str(value or ""),
            re.IGNORECASE,
        )
        if not match:
            return None
        return normalize_binding_price_value(match.group(1).replace(",", "."))

    if idx + 1 < len(lines):
        next_line = str(lines[idx + 1] or "").strip()
        if not next_line:
            return None
        if not binding_context_line_is_detail(lines, idx, idx + 1):
            return None
        price = extract_context_line_price(next_line)
        if price is not None:
            return price
    if idx + 2 < len(lines):
        next_line = str(lines[idx + 2] or "").strip()
        if not next_line:
            return None
        if not binding_context_line_is_detail(lines, idx, idx + 2):
            return None
        price = extract_context_line_price(next_line)
        if price is not None:
            return price
    return None


def extract_price(line):
    if not line: return None
    line_clean = str(line or "").replace("\xa0", " ")
    price_unit_pattern = r"(?:₽|руб\.?|р\.?|p\.?|rub|rur|usd|\$|€|к|k)"
    # Do not allow arbitrary spaces inside a number: "Cfi-7100 75000р" must yield 75000, not 710075000.
    price_number_pattern = r"(?:\d{1,3}(?:[ .]\d{3})+|\d{2,7})(?:[,.]\d{1,2})?"
    candidate_re = re.compile(
        rf"(?<![A-Za-zА-Яа-я0-9])(?P<number>{price_number_pattern})\s*(?P<unit>{price_unit_pattern})?(?![A-Za-zА-Яа-я0-9])",
        re.IGNORECASE,
    )
    all_numbers = list(candidate_re.finditer(line_clean))
    low_price_accessory = supplier_line_likely_low_price_accessory(line)

    if not all_numbers: return None

    for match in reversed(all_numbers):
        number_str = match.group('number').strip()
        try:
            price_value = parse_number(number_str)
            if price_value < (50 if low_price_accessory else 3000): continue
        except: continue

        end_pos = match.end()
        text_after = line_clean[end_pos:].strip().lower()
        if re.match(r'^[-]?\s*(шт|шт\.|штук|pcs|pc|item|ед|единиц)', text_after): continue

        text_before = line_clean[:match.start()].strip()
        has_unit = bool(match.group('unit'))
        has_price_separator = text_before.endswith(('-', '–', '—'))
        has_only_tail_noise = not text_after or not re.search(r"[A-Za-zА-Яа-я0-9]", text_after)
        if has_unit or has_price_separator or (price_value >= 1500 and has_only_tail_noise):
            return price_value

    for match in reversed(all_numbers):
        number_str = match.group('number').strip()
        try:
            price_value = parse_number(number_str)
            if price_value >= (50 if low_price_accessory else 1500):
                end_pos = match.end()
                text_after = line_clean[end_pos:].strip().lower()
                text_before = line_clean[:match.start()].strip()
                has_unit = bool(match.group('unit'))
                has_price_separator = text_before.endswith(('-', '–', '—'))
                has_only_tail_noise = not text_after or not re.search(r"[A-Za-zА-Яа-я0-9]", text_after)
                if (
                    not re.match(r'^[-]?\s*(шт|шт\.|штук|pcs|pc|item|ед|единиц)', text_after)
                    and (has_unit or has_price_separator or (price_value >= 1500 and has_only_tail_noise))
                ):
                    return price_value
        except: continue
    return None

LOGISTICS_MARKERS_RE = re.compile(r"[🛩🏎🚚🚙🚕🚗🚘🚖]")
LOGISTICS_WINDOW_RE = re.compile(
    r"(?:ориентировочно\s*)?(?:с\s*)?\d{1,2}[:.]\d{2}\s*(?:до|[-–—])\s*\d{1,2}[:.]\d{2}",
    re.IGNORECASE,
)
LINE_PRICE_SUFFIX_RE = re.compile(
    r"\s*[-–—]\s*(?:от\s*)?(?:\d{1,3}(?:[ .]\d{3})+|\d{4,})(?:[,.]\d{1,2})?\s*"
    r"(?:₽|руб\.?|р\.?|p\.?|rub|rur|usd|\$|€|к|k)?(?:\s+.*)?\s*$",
    re.IGNORECASE,
)
BARE_FINAL_PRICE_RE = re.compile(
    r"\s+(?:\d{1,3}(?:[ .]\d{3})+|\d{5,})(?:[,.]\d{1,2})?\s*"
    r"(?:₽|руб\.?|р\.?|p\.?|rub|rur|usd|\$|€|к|k)?\s*$",
    re.IGNORECASE,
)
LOW_PRICE_ACCESSORY_SUFFIX_RE = re.compile(
    r"\s*[-–—]\s*(?:от\s*)?\d{2,3}(?:[,.]\d{1,2})?\s*"
    r"(?:₽|руб\.?|р\.?|p\.?|rub|rur|usd|\$|€|к|k)?(?:\s+.*)?\s*$",
    re.IGNORECASE,
)
LOW_PRICE_ACCESSORY_BARE_RE = re.compile(
    r"\s+\d{2,3}(?:[,.]\d{1,2})?\s*"
    r"(?:₽|руб\.?|р\.?|p\.?|rub|rur|usd|\$|€|к|k)?\s*$",
    re.IGNORECASE,
)


def strip_supplier_line_noise(line):
    """Возвращает товарную часть строки без цены и логистического хвоста."""
    text = strip_supplier_line_decorations(line)
    if not text:
        return ""

    text = re.sub(r"\s+", " ", text)
    low_price_accessory = supplier_line_likely_low_price_accessory(text)
    for _ in range(3):
        before = text
        text = LOGISTICS_WINDOW_RE.sub(" ", text)
        text = LOGISTICS_MARKERS_RE.sub(" ", text)
        text = re.sub(r"\s+", " ", text).strip(" \t-–—")
        explicit_price_removed = False
        if low_price_accessory:
            low_before = text
            text = LOW_PRICE_ACCESSORY_SUFFIX_RE.sub("", text)
            text = LOW_PRICE_ACCESSORY_BARE_RE.sub("", text)
            explicit_price_removed = text != low_before
        suffix_before = text
        text = LINE_PRICE_SUFFIX_RE.sub("", text)
        explicit_price_removed = explicit_price_removed or text != suffix_before
        if not explicit_price_removed:
            text = BARE_FINAL_PRICE_RE.sub("", text)
        text = re.sub(r"\s+", " ", text).strip(" \t-–—")
        if explicit_price_removed:
            break
        if text == before:
            break
    return text


def get_binding_source_line(text, line_index):
    if not text:
        return ""
    try:
        idx = int(line_index)
    except (TypeError, ValueError):
        idx = -1
    if idx >= 0:
        lines = str(text).split("\n")
        if idx < len(lines):
            return lines[idx]
    return str(text)


DETAIL_LINE_PREFIX_RE = re.compile(
    r"^\s*(?:"
    r"(?:от\s*)?\d+\s*шт\b|"
    r"\d+\s*rev\b|\d+\s*рев\b|rev\b|рев\b|"
    r"cfi[\s-]?\d{2,4}[a-z0-9]*\b|"
    r"идеальн\w*|коробк\w*|мят\w*|чуть\b|витрин\w*|уценк\w*|"
    r"copy\b|копи\w*|аналог\w*|1\s*в\s*1|1к1|aaa?\+*|a\+*"
    r")",
    re.IGNORECASE,
)
IDENTITY_PRICE_FRAGMENT_RE = re.compile(
    r"\s*[-–—]\s*(?:от\s*)?(?:\d{1,3}(?:[.]\d{3})+|\d{2,7})(?:[,.]\d{1,2})?\s*"
    r"(?:₽|руб\.?|р\.?|p\.?|rub|rur|usd|\$|€|к|k)?(?![A-Za-zА-Яа-я0-9])",
    re.IGNORECASE,
)
IDENTITY_BARE_FINAL_PRICE_RE = re.compile(
    r"\s+(?:\d{1,3}(?:[ .]\d{3})+|\d{5,})(?:[,.]\d{1,2})?\s*"
    r"(?:₽|руб\.?|р\.?|p\.?|rub|rur|usd|\$|€|к|k)?\s*$",
    re.IGNORECASE,
)


def line_looks_like_separate_supplier_product_row(line):
    raw_line = str(line or "").strip()
    if not raw_line:
        return False
    cleaned = strip_supplier_line_decorations(raw_line)
    normalized = normalize_match_text(cleaned)
    if not normalized:
        return False
    if DETAIL_LINE_PREFIX_RE.search(normalized):
        return False
    if re.search(r"^\s*от\s*\d+\s*шт\b", normalized, re.IGNORECASE):
        return False
    if APPLE_PART_NUMBER_PATTERN.search(raw_line) and re.search(
        r"\b(?:iphone|ipad|macbook|mac\s*mini|imac|pro|air|studio|display|"
        r"silver|black|space|gray|grey|blue|gb|tb|гб|тб|cpu|gpu)\b",
        normalized,
        re.IGNORECASE,
    ):
        return True
    category = detect_match_category(raw_line)
    if not category:
        return False
    if category in {"playstation_console"} and re.search(r"^\s*(?:\d+\s*рев|rev|cfi)", normalized, re.IGNORECASE):
        return False
    return bool(extract_price(raw_line) or len(normalized.split()) >= 3)


def binding_context_line_is_detail(lines, source_index, candidate_index):
    if not lines:
        return False
    try:
        source_index = int(source_index)
        candidate_index = int(candidate_index)
    except (TypeError, ValueError):
        return False
    if candidate_index <= source_index or candidate_index >= len(lines):
        return False
    candidate = str(lines[candidate_index] or "").strip()
    if not candidate:
        return False
    for between in lines[source_index + 1:candidate_index]:
        if not str(between or "").strip():
            return False
    if re.fullmatch(r"[\s\-–—_=*•·➖]+", candidate):
        return False
    if line_looks_like_separate_supplier_product_row(candidate):
        return False
    source_line = str(lines[source_index] or "").strip() if 0 <= source_index < len(lines) else ""
    if extract_price(source_line) is not None:
        return True
    return bool(
        DETAIL_LINE_PREFIX_RE.search(candidate)
        or line_has_discount_condition_marker(candidate)
        or re.search(r"\b(?:шт|pcs|короб|box|rev|рев|cfi|copy|копи|аналог|1\s*в\s*1|1к1)\b", normalize_match_text(candidate), re.IGNORECASE)
    )


def strip_supplier_price_for_identity(value):
    text = strip_supplier_line_decorations(value)
    if not text:
        return ""
    text = LOGISTICS_WINDOW_RE.sub(" ", text)
    text = LOGISTICS_MARKERS_RE.sub(" ", text)
    text = IDENTITY_PRICE_FRAGMENT_RE.sub(" ", text)
    text = IDENTITY_BARE_FINAL_PRICE_RE.sub("", text)
    text = re.sub(r"\s+", " ", text).strip(" \t-–—:•·")
    return text


def normalize_binding_identity_text(value, *, strip_price=True):
    text = strip_supplier_price_for_identity(value) if strip_price else strip_supplier_line_decorations(value)
    text = normalize_match_text(text)
    text = re.sub(r"[₽$€]", " ", text)
    text = re.sub(r"\s+", " ", text).strip(" \t-–—:•·")
    return text


def binding_context_signature(text, line_index, *, strip_prices=True, max_lines=2):
    if not text:
        return ""
    try:
        idx = int(line_index)
    except (TypeError, ValueError):
        return ""
    if idx < 0:
        return ""
    lines = str(text or "").split("\n")
    if idx >= len(lines):
        return ""
    parts = []
    for next_index in range(idx + 1, min(len(lines), idx + 1 + max_lines)):
        if not binding_context_line_is_detail(lines, idx, next_index):
            break
        detail = normalize_binding_identity_text(lines[next_index], strip_price=strip_prices)
        if detail:
            parts.append(detail)
    return " | ".join(parts)


def binding_price_identity_value(value):
    price = normalize_binding_price_value(value)
    if price is None:
        return ""
    if float(price).is_integer():
        return str(int(price))
    return f"{price:.2f}".rstrip("0").rstrip(".")


def get_binding_fingerprint(chat_id, sender_name, text, line_index=-1, price=None, *, include_price=True):
    source_line = get_binding_source_line(text, line_index)
    base_line = normalize_binding_identity_text(source_line)
    detail_key = binding_context_signature(text, line_index, strip_prices=not include_price)
    price_key = binding_price_identity_value(price)
    if include_price and not price_key:
        try:
            idx = int(line_index)
        except (TypeError, ValueError):
            idx = -1
        if idx >= 0:
            price_key = binding_price_identity_value(extract_price_with_context(str(text or "").split("\n"), idx))
        if not price_key:
            price_key = binding_price_identity_value(extract_price(source_line))
    parts = [str(chat_id or ""), str(sender_name or ""), base_line]
    if detail_key:
        parts.append(detail_key)
    if include_price and price_key:
        parts.append(f"price:{price_key}")
    return "_".join(parts)


def normalize_binding_price_value(value):
    if value in (None, ""):
        return None
    try:
        price = float(value)
    except (TypeError, ValueError):
        return None
    if price <= 0:
        return None
    return price


def resolve_binding_display_line_and_price(text, line_index, extracted_price, *, require_price=True):
    raw_text = str(text or "")
    if not raw_text.strip():
        return "", None

    lines = raw_text.split("\n")
    source_line = str(get_binding_source_line(raw_text, line_index) or "").strip()
    resolved_price = normalize_binding_price_value(extracted_price)

    try:
        idx = int(line_index)
    except (TypeError, ValueError):
        idx = -1

    if idx >= 0:
        if not source_line:
            return "", None
        direct_price = extract_price(source_line)
        if direct_price is not None:
            return source_line, direct_price
        resolved_price = extract_price_with_context(lines, idx)
        if resolved_price is None and require_price:
            return "", None
        return source_line, resolved_price

    non_empty_lines = [str(line or "").strip() for line in lines if str(line or "").strip()]
    if len(non_empty_lines) == 1:
        source_line = non_empty_lines[0]
        if resolved_price is None:
            resolved_price = extract_price(source_line)
        if resolved_price is None and require_price:
            return "", None
        return source_line, resolved_price

    if resolved_price is None and require_price:
        return "", None

    compact_text = strip_supplier_line_noise(raw_text) or raw_text.strip()
    return compact_text, resolved_price


def sanitize_confirmed_binding_records(db, product_id=None, user_id=None):
    filters = ["pm.status = 'confirmed'"]
    params = []
    if product_id is not None:
        filters.append("pm.product_id = ?")
        params.append(product_id)
    if user_id is not None:
        filters.append("p.user_id = ?")
        params.append(user_id)

    rows = db.execute(f"""
        SELECT pm.id, pm.line_index, pm.extracted_price, m.text
        FROM product_messages pm
        JOIN messages m ON pm.message_id = m.id
        JOIN products p ON pm.product_id = p.id
        WHERE {' AND '.join(filters)}
    """, params).fetchall()

    changed = 0
    delete_ids = []
    for row in rows:
        source_line, resolved_price = resolve_binding_display_line_and_price(
            row["text"],
            row["line_index"],
            row["extracted_price"],
            require_price=False,
        )
        if not source_line:
            delete_ids.append(int(row["id"]))
            continue
        current_price = normalize_binding_price_value(row["extracted_price"])
        if resolved_price is not None and (current_price is None or abs(current_price - resolved_price) > 0.009):
            db.execute(
                "UPDATE product_messages SET extracted_price = ? WHERE id = ?",
                (resolved_price, row["id"]),
            )
            changed += 1

    for offset in range(0, len(delete_ids), 500):
        chunk = delete_ids[offset:offset + 500]
        placeholders = ",".join("?" for _ in chunk)
        db.execute(f"DELETE FROM product_messages WHERE id IN ({placeholders})", chunk)
    changed += len(delete_ids)
    return changed


def binding_row_has_explicit_line_identity(row, source_line):
    line_key = normalize_match_text(supplier_line_synonym(source_line) or strip_supplier_line_noise(source_line))
    source_key = normalize_match_text(source_line)
    product_model_key = normalize_match_text(row_value(row, "model_number", ""))
    if product_model_key and len(product_model_key) >= 4 and product_model_key in source_key:
        return True

    product_name_key = normalize_match_text(row_value(row, "name", ""))
    if product_name_key and len(product_name_key) >= 8:
        if product_name_key == line_key or product_name_key in line_key or line_key in product_name_key:
            return True

    for synonym in split_product_synonyms(row_value(row, "synonyms", "")):
        synonym_key = normalize_match_text(synonym)
        if not synonym_key or len(synonym_key) < 6:
            continue
        if synonym_key == line_key or synonym_key in line_key or line_key in synonym_key:
            return True
    return False


def deactivate_semantically_invalid_confirmed_bindings(db, product_id=None, user_id=None, limit=None):
    filters = [
        "pm.status = 'confirmed'",
        "COALESCE(pm.is_actual, 1) = 1",
    ]
    params = []
    if product_id is not None:
        filters.append("pm.product_id = ?")
        params.append(product_id)
    if user_id is not None:
        filters.append("p.user_id = ?")
        params.append(user_id)

    query = f"""
        SELECT
            pm.id,
            pm.line_index,
            p.name,
            p.country,
            p.synonyms,
            p.brand,
            p.model_number,
            p.color,
            p.storage,
            p.warranty,
            p.description,
            p.description_html,
            p.specs,
            m.text
        FROM product_messages pm
        JOIN products p ON p.id = pm.product_id
        JOIN messages m ON m.id = pm.message_id
        WHERE {' AND '.join(filters)}
    """
    if limit is not None:
        query += " LIMIT ?"
        params.append(int(limit))

    invalid_ids = []
    for row in db.execute(query, params).fetchall():
        source_line = get_binding_source_line(row["text"], row["line_index"])
        if source_line and binding_row_has_explicit_line_identity(row, source_line):
            continue
        if not source_line or not binding_product_row_matches_line(row, source_line, require_known_category=True):
            invalid_ids.append(int(row["id"]))

    for offset in range(0, len(invalid_ids), 500):
        chunk = invalid_ids[offset:offset + 500]
        placeholders = ",".join("?" for _ in chunk)
        db.execute(f"UPDATE product_messages SET is_actual = 0 WHERE id IN ({placeholders})", chunk)
    return len(invalid_ids)


def get_fingerprint(chat_id, sender_name, line):
    """Создает слепок товарной строки, игнорируя цену и доставочный хвост."""
    norm_line = strip_supplier_line_noise(line).lower()
    norm_line = re.sub(r"[₽$€рRkKкК]", "", norm_line).strip()
    norm_line = re.sub(r"\s+", " ", norm_line)
    return f"{chat_id}_{sender_name}_{norm_line}"


def find_existing_confirmed_binding_by_fingerprint(
    db,
    product_id,
    chat_id,
    sender_name,
    line,
    exclude_id=None,
    *,
    text=None,
    line_index=-1,
    price=None,
):
    target_text = text if text is not None else line
    target_fp = get_binding_fingerprint(chat_id, sender_name, target_text, line_index, price, include_price=True)
    target_tracking_fp = get_binding_fingerprint(chat_id, sender_name, target_text, line_index, include_price=False)
    rows = db.execute("""
        SELECT pm.id, pm.message_id, pm.line_index, pm.extracted_price, pm.is_actual, m.text, m.date,
               m.chat_id, m.sender_name
        FROM product_messages pm
        JOIN messages m ON pm.message_id = m.id
        WHERE pm.product_id = ? AND pm.status = 'confirmed'
    """, (product_id,)).fetchall()
    matches = []
    for row in rows:
        if exclude_id and int(row['id']) == int(exclude_id):
            continue
        row_line = get_binding_source_line(row['text'], row['line_index'])
        row_fp = get_binding_fingerprint(
            row['chat_id'],
            row['sender_name'],
            row['text'],
            row['line_index'],
            row['extracted_price'],
            include_price=True,
        )
        row_tracking_fp = get_binding_fingerprint(
            row['chat_id'],
            row['sender_name'],
            row['text'],
            row['line_index'],
            include_price=False,
        )
        row_has_price = normalize_binding_price_value(row['extracted_price']) is not None
        can_upgrade_priceless = normalize_binding_price_value(price) is not None and not row_has_price
        if row_fp == target_fp or (can_upgrade_priceless and row_tracking_fp == target_tracking_fp):
            matches.append(row)
    if not matches:
        return None
    return sorted(
        matches,
        key=lambda r: (
            1 if int(r['is_actual'] or 0) == 1 else 0,
            datetime_sort_value(r['date']),
            int(r['id'] or 0),
        ),
        reverse=True,
    )[0]


def get_duplicate_confirmed_binding_ids(db, product_id=None, user_id=None):
    filters = ["pm.status = 'confirmed'"]
    params = []
    if product_id is not None:
        filters.append("pm.product_id = ?")
        params.append(product_id)
    if user_id is not None:
        filters.append("p.user_id = ?")
        params.append(user_id)

    rows = db.execute(f"""
        SELECT pm.id, pm.product_id, pm.message_id, pm.line_index, pm.extracted_price, pm.is_actual,
               m.text, m.date, m.chat_id, m.sender_name
        FROM product_messages pm
        JOIN messages m ON pm.message_id = m.id
        JOIN products p ON pm.product_id = p.id
        WHERE {' AND '.join(filters)}
    """, params).fetchall()

    best_by_key = {}
    delete_ids = []
    for row in rows:
        line = get_binding_source_line(row['text'], row['line_index'])
        key = (
            row['product_id'],
            get_binding_fingerprint(
                row['chat_id'],
                row['sender_name'],
                row['text'],
                row['line_index'],
                row['extracted_price'],
                include_price=True,
            ),
        )
        current = best_by_key.get(key)
        row_rank = (
            1 if int(row['is_actual'] or 0) == 1 else 0,
            datetime_sort_value(row['date']),
            int(row['id'] or 0),
        )
        if current is None:
            best_by_key[key] = (row, row_rank)
            continue
        _, current_rank = current
        if row_rank > current_rank:
            delete_ids.append(current[0]['id'])
            best_by_key[key] = (row, row_rank)
        else:
            delete_ids.append(row['id'])

    return delete_ids


def prune_duplicate_confirmed_bindings(db, product_id=None, user_id=None):
    delete_ids = get_duplicate_confirmed_binding_ids(db, product_id=product_id, user_id=user_id)
    if delete_ids:
        for offset in range(0, len(delete_ids), 500):
            chunk = delete_ids[offset:offset + 500]
            placeholders = ",".join("?" for _ in chunk)
            db.execute(f"DELETE FROM product_messages WHERE id IN ({placeholders})", chunk)
    return len(delete_ids)


def binding_line_identity_score(row, source_line):
    line_key = normalize_match_text(supplier_line_synonym(source_line) or strip_supplier_line_noise(source_line))
    source_key = normalize_match_text(source_line)
    best = 0

    product_model_key = normalize_match_text(row_value(row, "model_number", ""))
    if product_model_key and len(product_model_key) >= 4 and product_model_key in source_key:
        best = max(best, 1000 + len(product_model_key))

    product_name_key = normalize_match_text(row_value(row, "name", ""))
    if product_name_key and len(product_name_key) >= 6:
        if product_name_key == line_key:
            best = max(best, 900 + len(product_name_key))
        elif product_name_key in line_key or line_key in product_name_key:
            best = max(best, 600 + min(len(product_name_key), len(line_key)))

    for synonym in split_product_synonyms(row_value(row, "synonyms", "")):
        synonym_key = normalize_match_text(synonym)
        if not synonym_key or len(synonym_key) < 6:
            continue
        if synonym_key == line_key:
            best = max(best, 850 + len(synonym_key))
        elif synonym_key in line_key or line_key in synonym_key:
            best = max(best, 550 + min(len(synonym_key), len(line_key)))

    if best:
        return best
    if product_supplier_line_matches(row, source_line, require_known_category=True, allow_article_conflict=True):
        return 100
    return 0


def deactivate_duplicate_actual_line_bindings(db, product_id=None, user_id=None):
    filters = [
        "pm.status = 'confirmed'",
        "COALESCE(pm.is_actual, 1) = 1",
    ]
    params = []
    if product_id is not None:
        filters.append("pm.product_id = ?")
        params.append(product_id)
    if user_id is not None:
        filters.append("p.user_id = ?")
        params.append(user_id)

    rows = db.execute(f"""
        SELECT
            pm.id,
            pm.product_id,
            pm.message_id,
            pm.line_index,
            pm.extracted_price,
            p.name,
            p.country,
            p.synonyms,
            p.brand,
            p.model_number,
            p.color,
            p.storage,
            m.text,
            m.date
        FROM product_messages pm
        JOIN products p ON p.id = pm.product_id
        JOIN messages m ON m.id = pm.message_id
        WHERE {' AND '.join(filters)}
    """, params).fetchall()

    groups = {}
    for row in rows:
        groups.setdefault((int(row["message_id"]), int(row["line_index"])), []).append(row)

    deactivate_ids = []
    for group_rows in groups.values():
        if len(group_rows) <= 1:
            continue
        ranked = []
        for row in group_rows:
            source_line = get_binding_source_line(row["text"], row["line_index"])
            score = binding_line_identity_score(row, source_line) if source_line else 0
            ranked.append((
                score,
                datetime_sort_value(row["date"]),
                int(row["id"] or 0),
                row,
            ))
        keep = max(ranked, key=lambda item: (item[0], item[1], item[2]))[3]
        keep_id = int(keep["id"])
        deactivate_ids.extend(int(row["id"]) for row in group_rows if int(row["id"]) != keep_id)

    for offset in range(0, len(deactivate_ids), 500):
        chunk = deactivate_ids[offset:offset + 500]
        placeholders = ",".join("?" for _ in chunk)
        db.execute(f"UPDATE product_messages SET is_actual = 0 WHERE id IN ({placeholders})", chunk)
    return len(deactivate_ids)


def get_discount_section_confirmed_binding_ids(db, product_id=None, user_id=None):
    filters = ["pm.status = 'confirmed'"]
    params = []
    if product_id is not None:
        filters.append("pm.product_id = ?")
        params.append(product_id)
    if user_id is not None:
        filters.append("p.user_id = ?")
        params.append(user_id)

    rows = db.execute(f"""
        SELECT pm.id, pm.line_index, m.text
        FROM product_messages pm
        JOIN messages m ON pm.message_id = m.id
        JOIN products p ON pm.product_id = p.id
        WHERE {' AND '.join(filters)}
    """, params).fetchall()
    return [
        int(row["id"])
        for row in rows
        if line_index_is_in_discount_section(row["text"], row["line_index"])
    ]


def remove_discount_section_confirmed_bindings(db, product_id=None, user_id=None):
    delete_ids = get_discount_section_confirmed_binding_ids(db, product_id=product_id, user_id=user_id)
    if delete_ids:
        for offset in range(0, len(delete_ids), 500):
            chunk = delete_ids[offset:offset + 500]
            placeholders = ",".join("?" for _ in chunk)
            db.execute(f"DELETE FROM product_messages WHERE id IN ({placeholders})", chunk)
    return len(delete_ids)


def refresh_confirmed_bindings_from_latest_messages(db, product_id=None, user_id=None):
    if user_id is None and product_id is not None:
        owner = db.execute("SELECT user_id FROM products WHERE id = ?", (product_id,)).fetchone()
        if owner:
            user_id = owner["user_id"]

    active_filters = [
        "(m.is_blocked IS NULL OR m.is_blocked = 0)",
        "(m.is_delayed IS NULL OR m.is_delayed = 0)",
    ]
    active_params = []
    if user_id is not None:
        active_filters.append("m.user_id = ?")
        active_params.append(user_id)

    active_rows = db.execute(f"""
        SELECT m.id, m.text, m.date, m.chat_id, m.sender_name
        FROM messages m
        WHERE {' AND '.join(active_filters)}
        ORDER BY m.date DESC, m.id DESC
    """, active_params).fetchall()

    latest_fingerprints = {}
    for msg in active_rows:
        if not msg["text"]:
            continue
        lines = str(msg["text"]).split("\n")
        in_discount_section = False
        for index, line in enumerate(lines):
            if not line.strip():
                continue
            if line_is_discount_section_marker(line):
                in_discount_section = True
                continue
            if in_discount_section:
                continue
            if line_has_discount_condition_marker(line):
                continue
            fp = get_binding_fingerprint(
                msg["chat_id"],
                msg["sender_name"],
                msg["text"],
                index,
                include_price=False,
            )
            candidate = {
                "message_id": msg["id"],
                "line_index": index,
                "full_text": msg["text"],
                "line": line,
                "price": extract_price_with_context(lines, index),
                "date": msg["date"],
                "rank": (datetime_sort_value(msg["date"]), int(msg["id"] or 0)),
            }
            current = latest_fingerprints.get(fp)
            if current is None or candidate["rank"] > current["rank"]:
                latest_fingerprints[fp] = candidate

    if not latest_fingerprints:
        return 0

    binding_filters = ["pm.status = 'confirmed'"]
    binding_params = []
    if product_id is not None:
        binding_filters.append("pm.product_id = ?")
        binding_params.append(product_id)
    if user_id is not None:
        binding_filters.append("p.user_id = ?")
        binding_params.append(user_id)

    rows = db.execute(f"""
        SELECT pm.id, pm.product_id, pm.message_id, pm.line_index, pm.extracted_price, pm.is_actual,
               m.text, m.chat_id, m.sender_name,
               p.name, p.country, p.synonyms, p.brand, p.model_number, p.color, p.storage
        FROM product_messages pm
        JOIN messages m ON pm.message_id = m.id
        JOIN products p ON pm.product_id = p.id
        WHERE {' AND '.join(binding_filters)}
    """, binding_params).fetchall()

    changed = 0
    for row in rows:
        line = get_binding_source_line(row["text"], row["line_index"])
        if not line:
            continue
        latest = latest_fingerprints.get(
            get_binding_fingerprint(
                row["chat_id"],
                row["sender_name"],
                row["text"],
                row["line_index"],
                include_price=False,
            )
        )
        if not latest:
            continue
        latest_line = latest.get("line") or line
        if not binding_product_row_matches_line(row, latest_line, require_known_category=True):
            if int(row["is_actual"] or 0) != 0:
                db.execute("UPDATE product_messages SET is_actual = 0 WHERE id = ?", (row["id"],))
                changed += 1
            continue
        latest_price = latest["price"] if latest["price"] is not None else row["extracted_price"]
        if int(row["is_actual"] or 0) != 1:
            occupied = db.execute(
                """
                SELECT id FROM product_messages
                WHERE message_id = ?
                  AND line_index = ?
                  AND product_id <> ?
                  AND status = 'confirmed'
                  AND COALESCE(is_actual, 1) = 1
                LIMIT 1
                """,
                (latest["message_id"], latest["line_index"], row["product_id"]),
            ).fetchone()
            if occupied:
                continue
        needs_update = (
            int(row["message_id"] or 0) != int(latest["message_id"] or 0)
            or int(row["line_index"] or 0) != int(latest["line_index"] or 0)
            or int(row["is_actual"] or 0) != 1
            or row["extracted_price"] != latest_price
        )
        if not needs_update:
            continue
        try:
            db.execute(
                """
                UPDATE product_messages
                SET message_id = ?, line_index = ?, extracted_price = ?, is_actual = 1
                WHERE id = ?
                """,
                (latest["message_id"], latest["line_index"], latest_price, row["id"]),
            )
        except sqlite3.IntegrityError:
            db.execute("DELETE FROM product_messages WHERE id = ?", (row["id"],))
        changed += 1

    changed += deactivate_duplicate_actual_line_bindings(db, product_id=product_id, user_id=user_id)
    return changed


def deactivate_confirmed_bindings_from_inactive_messages(db, product_id=None, user_id=None):
    filters = [
        "pm.status = 'confirmed'",
        "COALESCE(pm.is_actual, 1) = 1",
        "(COALESCE(m.is_blocked, 0) = 1 OR COALESCE(m.is_delayed, 0) = 1)",
    ]
    params = []
    if product_id is not None:
        filters.append("pm.product_id = ?")
        params.append(product_id)
    if user_id is not None:
        filters.append("p.user_id = ?")
        params.append(user_id)

    rows = db.execute(f"""
        SELECT pm.id
        FROM product_messages pm
        JOIN messages m ON pm.message_id = m.id
        JOIN products p ON pm.product_id = p.id
        WHERE {' AND '.join(filters)}
    """, params).fetchall()
    ids = [row['id'] for row in rows]
    for offset in range(0, len(ids), 500):
        chunk = ids[offset:offset + 500]
        placeholders = ",".join("?" for _ in chunk)
        db.execute(f"UPDATE product_messages SET is_actual = 0 WHERE id IN ({placeholders})", chunk)
    return len(ids)







@app.route('/api/product_messages/confirm', methods=['POST'])
@login_required
def confirm_product_message():
    data = request.get_json()
    db = get_db()
    try:
        line_idx = data.get('line_index', -1)
        price = data.get('price')
        if price == '': 
            price = None

        product = db.execute(
            """
            SELECT id, name, country, synonyms, brand, model_number, color, storage,
                   warranty, description, description_html, specs
            FROM products
            WHERE id = ? AND user_id = ?
            """,
            (data['product_id'], session['user_id']),
        ).fetchone()
        message = db.execute(
            """
            SELECT m.id, m.text, m.chat_id, m.sender_name, m.chat_title, tc.custom_name
            FROM messages m
            LEFT JOIN tracked_chats tc ON tc.chat_id = m.chat_id AND tc.user_id = m.user_id
            WHERE m.id = ? AND m.user_id = ?
            """,
            (data['message_id'], session['user_id']),
        ).fetchone()
        if not product or not message:
            return jsonify({'error': 'Товар или сообщение не найдены'}), 404
        source_line = get_binding_source_line(message['text'], line_idx)
        if not product_message_variants_compatible(
            product['name'],
            product['country'],
            source_line,
            product_condition_text=build_product_condition_text(product),
        ):
            return jsonify({'error': 'Строка не подходит к стране/SIM-варианту товара'}), 400
        if not binding_product_row_matches_line(product, source_line, require_known_category=True):
            return jsonify({'error': 'Строка не подходит к модели/цвету товара'}), 400
        source_line, resolved_price = resolve_binding_display_line_and_price(
            message['text'],
            line_idx,
            price,
            require_price=False,
        )
        if not source_line:
            return jsonify({'error': 'Не удалось определить строку товара для привязки'}), 400
            
        # --- НОВАЯ СТРОГАЯ ПРОВЕРКА: Занята ли уже эта строка другим товаром? ---
        if line_idx != -1:
            existing = db.execute("""
                SELECT product_id FROM product_messages 
                WHERE message_id = ? AND line_index = ? AND COALESCE(is_actual, 1) = 1
            """, (data['message_id'], line_idx)).fetchone()
            
            if existing and existing['product_id'] != data['product_id']:
                return jsonify({'error': 'Эта строка уже привязана к другому товару!'}), 400
        # ------------------------------------------------------------------------

        same_item_binding = find_existing_confirmed_binding_by_fingerprint(
            db,
            data['product_id'],
            message['chat_id'],
            message['sender_name'],
            source_line,
            text=message['text'],
            line_index=line_idx,
            price=resolved_price,
        )
        if same_item_binding:
            db.execute(
                """
                DELETE FROM product_messages
                WHERE product_id = ? AND message_id = ? AND line_index = ? AND id != ?
                """,
                (data['product_id'], data['message_id'], line_idx, same_item_binding['id']),
            )
            db.execute(
                """
                UPDATE product_messages
                SET message_id = ?, line_index = ?, status = 'confirmed', extracted_price = ?, is_actual = 1
                WHERE id = ?
                """,
                (data['message_id'], line_idx, resolved_price, same_item_binding['id']),
            )
        else:
            db.execute("""
                INSERT INTO product_messages (product_id, message_id, line_index, status, extracted_price, is_actual) 
                VALUES (?, ?, ?, 'confirmed', ?, 1)
                ON CONFLICT(product_id, message_id, line_index) DO UPDATE SET status='confirmed', extracted_price=excluded.extracted_price, is_actual=1
            """, (data['product_id'], data['message_id'], line_idx, resolved_price))
        learning_stats = learn_supplier_line_binding(
            db,
            session['user_id'],
            product,
            source_line,
            message=message,
        )
        prune_duplicate_confirmed_bindings(db, product_id=data['product_id'])
        maybe_sync_product_sort_from_message(db, session['user_id'], data['product_id'], data['message_id'], line_idx)
        db.commit()
        notify_clients()
         
        return jsonify({'success': True, 'learning': learning_stats})
    except Exception as e:
        logger.error(f"Ошибка привязки: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/product_messages/detach', methods=['POST'])
@login_required
def detach_product_message():
    data = request.get_json()
    db = get_db()
    # Ищем, состоит ли сообщение в группе. Если да — удаляем всю группу
    pm = db.execute("SELECT id, group_id FROM product_messages WHERE product_id=? AND message_id=?", 
                    (data['product_id'], data['message_id'])).fetchone()
    if pm:
        if pm['group_id']:
            db.execute("DELETE FROM product_messages WHERE group_id = ?", (pm['group_id'],))
        else:
            db.execute("DELETE FROM product_messages WHERE id = ?", (pm['id'],))
        db.commit()
        notify_clients()
         
    return jsonify({'success': True})

@app.route('/api/product_messages/update_price', methods=['POST'])
@login_required
def update_product_price():
    data = request.get_json()
    db = get_db()
    db.execute("UPDATE product_messages SET extracted_price = ? WHERE id = ?", 
               (data['price'], data['pm_id']))
    db.commit()
    notify_clients()
     
    return jsonify({'success': True})

def clean_for_search(text):
    if not text: return set()
    t = str(text or "").replace("", " apple ")
    t = re.sub(r'(?i)\bwi\s*[- ]\s*fi\b', 'wifi', t)
    t = re.sub(r'(?i)\btype\s*[- ]\s*c\b', 'usb c', t)
    t = re.sub(r'(?i)\bdual[\s-]+sen[cs]e\b', 'dualsense', t)
    t = re.sub(r'(?i)\bplay\s*station\s*5\b', 'ps5', t)
    t = re.sub(r'(?i)\bplaystation\s*5\b', 'ps5', t)
    t = re.sub(r'(?i)\bcfi[\s-]*(\d{2,4})\s*([ab])(?:0?1)?\b', r'cfi\1\2', t)
    t = re.sub(r'[()*.[\]{},;!|/+\-]', ' ', t)
    t = re.sub(r'\s+', ' ', t.replace('\xa0', ' ').replace('ё', 'е')).strip().lower()
    t = t.translate(str.maketrans('асеокрх', 'aceokpx'))
    words = set(re.findall(r'[a-zа-я0-9]+', t))
    unit_aliases = {
        "mm": ("mm", "мм", "м"),
        "мм": ("mm", "мм", "м"),
        "м": ("mm", "мм", "м"),
        "gb": ("gb", "гб"),
        "гб": ("gb", "гб"),
        "tb": ("tb", "тб"),
        "тб": ("tb", "тб"),
    }
    for token in list(words):
        compact_unit = re.fullmatch(r"(\d+(?:[.,]\d+)?)(gb|гб|tb|тб|mm|мм|м)", token)
        if compact_unit:
            amount = compact_unit.group(1).replace(",", ".")
            unit = compact_unit.group(2)
            words.add(amount)
            words.add(unit)
            for alias in unit_aliases.get(unit, (unit,)):
                words.add(f"{amount}{alias}")
    for amount, unit in re.findall(r"\b(\d+(?:[.,]\d+)?)\s*(gb|гб|tb|тб|mm|мм|м)\b", t):
        amount = amount.replace(",", ".")
        words.add(amount)
        words.add(unit)
        for alias in unit_aliases.get(unit, (unit,)):
            words.add(f"{amount}{alias}")
    for prefix, number in re.findall(r"\b(watch|buds|flip|fold|tab|note)\s+(\d{1,2})\b", t):
        words.add(f"{prefix}{number}")
    for prefix, number in re.findall(r"\b(watch|buds|flip|fold|tab|note)(\d{1,2})\b", t):
        words.add(prefix)
        words.add(number)
        words.add(f"{prefix}{number}")
    for number in re.findall(r"\bz\s+flip\s*(\d{1,2})\b", t):
        words.add("zflip")
        words.add(f"zflip{number}")
        words.add(f"flip{number}")
    if re.search(r"\bbuds\s+core\b|\bbudscore\b", t):
        words.add("buds")
        words.add("core")
        words.add("budscore")
    if re.search(r"\bdualsense\b", t):
        words.add("dualsense")
        words.add("dual")
        words.add("sense")
    if re.search(r"\bps5\b", t):
        words.add("ps5")
        words.add("playstation")
    words.update(token.lower() for token in extract_catalog_model_tokens(text))
    for color in extract_match_colors(text):
        words.add(f"color_{color}")
    for connectivity in extract_match_connectivity(text):
        words.add(f"conn_{connectivity}")
        if connectivity == "cellular":
            words.add("cellular")
            words.add("lte")
        elif connectivity == "wifi":
            words.add("wifi")
    category = detect_match_category(text)
    if category:
        words.add(f"cat_{category}")
    return words


PRODUCT_SEARCH_ALIAS_GROUPS = [
    ['геймпад', 'гейм пад', 'геймад', 'джойстик', 'джой', 'джостик', 'контроллер', 'controller', 'gamepad', 'game pad', 'joystick'],
    ['наушники', 'наушник', 'гарнитура', 'гарнитуры', 'headset', 'headsets', 'headphones', 'headphone', 'earbuds', 'earphones'],
    ['эйрподс', 'аирподс', 'airpods', 'air pods'],
    ['колонка', 'колонки', 'акустика', 'акустическая система', 'спикер', 'speaker', 'speakers'],
    ['зарядка', 'зарядки', 'зарядное', 'зарядное устройство', 'зарядник', 'сзу', 'charger', 'charging', 'power adapter', 'адаптер питания', 'блок питания'],
    ['кабель', 'кабели', 'шнур', 'провод', 'cable', 'cord', 'wire'],
    ['чехол', 'чехлы', 'кейс', 'кейсы', 'накладка', 'бампер', 'case', 'cover', 'sleeve'],
    ['пленка', 'пленки', 'плёнка', 'плёнки', 'стекло', 'стекла', 'стекло защитное', 'защитное стекло', 'glass', 'film'],
    ['ноутбук', 'ноутбуки', 'ноут', 'лэптоп', 'лаптоп', 'laptop', 'notebook'],
    ['макбук', 'мак', 'macbook', 'mac'],
    ['смартфон', 'смартфоны', 'телефон', 'телефоны', 'мобильник', 'phone', 'smartphone'],
    ['айфон', 'айфоны', 'iphone'],
    ['pro', 'про'],
    ['max', 'макс', 'мах'],
    ['plus', 'плюс'],
    ['ultra', 'ультра'],
    ['mini', 'мини'],
    ['air', 'эйр', 'аир'],
    ['планшет', 'планшеты', 'tablet'],
    ['айпад', 'айпеды', 'ipad'],
    ['часы', 'часики', 'умные часы', 'смарт часы', 'смарт-часы', 'smartwatch', 'watch'],
    ['монитор', 'мониторы', 'display', 'screen', 'экран'],
    ['мышь', 'мышка', 'мышки', 'mouse'],
    ['клавиатура', 'клава', 'keyboard'],
    ['роутер', 'роутеры', 'маршрутизатор', 'маршрутизаторы', 'router'],
    ['адаптер', 'адаптеры', 'переходник', 'переходники', 'adapter', 'hub', 'usb hub'],
    ['пауэрбанк', 'повербанк', 'powerbank', 'power bank', 'внешний аккумулятор', 'внешний аккум', 'аккум', 'акб'],
    ['dualsense', 'dual sense', 'dual-sense', 'дуалсенс', 'дуал сенс', 'дуал-сенс'],
    ['joycon', 'joy con', 'joy-con', 'джойкон', 'джой кон', 'джой- кон'],
    ['vr', 'виртуальная реальность', 'шлем', 'очки', 'гарнитура vr', 'vision pro'],
    ['камера', 'экшн камера', 'экшн-камера', 'action camera', 'gopro'],
    ['apple', 'эпл', 'эппл'],
    ['samsung', 'самсунг', 'самс', 'galaxy', 'галакси'],
    ['xiaomi', 'сяоми', 'ксиаоми', 'mi', 'redmi', 'редми', 'poco', 'поко'],
    ['huawei', 'хуавей'],
    ['honor', 'хонор'],
    ['dyson', 'дайсон'],
    ['sony', 'сони'],
    ['asus', 'асус'],
    ['lenovo', 'леново'],
    ['hp', 'хп', 'эйчпи'],
    ['nintendo', 'нинтендо', 'switch', 'свитч'],
    ['playstation', 'play station', 'плейстейшен', 'плейстейшн', 'плойка', 'сонька', 'ps', 'ps4', 'ps5', 'пс4', 'пс5'],
    ['xbox', 'x box', 'иксбокс'],
    ['steam deck', 'steamdeck', 'стим дек', 'стимдек'],
    ['type c', 'type-c', 'typec', 'usb c', 'usb-c', 'usbc', 'тайп си', 'тайпси', 'юсб си', 'юсбси'],
]

PRODUCT_SEARCH_STOP_WORDS = {
    'для', 'на', 'под', 'к', 'ко', 'от', 'с', 'со', 'из', 'у', 'в', 'во', 'по', 'and', 'the', 'for', 'with', 'of'
}

PRODUCT_SEARCH_NORMALIZATION_REPLACEMENTS = [
    (re.compile(r"\b(?:steam|стим)\s*deck\b", re.IGNORECASE), 'steamdeck'),
    (re.compile(r"\bстим\s*дек\b", re.IGNORECASE), 'steamdeck'),
    (re.compile(r"\b(?:type|тайп)\s*-?\s*(?:c|си)\b", re.IGNORECASE), 'typec'),
    (re.compile(r"\b(?:usb|юсб)\s*-?\s*(?:c|си)\b", re.IGNORECASE), 'usbc'),
    (re.compile(r"\bplay\s+station\b", re.IGNORECASE), 'playstation'),
    (re.compile(r"\bплей\s*стейш[её]н\b", re.IGNORECASE), 'плейстейшн'),
    (re.compile(r"\bdual[\s-]+sense\b", re.IGNORECASE), 'dualsense'),
    (re.compile(r"\bдуал[\s-]*сенс\b", re.IGNORECASE), 'дуалсенс'),
    (re.compile(r"\bjoy[\s-]*con\b", re.IGNORECASE), 'joycon'),
    (re.compile(r"\bair[\s-]*pods\b", re.IGNORECASE), 'airpods'),
    (re.compile(r"\bpower\s+bank\b", re.IGNORECASE), 'powerbank'),
    (re.compile(r"\bgame\s+pad\b", re.IGNORECASE), 'gamepad'),
    (re.compile(r"\bsmart[\s-]*watch\b", re.IGNORECASE), 'smartwatch'),
    (re.compile(r"\be\s*sim\b", re.IGNORECASE), 'esim'),
    (re.compile(r"\bps\s+([45])\b", re.IGNORECASE), r'ps\1'),
    (re.compile(r"\bпс\s+([45])\b", re.IGNORECASE), r'пс\1'),
    (re.compile(r"\b(?:playstation|плейстейшн|плейстейшен|плойка|сонька)\s*([45])\b", re.IGNORECASE), r'ps\1'),
]

PRODUCT_ACCESSORY_SEARCH_HINTS = [
    'аксессуар', 'аксессуары', 'чехол', 'чехлы', 'кейс', 'кейсы', 'бампер', 'накладка', 'накладки',
    'ремешок', 'ремешки', 'браслет', 'браслеты', 'стекло', 'пленка', 'пленки', 'зарядка', 'зарядное',
    'кабель', 'кабели', 'адаптер', 'адаптеры', 'переходник', 'переходники', 'док станция', 'dock',
    'станция', 'держатель', 'holder', 'mount', 'cover', 'strap', 'charger', 'cable',
    'ssd', 'накопитель', 'дисковод', 'панель', 'панели', 'игра', 'игры', 'game', 'subscription', 'подписка'
]

PRODUCT_COMMON_STORAGE_NUMBERS = {'32', '64', '128', '256', '512', '1024', '2048'}

PRODUCT_SQL_LITERAL_SEARCH_EXPANSIONS = {
    'steamdeck': ['Steam Deck', 'steam deck'],
    'applewatch': ['Watch', 'Apple Watch', 'apple watch'],
    'typec': ['Type-C', 'Type C', 'type-c', 'type c'],
    'usbc': ['USB-C', 'USB C', 'usb-c', 'usb c'],
}

RU_TO_EN_KEYBOARD = str.maketrans({
    'й': 'q', 'ц': 'w', 'у': 'e', 'к': 'r', 'е': 't', 'н': 'y', 'г': 'u', 'ш': 'i', 'щ': 'o', 'з': 'p', 'х': 'h', 'ъ': '',
    'ф': 'a', 'ы': 's', 'в': 'd', 'а': 'f', 'п': 'g', 'р': 'h', 'о': 'j', 'л': 'k', 'д': 'l', 'ж': '', 'э': '',
    'я': 'z', 'ч': 'x', 'с': 'c', 'м': 'v', 'и': 'b', 'т': 'n', 'ь': 'm', 'б': '', 'ю': '',
})
EN_TO_RU_KEYBOARD = str.maketrans({
    'q': 'й', 'w': 'ц', 'e': 'у', 'r': 'к', 't': 'е', 'y': 'н', 'u': 'г', 'i': 'ш', 'o': 'щ', 'p': 'з',
    'a': 'ф', 's': 'ы', 'd': 'в', 'f': 'а', 'g': 'п', 'h': 'р', 'j': 'о', 'k': 'л', 'l': 'д',
    'z': 'я', 'x': 'ч', 'c': 'с', 'v': 'м', 'b': 'и', 'n': 'т', 'm': 'ь',
})


def normalize_product_search_text(value):
    normalized = str(value or '').lower().replace('ё', 'е')
    normalized = re.sub(r'[+"\'`«»]', ' ', normalized)
    normalized = re.sub(r'[\u2010-\u2015]', '-', normalized)
    for pattern, replacement in PRODUCT_SEARCH_NORMALIZATION_REPLACEMENTS:
        normalized = pattern.sub(replacement, normalized)
    normalized = re.sub(r'[^a-zа-я0-9]+', ' ', normalized)
    return re.sub(r'\s+', ' ', normalized).strip()


def keyboard_layout_variants(value):
    normalized = normalize_product_search_text(value)
    if not normalized:
        return []
    variants = []
    if re.search(r'[а-я]', normalized):
        variants.append(normalize_product_search_text(normalized.translate(RU_TO_EN_KEYBOARD)))
    if re.search(r'[a-z]', normalized):
        variants.append(normalize_product_search_text(normalized.translate(EN_TO_RU_KEYBOARD)))
    return [variant for variant in variants if variant and variant != normalized]


def tokenize_product_search_query(value, keep_stop_words=False):
    normalized = normalize_product_search_text(value)
    if not normalized:
        return []
    words = [word for word in normalized.split() if word]
    if keep_stop_words:
        return words
    meaningful = [word for word in words if word not in PRODUCT_SEARCH_STOP_WORDS]
    return meaningful or words


def should_expand_alias_group_by_term(source_term, alias_term):
    normalized_source = normalize_product_search_text(source_term)
    normalized_alias = normalize_product_search_text(alias_term)
    if not normalized_source or not normalized_alias:
        return False
    if normalized_source == normalized_alias:
        return True
    return len(normalized_source) >= 3 and normalized_alias.startswith(normalized_source)


PRODUCT_SEARCH_ALIAS_INDEX = {}
for alias_group in PRODUCT_SEARCH_ALIAS_GROUPS:
    normalized_group = [normalize_product_search_text(item) for item in alias_group]
    normalized_group = [item for item in normalized_group if item]
    for item in normalized_group:
        PRODUCT_SEARCH_ALIAS_INDEX[item] = normalized_group


def collect_product_search_terms(value):
    normalized = normalize_product_search_text(value)
    if not normalized:
        return []

    layout_variants = keyboard_layout_variants(normalized)
    result = {normalized, normalized.replace(' ', ''), *layout_variants}
    result.update(variant.replace(' ', '') for variant in layout_variants)
    words = [word for word in normalized.split() if word]
    alias_candidates = list(dict.fromkeys([normalized, *layout_variants, *words]))

    canonical_units = {
        'мм': ['мм', 'mm', 'м'],
        'mm': ['мм', 'mm', 'м'],
        'м': ['мм', 'mm', 'м'],
        'gb': ['gb', 'гб'],
        'гб': ['gb', 'гб'],
        'tb': ['tb', 'тб'],
        'тб': ['tb', 'тб'],
        'hz': ['hz', 'гц'],
        'гц': ['hz', 'гц'],
        'mah': ['mah', 'мач'],
        'мач': ['mah', 'мач'],
        'вт': ['вт', 'w'],
        'w': ['вт', 'w'],
    }

    for word in [*words, *layout_variants]:
        result.add(word)

        if len(word) > 4 and word.endswith(('ы', 'и')):
            result.add(word[:-1])
        if len(word) > 5 and word.endswith(('а', 'е', 'о', 'у', 'я')):
            result.add(word[:-1])

        compact_unit_match = re.match(r'^(\d+)(мм|mm|м|gb|гб|tb|тб|hz|гц|mah|мач|вт|w)$', word, re.IGNORECASE)
        if compact_unit_match:
            numeric_value = compact_unit_match.group(1)
            raw_unit = compact_unit_match.group(2).lower()
            result.add(numeric_value)
            for unit in canonical_units.get(raw_unit, [raw_unit]):
                result.add(f'{numeric_value}{unit}')
                result.add(f'{numeric_value} {unit}')

        for alias in PRODUCT_SEARCH_ALIAS_INDEX.get(word, []):
            result.add(alias)

    for alias_group in PRODUCT_SEARCH_ALIAS_GROUPS:
        if any(should_expand_alias_group_by_term(candidate, item) for candidate in alias_candidates for item in alias_group):
            for item in alias_group:
                normalized_item = normalize_product_search_text(item)
                if normalized_item:
                    result.add(normalized_item)

    return [item for item in result if item]


def build_derived_product_search_aliases(product):
    combined = normalize_product_search_text(' '.join([
        str(product.get('name') or ''),
        str(product.get('brand') or ''),
        str(product.get('model_number') or ''),
        str(product.get('synonyms') or ''),
    ]))

    aliases = set()
    product_name = str(product.get('name') or '').strip()
    brand = sanitize_brand_value(product.get('brand'))
    storage = normalize_capacity(product.get('storage'), allow_tb=True)
    is_gamepad = bool(re.search(r'(геймпад|джойстик|контроллер|controller|gamepad|joystick|dualsense|joycon)', combined))
    is_playstation = bool(re.search(r'(ps5|ps4|playstation|плейстейшн|пс5|пс4|плойка|сонька)', combined))
    is_xbox = bool(re.search(r'(xbox|иксбокс)', combined))
    is_switch = bool(re.search(r'(switch|nintendo|нинтендо|joycon)', combined))

    if 'dualsense' in combined or (is_gamepad and is_playstation):
        aliases.update({
            'джойстик для ps5',
            'джой для ps5',
            'геймпад для ps5',
            'контроллер ps5',
            'джойстик для плойки',
            'геймпад для playstation',
            'sony dualsense',
            'dual sense',
            'дуал сенс',
        })

    if is_gamepad and is_xbox:
        aliases.update({
            'джойстик xbox',
            'геймпад xbox',
            'контроллер xbox',
            'xbox controller',
        })

    if ('joycon' in combined or 'joy con' in combined) and is_switch:
        aliases.update({
            'джойкон',
            'джойстик для switch',
            'геймпад для nintendo switch',
            'joy-con',
        })
    elif is_gamepad and is_switch:
        aliases.update({
            'джойстик для switch',
            'геймпад switch',
            'контроллер nintendo switch',
        })

    if 'airpods' in combined:
        aliases.update({
            'беспроводные наушники apple',
            'наушники для iphone',
            'air pods',
        })

    if 'buds' in combined and ('samsung' in combined or brand == 'Samsung'):
        aliases.update({
            'samsung buds',
            'galaxy buds',
            'samsung galaxy buds',
            'беспроводные наушники samsung',
        })
        if 'buds4 pro' in combined:
            aliases.update({'buds 4 pro', 'galaxy buds 4 pro', 'samsung buds 4 pro'})
        elif 'buds4' in combined:
            aliases.update({'buds 4', 'galaxy buds 4', 'samsung buds 4'})
        elif 'buds3 fe' in combined:
            aliases.update({'buds 3 fe', 'galaxy buds 3 fe', 'samsung buds 3 fe'})
        elif 'buds core' in combined:
            aliases.update({'buds core', 'galaxy buds core', 'samsung buds core'})

    if brand == 'Apple' and 'iphone' in combined and storage:
        normalized_name = normalize_match_text(product_name)
        color_key = ''
        for key, color_aliases in MATCH_COLOR_ALIASES.items():
            if any(normalize_match_text(alias) in normalized_name for alias in color_aliases):
                color_key = key
                break
        family_match = re.search(r'iphone\s+\d+(?:\s+(?:plus|pro|max|mini|e))?', combined)
        if family_match:
            family = family_match.group(0)
            aliases.add(f'{family} {storage}')
            if color_key:
                for alias in MATCH_COLOR_ALIASES.get(color_key, ()):
                    alias_norm = normalize_product_search_text(alias)
                    if re.search(r'[a-z]', alias_norm):
                        aliases.add(f'{family} {storage} {alias_norm}')

    raw_product_blob = normalize_match_text(" ".join([
        product_name,
        str(product.get("color") or ""),
        str(product.get("storage") or ""),
    ]))
    ram_storage_match = re.search(r"\b(\d{1,2})\s*/\s*(\d{2,4})\s*(?:gb|гб)?\b", raw_product_blob)
    ram_storage = f"{ram_storage_match.group(1)}/{ram_storage_match.group(2)}" if ram_storage_match else ""
    storage_number = ram_storage_match.group(2) if ram_storage_match else ""

    color_aliases = []
    color_keys = extract_match_colors(product_name, product.get("color"))
    for color_key in color_keys:
        for alias in MATCH_COLOR_ALIASES.get(color_key, ()):
            alias_norm = normalize_product_search_text(alias)
            if alias_norm and re.search(r"[a-z]", alias_norm) and alias_norm not in color_aliases:
                color_aliases.append(alias_norm)

    def add_short_device_aliases(family_values):
        if not ram_storage:
            return
        normalized_families = []
        for family_value in family_values:
            family_norm = normalize_product_search_text(family_value)
            if family_norm and family_norm not in normalized_families:
                normalized_families.append(family_norm)
        for family_norm in normalized_families:
            aliases.add(f"{family_norm} {ram_storage}")
            if storage_number:
                aliases.add(f"{family_norm} {storage_number}")
            for color_alias in color_aliases[:8]:
                aliases.add(f"{family_norm} {ram_storage} {color_alias}")
                if storage_number:
                    aliases.add(f"{family_norm} {storage_number} {color_alias}")

    poco_match = re.search(
        r"\bpoco\s+([a-z]?\d+[a-z0-9]*(?:\s+(?:pro\s*max|pro\+?|pro|plus|max|ultra))*)(?:\s+5g)?\b",
        raw_product_blob,
    )
    if poco_match:
        poco_family = f"poco {poco_match.group(1)}"
        add_short_device_aliases((poco_family, f"{poco_family} 5g"))

    realme_match = re.search(
        r"\brealme\s+(\d{1,2}[a-z]?(?:\s+(?:pro\s*plus|pro\+|pro|plus|max|ultra))?)(?:\s+5g)?\b",
        raw_product_blob,
    )
    if realme_match:
        realme_family = f"realme {realme_match.group(1)}"
        add_short_device_aliases((realme_family, f"{realme_family} 5g"))

    pixel_match = re.search(
        r"\bpixel\s+(\d{1,2}a?(?:\s+(?:pro\s*xl|pro|xl|fold))?)(?:\s+5g)?\b",
        raw_product_blob,
    )
    if pixel_match:
        pixel_family = f"pixel {pixel_match.group(1)}"
        add_short_device_aliases((pixel_family, f"{pixel_family} 5g", f"google {pixel_family}"))

    if re.search(r'(powerbank|пауэрбанк|повербанк|внешний аккумулятор|magnetic wireless charging)', combined):
        aliases.update({
            'power bank',
            'внешний аккум',
            'пауэрбанк',
            'повербанк',
        })

    return list(aliases)


def bounded_levenshtein_distance(left, right, max_distance):
    if left == right:
        return 0
    if abs(len(left) - len(right)) > max_distance:
        return max_distance + 1

    previous = list(range(len(right) + 1))
    for left_index, left_char in enumerate(left, 1):
        current = [left_index]
        row_min = current[0]
        for right_index, right_char in enumerate(right, 1):
            cost = 0 if left_char == right_char else 1
            value = min(
                previous[right_index] + 1,
                current[right_index - 1] + 1,
                previous[right_index - 1] + cost,
            )
            current.append(value)
            row_min = min(row_min, value)
        if row_min > max_distance:
            return max_distance + 1
        previous = current
    return previous[-1]


def is_product_search_typo_match(token, variant):
    if token.isdigit() or variant.isdigit():
        return False
    if len(token) < 4 or len(variant) < 4 or len(token) > 24 or len(variant) > 24:
        return False
    if token[0] != variant[0]:
        return False
    shorter_len = min(len(token), len(variant))
    max_typos = 1 if shorter_len < 8 else 2
    return bounded_levenshtein_distance(token, variant, max_typos) <= max_typos


def build_product_search_typo_keys(value):
    normalized = normalize_product_search_text(value).replace(' ', '')
    if len(normalized) < 4 or len(normalized) > 24 or normalized.isdigit():
        return []
    keys = {normalized}
    for index in range(len(normalized)):
        keys.add(normalized[:index] + normalized[index + 1:])
    return list(keys)


def token_matches_product_search_variant(token, variant):
    normalized_token = normalize_product_search_text(token)
    normalized_variant = normalize_product_search_text(variant)
    if not normalized_token or not normalized_variant:
        return False
    if re.fullmatch(r'\d+', normalized_variant):
        if re.fullmatch(r'\d+', normalized_token):
            return normalized_token == normalized_variant
        if normalized_token.endswith(normalized_variant) and len(normalized_token) <= len(normalized_variant) + 1:
            return True
        return False
    if normalized_token == normalized_variant:
        return True
    if normalized_token in PRODUCT_SEARCH_STOP_WORDS:
        return False
    if len(normalized_token) < 2 or len(normalized_variant) < 2:
        return False
    return (
        normalized_token.startswith(normalized_variant)
        or normalized_token in normalized_variant
        or normalized_variant in normalized_token
        or is_product_search_typo_match(normalized_token, normalized_variant)
    )


def token_list_matches_product_search_variant(token_list, variant):
    normalized_variant = normalize_product_search_text(variant)
    if not normalized_variant:
        return False
    return any(token_matches_product_search_variant(token, normalized_variant) for token in (token_list or []))


def search_text_matches_variant(text, variant):
    normalized_text = normalize_product_search_text(text)
    normalized_variant = normalize_product_search_text(variant)
    if not normalized_variant:
        return False
    if re.fullmatch(r'\d+', normalized_variant):
        return any(token_matches_product_search_variant(token, normalized_variant) for token in normalized_text.split())
    haystack = f' {normalized_text} '
    if f' {normalized_variant} ' in haystack:
        return True
    if token_list_matches_product_search_variant(normalized_text.split(), normalized_variant):
        return True
    return normalized_variant in haystack


def has_explicit_storage_query_intent(words):
    return any(re.match(r'^(?:\d+\s*(?:gb|tb|гб|тб)|\d+(?:gb|tb|гб|тб)|gb|tb|гб|тб)$', normalize_product_search_text(word)) for word in (words or []))


def extract_product_search_numbers(value):
    normalized = normalize_product_search_text(value)
    return list(dict.fromkeys(re.findall(r'\b\d+\b', normalized)))


def extract_storage_search_numbers(*values):
    numbers = []
    seen = set()
    for value in values:
        normalized = normalize_product_search_text(value)
        for match in re.finditer(r'\b(\d+)\s*(gb|tb|гб|тб)\b', normalized):
            number = match.group(1)
            if number and number not in seen:
                seen.add(number)
                numbers.append(number)
    return numbers


def is_accessory_search_intent(words):
    for word in (words or []):
        variants = collect_product_search_terms(word)
        if any(hint in variant for variant in variants for hint in PRODUCT_ACCESSORY_SEARCH_HINTS):
            return True
    return False


def is_gamepad_search_intent(words):
    hints = {
        'геймпад', 'джойстик', 'джой', 'контроллер', 'controller', 'gamepad',
        'joystick', 'dualsense', 'dual sense', 'joycon', 'joy con', 'джойкон'
    }
    for word in (words or []):
        variants = collect_product_search_terms(word)
        if any(hint in variant for variant in variants for hint in hints):
            return True
    return False


def looks_like_device_model_query(words):
    normalized_words = [normalize_product_search_text(word) for word in (words or []) if normalize_product_search_text(word)]
    if not normalized_words:
        return False

    device_brands = {
        'iphone', 'айфон', 'apple', 'galaxy', 'samsung', 'xiaomi', 'redmi',
        'poco', 'honor', 'huawei', 'oneplus', 'pixel', 'realme'
    }
    device_modifiers = {'pro', 'max', 'plus', 'mini', 'ultra'}
    numeric_words = [word for word in normalized_words if re.fullmatch(r'\d+', word)]

    return (
        any(word in device_brands for word in normalized_words)
        or any(word in device_modifiers for word in normalized_words)
        or (len(numeric_words) == 1 and len(normalized_words) >= 2)
    )


def get_product_search_sort_tuple(entry):
    sort_index = entry.get('sort_index')
    sort_bucket = 1 if sort_index is None else 0
    sort_value = sort_index if sort_index is not None else 0
    display_name = str(entry.get('display_name') or '')
    return (sort_bucket, sort_value, display_name.lower(), int(entry.get('id') or 0))


def build_search_token_set(*values):
    tokens = []
    seen = set()
    for value in values:
        normalized = normalize_product_search_text(value)
        for token in normalized.split():
            if not token or token in PRODUCT_SEARCH_STOP_WORDS:
                continue
            if token in seen:
                continue
            seen.add(token)
            tokens.append(token)
    return frozenset(tokens)


def build_compound_product_search_terms(chunks):
    compounds = []
    seen = set()
    for chunk in chunks:
        words = [
            word
            for word in normalize_product_search_text(chunk).split()
            if len(word) >= 2 and word not in PRODUCT_SEARCH_STOP_WORDS
        ]
        for size in (2, 3):
            if len(words) < size:
                continue
            for index in range(0, len(words) - size + 1):
                compound = ''.join(words[index:index + size])
                if len(compound) < 4 or len(compound) > 36 or compound in seen:
                    continue
                seen.add(compound)
                compounds.append(compound)
    return compounds


def build_server_product_search_entry(product_row):
    product = dict(product_row)
    derived_aliases = build_derived_product_search_aliases(product)
    combined_search_source = normalize_product_search_text(' '.join([
        str(product.get('name') or ''),
        str(product.get('brand') or ''),
        str(product.get('model_number') or ''),
        str(product.get('synonyms') or ''),
        ' '.join(derived_aliases),
    ]))
    normalized_chunks = [
        normalize_product_search_text(chunk)
        for chunk in (
            product.get('name'),
            product.get('brand'),
            product.get('model_number'),
            product.get('country'),
            product.get('storage'),
            product.get('color'),
            product.get('synonyms'),
            ' '.join(derived_aliases),
        )
        if chunk
    ]
    compound_chunks = build_compound_product_search_terms(normalized_chunks)
    compact_chunks = [
        chunk.replace(' ', '')
        for chunk in normalized_chunks
        if 4 <= len(chunk.replace(' ', '')) <= 36
    ]
    unique_chunks = list(dict.fromkeys([chunk for chunk in [*normalized_chunks, *compound_chunks, *compact_chunks] if chunk]))
    search_blob = f" {' '.join(unique_chunks)} "
    token_list = list(dict.fromkeys([
        token
        for chunk in [*normalized_chunks, *compound_chunks, *compact_chunks]
        for token in chunk.split()
        if len(token) >= 2 and token not in PRODUCT_SEARCH_STOP_WORDS
    ]))
    search_name = normalize_product_search_text(product.get('name'))
    search_brand = normalize_product_search_text(product.get('brand'))
    search_model = normalize_product_search_text(product.get('model_number'))
    search_synonyms = normalize_product_search_text(' '.join([
        str(product.get('synonyms') or ''),
        *derived_aliases,
    ]))
    typo_chunks = [search_name, search_brand, search_model]
    typo_token_set = frozenset([
        token
        for chunk in [*typo_chunks, *build_compound_product_search_terms(typo_chunks)]
        for token in normalize_product_search_text(chunk).split()
        if 4 <= len(token) <= 18
        and token not in PRODUCT_SEARCH_STOP_WORDS
        and re.fullmatch(r'[a-zа-я]+', token)
    ])
    sort_index_raw = product.get('sort_index')
    try:
        sort_index = int(sort_index_raw) if sort_index_raw is not None else None
    except (TypeError, ValueError):
        sort_index = None

    return {
        'id': int(product.get('id') or 0),
        'folder_id': int(product.get('folder_id') or 0) or None,
        'sort_index': sort_index,
        'display_name': str(product.get('name') or ''),
        'search_name': search_name,
        'search_brand': search_brand,
        'search_model': search_model,
        'search_synonyms': search_synonyms,
        'search_blob': search_blob,
        'search_token_list': token_list,
        'search_token_set': frozenset(token_list),
        'search_typo_token_set': typo_token_set,
        'search_name_tokens': build_search_token_set(search_name),
        'search_brand_tokens': build_search_token_set(search_brand),
        'search_model_tokens': build_search_token_set(search_model),
        'search_synonyms_tokens': build_search_token_set(search_synonyms),
        'search_model_numbers': extract_product_search_numbers(' '.join([
            str(product.get('model_number') or ''),
        ])),
        'search_storage_numbers': extract_storage_search_numbers(
            product.get('storage'),
            product.get('name'),
            product.get('synonyms'),
        ),
        'search_is_ps5_console': bool(re.match(r'^(?:sony )?(?:playstation 5|ps5)\b', search_name)),
        'search_is_accessory': any(hint in combined_search_source for hint in PRODUCT_ACCESSORY_SEARCH_HINTS),
        'search_is_gamepad': bool(re.search(r'(геймпад|джойстик|контроллер|controller|gamepad|joystick|dualsense|joy\s*con|joycon)', combined_search_source)),
        'search_is_game': bool(re.search(r'(^| )игра( |$)|(^| )game( |$)', combined_search_source)),
        'search_is_phone': bool(re.search(r'(iphone|айфон|смартфон|phone|galaxy|redmi|oneplus|pixel|xiaomi|honor|huawei|realme)', combined_search_source)),
    }


def get_cached_product_search_entries(db, user_id):
    cached_index = PRODUCT_SEARCH_CACHE.get(user_id)
    if cached_index is not None:
        return cached_index

    rows = db.execute(
        """
        SELECT id, name, synonyms, brand, model_number, country, storage, color, folder_id, sort_index
        FROM products
        WHERE user_id = ?
        ORDER BY CASE WHEN sort_index IS NULL THEN 1 ELSE 0 END, sort_index, id
        """,
        (user_id,),
    ).fetchall()
    entries = [build_server_product_search_entry(row) for row in rows]
    entry_by_id = {int(entry['id']): entry for entry in entries}
    token_index = {}
    typo_index = {}
    for entry in entries:
        product_id = int(entry['id'])
        for token in entry.get('search_token_set') or ():
            token_index.setdefault(token, set()).add(product_id)
        for token in entry.get('search_typo_token_set') or ():
            for typo_key in build_product_search_typo_keys(token):
                typo_index.setdefault(typo_key, set()).add(product_id)

    cached_index = {
        'entries': entries,
        'entry_by_id': entry_by_id,
        'token_index': token_index,
        'typo_index': typo_index,
    }
    PRODUCT_SEARCH_CACHE[user_id] = cached_index
    return cached_index


def get_product_search_entries_for_folder_ids(db, user_id, folder_ids):
    normalized_folder_ids = [
        int(folder_id)
        for folder_id in (folder_ids or [])
        if str(folder_id).strip()
    ]
    normalized_folder_ids = list(dict.fromkeys(normalized_folder_ids))
    if not normalized_folder_ids:
        return []

    placeholders = ','.join('?' for _ in normalized_folder_ids)
    rows = db.execute(
        f"""
        SELECT id, name, synonyms, brand, model_number, country, storage, color, folder_id, sort_index
        FROM products
        WHERE user_id = ?
          AND folder_id IN ({placeholders})
        ORDER BY CASE WHEN sort_index IS NULL THEN 1 ELSE 0 END, sort_index, id
        """,
        (user_id, *normalized_folder_ids),
    ).fetchall()
    return [build_server_product_search_entry(row) for row in rows]


def build_product_search_needles(query):
    words = tokenize_product_search_query(query)
    if not words:
        return []
    needles = []
    for word in words:
        variants = set(collect_product_search_terms(word))
        variants.add(normalize_product_search_text(word))
        needles.append([variant for variant in variants if variant])
    return needles


def token_set_matches_product_search_variant(token_set, variant):
    normalized_variant = normalize_product_search_text(variant)
    if not normalized_variant:
        return False
    tokens = token_set or ()
    if normalized_variant in tokens:
        return True

    if normalized_variant.isdigit():
        for token in tokens:
            if token == normalized_variant:
                return True
            # A16 / i16 should match query "16", but router model WE1626 should not.
            if token.endswith(normalized_variant) and len(token) <= len(normalized_variant) + 3:
                prefix = token[:-len(normalized_variant)]
                if prefix.isalpha():
                    return True
        return False

    if len(normalized_variant) < 3:
        return False

    for token in tokens:
        if len(token) < 2 or token in PRODUCT_SEARCH_STOP_WORDS:
            continue
        if token.startswith(normalized_variant):
            return True
        if normalized_variant in token or token in normalized_variant:
            return True
        if is_product_search_typo_match(token, normalized_variant):
            return True
    return False


def normalized_search_field_matches(field, token_set, variant):
    normalized_variant = normalize_product_search_text(variant)
    if not normalized_variant:
        return False
    normalized_field = str(field or '')
    if f' {normalized_variant} ' in f' {normalized_field} ':
        return True
    if token_set_matches_product_search_variant(token_set, normalized_variant):
        return True
    if normalized_variant.isdigit():
        return False
    return len(normalized_variant) >= 3 and normalized_variant in normalized_field


def normalized_search_field_matches_fast(field, token_set, variant):
    normalized_variant = normalize_product_search_text(variant)
    if not normalized_variant:
        return False
    normalized_field = str(field or '')
    tokens = token_set or set()
    if normalized_variant in tokens:
        return True
    if f' {normalized_variant} ' in f' {normalized_field} ':
        return True
    if normalized_variant.isdigit():
        return False
    return len(normalized_variant) >= 3 and normalized_variant in normalized_field


def token_matches_product_search_variant_fast(token, variant):
    normalized_token = normalize_product_search_text(token)
    normalized_variant = normalize_product_search_text(variant)
    if not normalized_token or not normalized_variant:
        return False
    if normalized_token == normalized_variant:
        return True
    if normalized_variant.isdigit():
        if normalized_token.endswith(normalized_variant) and len(normalized_token) <= len(normalized_variant) + 3:
            prefix = normalized_token[:-len(normalized_variant)]
            return prefix.isalpha()
        return False
    if len(normalized_variant) < 3 or len(normalized_token) < 2:
        return False
    return (
        normalized_token.startswith(normalized_variant)
        or normalized_variant in normalized_token
        or normalized_token in normalized_variant
        or is_product_search_typo_match(normalized_token, normalized_variant)
    )


def get_candidate_ids_from_search_index(search_index, needles):
    if not isinstance(search_index, dict):
        return None
    token_index = search_index.get('token_index') or {}
    typo_index = search_index.get('typo_index') or {}
    if not token_index or not needles:
        return None

    candidate_sets = []
    token_items = list(token_index.items())
    for variants in needles:
        ids_for_word = set()
        normalized_variants = [
            normalize_product_search_text(variant)
            for variant in (variants or [])
            if normalize_product_search_text(variant)
        ]
        for variant in normalized_variants:
            direct_ids = token_index.get(variant)
            if direct_ids:
                ids_for_word.update(direct_ids)
        if ids_for_word:
            candidate_sets.append(ids_for_word)
            continue

        for variant in normalized_variants:
            typo_ids = set()
            for typo_key in build_product_search_typo_keys(variant):
                found_ids = typo_index.get(typo_key)
                if found_ids:
                    typo_ids.update(found_ids)
            if typo_ids:
                ids_for_word.update(typo_ids)
        if ids_for_word:
            candidate_sets.append(ids_for_word)
            continue

        for variant in normalized_variants:
            for token, product_ids in token_items:
                if token_matches_product_search_variant_fast(token, variant):
                    ids_for_word.update(product_ids)

        if not ids_for_word:
            return set()
        candidate_sets.append(ids_for_word)

    if not candidate_sets:
        return None
    candidates = candidate_sets[0]
    for ids_for_word in candidate_sets[1:]:
        candidates = candidates.intersection(ids_for_word)
        if not candidates:
            break
    return candidates


def product_entry_matches_search(entry, needles):
    if not needles:
        return True
    haystack = str(entry.get('search_blob') or '')
    token_set = entry.get('search_token_set') or set()

    for variants in needles:
        matched = False
        for variant in variants:
            normalized_variant = normalize_product_search_text(variant)
            if not normalized_variant:
                continue
            if normalized_search_field_matches(haystack, token_set, normalized_variant):
                matched = True
                break
        if not matched:
            return False
    return True


def get_product_entry_search_relevance(entry, needles, query_words):
    name = str(entry.get('search_name') or '')
    brand = str(entry.get('search_brand') or '')
    model = str(entry.get('search_model') or '')
    synonyms = str(entry.get('search_synonyms') or '')
    name_tokens = entry.get('search_name_tokens') or set()
    brand_tokens = entry.get('search_brand_tokens') or set()
    model_tokens = entry.get('search_model_tokens') or set()
    synonym_tokens = entry.get('search_synonyms_tokens') or set()
    normalized_phrase = normalize_product_search_text(' '.join(query_words or []))
    accessory_intent = is_accessory_search_intent(query_words)
    gamepad_intent = is_gamepad_search_intent(query_words)
    device_model_intent = looks_like_device_model_query(query_words)
    accessory_product = bool(entry.get('search_is_accessory'))
    gamepad_product = bool(entry.get('search_is_gamepad'))
    game_product = bool(entry.get('search_is_game'))
    phone_product = bool(entry.get('search_is_phone'))
    ps5_console_product = bool(entry.get('search_is_ps5_console'))
    ps5_intent = is_ps5_search_intent(query_words)
    dualsense_product = 'dualsense' in name or 'dual sense' in name
    numeric_words = [word for word in (query_words or []) if re.match(r'^\d+$', word)]
    has_storage_intent = has_explicit_storage_query_intent(query_words)
    model_numbers = entry.get('search_model_numbers') or []
    storage_numbers = entry.get('search_storage_numbers') or []

    score = 0
    name_hits = 0
    direct_hits = 0

    for variants in needles:
        has_name_hit = any(normalized_search_field_matches(name, name_tokens, variant) for variant in variants)
        has_brand_hit = any(normalized_search_field_matches(brand, brand_tokens, variant) for variant in variants)
        has_model_hit = any(normalized_search_field_matches(model, model_tokens, variant) for variant in variants)
        has_synonym_hit = any(normalized_search_field_matches(synonyms, synonym_tokens, variant) for variant in variants)

        if has_name_hit:
            score += 260
            name_hits += 1
            direct_hits += 1
        if has_brand_hit:
            score += 180
            direct_hits += 1
        if has_model_hit:
            score += 170
            direct_hits += 1
        if has_synonym_hit:
            score += 120

    if normalized_phrase:
        if normalized_search_field_matches(name, name_tokens, normalized_phrase):
            score += 280
        elif normalized_search_field_matches(model, model_tokens, normalized_phrase):
            score += 220
        elif normalized_search_field_matches(synonyms, synonym_tokens, normalized_phrase):
            score += 140

    if (not accessory_intent) and (not has_storage_intent) and len(numeric_words) == 1 and numeric_words[0] not in PRODUCT_COMMON_STORAGE_NUMBERS:
        query_number = numeric_words[0]
        has_model_number_hit = query_number in model_numbers or normalized_search_field_matches(model, model_tokens, query_number)
        has_storage_only_hit = query_number in storage_numbers and not has_model_number_hit

        if has_model_number_hit:
            score += 260
        if has_storage_only_hit:
            score -= 520 if accessory_product else 180

    if needles and name_hits == len(needles):
        score += 320
    if needles and direct_hits == len(needles):
        score += 180

    if accessory_intent:
        if accessory_product:
            score += 140
    else:
        score += -360 if accessory_product else 180

    if gamepad_intent:
        if gamepad_product:
            score += 980
        if ps5_intent:
            score += 3600 if dualsense_product else 200
        if game_product:
            score -= 1180
    elif ps5_intent and not accessory_intent:
        if ps5_console_product:
            score += 5000
        elif gamepad_product:
            score += 1300
        elif accessory_product or game_product:
            score -= 1400
        else:
            score -= 700

    if device_model_intent and not accessory_intent:
        if accessory_product:
            score -= 980
        if phone_product:
            score += 260

    return score


def get_quick_product_entry_search_relevance(entry, query_words):
    name = str(entry.get('search_name') or '')
    brand = str(entry.get('search_brand') or '')
    model = str(entry.get('search_model') or '')
    synonyms = str(entry.get('search_synonyms') or '')
    name_tokens = entry.get('search_name_tokens') or set()
    brand_tokens = entry.get('search_brand_tokens') or set()
    model_tokens = entry.get('search_model_tokens') or set()
    synonym_tokens = entry.get('search_synonyms_tokens') or set()
    normalized_words = [normalize_product_search_text(word) for word in (query_words or []) if normalize_product_search_text(word)]
    normalized_phrase = normalize_product_search_text(' '.join(normalized_words))
    accessory_intent = is_accessory_search_intent(normalized_words)
    accessory_product = bool(entry.get('search_is_accessory'))
    gamepad_intent = is_gamepad_search_intent(normalized_words)
    gamepad_product = bool(entry.get('search_is_gamepad'))
    game_product = bool(entry.get('search_is_game'))
    phone_product = bool(entry.get('search_is_phone'))
    ps5_console_product = bool(entry.get('search_is_ps5_console'))
    ps5_intent = is_ps5_search_intent(normalized_words)
    dualsense_product = 'dualsense' in name or 'dual sense' in name
    device_model_intent = looks_like_device_model_query(normalized_words)

    score = 0
    if normalized_phrase:
        if normalized_search_field_matches_fast(name, name_tokens, normalized_phrase):
            score += 420
        if normalized_search_field_matches_fast(brand, brand_tokens, normalized_phrase):
            score += 360
        if normalized_search_field_matches_fast(model, model_tokens, normalized_phrase):
            score += 300
        if normalized_search_field_matches_fast(synonyms, synonym_tokens, normalized_phrase):
            score += 160

    for word in normalized_words:
        if normalized_search_field_matches_fast(brand, brand_tokens, word):
            score += 180
        if normalized_search_field_matches_fast(name, name_tokens, word):
            score += 150
        if normalized_search_field_matches_fast(model, model_tokens, word):
            score += 130
        if normalized_search_field_matches_fast(synonyms, synonym_tokens, word):
            score += 70

    if accessory_intent:
        if accessory_product:
            score += 140
    else:
        score += -360 if accessory_product else 180

    if gamepad_intent:
        if gamepad_product:
            score += 980
        if ps5_intent:
            score += 3600 if dualsense_product else 200
        if game_product:
            score -= 1180
    elif ps5_intent and not accessory_intent:
        if ps5_console_product:
            score += 5000
        elif gamepad_product:
            score += 1300
        elif accessory_product or game_product:
            score -= 1400
        else:
            score -= 700

    if device_model_intent and not accessory_intent:
        if accessory_product:
            score -= 980
        if phone_product:
            score += 260

    return score


def search_products_in_entries(entries, query, folder_ids=None, limit=200):
    needles = build_product_search_needles(query)
    query_words = tokenize_product_search_query(query)
    if not needles:
        return []

    normalized_folder_ids = {int(folder_id) for folder_id in (folder_ids or set()) if folder_id}
    indexed_candidate_mode = False
    if isinstance(entries, dict):
        entry_by_id = entries.get('entry_by_id') or {}
        candidate_ids = get_candidate_ids_from_search_index(entries, needles)
        if candidate_ids is None:
            filtered_entries = list(entry_by_id.values())
        else:
            filtered_entries = [entry_by_id[product_id] for product_id in candidate_ids if product_id in entry_by_id]
            indexed_candidate_mode = True
    else:
        filtered_entries = entries

    if normalized_folder_ids:
        filtered_entries = [
            entry for entry in filtered_entries
            if int(entry.get('folder_id') or 0) in normalized_folder_ids
        ]

    ranked_entries = []
    use_quick_score = indexed_candidate_mode and len(filtered_entries) > 250
    verify_candidates = (not indexed_candidate_mode) or len(filtered_entries) <= 1000
    for entry in filtered_entries:
        if verify_candidates and not product_entry_matches_search(entry, needles):
            continue
        ranked_entries.append({
            'id': entry['id'],
            'score': get_quick_product_entry_search_relevance(entry, query_words) if use_quick_score else get_product_entry_search_relevance(entry, needles, query_words),
            'sort_tuple': get_product_search_sort_tuple(entry),
        })

    ranked_entries.sort(key=lambda item: (-item['score'], *item['sort_tuple']))
    if limit is None:
        return [item['id'] for item in ranked_entries]
    return [item['id'] for item in ranked_entries[:limit]]


def product_line_matches_synonym_sets(line_words, synonyms_as_word_sets):
    for syn_set in synonyms_as_word_sets or ():
        if not synonym_word_set_is_informative(syn_set):
            continue
        if syn_set.issubset(line_words):
            if "pro" not in syn_set and "pro" in line_words:
                continue
            if "max" not in syn_set and "max" in line_words:
                continue
            if "plus" not in syn_set and "plus" in line_words:
                continue
            if "ultra" not in syn_set and "ultra" in line_words:
                continue
            if "mini" not in syn_set and "mini" in line_words:
                continue
            if "air" not in syn_set and "air" in line_words:
                continue
            if "fe" not in syn_set and "fe" in line_words:
                continue
            return True
    return False


MATCH_LOW_SIGNAL_TOKENS = {
    "apple", "samsung", "смартфон", "телефон", "моноблок", "ноутбук", "планшет",
    "gb", "гб", "tb", "тб", "sim", "esim", "e", "only", "global", "version",
    "версия", "китай", "сша", "usa", "us", "eu", "jp", "hk", "cn",
    "mm", "мм", "м", "умные", "часы", "титан",
    "black", "white", "silver", "blue", "red", "green", "pink", "purple",
    "orange", "yellow", "gold", "natural", "desert", "titanium", "midnight",
    "starlight", "черный", "черная", "черное", "чёрный", "чёрная", "чёрное",
    "белый", "белая", "белое", "синий", "синяя", "синее", "голубой",
    "голубая", "серый", "серая", "серебристый", "серебристая", "серебро",
    "красный", "красная", "зеленый", "зелёный", "зеленая", "розовый",
    "фиолетовый", "оранжевый", "желтый", "жёлтый", "золотой", "натуральный",
    "пустынный", "темная", "тёмная", "ночь",
}
MATCH_LOOKALIKE_TRANSLATION = str.maketrans('асеокрх', 'aceokpx')
MATCH_LOW_SIGNAL_SEARCH_TOKENS = {
    str(token).translate(MATCH_LOOKALIKE_TRANSLATION)
    for token in MATCH_LOW_SIGNAL_TOKENS
}

PRODUCT_SIGNATURE_DROP_TOKENS = {
    "apple", "смартфон", "телефон", "ноутбук", "моноблок", "планшет", "гаджет",
    "macbook", "макбук", "ipad", "айпад", "gb", "гб", "tb", "тб", "ssd",
    "cpu", "gpu", "core", "cores", "ram", "with", "and", "для", "the",
    "cellular", "version", "global", "only", "sim", "esim", "nano",
    "samsung", "huawei", "dyson", "xiaomi", "honor", "google", "sony",
    "умные", "часы", "выпрямитель", "волос", "беспроводные", "наушники",
}

MATCH_COLOR_ALIASES = {
    "black": ("black", "черный", "черная", "черное", "чёрный", "чёрная", "чёрное", "obsidian", "jetblack", "jet black", "nebula noir", "charcoal black", "🖤", "⚫", "⚫️"),
    "white": ("white", "белый", "белая", "белое", "blanco", "🤍"),
    "silver": ("silver", "серебристый", "серебристая", "серебро", "серебряный", "astral trail", "🤍", "🩶"),
    "gray": ("gray", "grey", "gris", "серый", "серая", "графит", "графитовый", "graphite", "graphite gray", "charcoal", "hazel", "moonstone", "storm titanium", "suede grey", "titanium colour", "🩶"),
    "space_gray": ("space gray", "space grey", "spacegray", "серый космос", "космический серый"),
    "space_black": ("space black", "черный космос", "чёрный космос"),
    "blue": ("blue", "синий", "синяя", "синее", "голубой", "голубая", "mist blue", "sky blue", "ice blue", "iceblue", "icyblue", "light blue", "blue shadow", "navy", "dark blue", "небесно", "indigo", "💙", "🔵"),
    "teal": ("teal", "бирюзовый", "бирюзовая"),
    "ultramarine": ("ultramarine", "ультрамарин", "ультрамариновый", "ультрамариновая"),
    "red": ("red", "красный", "красная", "nebula red"),
    "green": ("green", "зеленый", "зелёный", "зеленая", "зелёная", "mint", "mint green", "mint breeze", "мятный", "мятная", "мятно-зеленый", "мятно-зелёный"),
    "pink": ("pink", "розовый", "розовая", "dusty pink", "пыльно-розовый", "rose gold", "rose", "розовое", "🩷"),
    "purple": ("purple", "фиолетовый", "фиолетовая", "lilac", "lavender", "lavander", "лавандовый", "лавандовая", "violet", "cobalt violet", "💜"),
    "orange": ("orange", "оранжевый", "оранжевая", "apricot", "абрикосовый", "абрикосовая", "🧡", "🍊"),
    "yellow": ("yellow", "желтый", "жёлтый", "желтая", "жёлтая", "citrus", "⭐", "⭐️"),
    "gold": ("gold", "golden", "золотой", "золотая", "золотистый", "золотистая", "⭐", "⭐️"),
    "slate": ("slate", "сланцевый", "сланцевая"),
    "brown": ("brown", "terracotta", "tan", "коричневый", "коричневая", "терракотовый", "терракотовая"),
    "amber": ("amber", "amber silk", "янтарный", "янтарная"),
    "copper": ("copper", "медный", "медная"),
    "nickel": ("nickel", "никель"),
    "topaz": ("topaz", "топаз"),
    "natural": ("natural", "натуральный", "натуральная"),
    "desert": ("desert", "пустынный", "пустынная"),
    "titanium": ("titanium", "storm titanium", "titanium colour", "титановый", "титановая"),
    "midnight": ("midnight", "темная ночь", "тёмная ночь"),
    "starlight": ("starlight", "сияющая звезда", "⭐", "⭐️"),
}
MATCH_COLOR_ALIAS_SEARCH_TOKENS = {
    token
    for aliases in MATCH_COLOR_ALIASES.values()
    for alias in aliases
    for token in re.findall(
        r"[a-zа-я0-9]+",
        str(alias or "").replace("\xa0", " ").replace("ё", "е").lower().translate(MATCH_LOOKALIKE_TRANSLATION),
    )
    if token and not token.startswith("color_")
}

MATCH_CATEGORY_PATTERNS = (
    ("display", (r"\bstudio\s*display\b", r"\bpro\s*display\b", r"\bmonitor\b", r"\bмонитор\b")),
    ("imac", (r"\b(?:imac|аймак)\b", r"\bмоноблок\b")),
    ("macbook", (
        r"\bmacbook\b",
        r"\b(?:air|pro|neo)\s+1[1-9]\b.*\b[a-z]{2,4}[a-z0-9]{0,5}\d[a-z0-9]{0,5}\b",
        r"\b(?:air|pro)\s+1[1-9]\b.*\bm[1-9]\b",
        r"\bm[1-9]\b.*\b(?:air|pro)\s+1[1-9]\b",
        r"\bm[1-9]\b.*\b(?:cpu|gpu|core|ram)\b",
        r"\b\d+\s*(?:cpu|gpu)\b",
    )),
    ("ipad", (r"\bipad\b",)),
    ("apple_watch", (r"\bapple\s*watch\b", r"(?<!\w)(?:s|se|ul|ultra)\s*\d{1,2}\s+\d{2}\b")),
    ("airpods", (r"\bair\s*pods?\b", r"\bairpods?\b")),
    ("apple_keyboard", (r"\bmagic\s*keyboard\b",)),
    ("apple_mouse", (r"\bmagic\s*mouse\b",)),
    ("iphone", (
        r"\b(?:iphone|айфон)\b",
        r"(?<!\w)(?:1[67]e|17e|16e|se|1[1-7])\s+\d{2,4}\s*(?:gb|tb|гб|тб)\b",
        r"(?<!\w)(?:1[67]e|17e|16e|se|1[1-7])\s+(?:pro\s+max|pro|max|plus|mini|air)(?:\s+\d{2,4}\s*(?:gb|tb|гб|тб)?)?\b",
    )),
    ("apple_trackpad", (r"\bmagic\s*trackpad\b", r"\bтрекпад\b")),
    ("samsung_phone", (r"\bgalaxy\s+s\d{1,2}\b", r"\bz\s*(?:fold|flip)\s*\d*\b")),
    ("garmin_watch", (
        r"\bgarmin\b.*\b(?:fenix|forerunner|venu|epix|tactix|instinct|insticnt|approach|enduro|quatix|lily|vivoactive|descent|marq)\b",
    )),
    ("samsung_tablet", (r"\bgalaxy\s+tab\b", r"\btab\s+[as]\d{1,2}\b")),
    ("samsung_watch", (r"\bgalaxy\s+watch\s*\d*\b", r"\bwatch\s*\d+\s+\d{2}\b")),
    ("samsung_earbuds", (r"\bgalaxy\s+buds\s*\d*\b", r"\bbuds\s*(?:core|\d+(?:\s*(?:fe|pro))?)?\b", r"\bsm[-\s]?r\d{3}\b")),
    ("xiaomi_tablet", (
        r"\b(?:xiaomi|redmi|poco)\s+pad\b",
        r"\bredmi\s+pad\s*\d*(?:\s+pro|\s+4g)?\b",
        r"\bpoco\s+pad(?:\s+[a-z]\d+)?\b",
        r"\bxiaomi\s+pad\s*\d+(?:\s+pro)?\b",
    )),
    ("xiaomi_phone", (
        r"\bredmi\s+note\s*\d+(?:\s+pro(?:\s*\+|plus)?\s*5g|\s+pro(?:\s*\+|plus)?|\s+5g)?\b",
        r"(?<!\w)note\s+\d+(?:\s+pro(?:\s*\+|plus)?\s*5g|\s+pro(?:\s*\+|plus)?|\s+5g)?\b",
        r"\bpoco\s+[xmc]\d+(?:\s+pro(?:\s+max)?|\s+5g)?\b",
        r"\bxiaomi\s+\d{2}(?:t(?:\s+pro)?|\s+ultra)?\b",
    )),
    ("xiaomi_vacuum", (
        r"\b(?:xiaomi\s+)?dreame(?:\s+bot)?\s+[dflrwx]\d+[a-z]*(?:\s+(?:plus|pro|prime|ultra|complete|master|lite|essential|max(?:\s+gen\s*\d+)?))?\b",
        r"\b(?:xiaomi\s+)?roborock\s+(?:q\s*revo(?:\s+c)?|q\d+\s+max(?:\s+pro)?|saros\s+z\d+|s\d+)\b",
        r"\b(?:dreame|roborock)\b.*\b(?:vacuum|пылесос|robot|робот)\b",
    )),
    ("google_phone", (r"\bpixel\s+\d+(?:a|\s+pro(?:\s+xl|\s+fold)?|)\b",)),
    ("realme_phone", (r"\brealme\s+\d+(?:\s+pro)?\s*5g\b",)),
    ("huawei_phone", (r"\bhuawei\b.*\b(?:mate|pura|nova)\b",)),
    ("playstation_gamepad", (
        r"\bdual[\s-]*sen[cs]e\b",
        r"\bdualsense\b",
    )),
    ("playstation_console", (
        r"\b(?:sony\s+)?playstation\s*5\b",
        r"\bcfi[\s-]?\d{2,4}\s*[ab](?:0?1)?\b",
    )),
    ("dyson_airstrait", (r"\bairstrait\b", r"\bht[-\s]?01\b")),
    ("insta360", (r"\binsta\s*360\b",)),
    ("earbuds", (r"\bbuds?\b", r"\bearbuds?\b", r"\bнаушник",)),
    ("keyboard", (r"\bkeyboard\b", r"\bклавиатур",)),
    ("camera", (
        r"\bcanon\b",
        r"\bnikon\b",
        r"\bsony\b.*\b(?:alpha|zv[-\s]?e\d+|dsc[-\s]?zv1f?\b|a[67]\d{3}|a9\b|a1\b|ilce[-\s]?\d+|fx\d{1,2}|pz\s+e\s+\d{1,3}\s*-\s*\d{1,3}\s*mm)\b",
        r"\bfujifilm\b",
        r"\bolympus\b",
        r"\bom\s+system\b",
        r"\bpanasonic\b",
        r"\blumix\b",
        r"\bkodak\b",
        r"\bblackmagic\b",
        r"\bsigma\b",
        r"\btamron\b",
        r"\btokina\b",
        r"\bsamyang\b",
        r"\bnikkor\b",
        r"\beos\b",
        r"\bpowershot\b",
        r"\bpixpro\b",
        r"\bmirrorless\b",
        r"\bcamera\b",
        r"\bzv[-\s]?e\d+\b",
        r"\bdsc[-\s]?zv1f?\b",
        r"\ba[67]\d{3}\b",
        r"\bpz\s+e\s+\d{1,3}\s*-\s*\d{1,3}\s*mm\b",
        r"\bmark\s*[ivx]+\b",
        r"\bgopro\b",
        r"\bdji\s+osmo\b",
    )),
    ("purifier", (r"\bpurifier\b", r"\bочистител",)),
)

AUTO_CONFIRM_SOURCE_MARKERS = ("nistone", "гузун")
TELEGRAM_FLAG_PAIR_RE = re.compile(r"[\U0001F1E6-\U0001F1FF]{2}")
APPLE_COMPUTER_CHIP_CONTEXT_RE = re.compile(
    r"\bm[1-9]\b.*\b(?:cpu|gpu|core|ram)\b|\b\d+\s*(?:cpu|gpu)\b",
    re.IGNORECASE,
)


def synonym_word_set_is_informative(syn_set):
    informative = []
    for token in syn_set or ():
        normalized = str(token or "").strip().lower()
        if not normalized or normalized in MATCH_LOW_SIGNAL_TOKENS:
            continue
        if len(normalized) < 2:
            continue
        informative.append(normalized)
    if len(informative) >= 2:
        return True
    return any(re.fullmatch(r"[a-z]{1,4}\d{2,5}[a-z0-9/]*", token) for token in informative)


def normalize_match_text(value):
    text = str(value or "").replace("\xa0", " ").replace("ё", "е").casefold()
    return re.sub(r"\s+", " ", text).strip()


SUPPLIER_DECORATION_EMOJI_RE = re.compile(
    r"[\U0001F000-\U0001FAFF\u2600-\u27BF\uFE0F\u200D]+",
    re.UNICODE,
)


def strip_supplier_line_decorations(value):
    text = str(value or "").replace("\xa0", " ").strip()
    if not text:
        return ""
    text = TELEGRAM_FLAG_PAIR_RE.sub(" ", text)
    text = SUPPLIER_DECORATION_EMOJI_RE.sub(" ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip(" \t-–—:•·")


def line_is_discount_section_marker(line):
    normalized = normalize_match_text(line)
    if extract_price(line):
        return False
    return bool(re.search(
        r"\b(?:"
        r"у\s*ц\s*е\s*н\s*к\w*|уценк\w*|markdown|clearance|"
        r"дефект\w*|defect(?:ive)?|брак\w*|"
        r"чуть\s+мят(?!н)\w*|мят(?!н)\w*|"
        r"as[\s-]*is|asis|cpo|сро|refurb(?:ished)?|pre[\s-]*owned"
        r")\b",
        normalized,
    ))


SUPPLIER_CONDITION_DISCOUNT_RE = re.compile(
    r"\b(?:"
    r"уценк\w*|markdown|clearance|обменк\w*|распак\w*|"
    r"предактив\w*|актив\b|активк\w*|активирован\w*|активаци\w*|"
    r"от\s*:?\s*\d+\s*шт|"
    r"as[\s-]*is|asis|cpo|сро|refurb(?:ished)?|pre[\s-]*owned|"
    r"мят\w*\s+короб\w*|"
    r"чуть\s+мят(?!н)\w*|мят(?!н)\w*|"
    r"мят\w*\s+уг\w*|"
    r"гравиров\w*|engraving|engraved|"
    r"витрин\w*|open[\s-]*box|used|"
    r"потерт\w*|потертост\w*|косметическ\w*|"
    r"отсутств\w+\s+(?:кабел\w*|насад\w*|комплект\w*|короб\w*)|"
    r"замен\w+\s+(?:диспле\w*|экран\w*|аккумулятор\w*|батаре\w*)|"
    r"пробит\w*\s+короб\w*|"
    r"надорван\w*|порван\w*|дефект\w*|defect(?:ive)?|брак\w*|неисправ\w*|нерабоч\w*|"
    r"царап\w*|скол\w*|трещин\w*"
    r")\b",
    re.IGNORECASE,
)


def line_has_discount_condition_marker(line):
    normalized = normalize_match_text(line)
    return bool(SUPPLIER_CONDITION_DISCOUNT_RE.search(normalized))


def line_index_is_in_discount_section(text, line_index):
    try:
        target_index = int(line_index)
    except (TypeError, ValueError):
        return False
    if target_index < 0:
        return False

    in_discount_section = False
    for index, line in enumerate(str(text or "").split("\n")):
        if line_is_discount_section_marker(line):
            in_discount_section = True
        if index == target_index:
            return in_discount_section or line_has_discount_condition_marker(line)
    return False


def detect_match_brand(value):
    raw = str(value or "").replace("", " Apple ")
    normalized = normalize_match_text(raw)
    if not normalized:
        return ""
    if APPLE_PART_NUMBER_PATTERN.search(raw):
        return "Apple"
    apple_markers = (
        "iphone", "айфон", "ipad", "imac", "macbook", "apple watch",
        "airpods", "air pods", "magic keyboard", "magic mouse", "mac mini",
    )
    if any(marker in normalized for marker in apple_markers) or looks_like_iphone_shorthand(raw):
        return "Apple"
    if re.search(
        r"\b(?:galaxy\s+)?buds\s*(?:core|\d+(?:\s*(?:fe|pro))?)?\b|"
        r"\b(?:galaxy\s+)?watch\s*\d*\b|"
        r"\bsm[-\s]?[lr]\d{3,4}\b",
        normalized,
    ):
        return "Samsung"
    if re.search(
        r"\b(?:redmi|poco)\s+pad\b|"
        r"\bxiaomi\s+pad\s*\d+\b|"
        r"\b(?:xiaomi\s+)?(?:dreame|roborock)\b|"
        r"(?<!\w)note\s+\d+(?:\s+pro(?:\s*\+|plus)?(?:\s+5g)?|\s+5g)?\b|"
        r"\bpoco\s+[xmc]\d+\b|"
        r"\bxiaomi\s+\d{2}(?:t(?:\s+pro)?|\s+ultra)?\b",
        normalized,
    ):
        return "Xiaomi"
    for brand, markers in BRAND_KEYWORDS:
        if any(brand_marker_in_text(normalized, marker) for marker in markers):
            return brand
    return ""


BRAND_MATCH_EQUIVALENT_GROUPS = (
    {"Sony", "PlayStation"},
    {"Microsoft", "Xbox"},
    {"Meta", "Oculus"},
)


def brands_are_compatible_for_matching(left, right):
    left = sanitize_brand_value(left)
    right = sanitize_brand_value(right)
    if not left or not right:
        return True
    if left.casefold() == right.casefold():
        return True
    return any(left in group and right in group for group in BRAND_MATCH_EQUIVALENT_GROUPS)


def detect_match_category(value):
    normalized = normalize_match_text(strip_supplier_line_noise(value))
    if not normalized:
        return ""
    if looks_like_apple_watch_shorthand(normalized):
        return "apple_watch"
    if re.search(r"\bdual[\s-]*sen[cs]e\b|\bdualsense\b", normalized):
        if re.search(r"\b(?:заряд\w*|charging|charge|станц\w*|station|module|модул\w*|dock)\b", normalized):
            return "playstation_accessory"
        return "playstation_gamepad"
    if re.search(r"\b(?:meta\s+quest|oculus\s+quest|quest\s*3s?|quest\s*2)\b", normalized):
        if re.search(r"\b(?:case|strap|cover|grip|interface|stand|кабел|чех|наклад|креплен|подстав)\b", normalized):
            return "vr_accessory"
        return "vr_headset"
    if re.search(r"\bsteam\s*deck\b|\bsteamdeck\b", normalized):
        if re.search(r"\b(?:dock|case|cover|plate|modcase|charging|заряд|чех|наклад|панел)\b", normalized):
            return "steam_deck_accessory"
        return "steam_deck"
    if re.search(r"\b(?:x\s*box|xbox)\s+(?:series\s+)?[xs]\b|\bxbox\b", normalized):
        if re.search(r"\b(?:геймпад|controller|gamepad|джой|robot|carbon|pulse|deep|shock|velocity|camo|elite)\b", normalized):
            return "xbox_gamepad"
        if re.search(r"\b(?:series\s+)?[xs]\b|\b1\s*tb\b|\b512\s*(?:gb|гб)?\b|\bdigital\b|\bdisk\b|\bdisc\b", normalized):
            return "xbox_console"
    if re.search(r"\b(?:nintendo\s+)?switch\b|\bnsw2?\b", normalized):
        if re.search(r"\b(?:case|protector|screen|карта памяти|sd|чех|стекл|защит)\b", normalized):
            return "nintendo_accessory"
        if re.search(r"\b(?:joy[\s-]*con|controller|gamepad|геймпад|джой)\b", normalized):
            return "nintendo_gamepad"
        return "nintendo_console"
    if re.search(r"\bps\s*5\b|\bпс\s*5\b|\bplaystation\s*5\b", normalized):
        if re.search(r"\b(?:vr\s*2|vr2|portal|pulse|дисковод|disc\s*drive|disk\s*drive|vertical\s*stand|подстав|charging|заряд|station)\b", normalized):
            return "playstation_accessory"
        if re.search(r"\b(?:pro|slim|digital|disk|disc|cfi|без\s+диска|с\s+диском|\d+\s*(?:tb|тб|gb|гб))\b", normalized):
            return "playstation_console"
    if re.search(r"\bipad\b", normalized) or re.search(r"\b(?:air|pro)\s+(?:11|13)\s+m[1-9]\b.*\bwi\s*-?\s*fi\b", normalized):
        return "ipad"
    if re.search(r"\bmini\s*[67]\b", normalized) and re.search(r"\b(?:wi\s*-?\s*fi|wifi|cellular|lte|64|128|256|512)\b", normalized):
        return "ipad"
    detected_brand = detect_match_brand(normalized)
    if detected_brand == "Garmin" and re.search(r"\b(?:fenix|forerunner|venu|epix|tactix|instinct|insticnt|approach|enduro|quatix|lily|vivoactive|descent|marq)\b", normalized):
        return "garmin_watch"
    if re.search(r"\bgalaxy\s+s\d{1,2}\b", normalized) or re.search(r"\bz\s*(?:fold|flip)\s*\d*\b", normalized):
        return "samsung_phone"
    if re.search(r"(?<!\w)s\d{1,2}\+?\b", normalized):
        if detected_brand == "Samsung":
            return "samsung_phone"
        if re.search(r"\b(?:galaxy|ultra|plus|fe|max|5g|смартфон|phone|\d{2,4}\s*(?:gb|tb|гб|тб))\b", normalized):
            return "samsung_phone"
    for category, patterns in MATCH_CATEGORY_PATTERNS:
        if category == "iphone" and detected_brand and detected_brand != "Apple":
            continue
        if any(re.search(pattern, normalized, re.IGNORECASE) for pattern in patterns):
            return category
    return ""


APPLE_WATCH_CASE_SIZE_RE = re.compile(r"(?<!\d)(38|40|41|42|44|45|46|49)\s*(?:mm|мм)?\b", re.IGNORECASE)


def looks_like_apple_watch_shorthand(value):
    normalized = normalize_match_text(value)
    watch_size = r"(?:38|40|41|42|44|45|46|49)"
    size_gap = r"\s+(?:\(?20[1-3]\d\)?\s+)?"
    return bool(
        re.search(rf"(?<!\w)(?:se\s*[23]|s[0-9]{{1,2}}){size_gap}{watch_size}\b", normalized)
        or re.search(rf"(?<!\w)(?:ul|ultra)\s*\d?{size_gap}{watch_size}\b", normalized)
        or re.search(r"(?<!\w)(?:ul|ultra)\s*[23]\b", normalized)
    )


def extract_apple_watch_family(*values):
    normalized = normalize_match_text(" ".join(str(value or "") for value in values))
    if not normalized:
        return ""
    if re.search(r"\bultra\s*3\b|\bul\s*3\b", normalized):
        return "ultra3"
    if re.search(r"\bultra\s*2\b|\bul\s*2\b", normalized):
        return "ultra2"
    if re.search(r"\bultra\b|\bul\s+(?:49|46|45|44|42|41|40|38)\b", normalized):
        return "ultra"
    se_match = re.search(r"\bse\s*([23])\b", normalized)
    if se_match:
        return f"se{se_match.group(1)}"
    if re.search(r"\bse\b", normalized):
        return "se"
    series_match = re.search(r"\bseries\s*(\d{1,2})\b", normalized)
    if series_match:
        return f"s{series_match.group(1)}"
    short_match = re.search(r"\bs(\d{1,2})\s+(?:\(?20[1-3]\d\)?\s+)?(?:38|40|41|42|44|45|46|49)\b", normalized)
    if short_match:
        return f"s{short_match.group(1)}"
    return ""


def apple_watch_families_compatible(product_family, line_family, line_years=None):
    if not product_family or not line_family:
        return False
    if product_family == line_family:
        return True
    if product_family == "se3" and line_family == "se" and 2025 in set(line_years or ()):
        return True
    if {product_family, line_family} == {"se", "se2"} and 2024 in set(line_years or ()):
        return True
    return False


def extract_apple_watch_case_sizes(*values):
    sizes = set()
    for value in values:
        for match in APPLE_WATCH_CASE_SIZE_RE.finditer(str(value or "")):
            sizes.add(match.group(1))
    return sizes


def normalized_match_colors_without_material(colors):
    colors = set(colors or ())
    if len(colors) > 1 and "titanium" in colors:
        colors.discard("titanium")
    return colors


def normalized_apple_watch_band_colors(colors):
    colors = set(colors or ())
    colors.discard("titanium")
    return colors


def extract_apple_watch_product_case_colors(*values):
    text = normalize_match_text(" ".join(str(value or "") for value in values))
    if not text:
        return set()
    case_match = re.search(r"\b(.{0,160}?)(?:alumin(?:um|ium)|titanium)?\s+case\b", text)
    if case_match:
        colors = extract_match_colors(case_match.group(1))
        if colors:
            return normalized_match_colors_without_material(colors)
    return normalized_match_colors_without_material(extract_match_colors(text))


def match_color_hits(value):
    text = normalize_match_text(value)
    if not text:
        return []
    hits = []
    padded = f" {text} "
    for color, aliases in MATCH_COLOR_ALIASES.items():
        for alias in aliases:
            alias_norm = normalize_match_text(alias)
            if not alias_norm:
                continue
            pattern = re.escape(alias_norm)
            if " " in alias_norm:
                iterator = re.finditer(pattern, text)
                offset = 0
            else:
                iterator = re.finditer(rf"(?<![a-zа-я0-9]){pattern}(?![a-zа-я0-9])", padded)
                offset = 1
            for match in iterator:
                start = match.start() - offset
                end = match.end() - offset
                hits.append((start, end, -(end - start), color))
    hits.sort()
    return hits


def extract_first_match_colors(value):
    hits = match_color_hits(value)
    if not hits:
        return set()
    first_start = hits[0][0]
    return normalized_match_colors_without_material({color for start, _, _, color in hits if start == first_start})


def extract_apple_watch_line_case_colors(value):
    text = strip_supplier_line_noise(value)
    if "/" in text:
        before_slash = text.split("/", 1)[0]
        colors = extract_first_match_colors(before_slash)
        if colors:
            return colors
    colors = extract_first_match_colors(text)
    if colors and extract_apple_watch_family(text).startswith("ultra") and colors.isdisjoint({"black", "natural"}):
        return {"natural"}
    return colors


def extract_apple_watch_product_band_colors(*values):
    text = normalize_match_text(" ".join(str(value or "") for value in values))
    if not text or " case " not in f" {text} ":
        return set()
    after_case = re.split(r"\bcase\b", text, maxsplit=1)[-1]
    return normalized_apple_watch_band_colors(extract_match_colors(after_case))


def extract_apple_watch_line_band_colors(value):
    text = strip_supplier_line_noise(value)
    if not text:
        return set()
    if "/" in text:
        before_slash, after_slash = text.split("/", 1)
        colors = set(extract_match_colors(after_slash))
        before_hits = match_color_hits(before_slash)
        if before_hits:
            first_start = before_hits[0][0]
            first_end = max(end for start, end, _, _ in before_hits if start == first_start)
            colors.update(color for start, _, _, color in before_hits if start >= first_end)
        colors = normalized_apple_watch_band_colors(colors)
        if colors:
            return colors

    if not extract_apple_watch_band_kinds(text):
        return set()
    hits = match_color_hits(text)
    if not hits:
        return set()
    first_start = hits[0][0]
    first_end = max(end for start, end, _, _ in hits if start == first_start)
    colors = {color for start, _, _, color in hits if start >= first_end}
    colors = normalized_apple_watch_band_colors(colors)
    if colors:
        return colors
    return extract_first_match_colors(text)




def extract_apple_watch_band_sizes(*values):
    normalized = normalize_match_text(" ".join(str(value or "") for value in values))
    sizes = set()
    if re.search(r"\bs\s*/\s*[mм]\b|\bs\s+[mм]\b", normalized):
        sizes.add("sm")
    if re.search(r"\b[mм]\s*/\s*l\b|\b[mм]\s+l\b", normalized):
        sizes.add("ml")
    if re.search(r"\bone\s+size\b", normalized):
        sizes.add("one")
    watch_band_sized_context = re.search(
        r"\b(?:alpine|trail|ocean|milanese|milaneze|milenese|tml|loop|al|tl|ob|sb|sl)\b",
        normalized,
    )
    normalized_single = re.sub(r"\bs\s*/\s*[mм]\b|\b[mм]\s*/\s*l\b|\bs\s+[mм]\b|\b[mм]\s+l\b", " ", normalized)
    if watch_band_sized_context:
        if re.search(r"\b(?:small|size\s*s|s)\b", normalized_single):
            sizes.add("small")
        if re.search(r"\b(?:medium|size\s*[mм]|[mм])\b", normalized_single):
            sizes.add("medium")
        if re.search(r"\b(?:large|size\s*l|l)\b", normalized_single):
            sizes.add("large")
    return sizes


APPLE_WATCH_PROTECTIVE_ACCESSORY_RE = re.compile(
    r"\b(?:стекл\w*|пленк\w*|glass|protector|film|armor|бампер\w*|чехол\w*)\b",
    re.IGNORECASE,
)


def text_mentions_apple_watch_protective_accessory(*values):
    text = normalize_match_text(" ".join(str(value or "") for value in values))
    return bool(APPLE_WATCH_PROTECTIVE_ACCESSORY_RE.search(text))


def apple_watch_band_sizes_compatible(product_sizes, line_sizes):
    product_sizes = set(product_sizes or ())
    line_sizes = set(line_sizes or ())
    if not product_sizes or not line_sizes:
        return True
    if not product_sizes.isdisjoint(line_sizes):
        return True
    aliases = {
        "sm": {"sm", "small"},
        "small": {"sm", "small"},
        "ml": {"ml"},
        "medium": {"medium"},
        "large": {"large"},
        "one": {"one"},
    }
    expanded_product = set().union(*(aliases.get(size, {size}) for size in product_sizes))
    expanded_line = set().union(*(aliases.get(size, {size}) for size in line_sizes))
    return not expanded_product.isdisjoint(expanded_line)


def apple_watch_default_sport_band_colors(product_family, product_case_colors):
    case_colors = set(product_case_colors or ())
    defaults = {
        ("s11", "jet_black"): {"black"},
        ("s11", "space_gray"): {"black"},
        ("s11", "silver"): {"purple"},
        ("s11", "rose_gold"): {"pink"},
        ("s10", "jet_black"): {"black"},
        ("s10", "silver"): {"blue"},
        ("s10", "rose_gold"): {"pink"},
        ("se3", "midnight"): {"midnight"},
        ("se3", "starlight"): {"starlight"},
        ("se2", "midnight"): {"midnight"},
    }
    case_keys = set(case_colors)
    if "black" in case_keys:
        case_keys.add("jet_black")
    if {"pink", "gold"}.issubset(case_keys):
        case_keys.add("rose_gold")
    for case_key in case_keys:
        colors = defaults.get((product_family, case_key))
        if colors:
            return colors
    return set()


def extract_apple_watch_band_kinds(*values):
    normalized = normalize_match_text(" ".join(str(value or "") for value in values))
    kinds = set()
    if re.search(r"\bsb\b|\bsport\s+band\b", normalized):
        kinds.add("sport_band")
    if re.search(r"\bsl\b|\bsport\s+loop\b", normalized):
        kinds.add("sport_loop")
    if re.search(r"\bmila?ne[sz]e\b|\bmilenese\b|\btml\b", normalized):
        kinds.add("milanese")
    if re.search(r"\balpine\s+loop\b|\bal\b", normalized):
        kinds.add("alpine")
    if re.search(r"\blight\s+blue\s+loop\b", normalized):
        kinds.add("alpine")
    if re.search(r"\btrail\s+loop\b|\btl\b", normalized):
        kinds.add("trail")
    if re.search(r"\bcharcoal\s+loop\b", normalized):
        kinds.add("trail")
    if re.search(r"\bocean\s+band\b|\bob\b", normalized):
        kinds.add("ocean")
    if re.search(r"\bsolo\s+loop\b", normalized):
        kinds.add("solo_loop")
    if re.search(r"\blink\s+bracelet\b", normalized):
        kinds.add("link_bracelet")
    return kinds


def extract_match_capacities(*values):
    capacities = set()
    for value in values:
        normalized = normalize_match_text(value)
        for amount, unit in re.findall(r"\b(\d+(?:[.,]\d+)?)\s*(tb|тб|gb|гб)\b", normalized):
            try:
                number = float(amount.replace(",", "."))
            except ValueError:
                continue
            capacities.add(int(number * 1024 if unit in {"tb", "тб"} else number))
        for _, storage in re.findall(r"\b(\d{1,2})\s*/\s*(\d{2,4})\b", normalized):
            try:
                capacities.add(int(storage))
            except ValueError:
                continue
        if re.search(r"\b(?:iphone|айфон|1[67]e|17e|16e|se|1[1-9])\b", normalized):
            for storage in re.findall(
                r"\b(?:iphone\s+|айфон\s+)?(?:1[67]e|17e|16e|se|1[1-9])(?:\s+(?:pro\s+max|pro|max|plus|mini|air))?\s+(64|128|256|512)\b",
                normalized,
            ):
                capacities.add(int(storage))
        if re.search(r"\b(?:ipad|air|pro|mini)\b", normalized):
            for storage in re.findall(
                r"\b(?:ipad\s+)?(?:air|pro)\s+(?:11|13)\s+m[1-9]\s+(64|128|256|512)\b",
                normalized,
            ):
                capacities.add(int(storage))
            for storage in re.findall(
                r"\b(?:ipad\s+)?(?:air|pro)\s+(?:11|13)\s+m[1-9]\s+(?:\(?(?:20[1-3]\d)\)?\s+)?(64|128|256|512)\s*(?:gb|гб|b)?\b",
                normalized,
            ):
                capacities.add(int(storage))
            for storage in re.findall(
                r"\bipad\s*(?:\d{1,2}|mini(?:\s*\d+)?)\s+(64|128|256|512)\b",
                normalized,
            ):
                capacities.add(int(storage))
            for storage in re.findall(r"\bmini\s*\d+\s+(64|128|256|512)\b", normalized):
                capacities.add(int(storage))
        if re.search(r"\b(?:pixel|galaxy|redmi|poco|xiaomi|honor|huawei|oneplus|one\s+plus|realme|nothing|cmf)\b", normalized):
            for storage in re.findall(r"\b(?:5g\s+)?(64|128|256|512)\b", normalized):
                capacities.add(int(storage))
    return capacities


def extract_match_storage_values(*values):
    storage_values = set()
    for value in values:
        normalized = normalize_match_text(value)
        for _, storage in re.findall(r"\b(\d{1,3})\s*/\s*(\d{2,4})\b", normalized):
            try:
                storage_values.add(int(storage))
            except ValueError:
                continue
        for amount, unit in re.findall(r"\b(\d+(?:[.,]\d+)?)\s*(tb|тб)\b", normalized):
            try:
                storage_values.add(int(float(amount.replace(",", ".")) * 1024))
            except ValueError:
                continue
        for amount in re.findall(r"\b(\d{2,4})\s*(?:gb|гб)\s*(?:ssd|storage|памят|накоп)", normalized):
            try:
                storage_values.add(int(amount))
            except ValueError:
                continue
    return storage_values


def extract_match_ram_values(*values):
    ram_values = set()
    for value in values:
        normalized = normalize_match_text(value)
        for ram, _ in re.findall(r"\b(\d{1,3})\s*/\s*(\d{2,4})\b", normalized):
            try:
                ram_values.add(int(ram))
            except ValueError:
                continue
        for amount in re.findall(r"\b(\d{1,3})\s*(?:gb|гб)\s*(?:ram|озу)\b", normalized):
            try:
                ram_values.add(int(amount))
            except ValueError:
                continue
    return ram_values


def extract_match_connectivity(*values):
    text = normalize_product_search_text(" ".join(str(value or "") for value in values))
    found = set()
    if re.search(r"\b(?:wi\s*fi|wifi|вай\s*фай)\b", text):
        found.add("wifi")
    if re.search(r"\b(?:lte|4g|5g|cellular)\b", text):
        found.add("cellular")
    return found


CHARGER_MARKER_RE = re.compile(
    r"\b(?:charger|charging|power\s*adapter|adapter|заряд\w*|сзу|адаптер\s+питания)\b|"
    r"сетев\w*\s+заряд\w*",
    re.IGNORECASE,
)
CHARGER_BUNDLE_CABLE_RE = re.compile(
    r"(?:\+|\bplus\b)\s*(?:cable|кабель)\b|"
    r"\b(?:with|w/)\s*cable\b|"
    r"\bс\s+кабелем\b|"
    r"\b(?:cable|кабель)\s*(?:included|в\s+комплекте)\b",
    re.IGNORECASE,
)


def text_mentions_charger(*values):
    text = normalize_match_text(" ".join(str(value or "") for value in values))
    return bool(CHARGER_MARKER_RE.search(text))


def text_has_charger_cable_bundle_marker(*values):
    text = normalize_match_text(" ".join(str(value or "") for value in values))
    return bool(CHARGER_BUNDLE_CABLE_RE.search(text))


def extract_match_years(*values):
    text = normalize_match_text(" ".join(str(value or "") for value in values))
    return {int(match.group(0)) for match in re.finditer(r"\b20[1-3]\d\b", text)}


SAMSUNG_WATCH_CASE_SIZE_RE = re.compile(r"(?<!\d)(40|44|46|47)\s*(?:mm|мм)?\b", re.IGNORECASE)
SAMSUNG_WATCH_ACCESSORY_RE = re.compile(
    r"\b(?:ремеш\w*|браслет\w*|чехол\w*|бампер\w*|strap|band|case|cover|armor)\b",
    re.IGNORECASE,
)


def text_mentions_samsung_watch_accessory(*values):
    text = normalize_match_text(" ".join(str(value or "") for value in values))
    return bool(SAMSUNG_WATCH_ACCESSORY_RE.search(text))


def extract_samsung_watch_family(*values):
    text = normalize_match_text(" ".join(str(value or "") for value in values))
    if not text:
        return ""
    if re.search(r"\bwatch\s*8\s*classic\b|\bwatch8\s*classic\b", text):
        return "watch8_classic"
    if re.search(r"\bwatch\s*8\b|\bwatch8\b", text):
        return "watch8"
    if re.search(r"\bwatch\s*ultra\b|\bultra\b.*\bwatch\b", text):
        return "watch_ultra"
    match = re.search(r"\bwatch\s*(\d{1,2})\b|\bwatch(\d{1,2})\b", text)
    if match:
        return f"watch{match.group(1) or match.group(2)}"
    return ""


def extract_samsung_watch_case_sizes(*values):
    sizes = set()
    for value in values:
        for match in SAMSUNG_WATCH_CASE_SIZE_RE.finditer(str(value or "")):
            sizes.add(match.group(1))
    return sizes


def extract_insta360_model_key(*values):
    text = normalize_match_text(" ".join(str(value or "") for value in values))
    if "insta360" not in text and "insta 360" not in text:
        return ""
    patterns = (
        (r"\bx\s*3\b", "x3"),
        (r"\bx\s*4\b", "x4"),
        (r"\bx\s*5\b", "x5"),
        (r"\bgo\s*3s\b", "go3s"),
        (r"\bgo\s*ultra\b", "go_ultra"),
        (r"\bace\s*pro\s*2\b", "ace_pro_2"),
        (r"\bace\s*pro\b", "ace_pro"),
        (r"\bone\s*rs\b", "one_rs"),
    )
    for pattern, key in patterns:
        if re.search(pattern, text):
            return key
    return ""


def extract_match_apple_chips(*values):
    text = normalize_match_text(" ".join(str(value or "") for value in values))
    return {match.group(1).lower() for match in re.finditer(r"\b(m[1-9])\b", text)}


def has_pro_plus_marker(*values):
    text = normalize_match_text(" ".join(str(value or "") for value in values))
    return bool(re.search(r"\bpro\s*(?:\+|plus)(?=\b|[^a-zа-я0-9])", text))


def extract_match_colors(*values):
    text = normalize_match_text(" ".join(str(value or "") for value in values))
    if not text:
        return set()
    found = set()
    padded = f" {text} "
    for color, aliases in MATCH_COLOR_ALIASES.items():
        for alias in aliases:
            alias_norm = normalize_match_text(alias)
            if not alias_norm:
                continue
            if " " in alias_norm:
                if alias_norm in text:
                    found.add(color)
                    break
            elif not re.search(r"[a-zа-я0-9]", alias_norm):
                if alias_norm in text:
                    found.add(color)
                    break
            elif re.search(rf"(?<![a-zа-я0-9]){re.escape(alias_norm)}(?![a-zа-я0-9])", padded):
                found.add(color)
                break
    return found


def match_colors_compatible(product_colors, line_colors):
    product_colors = set(product_colors or ())
    line_colors = set(line_colors or ())
    if not product_colors or not line_colors:
        return True
    if not product_colors.isdisjoint(line_colors):
        return True

    blue_family = {"blue", "light_blue", "navy"}
    if line_colors == {"blue"} and product_colors.intersection(blue_family):
        return True
    if product_colors == {"blue"} and line_colors.intersection(blue_family):
        return True
    neutral_family = {"silver", "gray", "space_gray", "slate", "titanium", "natural"}
    if product_colors.intersection(neutral_family) and line_colors.intersection(neutral_family):
        return True
    return False


def detect_ipad_model_family(*values):
    normalized = normalize_match_text(" ".join(str(value or "") for value in values))
    if not normalized:
        return ""
    normalized = normalized.replace('"', " ")

    if re.search(r"\bmini\s*7\b", normalized) or (
        "ipad mini" in normalized and ("2024" in normalized or "a17 pro" in normalized)
    ):
        return "iPad Mini 7"
    if "ipad mini" in normalized and "2021" in normalized:
        return "iPad Mini 6"
    if "ipad mini" in normalized or re.search(r"\bmini\b", normalized):
        return "iPad Mini"

    if re.search(r"\b(?:ipad\s+)?pro\s*13\b", normalized):
        return "iPad Pro 13"
    if re.search(r"\b(?:ipad\s+)?pro\s*11\b", normalized):
        return "iPad Pro 11"
    if re.search(r"\b(?:ipad\s+)?air\s*13\b", normalized):
        return "iPad Air 13"
    if re.search(r"\b(?:ipad\s+)?air\s*11\b", normalized):
        return "iPad Air 11"
    if re.search(r"\bipad\s*11\b", normalized):
        return "iPad 11"
    if re.search(r"\bipad\s*10\b", normalized):
        return "iPad 10"
    if re.search(r"\bipad\s*9\b", normalized):
        return "iPad 9"
    return ""


def is_apple_ipad_product(product_name, brand=""):
    normalized_brand = sanitize_brand_value(brand)
    lower = str(product_name or "").lower()
    return "ipad" in lower and (normalized_brand in {"", "Apple"} or "apple" in lower)


def normalize_ipad_connectivity_label(product_name):
    normalized = normalize_match_text(product_name).replace(" ", "")
    if any(marker in normalized for marker in ("cellular", "lte", "esim", "1sim", "2sim")):
        return "Wi-Fi + Cellular (LTE)"
    return "Wi-Fi"


def normalize_ipad_catalog_name(product_name):
    raw_name = re.sub(r"\s+", " ", str(product_name or "").strip())
    if not raw_name:
        return ""
    connectivity = normalize_ipad_connectivity_label(raw_name)
    normalized = re.sub(
        r"\(\s*(?:1\s*sim\s*\+\s*e\s*sim|1sim\s*\+\s*esim|sim\s*\+\s*esim|e\s*sim(?:\s*only)?)\s*\)",
        "",
        raw_name,
        flags=re.IGNORECASE,
    )
    normalized = re.sub(
        r"\b(?:1\s*sim\s*\+\s*e\s*sim|1sim\s*\+\s*esim|sim\s*\+\s*esim|e\s*sim(?:\s*only)?)\b",
        "",
        normalized,
        flags=re.IGNORECASE,
    )
    normalized = re.sub(
        r"\bWi[\s-]*Fi(?:\s*\+\s*Cellular(?:\s*\(LTE\))?)?\b",
        connectivity,
        normalized,
        flags=re.IGNORECASE,
    )
    normalized = re.sub(r"\s+\(\s*LTE\s*\)\s+\(\s*LTE\s*\)", " (LTE)", normalized, flags=re.IGNORECASE)
    normalized = re.sub(r"\s{2,}", " ", normalized)
    return normalized.strip(" ,")


def extract_apple_root_model_no(*values):
    for value in values:
        model_no = extract_apple_model_no(value)
        if model_no:
            return model_no
    return ""


def build_ipad_default_specs(product_name, storage_value=""):
    family = detect_ipad_model_family(product_name)
    connectivity = normalize_ipad_connectivity_label(product_name)
    specs = {
        "connectivity": connectivity,
        "os": "iPadOS",
        "battery": "До 10 часов работы",
        "charging": "USB-C",
    }
    if family == "iPad Mini 6":
        specs.update({
            "display": '8.3" Liquid Retina',
            "processor": "Apple A15 Bionic",
            "camera": "12 Мп",
            "front_camera": "12 Мп",
            "video": "4K",
            "biometrics": "Touch ID",
        })
    elif family == "iPad Mini 7":
        specs.update({
            "display": '8.3" Liquid Retina',
            "processor": "Apple A17 Pro",
            "camera": "12 Мп",
            "front_camera": "12 Мп",
            "video": "4K",
            "biometrics": "Touch ID",
        })
    elif family == "iPad Air 13":
        specs.update({
            "display": '13" Liquid Retina',
            "processor": "Apple M2",
            "camera": "12 Мп",
            "front_camera": "12 Мп Center Stage",
            "video": "4K",
            "biometrics": "Touch ID",
        })
    elif family == "iPad Pro 13":
        specs.update({
            "display": '13" Ultra Retina XDR',
            "processor": "Apple M4",
            "camera": "12 Мп",
            "front_camera": "12 Мп",
            "video": "4K ProRes",
            "biometrics": "Face ID",
            "charging": "Thunderbolt / USB-C",
        })
    elif family == "iPad Pro 11":
        specs.update({
            "display": '11" Ultra Retina XDR',
            "processor": "Apple M4",
            "camera": "12 Мп",
            "front_camera": "12 Мп",
            "video": "4K ProRes",
            "biometrics": "Face ID",
            "charging": "Thunderbolt / USB-C",
        })
    else:
        specs.update({
            "display": "Liquid Retina",
            "processor": "Apple Silicon",
            "camera": "12 Мп",
            "front_camera": "12 Мп",
            "video": "4K",
            "biometrics": "Touch ID",
        })
    if storage_value:
        specs["storage"] = storage_value
    return specs


def build_ipad_description(product_name, storage_value=""):
    display_name = normalize_ipad_catalog_name(product_name) or str(product_name or "").strip()
    specs = build_ipad_default_specs(product_name, storage_value)
    storage_line = f" с {storage_value} встроенной памяти" if storage_value else ""
    return (
        f"{display_name} — планшет Apple для работы, учебы, чтения и мультимедиа.\n\n"
        f"Модель оснащена дисплеем {specs.get('display', 'Liquid Retina')} и работает на {specs.get('processor', 'Apple Silicon')}, "
        f"поэтому подходит для повседневных задач, общения, заметок и просмотра контента.\n\n"
        f"Версия{storage_line} и поддержкой {specs.get('connectivity', 'Wi-Fi')} удобно подходит для дома, поездок и мобильной работы."
    )


def build_ipad_description_html(product_name, storage_value=""):
    description = build_ipad_description(product_name, storage_value)
    specs = build_ipad_default_specs(product_name, storage_value)
    title = html.escape(normalize_ipad_catalog_name(product_name) or str(product_name or "").strip())
    paragraphs = [
        html.escape(part.strip())
        for part in description.split("\n\n")
        if part.strip()
    ]
    feature_lines = []
    if specs.get("display"):
        feature_lines.append(f"Дисплей: {specs['display']}")
    if specs.get("processor"):
        feature_lines.append(f"Процессор: {specs['processor']}")
    if storage_value:
        feature_lines.append(f"Память: {storage_value}")
    if specs.get("connectivity"):
        feature_lines.append(f"Связь: {specs['connectivity']}")
    features_html = "".join(f"<li>{html.escape(line)}</li>" for line in feature_lines)
    paragraphs_html = "".join(f"<p>{line}</p>" for line in paragraphs)
    return (
        '<div class="catalog-rich-description">'
        '<section class="catalog-rich-description__section">'
        f"<h2>{title}</h2>"
        f"{paragraphs_html}"
        f"<ul>{features_html}</ul>"
        "</section>"
        "</div>"
    )


def build_basic_description_html(title, description, specs_payload=None):
    safe_title = html.escape(str(title or "").strip())
    paragraphs = [
        html.escape(part.strip())
        for part in re.split(r"\n\s*\n", str(description or "").strip())
        if part.strip()
    ]
    bullets = []
    for key, value in list((specs_payload or {}).items())[:8]:
        clean_key = str(key or "").strip()
        clean_value = str(value or "").strip()
        if clean_key and clean_value:
            bullets.append(f"<li><strong>{html.escape(clean_key)}:</strong> {html.escape(clean_value)}</li>")
    paragraphs_html = ''.join(f"<p>{paragraph}</p>" for paragraph in paragraphs) or "<p></p>"
    bullets_html = f"<ul>{''.join(bullets)}</ul>" if bullets else ""
    return (
        '<div class="catalog-rich-description">'
        '<section class="catalog-rich-description__section">'
        f"<h2>{safe_title}</h2>"
        f"{paragraphs_html}"
        f"{bullets_html}"
        "</section>"
        "</div>"
    )


def build_generic_catalog_description(product):
    product_name = str(get_product_field_value(product, "name") or "").strip()
    brand = sanitize_brand_value(get_product_field_value(product, "brand"))
    kind = infer_description_kind_from_name(product_name)
    model_number = str(get_product_field_value(product, "model_number") or "").strip()
    color = str(get_product_field_value(product, "color") or "").strip()
    storage = normalize_capacity(get_product_field_value(product, "storage"), allow_tb=True)
    ram = normalize_capacity(get_product_field_value(product, "ram"), allow_tb=False)
    country = str(get_product_field_value(product, "country") or "").strip()
    warranty = str(get_product_field_value(product, "warranty") or "").strip()

    kind_label = {
        "smartphone": "смартфон",
        "tablet": "планшет",
        "laptop": "ноутбук",
        "headphones": "беспроводные наушники",
        "speaker": "портативная акустика",
        "watch": "умные часы",
        "charger": "зарядный аксессуар",
        "cable": "кабель",
        "mount": "аксессуар",
        "case": "аксессуар",
        "protection": "аксессуар",
        "strap": "аксессуар",
    }.get(kind, "товар")

    lead_brand = f" {brand}" if brand else ""
    lead = f"{product_name} — {kind_label}{lead_brand}, подготовленный для размещения в каталоге."

    details = []
    if model_number:
        details.append(f"модель {model_number}")
    if color:
        details.append(f"цвет {color}")
    if storage:
        details.append(f"память {storage}")
    if ram:
        details.append(f"ОЗУ {ram}")
    if country:
        details.append(f"версия поставки {country}")

    detail_line = ""
    if details:
        detail_line = " В карточке указаны основные параметры: " + ", ".join(details) + "."

    warranty_line = f" Гарантия: {warranty}." if warranty else ""
    usage_line = " Карточку можно использовать для публикации на сайте, подбора по каталогу и привязки к сообщениям поставщиков."
    return f"{lead}{detail_line}{warranty_line}\n\n{usage_line}".strip()


def build_generic_catalog_specs(product):
    product_name = str(get_product_field_value(product, "name") or "").strip()
    specs = {}

    model_number = str(get_product_field_value(product, "model_number") or "").strip()
    if model_number:
        specs["SKU / артикул"] = model_number

    color = str(get_product_field_value(product, "color") or "").strip()
    if color:
        specs["Цветовое исполнение"] = color

    storage = normalize_capacity(get_product_field_value(product, "storage"), allow_tb=True)
    if storage:
        specs["Конфигурация памяти"] = storage

    ram = normalize_capacity(get_product_field_value(product, "ram"), allow_tb=False)
    if ram:
        specs["Конфигурация ОЗУ"] = ram

    weight = normalize_autofill_weight(get_product_field_value(product, "weight"))
    if weight:
        specs["Масса, г"] = weight

    country = str(get_product_field_value(product, "country") or "").strip()
    if country:
        specs["Версия поставки"] = country

    warranty = str(get_product_field_value(product, "warranty") or "").strip()
    if warranty:
        specs["Сервисная гарантия"] = warranty

    normalized_name = normalize_match_text(product_name)
    if any(marker in normalized_name for marker in ("науш", "airpods", "buds", "freebuds")):
        specs.setdefault("Форм-фактор", "Беспроводные наушники")
        if "bluetooth" in normalized_name or "wireless" in normalized_name or "беспровод" in normalized_name or "true wireless" in normalized_name:
            specs.setdefault("Интерфейс подключения", "Bluetooth")
    elif any(marker in normalized_name for marker in ("iphone", "смартфон", "galaxy s", "pixel")):
        specs.setdefault("Тип устройства", "Смартфон")
    elif any(marker in normalized_name for marker in ("ipad", "планшет")):
        specs.setdefault("Тип устройства", "Планшет")
    elif any(marker in normalized_name for marker in ("macbook", "ноутбук")):
        specs.setdefault("Тип устройства", "Ноутбук")

    return specs


def detect_airpods_model_family(*values):
    normalized = normalize_match_text(" ".join(str(value or "") for value in values))
    if not normalized:
        return ""
    if "airpods" not in normalized and "air pods" not in normalized:
        return ""
    normalized = re.sub(r"\btype\s*[- ]?\s*c\b", "usb c", normalized)

    if re.search(r"\bmax\s*2\b", normalized) or (
        "max" in normalized and re.search(r"\b2026\b", normalized)
    ):
        return "AirPods Max 2"
    if "max" in normalized:
        if re.search(r"\b(?:lightning|mgym3|mgyn3|mgyh3|mgyj3|mgyl3)\b", normalized):
            return "AirPods Max Lightning"
        if re.search(r"\b(?:usb\s*c|usbc|mww43|mww53|mww63|mww73|mww83)\b", normalized):
            return "AirPods Max USB-C"
        return "AirPods Max"

    if re.search(r"\bpro\s*3\b", normalized):
        return "AirPods Pro 3"
    if re.search(r"\bpro\s*2\b", normalized):
        return "AirPods Pro 2"
    if re.search(r"\bairpods\s*4\b|\bair\s*pods\s*4\b", normalized):
        return "AirPods 4"
    if re.search(r"\bairpods\s*3\b|\bair\s*pods\s*3\b", normalized):
        return "AirPods 3"
    if re.search(r"\bairpods\s*2\b|\bair\s*pods\s*2\b", normalized):
        return "AirPods 2"
    return ""


def airpods_model_families_compatible(product_family, line_family):
    if not product_family or not line_family:
        return True
    if product_family == line_family:
        return True

    airpods_max_families = {"AirPods Max", "AirPods Max USB-C", "AirPods Max Lightning", "AirPods Max 2"}
    if product_family in airpods_max_families and line_family in airpods_max_families:
        if "AirPods Max 2" in {product_family, line_family}:
            return False
        if product_family != "AirPods Max" and line_family != "AirPods Max":
            return False
        return True

    return False


AIRPODS_CUSTOM_FINISH_RE = re.compile(
    r"\b(?:матов\w*|глян\w*|покраск\w*|крашен\w*|логотип|эмблем\w*|флаг|герб|"
    r"ral|полная\s+покраска|полная\s+прокраска)\b|"
    r"\(\d{4}(?:\s*,\s*\d{4})*\)",
    re.IGNORECASE,
)


def text_mentions_airpods_anc(*values):
    normalized = normalize_match_text(" ".join(str(value or "") for value in values))
    return bool(re.search(r"\b(?:anc|active\s+noise|noise\s+cancell\w*|шумодав|шумоподав\w*)\b", normalized))


def text_mentions_airpods_custom_finish(*values):
    normalized = normalize_match_text(" ".join(str(value or "") for value in values))
    return bool(AIRPODS_CUSTOM_FINISH_RE.search(normalized))


GARMIN_WATCH_CASE_SIZE_RE = re.compile(r"(?<!\d)(40|41|42|43|45|47|49|50|51)\s*(?:mm|мм)?\b", re.IGNORECASE)


def extract_garmin_watch_family(*values):
    normalized = normalize_match_text(" ".join(str(value or "") for value in values))
    if "garmin" not in normalized:
        return ""
    patterns = (
        (r"\bfenix\s*8\s*pro\b", "fenix_8_pro"),
        (r"\bfenix\s*8\b", "fenix_8"),
        (r"\btactix\s*8\b", "tactix_8"),
        (r"\benduro\s*3\b", "enduro_3"),
        (r"\binstinct\s*crossover\b", "instinct_crossover"),
        (r"\binstinct\s*3\b", "instinct_3"),
        (r"\binstinct\s*e\b", "instinct_e"),
        (r"\bquatix\s*7\b", "quatix_7"),
        (r"\bforerunner\s*\d{3,4}\b", "forerunner"),
        (r"\bvenu\s*x1\b", "venu_x1"),
        (r"\bepix\b", "epix"),
    )
    for pattern, family in patterns:
        if re.search(pattern, normalized):
            return family
    return ""


def extract_garmin_watch_case_sizes(*values):
    sizes = set()
    for value in values:
        for match in GARMIN_WATCH_CASE_SIZE_RE.finditer(str(value or "")):
            sizes.add(match.group(1))
    return sizes


def extract_garmin_watch_variant_markers(*values):
    normalized = normalize_match_text(" ".join(str(value or "") for value in values))
    markers = set()
    if re.search(r"\bamoled\b", normalized):
        markers.add("amoled")
    if re.search(r"\bsolar\b", normalized):
        markers.add("solar")
    if re.search(r"\bpro\b", normalized):
        markers.add("pro")
    return markers


CAMERA_BODY_FAMILY_PATTERNS = (
    (r"\bcanon\s+eos\s+r6\s+mark\s*ii\b", "canon_eos_r6_mark_ii"),
    (r"\bcanon\s+eos\s+r6\b", "canon_eos_r6"),
    (r"\bcanon\s+eos\s+r8\b", "canon_eos_r8"),
    (r"\bcanon\s+eos\s+r7\b", "canon_eos_r7"),
    (r"\bcanon\s+eos\s+r10\b", "canon_eos_r10"),
    (r"\bcanon\s+eos\s+r50\b", "canon_eos_r50"),
    (r"\bcanon\s+eos\s+rp\b", "canon_eos_rp"),
    (r"\bcanon\s+eos\s+250d\b", "canon_eos_250d"),
    (r"\bcanon\s+eos\s+2000d\b", "canon_eos_2000d"),
    (r"\bnikon\s+z30\b", "nikon_z30"),
    (r"\bnikon\s+z50\s*ii\b", "nikon_z50_ii"),
    (r"\bnikon\s+z50\b", "nikon_z50"),
    (r"\bnikon\s+z5\b", "nikon_z5"),
)

CAMERA_LENS_ONLY_RE = re.compile(
    r"\b(?:rf-s?|ef-s?|ef-m|rf|z\s*dx|dx|fx)\b.*\b\d{1,3}\s*[-/]\s*\d{1,3}(?:mm)?\b|"
    r"\b(?:rf-s?|ef-s?|ef-m|rf|z\s*dx|dx|fx)\b.*\bf\s*/\s*\d",
    re.IGNORECASE,
)


def extract_camera_body_family(*values):
    normalized = normalize_match_text(" ".join(str(value or "") for value in values))
    if not normalized:
        return ""
    for pattern, family in CAMERA_BODY_FAMILY_PATTERNS:
        if re.search(pattern, normalized):
            return family
    return ""


def detect_camera_listing_kind(*values):
    normalized = normalize_match_text(" ".join(str(value or "") for value in values))
    if not normalized:
        return ""
    if re.search(r"\bmount\s+adapter\b|\badapter\b", normalized):
        return "accessory"
    if CAMERA_LENS_ONLY_RE.search(normalized) and not extract_camera_body_family(normalized):
        return "lens"
    if re.search(r"\bkit\b", normalized):
        return "kit"
    if re.search(r"\bbody\b", normalized):
        return "body"
    if extract_camera_body_family(normalized):
        return "camera"
    return ""


def extract_camera_kit_lens_signature(*values):
    normalized = normalize_match_text(" ".join(str(value or "") for value in values))
    if not normalized:
        return ""
    match = re.search(
        r"\b(rf-s?|ef-s?|ef-m|z\s*dx|dx|fx)\b[^\d]{0,20}(\d{1,3})\s*[-/]\s*(\d{1,3})(?:mm)?\b",
        normalized,
        re.IGNORECASE,
    )
    if not match:
        return ""
    mount = re.sub(r"\s+", "", match.group(1).lower())
    return f"{mount}_{match.group(2)}_{match.group(3)}"


def product_line_semantics_compatible(product, line_text, require_known_category=False):
    product_name = row_value(product, "name", "") or ""
    product_country = row_value(product, "country", "") or ""
    product_brand = row_value(product, "brand", "") or ""
    product_model = row_value(product, "model_number", "") or ""
    product_color = row_value(product, "color", "") or ""
    product_storage = row_value(product, "storage", "") or ""
    product_condition_text = build_product_condition_text(product)
    line_clean = strip_supplier_line_noise(line_text)
    product_model_tokens = get_product_model_tokens(product)
    line_model_tokens = extract_catalog_model_tokens(line_clean)
    explicit_product_model_tokens = extract_catalog_model_tokens(product_model, product_name)

    if explicit_product_model_tokens and line_model_tokens and explicit_product_model_tokens.isdisjoint(line_model_tokens):
        return False
    if product_model_tokens and line_model_tokens and product_model_tokens.isdisjoint(line_model_tokens):
        return False

    if not product_message_variants_compatible(
        product_name,
        product_country,
        line_text,
        product_condition_text=product_condition_text,
    ):
        return False

    detected_product_brand = sanitize_brand_value(product_brand)
    detected_name_brand = explicit_brand_from_product_name(product_name, product_model) or detect_match_brand(product_name)
    if detected_product_brand in PLATFORM_COMPATIBILITY_BRANDS and detected_name_brand:
        detected_product_brand = detected_name_brand
    detected_product_brand = detected_product_brand or detected_name_brand
    detected_line_brand = explicit_brand_from_product_name(line_clean) or detect_match_brand(line_clean)
    if not brands_are_compatible_for_matching(detected_product_brand, detected_line_brand):
        return False

    product_category = detect_match_category(f"{product_name} {product_brand}")
    line_category = detect_match_category(line_clean)
    if product_category == "iphone" and APPLE_COMPUTER_CHIP_CONTEXT_RE.search(normalize_match_text(line_clean)):
        return False
    if product_category and line_category and product_category != line_category:
        return False
    if require_known_category and product_category and not line_category:
        model_key = normalize_match_text(product_model)
        if not (model_key and model_key in normalize_match_text(line_clean)):
            return False
    if product_category == "playstation_gamepad" and line_category == "playstation_gamepad":
        product_gamepad_variant = extract_playstation_gamepad_variant_key(
            product_name,
            product_model,
            row_value(product, "synonyms", ""),
        )
        line_gamepad_variant = extract_playstation_gamepad_variant_key(line_clean)
        if product_gamepad_variant or line_gamepad_variant:
            if product_gamepad_variant != line_gamepad_variant:
                return False

    product_has_cable_bundle = text_has_charger_cable_bundle_marker(
        product_name,
        product_model,
        row_value(product, "synonyms", ""),
    )
    line_has_cable_bundle = text_has_charger_cable_bundle_marker(line_clean)
    if product_has_cable_bundle != line_has_cable_bundle:
        if text_mentions_charger(product_name, product_model, row_value(product, "synonyms", "")) or text_mentions_charger(line_clean):
            return False

    if has_pro_plus_marker(line_clean) and not has_pro_plus_marker(product_name, product_model, row_value(product, "synonyms", "")):
        return False
    if has_pro_plus_marker(product_name, product_model) and re.search(r"\bpro\b", normalize_match_text(line_clean)) and not has_pro_plus_marker(line_clean):
        return False

    if product_category == "ipad":
        product_ipad_family = detect_ipad_model_family(product_name, product_model)
        line_ipad_family = detect_ipad_model_family(line_clean)
        if product_ipad_family and line_ipad_family and product_ipad_family != line_ipad_family:
            return False
    if detected_product_brand == "Garmin" or detected_line_brand == "Garmin":
        product_garmin_family = extract_garmin_watch_family(product_name, product_model)
        line_garmin_family = extract_garmin_watch_family(line_clean)
        if product_garmin_family and line_garmin_family and product_garmin_family != line_garmin_family:
            return False

        product_garmin_sizes = extract_garmin_watch_case_sizes(product_name, product_model)
        line_garmin_sizes = extract_garmin_watch_case_sizes(line_clean)
        if product_garmin_sizes and line_garmin_sizes and product_garmin_sizes.isdisjoint(line_garmin_sizes):
            return False

        product_garmin_variants = extract_garmin_watch_variant_markers(product_name, product_model)
        line_garmin_variants = extract_garmin_watch_variant_markers(line_clean)
        if product_garmin_variants and line_garmin_variants and product_garmin_variants.isdisjoint(line_garmin_variants):
            return False
        if line_garmin_variants and not product_garmin_variants and product_model_tokens.isdisjoint(line_model_tokens):
            return False
    if product_category == "camera" or line_category == "camera":
        product_camera_kind = detect_camera_listing_kind(product_name, product_model)
        line_camera_kind = detect_camera_listing_kind(line_clean)
        if "lens" in {product_camera_kind, line_camera_kind} and product_camera_kind != line_camera_kind:
            return False
        if line_camera_kind == "accessory" and product_camera_kind != "accessory":
            return False

        product_camera_family = extract_camera_body_family(product_name, product_model)
        line_camera_family = extract_camera_body_family(line_clean)
        if product_camera_family and line_camera_family and product_camera_family != line_camera_family:
            return False

        product_lens_signature = extract_camera_kit_lens_signature(product_name)
        line_lens_signature = extract_camera_kit_lens_signature(line_clean)
        if product_lens_signature and line_lens_signature and product_lens_signature != line_lens_signature:
            return False
        if product_camera_kind == "kit" and line_camera_kind == "kit" and product_lens_signature and not line_lens_signature:
            return False
    if product_category == "airpods":
        product_airpods_family = detect_airpods_model_family(product_name, product_model)
        line_airpods_family = detect_airpods_model_family(line_clean)
        if not airpods_model_families_compatible(product_airpods_family, line_airpods_family):
            return False
        if product_airpods_family == "AirPods 4" and line_airpods_family == "AirPods 4":
            if text_mentions_airpods_anc(product_name, product_model, row_value(product, "synonyms", "")) != text_mentions_airpods_anc(line_clean):
                return False
        if text_mentions_airpods_custom_finish(product_name, product_model, row_value(product, "synonyms", "")) and not text_mentions_airpods_custom_finish(line_clean):
            return False
        airpods_product_colors = extract_match_colors(product_name, product_color)
        airpods_line_colors = extract_match_colors(line_clean)
        if airpods_product_colors and not airpods_line_colors and "white" not in airpods_product_colors:
            return False
    if product_category == "samsung_watch" or line_category == "samsung_watch":
        product_is_accessory = text_mentions_samsung_watch_accessory(
            product_name,
            product_model,
            row_value(product, "synonyms", ""),
        )
        line_is_accessory = text_mentions_samsung_watch_accessory(line_clean)
        if product_is_accessory != line_is_accessory:
            return False

        product_watch_family = extract_samsung_watch_family(product_name, product_model)
        line_watch_family = extract_samsung_watch_family(line_clean)
        if product_watch_family and line_watch_family and product_watch_family != line_watch_family:
            return False

        product_watch_sizes = extract_samsung_watch_case_sizes(product_name, product_model)
        line_watch_sizes = extract_samsung_watch_case_sizes(line_clean)
        if product_watch_sizes and line_watch_sizes and product_watch_sizes.isdisjoint(line_watch_sizes):
            return False

    product_insta360_model = extract_insta360_model_key(product_name, product_model)
    line_insta360_model = extract_insta360_model_key(line_clean)
    if product_insta360_model and line_insta360_model and product_insta360_model != line_insta360_model:
        return False
    product_mentions_insta360 = "insta360" in normalize_match_text(f"{product_name} {product_brand} {product_model}")
    if line_insta360_model and product_mentions_insta360 and not product_insta360_model:
        return False

    year_sensitive_categories = {"macbook", "imac", "ipad", "apple_watch", "samsung_watch"}
    product_years = extract_match_years(product_name, product_model)
    line_years = extract_match_years(line_clean)
    if line_years and product_category in year_sensitive_categories:
        product_watch_family = extract_apple_watch_family(product_name, product_model) if product_category == "apple_watch" else ""
        line_watch_family = extract_apple_watch_family(line_clean) if product_category == "apple_watch" else ""
        has_watch_family_match = product_category == "apple_watch" and not product_years and apple_watch_families_compatible(
            product_watch_family,
            line_watch_family,
            line_years,
        )
        if not has_watch_family_match and (not product_years or product_years.isdisjoint(line_years)):
            return False
    elif product_years and line_years and product_years.isdisjoint(line_years):
        return False

    product_capacities = extract_match_capacities(product_name, product_storage)
    line_capacities = extract_match_capacities(line_clean)
    if product_capacities and line_capacities and product_capacities.isdisjoint(line_capacities):
        return False

    product_storage_values = extract_match_storage_values(product_name, product_storage)
    line_storage_values = extract_match_storage_values(line_clean)
    if product_storage_values and line_storage_values and product_storage_values.isdisjoint(line_storage_values):
        return False

    product_ram_values = extract_match_ram_values(product_name, product_storage)
    line_ram_values = extract_match_ram_values(line_clean)
    if product_ram_values and line_ram_values and product_ram_values.isdisjoint(line_ram_values):
        return False

    product_connectivity = extract_match_connectivity(product_name)
    line_connectivity = extract_match_connectivity(line_clean)
    if product_connectivity and line_connectivity:
        if product_connectivity.isdisjoint(line_connectivity):
            return False
        if "cellular" in product_connectivity and "cellular" not in line_connectivity:
            return False
        if "cellular" in line_connectivity and "cellular" not in product_connectivity:
            return False

    product_chips = extract_match_apple_chips(product_name)
    line_chips = extract_match_apple_chips(line_clean)
    if product_chips and line_chips and product_chips.isdisjoint(line_chips):
        return False

    product_colors = extract_match_colors(product_name) or extract_match_colors(product_color)
    line_colors = extract_match_colors(line_clean)
    if product_colors and line_colors and not match_colors_compatible(product_colors, line_colors):
        return False

    model_key = normalize_match_text(product_model)
    if model_key and len(model_key) >= 4 and model_key in normalize_match_text(line_clean):
        return True

    return True


def iphone_shorthand_product_line_matches(product, line_text):
    product_name = row_value(product, "name", "") or ""
    product_brand = row_value(product, "brand", "") or ""
    if not is_apple_iphone_product(product_name, product_brand or "Apple"):
        return False

    line_clean = strip_supplier_line_noise(line_text)
    product_family = detect_iphone_model_family(product_name)
    line_family = detect_iphone_model_family(line_clean)
    if not product_family or not line_family or product_family != line_family:
        return False

    if not product_line_semantics_compatible(product, line_text, require_known_category=True):
        return False

    product_storage = row_value(product, "storage", "") or ""
    product_capacities = extract_match_capacities(product_name, product_storage)
    line_capacities = extract_match_capacities(line_clean)
    if not product_capacities or not line_capacities or product_capacities.isdisjoint(line_capacities):
        return False

    product_color = row_value(product, "color", "") or ""
    product_colors = extract_match_colors(product_name) or extract_match_colors(product_color)
    line_colors = extract_match_colors(line_clean)
    if not product_colors or not line_colors or not match_colors_compatible(product_colors, line_colors):
        return False

    return True


def ipad_shorthand_product_line_matches(product, line_text):
    product_name = row_value(product, "name", "") or ""
    product_brand = row_value(product, "brand", "") or ""
    detected_product_brand = sanitize_brand_value(product_brand) or detect_match_brand(product_name)
    if detected_product_brand and detected_product_brand != "Apple":
        return False
    if detect_match_category(product_name) != "ipad":
        return False

    product_family = detect_ipad_model_family(product_name, row_value(product, "model_number", ""))
    line_family = detect_ipad_model_family(line_text)
    if not product_family or not line_family or product_family != line_family:
        return False

    if not product_line_semantics_compatible(product, line_text, require_known_category=True):
        return False

    product_storage = row_value(product, "storage", "") or ""
    product_capacities = extract_match_capacities(product_name, product_storage)
    line_capacities = extract_match_capacities(line_text)
    if not product_capacities or not line_capacities or product_capacities.isdisjoint(line_capacities):
        return False

    product_connectivity = extract_match_connectivity(product_name, row_value(product, "synonyms", ""))
    line_connectivity = extract_match_connectivity(line_text)
    if not product_connectivity or not line_connectivity or product_connectivity.isdisjoint(line_connectivity):
        return False

    product_color = row_value(product, "color", "") or ""
    product_colors = extract_match_colors(product_name) or extract_match_colors(product_color)
    line_colors = extract_match_colors(line_text)
    if not product_colors or not line_colors or not match_colors_compatible(product_colors, line_colors):
        return False

    return True


def extract_xiaomi_tablet_family(*values):
    normalized = normalize_match_text(" ".join(str(value or "") for value in values))
    if not normalized:
        return ""
    patterns = (
        (r"\bredmi\s+pad\s*2\s+pro\b", "redmi_pad_2_pro"),
        (r"\bredmi\s+pad\s*2\b", "redmi_pad_2"),
        (r"\bpoco\s+pad\s+x1\b", "poco_pad_x1"),
        (r"\bpoco\s+pad\s+m1\b", "poco_pad_m1"),
        (r"\bpoco\s+pad\b", "poco_pad"),
        (r"\bxiaomi\s+pad\s*7\s+pro\b", "xiaomi_pad_7_pro"),
        (r"\bxiaomi\s+pad\s*7\b", "xiaomi_pad_7"),
    )
    for pattern, family in patterns:
        if re.search(pattern, normalized):
            return family
    return ""


def xiaomi_tablet_shorthand_product_line_matches(product, line_text):
    product_name = row_value(product, "name", "") or ""
    product_model = row_value(product, "model_number", "") or ""
    if detect_match_category(product_name) != "xiaomi_tablet" or detect_match_category(line_text) != "xiaomi_tablet":
        return False

    product_family = extract_xiaomi_tablet_family(product_name, product_model, row_value(product, "synonyms", ""))
    line_family = extract_xiaomi_tablet_family(line_text)
    if not product_family or not line_family or product_family != line_family:
        return False

    if not product_line_semantics_compatible(product, line_text, require_known_category=True):
        return False

    product_storage = row_value(product, "storage", "") or ""
    product_capacities = extract_match_capacities(product_name, product_storage)
    line_capacities = extract_match_capacities(line_text)
    if not product_capacities or not line_capacities or product_capacities.isdisjoint(line_capacities):
        return False

    product_ram_values = extract_match_ram_values(product_name, product_storage)
    line_ram_values = extract_match_ram_values(line_text)
    if product_ram_values and line_ram_values and product_ram_values.isdisjoint(line_ram_values):
        return False

    product_color = row_value(product, "color", "") or ""
    product_colors = extract_match_colors(product_name) or extract_match_colors(product_color)
    line_colors = extract_match_colors(line_text)
    if not product_colors or not line_colors or not match_colors_compatible(product_colors, line_colors):
        return False

    return True


def extract_xiaomi_vacuum_family(*values):
    normalized = normalize_match_text(" ".join(str(value or "") for value in values))
    if not normalized:
        return ""
    normalized = normalized.replace("dreame bot", "dreame")
    patterns = (
        (r"\bdreame\s+d20\s+plus\b", "dreame_d20_plus"),
        (r"\bdreame\s+d20\s+pro\b", "dreame_d20_pro"),
        (r"\bdreame\s+d20\b", "dreame_d20"),
        (r"\bdreame\s+f10\s+plus\b", "dreame_f10_plus"),
        (r"\bdreame\s+r10s\s+lite\b", "dreame_r10s_lite"),
        (r"\bdreame\s+r10s\s+essential\b", "dreame_r10s_essential"),
        (r"\bdreame\s+d10s\s+plus\b", "dreame_d10s_plus"),
        (r"\bdreame\s+d10s\s+pro\b", "dreame_d10s_pro"),
        (r"\bdreame\s+d10s\b", "dreame_d10s"),
        (r"\bdreame\s+l10s\s+ultra\b", "dreame_l10s_ultra"),
        (r"\bdreame\s+l10s\s+plus\b", "dreame_l10s_plus"),
        (r"\bdreame\s+l10\s+prime\b", "dreame_l10_prime"),
        (r"\bdreame\s+l10\s+pro\b", "dreame_l10_pro"),
        (r"\bdreame\s+w10\s+pro\b", "dreame_w10_pro"),
        (r"\bdreame\s+w10\b", "dreame_w10"),
        (r"\bdreame\s+d9\s+max\s+gen\s*2\b", "dreame_d9_max_gen2"),
        (r"\bdreame\s+l20\s+ultra(?:\s+complete)?\b", "dreame_l20_ultra"),
        (r"\bdreame\s+l40s\s+pro\s+ultra\b", "dreame_l40s_pro_ultra"),
        (r"\bdreame\s+x40\s+ultra(?:\s+complete)?\b", "dreame_x40_ultra"),
        (r"\bdreame\s+x50\s+master\b", "dreame_x50_master"),
        (r"\bdreame\s+x50\s+ultra(?:\s+complete)?\b", "dreame_x50_ultra"),
        (r"\broborock\s+q\s*revo\s+c\b", "roborock_q_revo_c"),
        (r"\broborock\s+q8\s+max\s+pro\b", "roborock_q8_max_pro"),
        (r"\broborock\s+saros\s+z70\b", "roborock_saros_z70"),
        (r"\broborock\s+s8\b", "roborock_s8"),
    )
    for pattern, family in patterns:
        if re.search(pattern, normalized):
            return family
    return ""


def xiaomi_vacuum_shorthand_product_line_matches(product, line_text):
    product_name = row_value(product, "name", "") or ""
    product_model = row_value(product, "model_number", "") or ""
    if detect_match_category(product_name) != "xiaomi_vacuum" or detect_match_category(line_text) != "xiaomi_vacuum":
        return False

    product_family = extract_xiaomi_vacuum_family(product_name, product_model, row_value(product, "synonyms", ""))
    line_family = extract_xiaomi_vacuum_family(line_text)
    if not product_family or not line_family or product_family != line_family:
        return False

    if not product_line_semantics_compatible(product, line_text, require_known_category=True):
        return False

    product_color = row_value(product, "color", "") or ""
    product_colors = extract_match_colors(product_name) or extract_match_colors(product_color)
    line_colors = extract_match_colors(line_text)
    if product_colors and line_colors and not match_colors_compatible(product_colors, line_colors):
        return False

    return True


def extract_playstation_console_family(*values):
    normalized = normalize_match_text(" ".join(str(value or "") for value in values))
    if not normalized:
        return ""
    match = re.search(r"\bcfi[\s-]?(\d{2,4})\s*([ab])(?:0?1)?\b", normalized)
    if match:
        return f"cfi{match.group(1)}{match.group(2)}"
    match = re.search(r"\bcfi[\s-]?(\d{2,4})\b", normalized)
    if match:
        return f"cfi{match.group(1)}"
    if re.search(r"\bps\s*5\b|\bplaystation\s*5\b", normalized):
        match = re.search(r"\b(7[01]\d{2}|2[01]\d{2})\s*([abа])?\b", normalized)
        if match:
            suffix = (match.group(2) or "").replace("а", "a")
            return f"cfi{match.group(1)}{suffix}"
    return ""


def is_playstation_digital_text(*values):
    normalized = normalize_match_text(" ".join(str(value or "") for value in values))
    return bool(re.search(r"\bdigital(?:\s+edition)?\b", normalized))


def is_playstation_pro_text(*values):
    normalized = normalize_match_text(" ".join(str(value or "") for value in values))
    return bool(re.search(r"\bps\s*5\b|\bplaystation\s*5\b", normalized) and re.search(r"\bpro\b", normalized))


def is_playstation_gamepad_special_edition(*values):
    normalized = normalize_match_text(" ".join(str(value or "") for value in values))
    markers = (
        "astro bot", "marathon", "ghost", "helldivers", "spider man", "spiderman",
        "the last of us", "fortnite", "death stranding", "anniversary", "леброн",
        "альпийский", "ремикс", "хром", "limited edition", "edge",
    )
    return any(marker in normalized for marker in markers)


def is_playstation_gamepad_accessory_text(*values):
    normalized = normalize_match_text(" ".join(str(value or "") for value in values))
    markers = ("заряд", "charging", "charge", "станц", "station", "module", "модул", "dock")
    return any(marker in normalized for marker in markers)


def extract_playstation_gamepad_color_family(*values):
    normalized = normalize_match_text(" ".join(str(value or "") for value in values))
    if not normalized:
        return ""
    if re.search(r"\bcamo(?:uflage)?\b", normalized) or "камуфляж" in normalized:
        return "camouflage"
    if "chrome teal" in normalized or "бирюз" in normalized:
        return "teal"
    if "chrome indigo" in normalized or re.search(r"\bindigo\b", normalized) or "индиго" in normalized:
        return "indigo"
    if "chrome pearl" in normalized or "pearl" in normalized or "жемчуж" in normalized:
        return "pearl"
    if "sterling silver" in normalized or "серебрист" in normalized:
        return "silver"
    if "alpine green" in normalized or "альпийск" in normalized:
        return "alpine_green"
    if "midnight black" in normalized or "черный" in normalized or "чёрный" in normalized:
        return "black"
    if "white" in normalized or "белый" in normalized:
        return "white"
    if "cobalt blue" in normalized or "кобальтово" in normalized:
        return "blue"
    if "nova pink" in normalized or "розовый" in normalized:
        return "pink"
    if "galactic purple" in normalized or "фиолетовый" in normalized:
        return "purple"
    if "volcanic red" in normalized or "вулканический" in normalized:
        return "volcanic_red"
    if "cosmic red" in normalized or re.search(r"\bкрасный\b", normalized):
        return "red"
    return ""


def extract_playstation_gamepad_variant_key(*values):
    normalized = normalize_match_text(" ".join(str(value or "") for value in values))
    if not normalized:
        return ""
    patterns = (
        (r"\bastro\s*bot\b", "astro_bot"),
        (r"\bghost\s+of\s+yotei\b", "ghost_of_yotei"),
        (r"\bmarathon\b", "marathon"),
        (r"\bfortnite\b", "fortnite"),
        (r"\bthe\s+last\s+of\s+us\b", "the_last_of_us"),
        (r"\bhelldivers\b", "helldivers"),
        (r"\bspider[\s-]*man\b|\bspiderman\b", "spider_man"),
        (r"\bdeath\s+stranding\b", "death_stranding"),
        (r"\banniversary\b", "anniversary"),
        (r"\bedge\b", "edge"),
        (r"\bcamo(?:uflage)?\b|камуфляж", "camouflage"),
        (r"\balpine\s+green\b|альпийск", "alpine_green"),
        (r"\bgreen\s+remix\b|зелен\w+\s+ремикс|зелён\w+\s+ремикс", "green_remix"),
        (r"\bchrome\s+teal\b|бирюз\w+\s+хром", "chrome_teal"),
        (r"\bchrome\s+indigo\b|индиго\s+хром", "chrome_indigo"),
        (r"\bchrome\s+pearl\b|жемчуж\w+\s+хром", "chrome_pearl"),
        (r"\bsterling\s+silver\b|серебрист", "sterling_silver"),
        (r"\bvolcanic\s+red\b|вулканическ", "volcanic_red"),
        (r"\bcobalt\s+blue\b|кобальтов", "cobalt_blue"),
    )
    for pattern, key in patterns:
        if re.search(pattern, normalized):
            return key
    return ""


def is_playstation_game_listing_text(*values):
    normalized = normalize_match_text(" ".join(str(value or "") for value in values))
    if not normalized:
        return False
    return bool(
        re.search(r"\b(?:игра|игры|game)\s+(?:для\s+)?ps\s*5\b", normalized)
        or re.search(r"\b(?:игра|игры|game)\s+(?:для\s+)?playstation\s*5\b", normalized)
    )


def playstation_console_shorthand_product_line_matches(product, line_text):
    product_name = row_value(product, "name", "") or ""
    product_model = row_value(product, "model_number", "") or ""
    if is_playstation_game_listing_text(product_name, row_value(product, "synonyms", "")):
        return False
    if detect_match_category(product_name) != "playstation_console" or detect_match_category(line_text) != "playstation_console":
        return False

    product_family = extract_playstation_console_family(product_name, product_model)
    line_family = extract_playstation_console_family(line_text)
    if product_family and line_family and product_family != line_family:
        return False

    product_is_digital = is_playstation_digital_text(product_name, product_model)
    line_is_digital = is_playstation_digital_text(line_text)
    pro_digital_listing = is_playstation_pro_text(product_name, product_model) and is_playstation_pro_text(line_text)
    if product_is_digital != line_is_digital and not pro_digital_listing:
        return False

    if not product_line_semantics_compatible(product, line_text, require_known_category=True):
        return False
    return True


def playstation_gamepad_shorthand_product_line_matches(product, line_text):
    product_name = row_value(product, "name", "") or ""
    product_model = row_value(product, "model_number", "") or ""
    if detect_match_category(product_name) != "playstation_gamepad" or detect_match_category(line_text) != "playstation_gamepad":
        return False

    if is_playstation_gamepad_accessory_text(product_name, product_model, row_value(product, "synonyms", "")) or is_playstation_gamepad_accessory_text(line_text):
        return False

    product_variant_key = extract_playstation_gamepad_variant_key(
        product_name,
        product_model,
        row_value(product, "synonyms", ""),
    )
    line_variant_key = extract_playstation_gamepad_variant_key(line_text)
    if product_variant_key or line_variant_key:
        if product_variant_key != line_variant_key:
            return False

    if is_playstation_gamepad_special_edition(product_name, product_model) and not is_playstation_gamepad_special_edition(line_text):
        return False

    product_color_family = extract_playstation_gamepad_color_family(product_name, row_value(product, "color", ""), row_value(product, "synonyms", ""))
    line_color_family = extract_playstation_gamepad_color_family(line_text)
    if product_color_family and line_color_family and product_color_family != line_color_family:
        return False

    if not product_line_semantics_compatible(product, line_text, require_known_category=True):
        return False
    return True


def gaming_match_text(*values):
    return normalize_product_search_text(strip_supplier_line_decorations(" ".join(str(value or "") for value in values)))


def extract_gaming_storage_values(*values):
    storage_values = set(extract_match_storage_values(*values))
    storage_values.update(extract_match_capacities(*values))
    text = gaming_match_text(*values)
    if re.search(r"\b(?:quest|oculus|meta|steamdeck|steam\s*deck|switch|xbox|ps5|playstation)\b", text):
        for amount in re.findall(r"\b(64|128|256|512|825)\b", text):
            storage_values.add(int(amount))
    return storage_values


def gaming_storage_compatible(product_values, line_values):
    product_values = set(product_values or ())
    line_values = set(line_values or ())
    return not product_values or not line_values or not product_values.isdisjoint(line_values)


def gaming_variant_markers_compatible(product_name, line_text):
    product_text = gaming_match_text(product_name)
    line_text_norm = gaming_match_text(line_text)
    raw_line = str(line_text or "")
    if "gotov k rabote" in product_text or "gotov k rabote v rf" in product_text or "готов" in normalize_match_text(product_name):
        if not re.search(r"\b(?:готов|rf|рф)\b", line_text_norm):
            return False
    if "yaponskaya versiya" in product_text or "япон" in normalize_match_text(product_name):
        if not ("🇯🇵" in raw_line or re.search(r"\b(?:japan|jp|япон)\b", line_text_norm)):
            return False
    if re.search(r"\b(?:limited|anniversary|spider\s*man|fortnite|helldivers|death\s*stranding|marathon|god\s*of\s*war)\b", product_text):
        if not re.search(r"\b(?:limited|anniversary|spider\s*man|fortnite|helldivers|death\s*stranding|marathon|god\s*of\s*war)\b", line_text_norm):
            return False
    if "mario" in product_text and "mario" not in line_text_norm and "марио" not in normalize_match_text(line_text):
        return False
    if "xbox edition" in product_text and "xbox" not in line_text_norm:
        return False
    return True


def gaming_colors_compatible(product, line_text):
    product_colors = extract_match_colors(row_value(product, "name", ""), row_value(product, "color", ""))
    line_colors = extract_match_colors(strip_supplier_line_noise(line_text))
    return match_colors_compatible(product_colors, line_colors)


def extract_meta_quest_family(*values):
    text = gaming_match_text(*values)
    if re.search(r"\b(?:meta\s*)?(?:oculus\s*)?quest\b.*\b3s\b|\bquest3s\b", text):
        return "quest3s"
    if re.search(r"\b(?:meta\s*)?(?:oculus\s*)?quest\b.*\b3\b|\bquest3\b", text):
        return "quest3"
    if re.search(r"\b(?:meta\s*)?(?:oculus\s*)?quest\b.*\b2\b|\bquest2\b", text):
        return "quest2"
    return ""


def extract_steam_deck_family(*values):
    text = gaming_match_text(*values)
    if "steamdeck" in text or "steam deck" in text:
        return "steam_deck_oled" if "oled" in text else "steam_deck"
    return ""


def extract_nintendo_switch_family(*values):
    text = gaming_match_text(*values)
    if not re.search(r"\b(?:nintendo\s+)?switch\b|\bnsw2?\b", text):
        return ""
    if re.search(r"\bswitch\s*2\b|\bnsw2\b", text):
        return "switch2"
    if "oled" in text:
        return "switch_oled"
    if "lite" in text:
        return "switch_lite"
    return "switch"


def extract_xbox_family(*values):
    text = gaming_match_text(*values)
    if re.search(r"\b(?:xbox|x\s*box)\s+(?:series\s+)?x\b|\bseries\s+x\b", text):
        return "xbox_series_x"
    if re.search(r"\b(?:xbox|x\s*box)\s+(?:series\s+)?s\b|\bseries\s+s\b", text):
        return "xbox_series_s"
    return ""


def is_xbox_digital_text(*values):
    return bool(re.search(r"\bdigital\b", gaming_match_text(*values)))


def is_xbox_disk_text(*values):
    return bool(re.search(r"\b(?:disk|disc)\b", gaming_match_text(*values)))


def extract_playstation_accessory_key(*values):
    text = gaming_match_text(*values)
    if re.search(r"\bpulse\s*elite\b", text):
        return "pulse_elite"
    if re.search(r"\bpulse\s*3d\b|\b3d\s*pulse\b", text):
        return "pulse_3d"
    if re.search(r"\bps\s*portal\b|\bps5\s*portal\b|\bplaystation\s*portal\b", text):
        return "ps_portal"
    if re.search(r"\bvr\s*2\b|\bvr2\b|playstation\s*vr\s*2", text):
        return "ps_vr2"
    if re.search(r"\bdual\s*sense\b|\bdualsense\b", text) and re.search(r"\b(?:charging|charge|station|заряд\w*|станц\w*)\b", text):
        return "dualsense_charging_station"
    if re.search(r"\b(?:vertical\s*stand|подстав)\b", text) and re.search(r"\b(?:ps5|playstation|slim|pro)\b", text):
        return "ps5_vertical_stand"
    if re.search(r"\b(?:disc\s*drive|disk\s*drive|дисковод)\b", text):
        return "ps5_disc_drive"
    if re.search(r"\blogitech\s*g29\b|\bg29\b", text):
        return "logitech_g29"
    return ""


def playstation_game_title_key(*values):
    text = gaming_match_text(*values)
    text = text.replace("тhe", "the")
    patterns = (
        (r"\bfc\s*26\b|\bfc26\b", "fc26"),
        (r"\b(?:the\s+)?last\s+of\s+us\s*(?:part\s*)?(?:2|ii)\b", "last_of_us_2"),
        (r"\bufc\s*5\b|\bufc5\b", "ufc5"),
        (r"\bmk\s*1\b|\bmortal\s*(?:kombat\s*)?1\b", "mortal_kombat_1"),
        (r"\bmk\s*11\b|\bmortal\s*kombat\s*11\b", "mortal_kombat_11"),
        (r"\bresident\s*evil\s*requiem\b", "resident_evil_requiem"),
        (r"\bgta\s*5\b|\bgrand\s*theft\s*auto\s*v\b", "gta_5"),
        (r"\bminecraft\s*legends\b", "minecraft_legends"),
        (r"\bminecraft\b", "minecraft"),
        (r"\bspider[\s-]*man\s*2\b", "spider_man_2"),
        (r"\bgran\s*turismo\s*7\b", "gran_turismo_7"),
    )
    for pattern, key in patterns:
        if re.search(pattern, text):
            return key
    return ""


def product_is_account_game_variant(product_name):
    return bool(re.search(r"\b(?:account|учетн|учётн|оформлен)\b", gaming_match_text(product_name)))


GAMING_MATCH_MARKER_RE = re.compile(
    r"\b(?:ps5|playstation|dualsense|xbox|switch|nintendo|steamdeck|steam\s*deck|"
    r"quest|oculus|meta|pulse|portal|vr2|vr\s*2|g29|logitech)\b",
    re.IGNORECASE,
)


def gaming_shorthand_product_line_matches(product, line_text):
    product_name = row_value(product, "name", "") or ""
    product_model = row_value(product, "model_number", "") or ""
    product_probe = gaming_match_text(
        product_name,
        product_model,
        row_value(product, "brand", ""),
        row_value(product, "synonyms", ""),
    )
    line_probe = gaming_match_text(line_text)
    if not GAMING_MATCH_MARKER_RE.search(line_probe) or not GAMING_MATCH_MARKER_RE.search(product_probe):
        return False
    product_category = detect_match_category(
        f"{product_name} {row_value(product, 'brand', '')} {row_value(product, 'synonyms', '')}"
    )
    line_category = detect_match_category(line_text)

    product_game_key = playstation_game_title_key(product_name, row_value(product, "synonyms", ""))
    line_game_key = playstation_game_title_key(line_text)
    if product_game_key and line_game_key:
        if product_game_key != line_game_key:
            return False
        if product_is_account_game_variant(product_name) and not product_is_account_game_variant(line_text):
            return False
        return product_line_semantics_compatible(product, line_text, require_known_category=False)

    product_accessory_key = extract_playstation_accessory_key(product_name, product_model, row_value(product, "synonyms", ""))
    line_accessory_key = extract_playstation_accessory_key(line_text)
    if product_accessory_key and line_accessory_key:
        if product_accessory_key != line_accessory_key:
            return False
        return (
            gaming_variant_markers_compatible(product_name, line_text)
            and gaming_colors_compatible(product, line_text)
            and product_line_semantics_compatible(product, line_text, require_known_category=False)
        )

    product_quest = extract_meta_quest_family(product_name, product_model, row_value(product, "synonyms", ""))
    line_quest = extract_meta_quest_family(line_text)
    if product_quest and line_quest:
        if product_quest != line_quest:
            return False
        return (
            gaming_variant_markers_compatible(product_name, line_text)
            and gaming_storage_compatible(
                extract_gaming_storage_values(product_name, row_value(product, "storage", "")),
                extract_gaming_storage_values(line_text),
            )
            and gaming_colors_compatible(product, line_text)
            and product_line_semantics_compatible(product, line_text, require_known_category=False)
        )

    product_steam = extract_steam_deck_family(product_name, product_model, row_value(product, "synonyms", ""))
    line_steam = extract_steam_deck_family(line_text)
    if product_steam and line_steam:
        if line_steam == "steam_deck_oled" and product_steam != "steam_deck_oled":
            return False
        return (
            gaming_variant_markers_compatible(product_name, line_text)
            and gaming_storage_compatible(
                extract_gaming_storage_values(product_name, row_value(product, "storage", "")),
                extract_gaming_storage_values(line_text),
            )
            and gaming_colors_compatible(product, line_text)
            and product_line_semantics_compatible(product, line_text, require_known_category=False)
        )

    product_xbox = extract_xbox_family(product_name, product_model, row_value(product, "synonyms", ""))
    line_xbox = extract_xbox_family(line_text)
    if product_xbox and line_xbox and product_category in {"xbox_console", "xbox_gamepad"}:
        if product_xbox != line_xbox:
            return False
        if product_category == "xbox_console":
            if is_xbox_digital_text(line_text) and not is_xbox_digital_text(product_name, product_model):
                return False
            if is_xbox_disk_text(line_text) and is_xbox_digital_text(product_name, product_model):
                return False
            if not gaming_storage_compatible(
                extract_gaming_storage_values(product_name, row_value(product, "storage", "")),
                extract_gaming_storage_values(line_text),
            ):
                return False
        return (
            gaming_variant_markers_compatible(product_name, line_text)
            and gaming_colors_compatible(product, line_text)
            and product_line_semantics_compatible(product, line_text, require_known_category=False)
        )

    if product_category == "xbox_gamepad" and line_category == "xbox_gamepad":
        return (
            gaming_variant_markers_compatible(product_name, line_text)
            and gaming_colors_compatible(product, line_text)
            and product_line_semantics_compatible(product, line_text, require_known_category=False)
        )

    product_switch = extract_nintendo_switch_family(product_name, product_model, row_value(product, "synonyms", ""))
    line_switch = extract_nintendo_switch_family(line_text)
    if product_switch and line_switch:
        if product_switch != line_switch:
            return False
        return (
            gaming_variant_markers_compatible(product_name, line_text)
            and gaming_storage_compatible(
                extract_gaming_storage_values(product_name, row_value(product, "storage", "")),
                extract_gaming_storage_values(line_text),
            )
            and gaming_colors_compatible(product, line_text)
            and product_line_semantics_compatible(product, line_text, require_known_category=False)
        )

    return False


def detect_earpods_connector(*values):
    text = normalize_product_search_text(" ".join(str(value or "") for value in values))
    if re.search(r"\b(?:usb\s*c|usbc|type\s*c|typec)\b", text):
        return "usb_c"
    if "lightning" in text:
        return "lightning"
    if re.search(r"\b3\s*5\b|\bjack\b|\baux\b", text):
        return "3_5"
    return ""


def earpods_shorthand_product_line_matches(product, line_text):
    product_name = row_value(product, "name", "") or ""
    product_blob = f"{product_name} {row_value(product, 'synonyms', '')} {row_value(product, 'model_number', '')}"
    line_blob = strip_supplier_line_noise(line_text)
    if "earpods" not in normalize_product_search_text(product_blob) or "earpods" not in normalize_product_search_text(line_blob):
        return False
    product_connector = detect_earpods_connector(product_blob)
    line_connector = detect_earpods_connector(line_blob)
    if product_connector and line_connector and product_connector != line_connector:
        return False
    return product_line_semantics_compatible(product, line_text, require_known_category=False)


def airpods_shorthand_product_line_matches(product, line_text):
    product_name = row_value(product, "name", "") or ""
    product_model = row_value(product, "model_number", "") or ""
    if detect_match_category(product_name) != "airpods" or detect_match_category(line_text) != "airpods":
        return False

    product_family = detect_airpods_model_family(product_name, product_model)
    line_family = detect_airpods_model_family(line_text)
    if not product_family or not line_family:
        return False
    if not airpods_model_families_compatible(product_family, line_family):
        return False

    if not product_line_semantics_compatible(product, line_text, require_known_category=True):
        return False

    if "AirPods Max" in product_family or "AirPods Max" in line_family:
        product_color = row_value(product, "color", "") or ""
        product_colors = extract_match_colors(product_name) or extract_match_colors(product_color)
        line_colors = extract_match_colors(line_text)
        if not product_colors or not line_colors or not match_colors_compatible(product_colors, line_colors):
            return False

    return True


def apple_watch_shorthand_product_line_matches(product, line_text):
    product_name = row_value(product, "name", "") or ""
    product_model = row_value(product, "model_number", "") or ""
    if detect_match_category(product_name) != "apple_watch" or detect_match_category(line_text) != "apple_watch":
        return False
    if text_mentions_apple_watch_protective_accessory(product_name, product_model, row_value(product, "synonyms", "")):
        return False
    if text_mentions_apple_watch_protective_accessory(line_text):
        return False

    product_family = extract_apple_watch_family(product_name, product_model)
    line_family = extract_apple_watch_family(line_text)
    line_years = extract_match_years(line_text)
    if not apple_watch_families_compatible(product_family, line_family, line_years):
        return False

    product_sizes = extract_apple_watch_case_sizes(product_name, product_model)
    line_sizes = extract_apple_watch_case_sizes(line_text)
    if not line_sizes and line_family.startswith("ultra"):
        line_sizes.add("49")
    if not product_sizes or not line_sizes or product_sizes.isdisjoint(line_sizes):
        return False

    product_color = row_value(product, "color", "") or ""
    product_case_colors = extract_apple_watch_product_case_colors(product_name, product_model, product_color)
    line_case_colors = extract_apple_watch_line_case_colors(line_text)
    if not product_case_colors or not line_case_colors:
        return False
    if not match_colors_compatible(product_case_colors, line_case_colors):
        return False

    normalized_product_name = normalize_match_text(product_name)
    normalized_line = normalize_match_text(line_text)
    if "hermes" in normalized_product_name and "hermes" not in normalized_line:
        return False
    product_is_titanium_case = "titanium case" in normalized_product_name
    line_has_titanium_context = bool(re.search(r"\b(?:ti|titanium|титан)\b", normalized_line))
    line_has_premium_band_context = bool(
        re.search(r"\b(?:milanese|milaneze|milenese|tml|link\s+bracelet)\b", normalized_line)
    )
    if product_is_titanium_case and not product_family.startswith("ultra") and not (line_has_titanium_context or line_has_premium_band_context):
        return False

    product_band_colors = extract_apple_watch_product_band_colors(product_name)
    line_band_colors = extract_apple_watch_line_band_colors(line_text)
    if product_band_colors and line_band_colors:
        if len(product_band_colors) >= 2 and len(line_band_colors) == 1:
            if not product_band_colors.issubset(line_band_colors):
                return False
        elif len(line_band_colors) >= 2:
            if not product_band_colors.issuperset(line_band_colors):
                return False
        elif not match_colors_compatible(product_band_colors, line_band_colors):
            return False

    product_band_sizes = extract_apple_watch_band_sizes(product_name, product_model)
    line_band_sizes = extract_apple_watch_band_sizes(line_text)
    if not apple_watch_band_sizes_compatible(product_band_sizes, line_band_sizes):
        return False

    product_band_kinds = extract_apple_watch_band_kinds(product_name, product_model)
    line_band_kinds = extract_apple_watch_band_kinds(line_text)
    if line_band_sizes and not line_band_colors and not line_band_kinds and not product_family.startswith("ultra"):
        default_band_colors = apple_watch_default_sport_band_colors(product_family, product_case_colors)
        if product_band_kinds and "sport_band" not in product_band_kinds:
            return False
        if default_band_colors and product_band_colors and not match_colors_compatible(product_band_colors, default_band_colors):
            return False
    if not line_band_sizes and "milanese" in product_band_kinds and product_band_sizes and "one" not in product_band_sizes:
        return False
    if product_band_kinds and line_band_kinds and product_band_kinds.isdisjoint(line_band_kinds):
        return False

    return product_line_semantics_compatible(product, line_text, require_known_category=True)


def product_supplier_line_matches(product, line_text, line_words=None, synonym_sets=None, require_known_category=False, allow_article_conflict=False):
    line_words = line_words if line_words is not None else clean_for_search(line_text)
    synonym_sets = synonym_sets if synonym_sets is not None else (
        row_value(product, "_synonym_sets", None) or build_synonym_word_sets(row_value(product, "synonyms", ""))
    )
    if not allow_article_conflict and product_line_has_strict_article_conflict(product, line_text):
        return False
    product_model_tokens = get_product_model_tokens(product)
    line_model_tokens = extract_catalog_model_tokens(line_text)
    if (
        product_model_tokens
        and line_model_tokens
        and product_model_tokens.intersection(line_model_tokens)
        and product_line_semantics_compatible(product, line_text, require_known_category=require_known_category)
    ):
        return True
    if (
        apple_model_no_in_text(row_value(product, "model_number", ""), line_text)
        and product_line_semantics_compatible(product, line_text, require_known_category=require_known_category)
    ):
        return True
    if iphone_shorthand_product_line_matches(product, line_text):
        return True
    if ipad_shorthand_product_line_matches(product, line_text):
        return True
    if xiaomi_tablet_shorthand_product_line_matches(product, line_text):
        return True
    if xiaomi_vacuum_shorthand_product_line_matches(product, line_text):
        return True
    if playstation_console_shorthand_product_line_matches(product, line_text):
        return True
    if playstation_gamepad_shorthand_product_line_matches(product, line_text):
        return True
    if gaming_shorthand_product_line_matches(product, line_text):
        return True
    if earpods_shorthand_product_line_matches(product, line_text):
        return True
    if airpods_shorthand_product_line_matches(product, line_text):
        return True
    if apple_watch_shorthand_product_line_matches(product, line_text):
        return True
    if detect_match_category(row_value(product, "name", "")) == "apple_watch" or detect_match_category(line_text) == "apple_watch":
        return False
    if not product_line_matches_synonym_sets(line_words, synonym_sets):
        return False
    return product_line_semantics_compatible(product, line_text, require_known_category=require_known_category)


def source_label_matches_auto_confirm(value):
    normalized = str(value or "").casefold()
    return any(marker in normalized for marker in AUTO_CONFIRM_SOURCE_MARKERS)


def row_value(row, key, default=None):
    if row is None:
        return default
    if isinstance(row, dict):
        return row.get(key, default)
    try:
        return row[key]
    except (KeyError, IndexError, TypeError):
        return default


def source_row_is_auto_confirm_enabled(row):
    if any(source_label_matches_auto_confirm(row_value(row, key)) for key in ("sender_name", "chat_title", "custom_name")):
        return True
    return any(str(row_value(row, key, "") or "").strip() for key in ("custom_name", "chat_title", "sender_name"))


def binding_product_row_matches_line(binding_row, line_text, *, require_known_category=False):
    if not line_text:
        return False
    product_payload = {
        "name": row_value(binding_row, "name", ""),
        "country": row_value(binding_row, "country", ""),
        "synonyms": row_value(binding_row, "synonyms", ""),
        "brand": row_value(binding_row, "brand", ""),
        "model_number": row_value(binding_row, "model_number", ""),
        "color": row_value(binding_row, "color", ""),
        "storage": row_value(binding_row, "storage", ""),
        "warranty": row_value(binding_row, "warranty", ""),
        "description": row_value(binding_row, "description", ""),
        "description_html": row_value(binding_row, "description_html", ""),
        "specs": row_value(binding_row, "specs", ""),
    }
    if binding_row_has_explicit_line_identity(product_payload, line_text):
        return product_condition_markers_compatible(product_payload, line_text)
    return product_supplier_line_matches(
        product_payload,
        line_text,
        line_words=clean_for_search(line_text),
        require_known_category=require_known_category,
    )


def split_product_synonyms(value):
    return [part.strip() for part in str(value or "").split(",") if part.strip()]


def get_product_canonical_article_tokens(product):
    direct_tokens = extract_exact_catalog_article_tokens(
        row_value(product, "model_number", ""),
        row_value(product, "name", ""),
    )
    if direct_tokens:
        return direct_tokens
    synonym_tokens = set()
    for synonym in split_product_synonyms(row_value(product, "synonyms", ""))[:20]:
        synonym_tokens.update(extract_exact_catalog_article_tokens(synonym))
        if len(synonym_tokens) > 1:
            return set()
    return synonym_tokens


def product_line_has_strict_article_conflict(product, line_text):
    line_tokens = extract_exact_catalog_article_tokens(line_text)
    if not line_tokens:
        return False
    product_tokens = get_product_canonical_article_tokens(product)
    if not product_tokens:
        return False
    return product_tokens.isdisjoint(line_tokens)


def normalize_synonym_key(value):
    return " ".join(str(value or "").replace("ё", "е").casefold().split())


def supplier_line_synonym(line):
    text = strip_supplier_line_noise(line)
    text = TELEGRAM_FLAG_PAIR_RE.sub(" ", text)
    text = re.sub(r"\s+", " ", text).strip(" \t-–—")
    return text


def make_supplier_missing_position_key(line):
    return normalize_synonym_key(supplier_line_synonym(line))


def serialize_auto_confirm_candidates(candidates, limit=8):
    payload = []
    for item in list(candidates or ())[:limit]:
        payload.append({
            "product_id": row_value(item, "id"),
            "product_name": row_value(item, "name", ""),
        })
    return json.dumps(payload, ensure_ascii=False)


def ensure_supplier_missing_positions_table(db):
    db.executescript(
        """
        CREATE TABLE IF NOT EXISTS supplier_missing_positions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            source_chat_id TEXT,
            source_name TEXT,
            message_id INTEGER,
            line_index INTEGER DEFAULT -1,
            line_text TEXT NOT NULL,
            match_line TEXT,
            normalized_key TEXT NOT NULL,
            brand TEXT,
            category TEXT,
            price_value REAL,
            status TEXT DEFAULT 'unmatched',
            candidate_count INTEGER DEFAULT 0,
            candidate_snapshot TEXT,
            seen_count INTEGER DEFAULT 1,
            first_seen_at TEXT,
            last_seen_at TEXT,
            resolved_product_id INTEGER,
            resolved_at TEXT,
            FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE,
            UNIQUE(user_id, normalized_key)
        );
        CREATE INDEX IF NOT EXISTS idx_supplier_missing_positions_user_status
            ON supplier_missing_positions(user_id, status, last_seen_at);
        """
    )


def ensure_supplier_line_memory_table(db):
    db.executescript(
        """
        CREATE TABLE IF NOT EXISTS supplier_line_memory (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            normalized_key TEXT NOT NULL,
            product_id INTEGER NOT NULL,
            source_name TEXT,
            message_id INTEGER,
            line_text TEXT,
            seen_count INTEGER DEFAULT 1,
            first_confirmed_at TEXT,
            last_confirmed_at TEXT,
            FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE,
            UNIQUE(user_id, normalized_key)
        );
        CREATE INDEX IF NOT EXISTS idx_supplier_line_memory_user_product
            ON supplier_line_memory(user_id, product_id);
        """
    )


def remember_supplier_line_binding(db, user_id, product_id, line, *, message=None, increment_seen=True):
    normalized_key = make_supplier_missing_position_key(line)
    if not normalized_key or not product_id:
        return 0
    ensure_supplier_line_memory_table(db)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    source_name = ""
    message_id = None
    if message is not None:
        source_name = (
            row_value(message, "custom_name", "")
            or row_value(message, "chat_title", "")
            or row_value(message, "sender_name", "")
            or str(row_value(message, "chat_id", ""))
        )
        message_id = row_value(message, "id")
    existing = db.execute(
        """
        SELECT id, product_id
        FROM supplier_line_memory
        WHERE user_id = ? AND normalized_key = ?
        LIMIT 1
        """,
        (user_id, normalized_key),
    ).fetchone()
    if existing:
        if not increment_seen:
            try:
                existing_product_id = int(existing["product_id"])
            except (TypeError, ValueError):
                existing_product_id = 0
            if existing_product_id == int(product_id):
                return 0
        seen_count_sql = (
            "seen_count = COALESCE(seen_count, 0) + 1,"
            if increment_seen
            else "seen_count = MAX(COALESCE(seen_count, 1), 1),"
        )
        db.execute(
            f"""
            UPDATE supplier_line_memory
            SET product_id = ?,
                source_name = ?,
                message_id = ?,
                line_text = ?,
                {seen_count_sql}
                last_confirmed_at = ?
            WHERE id = ?
            """,
            (product_id, source_name, message_id, str(line or ""), now, int(existing["id"])),
        )
        return 1
    db.execute(
        """
        INSERT INTO supplier_line_memory (
            user_id, normalized_key, product_id, source_name,
            message_id, line_text, seen_count, first_confirmed_at, last_confirmed_at
        )
        VALUES (?, ?, ?, ?, ?, ?, 1, ?, ?)
        """,
        (user_id, normalized_key, product_id, source_name, message_id, str(line or ""), now, now),
    )
    return 1


def get_supplier_memory_product(db, user_id, line, products_by_id):
    normalized_key = make_supplier_missing_position_key(line)
    if not normalized_key:
        return None
    ensure_supplier_line_memory_table(db)
    row = db.execute(
        """
        SELECT product_id
        FROM supplier_line_memory
        WHERE user_id = ? AND normalized_key = ?
        LIMIT 1
        """,
        (user_id, normalized_key),
    ).fetchone()
    if not row:
        return None
    try:
        product_id = int(row["product_id"])
    except (TypeError, ValueError):
        return None
    product = (products_by_id or {}).get(product_id)
    if not product:
        return None
    if not product_supplier_line_matches(
        product,
        line,
        line_words=clean_for_search(line),
        synonym_sets=product.get("_synonym_sets"),
        require_known_category=True,
    ):
        return None
    return product


def upsert_supplier_missing_position(
    db,
    user_id,
    message,
    line_index,
    line,
    *,
    match_line="",
    price=None,
    status="unmatched",
    candidates=None,
):
    normalized_key = make_supplier_missing_position_key(line)
    if not normalized_key:
        return 0
    ensure_supplier_missing_positions_table(db)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    source_name = (
        row_value(message, "custom_name", "")
        or row_value(message, "chat_title", "")
        or row_value(message, "sender_name", "")
        or str(row_value(message, "chat_id", ""))
    )
    normalized_match_line = supplier_line_synonym(match_line or line)
    detected_brand = detect_match_brand(normalized_match_line)
    detected_category = detect_match_category(normalized_match_line)
    candidate_list = list(candidates or ())
    candidate_snapshot = serialize_auto_confirm_candidates(candidate_list)
    candidate_count = len(candidate_list)
    existing = db.execute(
        """
        SELECT id
        FROM supplier_missing_positions
        WHERE user_id = ? AND normalized_key = ?
        LIMIT 1
        """,
        (user_id, normalized_key),
    ).fetchone()
    payload = (
        row_value(message, "chat_id"),
        source_name,
        int(message["id"]),
        int(line_index),
        str(line or ""),
        normalized_match_line,
        normalized_key,
        detected_brand,
        detected_category,
        price,
        str(status or "unmatched"),
        candidate_count,
        candidate_snapshot,
        now,
    )
    if existing:
        db.execute(
            """
            UPDATE supplier_missing_positions
            SET source_chat_id = ?,
                source_name = ?,
                message_id = ?,
                line_index = ?,
                line_text = ?,
                match_line = ?,
                normalized_key = ?,
                brand = ?,
                category = ?,
                price_value = ?,
                status = ?,
                candidate_count = ?,
                candidate_snapshot = ?,
                last_seen_at = ?,
                seen_count = COALESCE(seen_count, 0) + 1,
                resolved_product_id = NULL,
                resolved_at = NULL
            WHERE id = ?
            """,
            (*payload, int(existing["id"])),
        )
        return 1
    db.execute(
        """
        INSERT INTO supplier_missing_positions (
            user_id, source_chat_id, source_name, message_id, line_index,
            line_text, match_line, normalized_key, brand, category,
            price_value, status, candidate_count, candidate_snapshot,
            first_seen_at, last_seen_at, seen_count
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
        """,
        (
            user_id,
            *payload,
            now,
        ),
    )
    return 1


def resolve_supplier_missing_position(db, user_id, line, *, product_id=None):
    normalized_key = make_supplier_missing_position_key(line)
    if not normalized_key:
        return 0
    ensure_supplier_missing_positions_table(db)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cursor = db.execute(
        """
        UPDATE supplier_missing_positions
        SET status = 'resolved',
            resolved_product_id = ?,
            resolved_at = ?,
            last_seen_at = ?
        WHERE user_id = ?
          AND normalized_key = ?
          AND COALESCE(status, '') != 'resolved'
        """,
        (product_id, now, now, user_id, normalized_key),
    )
    return cursor.rowcount or 0


def learn_supplier_line_binding(
    db,
    user_id,
    product,
    line,
    *,
    message=None,
    add_synonym=True,
    increment_memory_seen=True,
):
    result = {"memory_written": 0, "missing_resolved": 0, "synonyms_added": 0}
    try:
        product_id = int(row_value(product, "id") or 0)
    except (TypeError, ValueError):
        product_id = 0
    if not product_id or not str(line or "").strip():
        return result

    result["memory_written"] = remember_supplier_line_binding(
        db,
        user_id,
        product_id,
        line,
        message=message,
        increment_seen=increment_memory_seen,
    )
    result["missing_resolved"] = resolve_supplier_missing_position(
        db,
        user_id,
        line,
        product_id=product_id,
    )

    if add_synonym:
        synonym = supplier_line_synonym(line)
        if synonym:
            current = db.execute(
                "SELECT synonyms FROM products WHERE id = ? AND user_id = ?",
                (product_id, user_id),
            ).fetchone()
            existing_synonyms = row_value(current, "synonyms", row_value(product, "synonyms", ""))
            added, _ = append_product_synonym(db, product_id, existing_synonyms, synonym)
            result["synonyms_added"] = added
    return result


def backfill_supplier_learning_from_confirmed_bindings(db, user_id=None, limit=5000):
    ensure_supplier_line_memory_table(db)
    ensure_supplier_missing_positions_table(db)
    filters = [
        "pm.status = 'confirmed'",
        "COALESCE(pm.is_actual, 1) = 1",
        "m.text IS NOT NULL",
        "(COALESCE(m.is_blocked, 0) = 0)",
        "(COALESCE(m.is_delayed, 0) = 0)",
    ]
    params = []
    if user_id is not None:
        filters.append("p.user_id = ?")
        params.append(user_id)
    params.append(int(limit or 5000))

    rows = db.execute(
        f"""
        SELECT
            p.id AS product_id,
            p.user_id AS product_user_id,
            p.name AS product_name,
            p.country AS product_country,
            p.synonyms AS product_synonyms,
            p.brand AS product_brand,
            p.model_number AS product_model_number,
            p.color AS product_color,
            p.storage AS product_storage,
            p.warranty AS product_warranty,
            p.description AS product_description,
            p.description_html AS product_description_html,
            p.specs AS product_specs,
            pm.message_id,
            pm.line_index,
            pm.extracted_price,
            m.text AS message_text,
            m.chat_id,
            m.sender_name,
            m.chat_title,
            tc.custom_name
        FROM product_messages pm
        JOIN products p ON p.id = pm.product_id
        JOIN messages m ON m.id = pm.message_id
        LEFT JOIN tracked_chats tc ON tc.chat_id = m.chat_id AND tc.user_id = p.user_id
        WHERE {' AND '.join(filters)}
        ORDER BY m.date DESC, pm.id DESC
        LIMIT ?
        """,
        params,
    ).fetchall()

    stats = {
        "rows_seen": 0,
        "rows_learned": 0,
        "memory_written": 0,
        "missing_resolved": 0,
        "synonyms_added": 0,
        "skipped": 0,
    }
    for row in rows:
        stats["rows_seen"] += 1
        source_line, resolved_price = resolve_binding_display_line_and_price(
            row["message_text"],
            row["line_index"],
            row["extracted_price"],
        )
        if not source_line or resolved_price is None:
            stats["skipped"] += 1
            continue
        product = {
            "id": row["product_id"],
            "name": row["product_name"],
            "country": row["product_country"],
            "synonyms": row["product_synonyms"],
            "brand": row["product_brand"],
            "model_number": row["product_model_number"],
            "color": row["product_color"],
            "storage": row["product_storage"],
            "warranty": row["product_warranty"],
            "description": row["product_description"],
            "description_html": row["product_description_html"],
            "specs": row["product_specs"],
        }
        if not binding_product_row_matches_line(product, source_line, require_known_category=True):
            stats["skipped"] += 1
            continue
        message = {
            "id": row["message_id"],
            "chat_id": row["chat_id"],
            "sender_name": row["sender_name"],
            "chat_title": row["chat_title"],
            "custom_name": row["custom_name"],
        }
        learned = learn_supplier_line_binding(
            db,
            row["product_user_id"],
            product,
            source_line,
            message=message,
            increment_memory_seen=False,
        )
        stats["rows_learned"] += 1
        stats["memory_written"] += learned["memory_written"]
        stats["missing_resolved"] += learned["missing_resolved"]
        stats["synonyms_added"] += learned["synonyms_added"]
    return stats


def supplier_line_has_price_marker(line):
    return extract_price(line) is not None


SUPPLIER_CONTEXT_HEADER_STOP_WORDS = {
    "показать полностью",
    "свернуть",
    "show more",
    "profile",
    "профиль",
}
SUPPLIER_CONTEXT_HEADER_SECTION_STOP_WORDS = {
    "accessories",
    "chair",
    "chairs",
    "console",
    "consoles",
    "controller",
    "controllers",
    "gamepad",
    "gamepads",
    "gaming",
    "gaming chair",
    "gaming chairs",
    "nintendo",
    "nintendo switch",
    "playstation",
    "ps",
    "ps4",
    "ps5",
    "seat",
    "seats",
    "sony",
    "switch",
    "xbox",
    "аксессуары",
    "гейминг",
    "геймпады",
    "игровые аксессуары",
    "консоли",
    "кресла",
    "приставки",
    "стулья",
}


def supplier_context_header(line):
    if not line:
        return ""
    if (supplier_line_has_price_marker(line) or extract_price(line)) and re.search(
        r"(?:[—–-]\s*\d{3,}|\d[\d\s]*(?:₽|руб\.?|р\.|usd|eur|\$|€))",
        str(line or ""),
        re.IGNORECASE,
    ):
        return ""
    if line_is_discount_section_marker(line):
        return ""

    header = supplier_line_synonym(line)
    header = re.sub(r"[_=*~`]+", " ", header)
    header = re.sub(r"^[^\w]+|[^\w]+$", " ", header, flags=re.UNICODE)
    header = re.sub(r"\s+", " ", header).strip(" \t-–—:•·")
    normalized = normalize_match_text(header)
    if len(normalized) < 3 or len(normalized) > 90:
        return ""
    if normalized in SUPPLIER_CONTEXT_HEADER_STOP_WORDS:
        return ""
    compact_normalized = re.sub(r"[^a-zа-я0-9]+", " ", normalized).strip()
    if compact_normalized in SUPPLIER_CONTEXT_HEADER_SECTION_STOP_WORDS:
        return ""
    header_words = set(compact_normalized.split())
    if header_words and header_words.issubset(SUPPLIER_CONTEXT_HEADER_SECTION_STOP_WORDS):
        return ""
    if any(marker in compact_normalized for marker in ("chair", "chairs", "кресл", "стул", "racing cockpit")):
        return ""
    if "http://" in normalized or "https://" in normalized or "t.me/" in normalized:
        return ""
    if extract_match_capacities(header) or extract_match_storage_values(header) or extract_match_ram_values(header):
        return ""
    if not re.search(r"[a-zа-я0-9]", normalized):
        return ""
    return header


def update_supplier_context_headers(headers, line, max_headers=2):
    header = supplier_context_header(line)
    if not header:
        return headers
    return append_supplier_context_header(headers, header, max_headers=max_headers)


def append_supplier_context_header(headers, header, max_headers=2):
    key = normalize_synonym_key(header)
    kept = [item for item in headers if normalize_synonym_key(item) != key]
    kept.append(header)
    return kept[-max_headers:]


def make_supplier_contextual_line(headers, line):
    clean_line = str(line or "").strip()
    if not clean_line:
        return clean_line
    normalized_line = normalize_match_text(clean_line)
    useful_headers = [
        header for header in headers or ()
        if header and normalize_match_text(header) not in normalized_line
    ]
    if not useful_headers:
        return clean_line
    return f"{' '.join(useful_headers)} {clean_line}"


def line_is_supplier_catalog_section_header(line):
    return bool(supplier_catalog_section_header_text(line))


def supplier_catalog_section_header_text(line):
    raw_line = str(line or "").strip()
    if not raw_line:
        return ""
    if re.search(r"(?:[—–-]\s*\d{3,}|\d[\d\s]*(?:₽|руб\.?|р\.|usd|eur|\$|€))", raw_line, re.IGNORECASE):
        return ""
    header = supplier_line_synonym(raw_line)
    header = re.sub(r"[_=*~`■□▪▫●•]+", " ", header)
    header = re.sub(r"^[^\w]+|[^\w]+$", " ", header, flags=re.UNICODE)
    header = re.sub(r"\s+", " ", header).strip(" \t-–—:•·")
    normalized = normalize_match_text(header)
    if len(normalized) < 2 or len(normalized) > 60:
        return ""
    if normalized in SUPPLIER_CONTEXT_HEADER_STOP_WORDS:
        return ""
    if not re.search(r"\d", normalized):
        return ""
    if len(normalized.split()) > 6:
        return ""
    return header


def source_text_has_supplier_price_lines(text):
    in_discount_section = False
    for raw_line in str(text or "").split("\n"):
        line = raw_line.strip()
        if line_is_discount_section_marker(line):
            in_discount_section = True
            continue
        if in_discount_section:
            continue
        if line_has_discount_condition_marker(line):
            continue
        if line_is_supplier_catalog_section_header(line):
            continue
        if line and supplier_line_has_price_marker(line) and extract_price(line):
            return True
    return False


def source_text_has_supplier_stop_signal(text):
    normalized = normalize_match_text(str(text or ""))
    if not normalized:
        return False
    stop_patterns = (
        r"\b(?:продажи|заказы)\s+(?:временно\s+)?закрыт",
        r"\bпри[её]м\s+заказов\s+(?:временно\s+)?закрыт",
        r"\bпри[её]м\s+(?:временно\s+)?закрыт",
        r"\b(?:продажи|заказы)\s+(?:остановлен|приостановлен|завершен)",
        r"\bкаталог\s+(?:временно\s+)?недоступ",
        r"\bожидайте\s+старта\s+продаж",
    )
    return any(re.search(pattern, normalized, re.IGNORECASE) for pattern in stop_patterns)


def append_product_synonym(db, product_id, existing_synonyms, synonym):
    synonym = str(synonym or "").strip()
    if not synonym:
        return 0, existing_synonyms or ""
    items = split_product_synonyms(existing_synonyms)
    keys = {normalize_synonym_key(item) for item in items}
    key = normalize_synonym_key(synonym)
    if not key or key in keys:
        return 0, existing_synonyms or ""
    items.append(synonym)
    updated = ", ".join(items)
    db.execute("UPDATE products SET synonyms = ? WHERE id = ?", (updated, product_id))
    return 1, updated


def get_product_model_tokens(product):
    existing = row_value(product, "_model_tokens", None)
    if existing is not None:
        return {str(token).upper() for token in existing if str(token or "").strip()}
    return extract_catalog_model_tokens(
        row_value(product, "model_number", ""),
        row_value(product, "name", ""),
        row_value(product, "synonyms", ""),
    )


def signature_token_is_useful(token, model_tokens):
    token = str(token or "").strip().lower()
    if not token:
        return False
    if token in MATCH_COLOR_ALIAS_SEARCH_TOKENS:
        return False
    if token in PRODUCT_SIGNATURE_DROP_TOKENS:
        return False
    if token in MATCH_LOW_SIGNAL_TOKENS or token.translate(MATCH_LOOKALIKE_TRANSLATION) in MATCH_LOW_SIGNAL_SEARCH_TOKENS:
        return False
    if token.upper() in model_tokens:
        return False
    if re.fullmatch(r"\d+(?:gb|гб|tb|тб|mm|мм|м)", token):
        return False
    if re.fullmatch(r"20\d{2}", token):
        return False
    if re.fullmatch(r"\d{5,}", token):
        return False
    if len(token) < 2:
        return False
    return True


def auto_confirm_seed_token_is_useful(token, line_words):
    token = str(token or "").strip().lower()
    if not token:
        return False
    if token.startswith("color_") or token.startswith("conn_"):
        return False
    if token in MATCH_LOW_SIGNAL_TOKENS or token.translate(MATCH_LOOKALIKE_TRANSLATION) in MATCH_LOW_SIGNAL_SEARCH_TOKENS:
        return False
    line_category = ""
    for word in line_words or ():
        if str(word).startswith("cat_"):
            line_category = str(word)[4:]
            break
    if line_category == "samsung_watch" and re.fullmatch(r"(?:sm)?l\d{3,4}[a-z]?", token, re.IGNORECASE):
        return False
    return True


def normalize_signature_words(text, model_tokens):
    words = clean_for_search(strip_supplier_line_noise(text))
    return {
        token for token in words
        if signature_token_is_useful(token, model_tokens)
    }


def signature_word_set_is_informative(words):
    words = set(words or ())
    if len(words) < 4:
        return False
    has_numberish = any(re.search(r"\d", token) for token in words)
    has_named_detail = any(re.search(r"[a-zа-я]", token) for token in words)
    return has_numberish and has_named_detail


def build_product_signature_word_sets(product, model_tokens):
    sets = []
    sources = [row_value(product, "name", "")]
    sources.extend(split_product_synonyms(row_value(product, "synonyms", ""))[:12])
    sources.extend(build_derived_product_search_aliases(product)[:40])
    for source in sources:
        words = normalize_signature_words(source, model_tokens)
        if signature_word_set_is_informative(words):
            sets.append(words)
            relaxed = set(words)
            relaxed.discard("macbook")
            relaxed.discard("apple")
            if signature_word_set_is_informative(relaxed):
                sets.append(relaxed)

    unique = []
    seen = set()
    for words in sets:
        key = tuple(sorted(words))
        if key in seen:
            continue
        seen.add(key)
        unique.append(words)
    return unique


def prepare_auto_confirm_product(product):
    product["_match_category"] = detect_match_category(
        " ".join(
            str(row_value(product, key, "") or "")
            for key in ("name", "brand", "model_number")
        )
    )
    model_tokens = get_product_model_tokens(product)
    product["_model_tokens"] = model_tokens
    base_product = dict(product or {})
    base_product["synonyms"] = ""
    synonym_sets = []
    for synonym in split_product_synonyms(product.get("synonyms"))[:60]:
        if not product_line_semantics_compatible(base_product, synonym, require_known_category=False):
            continue
        syn_set = clean_for_search(synonym)
        if synonym_word_set_is_informative(syn_set):
            synonym_sets.append(syn_set)
    synonym_sets.extend(build_product_signature_word_sets(product, model_tokens))

    unique_sets = []
    seen = set()
    for syn_set in synonym_sets:
        key = tuple(sorted(syn_set))
        if key in seen:
            continue
        seen.add(key)
        unique_sets.append(syn_set)
    product["_synonym_sets"] = unique_sets
    seed_tokens = set()
    for syn_set in unique_sets:
        seed_tokens.update(str(token or "").strip().lower() for token in syn_set if str(token or "").strip())
    seed_tokens.update(str(token or "").strip().lower() for token in model_tokens if str(token or "").strip())
    product["_auto_confirm_seed_tokens"] = seed_tokens
    product["_model_token"] = next(iter(sorted(model_tokens)), "").lower()
    return product


def find_existing_folder_path_id(db, user_id, folder_names):
    parent_id = None
    for folder_name in normalize_folder_path_parts(folder_names):
        row = find_folder_by_name_and_parent(db, user_id, folder_name, parent_id)
        if not row:
            return None
        parent_id = int(row["id"])
    return parent_id


def get_descendant_folder_ids(db, user_id, folder_id):
    normalized_folder_id = normalize_folder_id(folder_id)
    if normalized_folder_id is None:
        return []
    rows = db.execute(
        """
        WITH RECURSIVE subtree(id) AS (
            SELECT id
            FROM folders
            WHERE id = ? AND user_id = ?
            UNION ALL
            SELECT f.id
            FROM folders f
            JOIN subtree s ON f.parent_id = s.id
            WHERE f.user_id = ?
        )
        SELECT id FROM subtree
        """,
        (normalized_folder_id, user_id, user_id),
    ).fetchall()
    return [int(row["id"]) for row in rows]


AUTO_CONFIRM_CATEGORY_SQL_KEYWORDS = {
    "display": ("display", "monitor", "монитор"),
    "imac": ("imac", "аймак", "моноблок"),
    "macbook": ("macbook", "макбук"),
    "ipad": ("ipad", "айпад"),
    "apple_watch": ("apple watch", "watch"),
    "airpods": ("airpods", "air pods", "аирподс", "эйрподс"),
    "apple_keyboard": ("magic keyboard", "keyboard", "клавиатур"),
    "apple_mouse": ("magic mouse", "mouse", "мыш"),
    "apple_trackpad": ("magic trackpad", "trackpad", "трекпад"),
    "iphone": ("iphone", "айфон"),
    "samsung_phone": ("samsung", "galaxy"),
    "samsung_tablet": ("samsung", "galaxy tab", "tab"),
    "samsung_watch": ("samsung", "galaxy watch", "watch"),
    "samsung_earbuds": ("samsung", "galaxy buds", "buds"),
    "xiaomi_tablet": ("xiaomi", "redmi pad", "poco pad", "pad"),
    "xiaomi_phone": ("xiaomi", "redmi", "poco"),
    "xiaomi_vacuum": ("xiaomi", "dreame", "roborock", "пылесос"),
    "google_phone": ("pixel", "google"),
    "realme_phone": ("realme",),
    "huawei_phone": ("huawei", "mate", "pura", "nova"),
    "garmin_watch": ("garmin",),
    "playstation_gamepad": ("dualsense", "dual sense", "геймпад"),
    "playstation_console": ("playstation", "ps5", "cfi"),
    "dyson_airstrait": ("dyson", "airstrait"),
    "insta360": ("insta360", "insta 360"),
    "earbuds": ("buds", "earbuds", "наушник"),
    "keyboard": ("keyboard", "клавиатур"),
    "camera": ("camera", "canon", "nikon", "sony", "fujifilm", "lumix", "gopro", "dji"),
    "purifier": ("purifier", "очистител"),
}


def collect_auto_confirm_message_categories(lines):
    categories = set()
    in_discount_section = False
    context_headers = []
    for raw_line in lines:
        line = str(raw_line or "").strip()
        if not line:
            continue
        if line_is_discount_section_marker(line):
            in_discount_section = True
            continue
        if in_discount_section or line_has_discount_condition_marker(line):
            continue
        section_header = supplier_catalog_section_header_text(line)
        if section_header:
            context_headers = append_supplier_context_header(context_headers, section_header)
            continue
        if not supplier_line_has_price_marker(line) or not extract_price(line):
            context_headers = update_supplier_context_headers(context_headers, line)
            continue
        category = detect_match_category(line)
        if category:
            categories.add(category)
        context_line = make_supplier_contextual_line(context_headers, line)
        if context_line != line:
            context_category = detect_match_category(context_line)
            if context_category:
                categories.add(context_category)
    return categories


def build_auto_confirm_message_scope_sql_filter(scope_keywords):
    scope_keywords = [
        str(keyword or "").casefold().strip()
        for keyword in (scope_keywords or [])
        if len(str(keyword or "").strip()) >= 2
    ]
    scope_keywords = list(dict.fromkeys(scope_keywords))
    if not scope_keywords:
        return "", []

    haystack = (
        "LOWER(COALESCE(name, '') || ' ' || COALESCE(brand, '') || ' ' || "
        "COALESCE(model_number, '') || ' ' || COALESCE(synonyms, ''))"
    )
    clauses = [f"{haystack} LIKE ?" for _ in scope_keywords]
    return " AND (" + " OR ".join(clauses) + ")", [f"%{keyword}%" for keyword in scope_keywords]


def build_auto_confirm_category_sql_filter(categories):
    categories = {str(category or "").strip() for category in (categories or set()) if str(category or "").strip()}
    if not categories:
        return "", []

    keywords = []
    for category in sorted(categories):
        category_keywords = AUTO_CONFIRM_CATEGORY_SQL_KEYWORDS.get(category)
        if not category_keywords:
            return "", []
        keywords.extend(category_keywords)

    keywords = list(dict.fromkeys(str(keyword or "").casefold() for keyword in keywords if str(keyword or "").strip()))
    if not keywords:
        return "", []

    haystack = (
        "LOWER(COALESCE(name, '') || ' ' || COALESCE(brand, '') || ' ' || "
        "COALESCE(model_number, '') || ' ' || COALESCE(synonyms, ''))"
    )
    clauses = [f"{haystack} LIKE ?" for _ in keywords]
    return " AND (" + " OR ".join(clauses) + ")", [f"%{keyword}%" for keyword in keywords]


def collect_auto_confirm_message_scope_keywords(lines, categories):
    categories = {str(category or "").strip() for category in (categories or set()) if str(category or "").strip()}
    scope_keywords = set()
    if "iphone" not in categories:
        return scope_keywords

    in_discount_section = False
    context_headers = []
    for raw_line in lines:
        line = str(raw_line or "").strip()
        if not line:
            continue
        if line_is_discount_section_marker(line):
            in_discount_section = True
            continue
        if in_discount_section or line_has_discount_condition_marker(line):
            continue
        section_header = supplier_catalog_section_header_text(line)
        if section_header:
            context_headers = append_supplier_context_header(context_headers, section_header)
            continue
        if not supplier_line_has_price_marker(line) or not extract_price(line):
            context_headers = update_supplier_context_headers(context_headers, line)
            continue

        match_line = make_supplier_contextual_line(context_headers, line)
        family = detect_iphone_model_family(match_line) or detect_iphone_model_family(line)
        if not family:
            continue
        family_keyword = str(family).casefold().strip()
        if not family_keyword:
            continue
        scope_keywords.add(family_keyword)
        scope_keywords.add(f"iphone {family_keyword}")

    return scope_keywords


def infer_auto_confirm_source_folder_ids(db, user_id, text):
    raw_text = str(text or "")
    normalized = normalize_match_text(raw_text)
    if "iphone" not in normalized:
        return []
    # Bot catalog responses use explicit phone section headers. Do not scope
    # accessory price lists that only mention iPhone cases or glass.
    if not re.search(r"^\s*📱\s*i\s*phone\b", raw_text, re.IGNORECASE | re.MULTILINE):
        return []
    iphone_folder_id = find_existing_folder_path_id(db, user_id, IPHONE_CATALOG_PATH)
    return get_descendant_folder_ids(db, user_id, iphone_folder_id)


def load_products_for_source_auto_confirm(db, user_id, folder_ids=None, categories=None, scope_keywords=None):
    normalized_folder_ids = [
        int(folder_id)
        for folder_id in (folder_ids or [])
        if normalize_folder_id(folder_id) is not None
    ]
    normalized_folder_ids = list(dict.fromkeys(normalized_folder_ids))
    folder_filter = ""
    params = [user_id]
    if normalized_folder_ids:
        placeholders = ",".join("?" for _ in normalized_folder_ids)
        folder_filter = f" AND folder_id IN ({placeholders})"
        params.extend(normalized_folder_ids)
    category_filter, category_params = build_auto_confirm_category_sql_filter(categories)
    params.extend(category_params)
    scope_filter, scope_params = build_auto_confirm_message_scope_sql_filter(scope_keywords)
    params.extend(scope_params)

    rows = db.execute(
        f"""
        SELECT id, name, country, synonyms, brand, model_number, color, storage,
               warranty, description, description_html, specs, folder_id
        FROM products
        WHERE user_id = ?
        {folder_filter}
        {category_filter}
        {scope_filter}
        ORDER BY id
        """,
        tuple(params),
    ).fetchall()
    products = []
    for row in rows:
        product = dict(row)
        prepare_auto_confirm_product(product)
        if product["_synonym_sets"] or product["_model_tokens"]:
            products.append(product)
    return products


def build_auto_confirm_token_index(products):
    token_index = {}
    for index, product in enumerate(products):
        tokens = set()
        for synonym_set in product.get("_synonym_sets") or ():
            tokens.update(synonym_set)
        for model_token in product.get("_model_tokens") or ():
            tokens.add(str(model_token).lower())
        for token in tokens:
            token = str(token or "").strip()
            if len(token) < 2:
                continue
            token_index.setdefault(token, set()).add(index)
    return token_index


def get_auto_confirm_candidate_pool(products, token_index, line_words, max_seed_tokens=8):
    hits = []
    for token in line_words:
        product_indexes = token_index.get(token)
        if product_indexes:
            hits.append((len(product_indexes), token, product_indexes))
    if not hits:
        return []
    seed_hits = [
        hit for hit in hits
        if auto_confirm_seed_token_is_useful(hit[1], line_words)
    ] or hits
    seed_hits.sort(key=lambda item: item[0])
    candidate_indexes = set(seed_hits[0][2])
    for _, _, product_indexes in seed_hits[1:max_seed_tokens]:
        narrowed = candidate_indexes.intersection(product_indexes)
        if narrowed:
            candidate_indexes = narrowed
            if len(candidate_indexes) <= 300:
                break

    return [products[index] for index in sorted(candidate_indexes)]


def get_auto_confirm_search_fallback_pool(db, user_id, products_by_id, line_text, limit=160):
    query = strip_supplier_line_noise(line_text) or str(line_text or "").strip()
    if not query:
        return []
    try:
        search_index = get_cached_product_search_entries(db, user_id)
        candidate_ids = search_products_in_entries(search_index, query, limit=limit)
    except Exception:
        logger.exception("Auto-confirm search fallback failed for query: %s", query)
        return []

    candidates = []
    for product_id in candidate_ids:
        product = products_by_id.get(int(product_id))
        if product:
            candidates.append(product)
    return candidates


def find_auto_confirm_synonym_fallback_candidates(products, line, line_words=None, limit=160, allow_article_conflict=False):
    line_words = line_words if line_words is not None else clean_for_search(line)
    if not line_words:
        return []
    strong_tokens = {
        str(token or "").strip().lower()
        for token in line_words
        if auto_confirm_seed_token_is_useful(token, line_words)
    }
    candidates = []
    for product in products:
        seed_tokens = product.get("_auto_confirm_seed_tokens") or set()
        if strong_tokens and seed_tokens and not seed_tokens.intersection(strong_tokens):
            continue
        if product_supplier_line_matches(
            product,
            line,
            line_words=line_words,
            synonym_sets=product.get("_synonym_sets"),
            require_known_category=True,
            allow_article_conflict=allow_article_conflict,
        ):
            candidates.append(product)
            if len(candidates) >= limit:
                break
    return candidates


def find_auto_confirm_product_candidates(products, line, line_words=None):
    line_words = line_words if line_words is not None else clean_for_search(line)
    if not line_words:
        return []
    line_clean = strip_supplier_line_noise(line)
    line_family = detect_iphone_model_family(line_clean)
    candidates = []
    if line_family:
        for product in products:
            if not is_apple_iphone_product(
                row_value(product, "name", ""),
                row_value(product, "brand", "") or "Apple",
            ):
                continue
            product_family = detect_iphone_model_family(row_value(product, "name", ""))
            if product_family and product_family != line_family:
                continue
            if iphone_shorthand_product_line_matches(product, line):
                candidates.append(product)
        return candidates

    for product in products:
        if line_family and not is_apple_iphone_product(product.get("name"), "Apple"):
            continue
        if product_supplier_line_matches(
            product,
            line,
            line_words=line_words,
            synonym_sets=product.get("_synonym_sets"),
            require_known_category=True,
        ):
            candidates.append(product)
    return candidates


def find_auto_confirm_article_conflict_candidates(products, line, line_words=None):
    line_words = line_words if line_words is not None else clean_for_search(line)
    if not line_words:
        return []
    conflicts = []
    for product in products:
        if not product_line_has_strict_article_conflict(product, line):
            continue
        if product_supplier_line_matches(
            product,
            line,
            line_words=line_words,
            synonym_sets=product.get("_synonym_sets"),
            require_known_category=True,
            allow_article_conflict=True,
        ):
            conflicts.append(product)
    return conflicts


def line_already_bound_to_other_product(db, message_id, line_index, product_id):
    row = db.execute(
        """
        SELECT product_id
        FROM product_messages
        WHERE message_id = ? AND line_index = ? AND product_id != ?
          AND COALESCE(is_actual, 1) = 1
        LIMIT 1
        """,
        (message_id, line_index, product_id),
    ).fetchone()
    return bool(row)


def upsert_auto_confirmed_binding(db, product, message, line_index, line, price, apply=True):
    product_id = int(product["id"])
    message_id = int(message["id"])
    source_line, resolved_price = resolve_binding_display_line_and_price(
        row_value(message, "text", ""),
        line_index,
        price,
        require_price=False,
    )
    if not source_line:
        return "invalid"
    if line_already_bound_to_other_product(db, message_id, line_index, product_id):
        return "occupied"

    same_item_binding = find_existing_confirmed_binding_by_fingerprint(
        db,
        product_id,
        message["chat_id"],
        message["sender_name"],
        source_line,
        text=row_value(message, "text", ""),
        line_index=line_index,
        price=resolved_price,
    )
    if not apply:
        return "updated" if same_item_binding else "inserted"

    if same_item_binding:
        db.execute(
            """
            DELETE FROM product_messages
            WHERE product_id = ? AND message_id = ? AND line_index = ? AND id != ?
            """,
            (product_id, message_id, line_index, same_item_binding["id"]),
        )
        db.execute(
            """
            UPDATE product_messages
            SET message_id = ?, line_index = ?, status = 'confirmed', extracted_price = ?, is_actual = 1
            WHERE id = ?
            """,
            (message_id, line_index, resolved_price, same_item_binding["id"]),
        )
        return "updated"

    db.execute(
        """
        INSERT INTO product_messages (product_id, message_id, line_index, status, extracted_price, is_actual)
        VALUES (?, ?, ?, 'confirmed', ?, 1)
        ON CONFLICT(product_id, message_id, line_index)
        DO UPDATE SET status='confirmed', extracted_price=excluded.extracted_price, is_actual=1
        """,
        (product_id, message_id, line_index, resolved_price),
    )
    return "inserted"


def infer_conflict_preferred_folder_id(conflict_candidates):
    folder_ids = [
        normalize_folder_id(row_value(product, "folder_id"))
        for product in list(conflict_candidates or ())
        if normalize_folder_id(row_value(product, "folder_id")) is not None
    ]
    if not folder_ids:
        return None
    counts = Counter(folder_ids)
    return counts.most_common(1)[0][0]


def create_product_for_supplier_line(db, user_id, line, preferred_folder_id=None):
    product_name = supplier_line_synonym(line)
    if not product_name:
        return None

    article_tokens = sorted(
        extract_exact_catalog_article_tokens(product_name),
        key=lambda item: (-len(str(item or "")), str(item or "")),
    )
    article = article_tokens[0] if article_tokens else ""
    payload, _ = get_or_generate_autofill_payload(db, user_id, product_name, article, use_cache=True)
    resolved_brand = resolve_product_brand_value(product_name, article, payload.get("brand"))
    resolved_folder_id = resolve_product_folder_assignment(
        db,
        user_id,
        product_name,
        preferred_folder_id,
        payload.get("storage"),
        resolved_brand,
    )
    merged_specs = build_specs_payload_from_autofill({}, payload)
    resolved_country = (
        payload.get("country_sim")
        or infer_brand_country_fallback(resolved_brand)
        or infer_accessory_country_fallback(product_name, resolved_brand)
    )
    resolved_model_no = normalize_model_no(payload.get("model_no"), resolved_brand)
    if not resolved_model_no:
        resolved_model_no = normalize_model_no(article, resolved_brand)
    if not resolved_model_no and article:
        resolved_model_no = article

    synonyms = merge_synonyms("", payload.get("synonyms", []))
    synonyms = merge_synonyms(synonyms, [product_name])

    cursor = db.execute(
        """
        INSERT INTO products (
            user_id, name, synonyms, price, folder_id, photo_url, brand, country, weight,
            model_number, is_on_request, color, storage, ram, warranty, description,
            description_html, specs, source_url, sort_index
        )
        VALUES (?, ?, ?, 0, ?, '', ?, ?, ?, ?, 0, ?, ?, ?, ?, ?, '', ?, ?, ?)
        RETURNING id
        """,
        (
            user_id,
            product_name,
            synonyms,
            resolved_folder_id,
            resolved_brand,
            resolved_country,
            payload.get("weight") or "",
            resolved_model_no,
            payload.get("color") or "",
            payload.get("storage") or "",
            payload.get("ram") or "",
            str(payload.get("warranty") or STORE_WARRANTY_DEFAULT).strip(),
            payload.get("description") or "",
            json.dumps(merged_specs, ensure_ascii=False) if merged_specs else "",
            payload.get("source_url") or "",
            get_next_product_sort_index(db, user_id),
        ),
    )
    new_id = int(cursor.fetchone()["id"])
    row = db.execute(
        """
        SELECT id, name, country, synonyms, brand, model_number, color, storage, folder_id
        FROM products
        WHERE id = ? AND user_id = ?
        """,
        (new_id, user_id),
    ).fetchone()
    if not row:
        return None
    product = dict(row)
    prepare_auto_confirm_product(product)
    return product


def auto_confirm_source_message_matches(db, user_id, message_id, apply=True, add_synonyms=True, max_examples=30, max_seconds=None):
    started_at = time.monotonic()
    stats = {
        "message_id": message_id,
        "source_enabled": False,
        "lines_seen": 0,
        "price_lines": 0,
        "inserted": 0,
        "updated": 0,
        "occupied": 0,
        "ambiguous": 0,
        "unmatched": 0,
        "memory_hits": 0,
        "memory_written": 0,
        "search_fallback_hits": 0,
        "synonym_fallback_hits": 0,
        "missing_logged": 0,
        "missing_resolved": 0,
        "synonyms_added": 0,
        "discount_lines_skipped": 0,
        "invalid": 0,
        "created_products": 0,
        "timed_out": 0,
        "message_categories": 0,
        "message_scope_keywords": 0,
        "loaded_products": 0,
    }
    examples = {"confirmed": [], "ambiguous": [], "unmatched": []}
    message = db.execute(
        """
        SELECT m.id, m.text, m.date, m.chat_id, m.sender_name, m.chat_title, tc.custom_name
        FROM messages m
        LEFT JOIN tracked_chats tc ON tc.chat_id = m.chat_id AND tc.user_id = m.user_id
        WHERE m.id = ? AND m.user_id = ?
        """,
        (message_id, user_id),
    ).fetchone()
    if not message or not source_row_is_auto_confirm_enabled(message):
        return {"stats": stats, "examples": examples}

    stats["source_enabled"] = True
    lines = str(message["text"] or "").split("\n")
    message_categories = collect_auto_confirm_message_categories(lines)
    stats["message_categories"] = len(message_categories)
    message_scope_keywords = collect_auto_confirm_message_scope_keywords(lines, message_categories)
    stats["message_scope_keywords"] = len(message_scope_keywords)
    source_folder_ids = infer_auto_confirm_source_folder_ids(db, user_id, message["text"])
    if message_categories and (message_categories - {"iphone"}):
        source_folder_ids = []
    source_is_scoped = bool(source_folder_ids)
    products = load_products_for_source_auto_confirm(
        db,
        user_id,
        folder_ids=source_folder_ids,
        categories=message_categories,
        scope_keywords=message_scope_keywords,
    )
    if not products and message_scope_keywords:
        products = load_products_for_source_auto_confirm(
            db,
            user_id,
            folder_ids=source_folder_ids,
            categories=message_categories,
        )
    if not products and (source_folder_ids or message_categories):
        products = load_products_for_source_auto_confirm(db, user_id)
    stats["loaded_products"] = len(products)
    products_by_id = {int(product["id"]): product for product in products}
    token_index = build_auto_confirm_token_index(products)
    in_discount_section = False
    context_headers = []
    for line_index, raw_line in enumerate(lines):
        if max_seconds and (time.monotonic() - started_at) >= max_seconds:
            stats["timed_out"] = 1
            logger.warning(
                "Автопривязка message_id=%s остановлена по time budget %ss",
                message_id,
                max_seconds,
            )
            break
        line = raw_line.strip()
        if not line:
            continue
        if line_is_discount_section_marker(line):
            in_discount_section = True
            stats["discount_lines_skipped"] += 1
            continue
        if in_discount_section:
            stats["discount_lines_skipped"] += 1
            continue
        if line_has_discount_condition_marker(line):
            stats["discount_lines_skipped"] += 1
            continue
        stats["lines_seen"] += 1
        section_header = supplier_catalog_section_header_text(line)
        if section_header:
            context_headers = append_supplier_context_header(context_headers, section_header)
            continue
        if not supplier_line_has_price_marker(line):
            context_headers = update_supplier_context_headers(context_headers, line)
            continue
        price = extract_price(line)
        if not price:
            continue
        stats["price_lines"] += 1
        match_line = line
        line_words = clean_for_search(line)
        article_conflict_candidates = []
        remembered_product = get_supplier_memory_product(db, user_id, line, products_by_id)
        if remembered_product:
            candidates = [remembered_product]
            stats["memory_hits"] += 1
        else:
            product_pool = get_auto_confirm_candidate_pool(products, token_index, line_words)
            candidates = find_auto_confirm_product_candidates(product_pool, line, line_words=line_words)
            if not candidates:
                article_conflict_candidates = find_auto_confirm_article_conflict_candidates(product_pool, line, line_words=line_words)
            if not candidates and not source_is_scoped:
                fallback_pool = get_auto_confirm_search_fallback_pool(db, user_id, products_by_id, line)
                if fallback_pool:
                    candidates = find_auto_confirm_product_candidates(fallback_pool, line, line_words=line_words)
                    if not candidates and not article_conflict_candidates:
                        article_conflict_candidates = find_auto_confirm_article_conflict_candidates(fallback_pool, line, line_words=line_words)
                    if candidates:
                        stats["search_fallback_hits"] += 1
            if not candidates and not source_is_scoped:
                candidates = find_auto_confirm_synonym_fallback_candidates(products, line, line_words=line_words)
                if not candidates and not article_conflict_candidates:
                    article_conflict_candidates = find_auto_confirm_synonym_fallback_candidates(
                        products,
                        line,
                        line_words=line_words,
                        allow_article_conflict=True,
                    )
                if candidates:
                    stats["synonym_fallback_hits"] += 1
        if len(candidates) != 1:
            context_line = make_supplier_contextual_line(context_headers, line)
            if context_line != line:
                context_words = clean_for_search(context_line)
                context_article_conflict_candidates = []
                remembered_product = get_supplier_memory_product(db, user_id, context_line, products_by_id)
                if remembered_product:
                    context_candidates = [remembered_product]
                    stats["memory_hits"] += 1
                else:
                    context_pool = get_auto_confirm_candidate_pool(products, token_index, context_words)
                    context_candidates = find_auto_confirm_product_candidates(context_pool, context_line, line_words=context_words)
                    if not context_candidates:
                        context_article_conflict_candidates = find_auto_confirm_article_conflict_candidates(context_pool, context_line, line_words=context_words)
                    if not context_candidates and not source_is_scoped:
                        fallback_pool = get_auto_confirm_search_fallback_pool(db, user_id, products_by_id, context_line)
                        if fallback_pool:
                            context_candidates = find_auto_confirm_product_candidates(fallback_pool, context_line, line_words=context_words)
                            if not context_candidates and not context_article_conflict_candidates:
                                context_article_conflict_candidates = find_auto_confirm_article_conflict_candidates(fallback_pool, context_line, line_words=context_words)
                            if context_candidates:
                                stats["search_fallback_hits"] += 1
                    if not context_candidates and not source_is_scoped:
                        context_candidates = find_auto_confirm_synonym_fallback_candidates(
                            products,
                            context_line,
                            line_words=context_words,
                        )
                        if not context_candidates and not context_article_conflict_candidates:
                            context_article_conflict_candidates = find_auto_confirm_synonym_fallback_candidates(
                                products,
                                context_line,
                                line_words=context_words,
                                allow_article_conflict=True,
                            )
                        if context_candidates:
                            stats["synonym_fallback_hits"] += 1
                    context_candidates = [
                        product for product in context_candidates
                        if product_line_semantics_compatible(product, line, require_known_category=False)
                    ]
                if len(context_candidates) == 1:
                    match_line = context_line
                    line_words = context_words
                    candidates = context_candidates
                elif context_article_conflict_candidates:
                    article_conflict_candidates = context_article_conflict_candidates
        if len(candidates) != 1:
            if not candidates and apply and article_conflict_candidates:
                created_product = create_product_for_supplier_line(
                    db,
                    user_id,
                    line,
                    preferred_folder_id=infer_conflict_preferred_folder_id(article_conflict_candidates),
                )
                if created_product:
                    products.append(created_product)
                    products_by_id[int(created_product["id"])] = created_product
                    token_index = build_auto_confirm_token_index(products)
                    candidates = [created_product]
                    stats["created_products"] += 1
            if len(candidates) == 1:
                product = candidates[0]
                result = upsert_auto_confirmed_binding(db, product, message, line_index, line, price, apply=apply)
                if result == "invalid":
                    stats["invalid"] += 1
                    continue
                stats[result] += 1
                if apply and result in ("inserted", "updated"):
                    stats["memory_written"] += remember_supplier_line_binding(
                        db,
                        user_id,
                        product["id"],
                        line,
                        message=message,
                    )
                    stats["missing_resolved"] += resolve_supplier_missing_position(
                        db,
                        user_id,
                        line,
                        product_id=product["id"],
                    )
                if result in ("inserted", "updated") and add_synonyms:
                    synonym = supplier_line_synonym(match_line)
                    if apply:
                        added, updated_synonyms = append_product_synonym(db, product["id"], product.get("synonyms"), synonym)
                        if added:
                            product["synonyms"] = updated_synonyms
                            prepare_auto_confirm_product(product)
                            token_index = build_auto_confirm_token_index(products)
                            stats["synonyms_added"] += added
                    else:
                        existing_keys = {normalize_synonym_key(item) for item in split_product_synonyms(product.get("synonyms"))}
                        if normalize_synonym_key(synonym) not in existing_keys:
                            stats["synonyms_added"] += 1
                if result in ("inserted", "updated") and len(examples["confirmed"]) < max_examples:
                    examples["confirmed"].append({
                        "line_index": line_index,
                        "line": line,
                        "match_line": match_line if match_line != line else "",
                        "product_id": product["id"],
                        "product_name": product["name"],
                        "action": result,
                    })
                continue
            key = "ambiguous" if candidates else "unmatched"
            stats[key] += 1
            if apply:
                stats["missing_logged"] += upsert_supplier_missing_position(
                    db,
                    user_id,
                    message,
                    line_index,
                    line,
                    match_line=match_line,
                    price=price,
                    status=key,
                    candidates=candidates,
                )
            if len(examples[key]) < max_examples:
                examples[key].append({
                    "line_index": line_index,
                    "line": line,
                    "match_line": match_line if match_line != line else "",
                    "candidate_count": len(candidates),
                    "candidates": [
                        {"product_id": item["id"], "product_name": item["name"]}
                        for item in candidates[:8]
                    ],
                })
            continue

        product = candidates[0]
        result = upsert_auto_confirmed_binding(db, product, message, line_index, line, price, apply=apply)
        if result == "invalid":
            stats["invalid"] += 1
            continue
        stats[result] += 1
        if apply and result in ("inserted", "updated"):
            stats["memory_written"] += remember_supplier_line_binding(
                db,
                user_id,
                product["id"],
                line,
                message=message,
            )
            stats["missing_resolved"] += resolve_supplier_missing_position(
                db,
                user_id,
                line,
                product_id=product["id"],
            )
        if result in ("inserted", "updated") and add_synonyms:
            synonym = supplier_line_synonym(match_line)
            if apply:
                added, updated_synonyms = append_product_synonym(db, product["id"], product.get("synonyms"), synonym)
                if added:
                    product["synonyms"] = updated_synonyms
                    prepare_auto_confirm_product(product)
                    stats["synonyms_added"] += added
            else:
                existing_keys = {normalize_synonym_key(item) for item in split_product_synonyms(product.get("synonyms"))}
                if normalize_synonym_key(synonym) not in existing_keys:
                    stats["synonyms_added"] += 1
        if result in ("inserted", "updated") and len(examples["confirmed"]) < max_examples:
            examples["confirmed"].append({
                "line_index": line_index,
                "line": line,
                "match_line": match_line if match_line != line else "",
                "product_id": product["id"],
                "product_name": product["name"],
                "action": result,
            })

    if apply and (stats["inserted"] or stats["updated"]):
        prune_duplicate_confirmed_bindings(db, user_id=user_id)
    return {"stats": stats, "examples": examples}


def sweep_auto_binding_for_recent_messages(db, user_id=None, limit=AUTO_BINDING_SWEEP_MESSAGE_LIMIT):
    filters = [
        "m.text IS NOT NULL",
        "TRIM(m.text) != ''",
        "COALESCE(m.is_blocked, 0) = 0",
        "COALESCE(m.is_delayed, 0) = 0",
    ]
    params = []
    if user_id is not None:
        filters.append("m.user_id = ?")
        params.append(user_id)

    params.extend([int(AUTO_BINDING_SWEEP_SOURCE_WINDOW_MINUTES), int(limit or AUTO_BINDING_SWEEP_MESSAGE_LIMIT)])
    rows = db.execute(
        f"""
        WITH source_latest AS (
            SELECT user_id, chat_id, MAX(date) AS latest_date
            FROM messages
            WHERE text IS NOT NULL
              AND TRIM(text) != ''
              AND COALESCE(is_blocked, 0) = 0
              AND COALESCE(is_delayed, 0) = 0
            GROUP BY user_id, chat_id
        )
        SELECT m.id, m.user_id, m.text
        FROM messages m
        JOIN source_latest sl
          ON sl.user_id = m.user_id
         AND CAST(sl.chat_id AS TEXT) = CAST(m.chat_id AS TEXT)
        LEFT JOIN product_messages pm
          ON pm.message_id = m.id
         AND pm.status = 'confirmed'
        WHERE {' AND '.join(filters)}
          AND m.date >= datetime(sl.latest_date, '-' || ? || ' minutes')
        GROUP BY m.id, m.user_id, m.text, m.date
        HAVING COALESCE(SUM(CASE WHEN pm.is_actual = 1 THEN 1 ELSE 0 END), 0) = 0
        ORDER BY m.date DESC, m.id DESC
        LIMIT ?
        """,
        params,
    ).fetchall()

    totals = Counter()
    processed = 0
    now_monotonic = time.monotonic()
    if len(auto_binding_sweep_attempts) > 5000:
        cutoff = now_monotonic - AUTO_BINDING_SWEEP_RETRY_SECONDS
        stale_keys = [key for key, value in auto_binding_sweep_attempts.items() if value < cutoff]
        for key in stale_keys:
            auto_binding_sweep_attempts.pop(key, None)

    for row in rows:
        text = str(row["text"] or "")
        if not any(
            supplier_line_has_price_marker(line)
            and extract_price(line)
            and not line_is_supplier_catalog_section_header(line)
            for line in text.splitlines()
        ):
            continue
        attempt_key = (int(row["user_id"]), int(row["id"]))
        last_attempt = auto_binding_sweep_attempts.get(attempt_key)
        if last_attempt and (now_monotonic - last_attempt) < AUTO_BINDING_SWEEP_RETRY_SECONDS:
            continue
        auto_binding_sweep_attempts[attempt_key] = now_monotonic

        result = auto_confirm_source_message_matches(
            db,
            row["user_id"],
            row["id"],
            apply=True,
            max_seconds=AUTO_CONFIRM_MESSAGE_MAX_SECONDS,
        )
        stats = result.get("stats") or {}
        if not stats.get("source_enabled"):
            continue
        processed += 1
        for key, value in stats.items():
            if key == "message_id" or not isinstance(value, int):
                continue
            totals[key] += value

    if processed:
        totals["processed_messages"] = processed
    if totals.get("inserted") or totals.get("updated"):
        prune_duplicate_confirmed_bindings(db, user_id=user_id)
        deactivate_duplicate_actual_line_bindings(db, user_id=user_id)
        sanitize_confirmed_binding_records(db, user_id=user_id)
        refresh_confirmed_bindings_from_latest_messages(db, user_id=user_id)
    return dict(totals)


def auto_binding_sweep_scheduler():
    time.sleep(90)
    while True:
        try:
            with app.app_context():
                db = get_db()
                with sqlite_write_lock:
                    stats = sweep_auto_binding_for_recent_messages(db)
                    db.commit()
                if stats.get("inserted") or stats.get("updated"):
                    logger.info("Фоновая автопривязка обновила прайсы: %s", stats)
                    notify_clients("products")
        except Exception as e:
            logger.error("Ошибка фоновой автопривязки: %s", e, exc_info=True)
        time.sleep(AUTO_BINDING_SWEEP_INTERVAL_SECONDS)


def build_synonym_word_sets(synonyms):
    synonym_sets = []
    for synonym in str(synonyms or "").split(","):
        if synonym.strip():
            words = clean_for_search(synonym)
            if words:
                synonym_sets.append(words)
    return synonym_sets


def get_recent_message_lines_for_product_matching(db, user_id, limit=1000):
    occupied_rows = db.execute("""
        SELECT pm.message_id, pm.line_index
        FROM product_messages pm
        JOIN messages m ON pm.message_id = m.id
        WHERE pm.status = 'confirmed'
          AND m.user_id = ?
    """, (user_id,)).fetchall()
    occupied = {(row['message_id'], row['line_index']) for row in occupied_rows}

    rows = db.execute(
        "SELECT id, text FROM messages WHERE user_id = ? ORDER BY date DESC LIMIT ?",
        (user_id, limit),
    ).fetchall()
    recent_lines = []
    for row in rows:
        lines = row['text'].split('\n') if row['text'] else []
        in_discount_section = False
        for index, line in enumerate(lines):
            if not line.strip():
                continue
            if line_is_discount_section_marker(line):
                in_discount_section = True
                continue
            if in_discount_section:
                continue
            if line_has_discount_condition_marker(line):
                continue
            if (row['id'], index) in occupied or (row['id'], -1) in occupied:
                continue
            if line_is_supplier_catalog_section_header(line):
                continue
            if not supplier_line_has_price_marker(line) or not extract_price(line):
                continue
            line_words = clean_for_search(line)
            if line_words:
                recent_lines.append({
                    'text': line,
                    'words': line_words,
                })
    return recent_lines


def product_has_dynamic_proposed_match(product, recent_message_lines):
    if int(product.get('confirmed_count') or 0) > 0:
        return False

    prepared_product = dict(product or {})
    prepare_auto_confirm_product(prepared_product)
    synonym_sets = prepared_product.get('_synonym_sets') or []
    if not synonym_sets and not prepared_product.get('_model_tokens'):
        return False

    for line in recent_message_lines or ():
        line_text = line.get('text') or ""
        if product_supplier_line_matches(
            prepared_product,
            line_text,
            line_words=line.get('words') or set(),
            synonym_sets=synonym_sets,
        ):
            return True
    return False


CATALOG_FALLBACK_DROP_TOKENS = {
    "apple", "garmin", "samsung", "honor", "huawei", "xiaomi", "dji", "dyson",
    "wifi", "wi", "fi", "cellular", "only", "sim", "esim", "with", "and", "usb",
    "standard", "glass", "nano", "texture", "gb", "tb",
}


def build_catalog_family_search_pattern(product_name):
    text = re.sub(r"\s+", " ", str(product_name or "").strip())
    if not text:
        return ""
    text = re.sub(r"\b\d+(?:[.,]\d+)?\s*(?:GB|TB|ГБ|ТБ)\b", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\bWi[\s-]*Fi(?:\s*\+\s*Cellular)?\b", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\bCellular\b", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\((?:[^)]*e\s*sim[^)]*|[^)]*1\s*sim[^)]*|[^)]*2\s*sim[^)]*)\)", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"[(),]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def extract_catalog_fallback_tokens(product_name):
    tokens = []
    for token in re.findall(r"[A-Za-zА-Яа-я0-9]+", normalize_match_text(product_name)):
        compact = token.replace(" ", "")
        if not compact or compact in CATALOG_FALLBACK_DROP_TOKENS:
            continue
        if re.fullmatch(r"\d+(?:gb|tb|гб|тб)", compact):
            continue
        if len(compact) < 2:
            continue
        if compact not in tokens:
            tokens.append(compact)
    return tokens[:6]


def extract_catalog_variant_markers(product_name):
    compact = normalize_match_text(product_name).replace(" ", "")
    markers = set()
    if "wifi" in compact:
        markers.add("wifi")
    if "cellular" in compact:
        markers.add("cellular")
    if "esim" in compact:
        markers.add("esim")
    if "1sim" in compact:
        markers.add("1sim")
    if "2sim" in compact:
        markers.add("2sim")
    if "nanotexture" in compact:
        markers.add("nano")
    if "standardglass" in compact:
        markers.add("standard_glass")
    return markers


def find_catalog_product_fallback_row(db, user_id, product):
    product_id = int(product.get("id") or 0)
    cache_key = (
        product_id,
        str(product.get("name") or "").strip(),
        str(product.get("brand") or "").strip(),
        str(product.get("color") or "").strip(),
        str(product.get("storage") or "").strip(),
    )
    cache = getattr(g, "_catalog_product_fallback_cache", None)
    if cache is None:
        cache = {}
        g._catalog_product_fallback_cache = cache
    if cache_key in cache:
        return cache[cache_key]

    brand = sanitize_brand_value(product.get("brand"))
    family_tokens = extract_catalog_fallback_tokens(build_catalog_family_search_pattern(product.get("name")))
    if not brand or not family_tokens:
        cache[cache_key] = None
        return None

    query = """
        SELECT id, name, photo_url, country, weight, color, storage, ram, model_number,
               synonyms, warranty, description, description_html, specs
        FROM products
        WHERE user_id = ?
          AND id != ?
          AND brand = ?
          AND (
              COALESCE(TRIM(photo_url), '') != ''
              OR COALESCE(TRIM(country), '') != ''
              OR COALESCE(TRIM(weight), '') != ''
          )
    """
    params = [user_id, product_id, brand]
    for token in family_tokens[:4]:
        query += " AND LOWER(name) LIKE ?"
        params.append(f"%{token.lower()}%")
    query += " ORDER BY id DESC LIMIT 25"

    current_color = normalize_match_text(product.get("color"))
    current_storage = normalize_capacity(product.get("storage"), allow_tb=True)
    current_markers = extract_catalog_variant_markers(product.get("name"))
    current_model = str(product.get("model_number") or "").strip().upper()
    best_row = None
    best_score = -1

    for row in db.execute(query, tuple(params)).fetchall():
        candidate = dict(row)
        score = len(set(family_tokens) & set(extract_catalog_fallback_tokens(candidate.get("name"))))

        candidate_color = normalize_match_text(candidate.get("color"))
        if current_color and candidate_color and current_color == candidate_color:
            score += 6

        candidate_storage = normalize_capacity(candidate.get("storage"), allow_tb=True)
        if current_storage and candidate_storage and current_storage == candidate_storage:
            score += 4

        candidate_model = str(candidate.get("model_number") or "").strip().upper()
        if current_model and candidate_model and current_model == candidate_model:
            score += 5

        score += len(current_markers & extract_catalog_variant_markers(candidate.get("name"))) * 2

        if score > best_score:
            best_score = score
            best_row = candidate

    cache[cache_key] = best_row
    return best_row


def normalize_catalog_product_snapshot(db, user_id, product_row):
    product = dict(product_row)
    product_name = str(product.get("name") or "").strip()
    brand = sanitize_brand_value(product.get("brand")) or sanitize_brand_value(detect_brand_from_name(product_name))
    product["brand"] = brand or str(product.get("brand") or "").strip()
    is_ipad = is_apple_ipad_product(product_name, brand)

    if is_ipad:
        normalized_name = normalize_ipad_catalog_name(product_name)
        if normalized_name:
            product_name = normalized_name
            product["name"] = normalized_name

    storage_value = normalize_capacity(product.get("storage"), allow_tb=True)
    ram_value = normalize_capacity(product.get("ram"), allow_tb=False)
    inferred_ram, inferred_storage = extract_ram_storage_from_name(product_name)
    if inferred_storage and (not storage_value or (ram_value and storage_value and storage_value == ram_value)):
        storage_value = inferred_storage
    if is_ipad:
        ram_value = ""
    elif inferred_ram:
        if not ram_value or (storage_value and ram_value == storage_value):
            ram_value = inferred_ram
    elif ram_value and storage_value and ram_value == storage_value:
        ram_value = ""

    if storage_value:
        product["storage"] = storage_value
    if is_ipad or ram_value or str(product.get("ram") or "").strip():
        product["ram"] = ram_value

    model_number_value = str(product.get("model_number") or "").strip()
    if brand == "Apple":
        extracted_model_number = extract_apple_root_model_no(
            model_number_value,
            product.get("synonyms"),
            product_name,
        )
        if extracted_model_number:
            product["model_number"] = extracted_model_number

    country_value = str(product.get("country") or "").strip()
    normalized_country_compact = normalize_product_search_text(country_value).replace(" ", "")
    if is_ipad and normalized_country_compact in {"esim", "esimonly", "1sim", "2sim", "1sim+esim", "sim+esim"}:
        country_value = "Глобальная версия" if "Cellular" in normalize_ipad_connectivity_label(product_name) else ""
    if not country_value:
        country_value = detect_country_variant(product_name) or infer_accessory_country_fallback(product_name, brand)

    weight_value = normalize_autofill_weight(product.get("weight"))
    photo_value = str(product.get("photo_url") or "").strip()
    fallback = None

    if not photo_value or not country_value or not weight_value:
        fallback = find_catalog_product_fallback_row(db, user_id, product)
        if fallback:
            if not photo_value:
                photo_value = str(fallback.get("photo_url") or "").strip()
            if not country_value:
                country_value = str(fallback.get("country") or "").strip()
            if not weight_value:
                weight_value = normalize_autofill_weight(fallback.get("weight"))
            if not str(product.get("model_number") or "").strip():
                product["model_number"] = str(fallback.get("model_number") or "").strip()
    elif is_ipad:
        fallback = find_catalog_product_fallback_row(db, user_id, product)

    if is_ipad:
        model_number = normalize_model_no(product.get("model_number"), "Apple")
        if not model_number:
            model_number = extract_apple_root_model_no(
                product.get("model_number"),
                product.get("synonyms"),
                product.get("name"),
                product_name,
                fallback.get("synonyms") if fallback else "",
                fallback.get("model_number") if fallback else "",
                fallback.get("name") if fallback else "",
            )
        if model_number:
            product["model_number"] = model_number

        if not str(product.get("warranty") or "").strip():
            product["warranty"] = "12 месяцев от Apple"

        existing_description = str(product.get("description") or "").strip()
        existing_description_html = str(product.get("description_html") or "").strip()
        specs_payload = parse_specs_payload(product.get("specs"), product)
        if fallback:
            fallback_specs = parse_specs_payload(fallback.get("specs"), fallback)
            if not specs_payload:
                specs_payload = fallback_specs

        expected_specs = build_ipad_default_specs(product_name, storage_value)
        if not specs_payload or not any(str(specs_payload.get(key) or "").strip() for key in ("display", "processor", "connectivity", "os")):
            specs_payload = {}
        for key, value in expected_specs.items():
            if key == "storage":
                continue
            if key == "connectivity" or not str(specs_payload.get(key) or "").strip():
                specs_payload[key] = value
        product["specs"] = sanitize_specs_payload(specs_payload, product)

        if not existing_description:
            existing_description = build_ipad_description(product_name, storage_value)
        if not existing_description_html:
            existing_description_html = build_ipad_description_html(product_name, storage_value)
        product["description"] = existing_description
        product["description_html"] = existing_description_html

    specs_payload = parse_specs_payload(product.get("specs"), product)
    if not specs_payload:
        generic_specs = build_generic_catalog_specs(product)
        if generic_specs:
            product["specs"] = generic_specs
            specs_payload = generic_specs

    description_value = str(product.get("description") or "").strip()
    if not description_value:
        description_value = build_generic_catalog_description(product)
        if description_value:
            product["description"] = description_value
    if description_value and not str(product.get("description_html") or "").strip():
        product["description_html"] = build_basic_description_html(
            product.get("name"),
            description_value,
            specs_payload,
        )

    if photo_value:
        product["photo_url"] = photo_value
    if country_value:
        product["country"] = country_value
    if weight_value:
        product["weight"] = weight_value
    return product


def build_compact_product_payload(db, user_id, product_row):
    product = normalize_catalog_product_snapshot(db, user_id, product_row)
    photo_url = str(product.get('photo_url') or '').strip()
    if photo_url:
        product['photo_url'] = photo_url.split(',')[0].strip()
    if 'proposed_count' not in product:
        product['proposed_count'] = 1 if int(product.get('pending_count') or 0) > 0 and int(product.get('confirmed_count') or 0) == 0 else 0
    return product


def build_compact_product_payloads(db, user_id, product_rows, include_dynamic_proposed=True):
    products = [dict(product_row) for product_row in (product_rows or [])]
    if not products:
        return []

    product_ids = [int(product.get('id') or 0) for product in products if product.get('id')]
    counts_by_id = {
        product_id: {'confirmed_count': 0, 'pending_count': 0}
        for product_id in product_ids
    }
    if product_ids:
        placeholders = ','.join('?' for _ in product_ids)
        confirmed_rows = db.execute(f"""
            SELECT pm.product_id, COUNT(*) AS confirmed_count
            FROM product_messages pm
            JOIN messages m ON pm.message_id = m.id
            WHERE pm.status = 'confirmed'
              AND COALESCE(pm.is_actual, 1) = 1
              AND m.user_id = ?
              AND pm.product_id IN ({placeholders})
            GROUP BY pm.product_id
        """, (user_id, *product_ids)).fetchall()
        for row in confirmed_rows:
            counts_by_id[int(row['product_id'])]['confirmed_count'] = int(row['confirmed_count'] or 0)

        pending_rows = db.execute(f"""
            SELECT product_id, COUNT(*) AS pending_count
            FROM product_messages
            WHERE status = 'pending'
              AND product_id IN ({placeholders})
            GROUP BY product_id
        """, tuple(product_ids)).fetchall()
        for row in pending_rows:
            counts_by_id[int(row['product_id'])]['pending_count'] = int(row['pending_count'] or 0)

    for product in products:
        counts = counts_by_id.get(int(product.get('id') or 0), {})
        product['confirmed_count'] = int(counts.get('confirmed_count') or 0)
        product['pending_count'] = int(counts.get('pending_count') or 0)
        product['proposed_count'] = 0

    if include_dynamic_proposed:
        dynamic_candidates = []
        for product in products:
            if int(product.get('confirmed_count') or 0) != 0 or int(product.get('pending_count') or 0) != 0:
                continue
            prepare_auto_confirm_product(product)
            if product.get('_synonym_sets') or product.get('_model_tokens'):
                dynamic_candidates.append(product)
        if dynamic_candidates:
            recent_message_lines = get_recent_message_lines_for_product_matching(db, user_id)
            for product in dynamic_candidates:
                if product_has_dynamic_proposed_match(product, recent_message_lines):
                    product['proposed_count'] = 1

    for product in products:
        if int(product.get('proposed_count') or 0) == 0 and int(product.get('pending_count') or 0) > 0 and int(product.get('confirmed_count') or 0) == 0:
            product['proposed_count'] = 1

    return [build_compact_product_payload(db, user_id, product) for product in products]


def build_sql_search_variants_for_word(word, max_variants=14):
    normalized_word = normalize_product_search_text(word)
    if not normalized_word:
        return []
    candidates = [normalized_word, *collect_product_search_terms(normalized_word)]
    variants = []
    seen = set()
    for candidate in candidates:
        normalized = normalize_product_search_text(candidate)
        if len(normalized) < 2 or normalized in PRODUCT_SEARCH_STOP_WORDS or normalized in seen:
            continue
        seen.add(normalized)
        variants.append(normalized)
        if len(variants) >= max_variants:
            break
    for literal in PRODUCT_SQL_LITERAL_SEARCH_EXPANSIONS.get(normalized_word, []):
        if literal not in seen and len(variants) < max_variants:
            seen.add(literal)
            variants.append(literal)
    return variants


def build_sql_search_needles(query):
    return [
        variants
        for variants in (build_sql_search_variants_for_word(word) for word in tokenize_product_search_query(query))
        if variants
    ]


def compact_sql_search_value(value):
    return normalize_product_search_text(value).replace(' ', '')


def compact_sql_search_field_sql(field):
    return (
        "LOWER("
        f"REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(COALESCE({field}, ''), ' ', ''), '-', ''), '/', ''), '.', ''), '+', '')"
        ")"
    )


def fetch_ranked_product_rows_for_ids(db, select_fields, base_from, base_params, ranked_ids):
    page_ids = [int(product_id) for product_id in (ranked_ids or []) if product_id]
    if not page_ids:
        return []
    placeholders = ','.join('?' for _ in page_ids)
    order_case = ' '.join(f"WHEN {product_id} THEN {index}" for index, product_id in enumerate(page_ids))
    query = f"""
        SELECT {select_fields}
        {base_from}
        AND p.id IN ({placeholders})
        ORDER BY CASE p.id {order_case} ELSE {len(page_ids)} END
    """
    return db.execute(query, (*base_params, *page_ids)).fetchall()


def append_sql_search_field_match(parts, params, field, like_variant):
    like_value = f"%{like_variant}%"
    parts.append(f"{field} LIKE ?")
    params.append(like_value)


def build_fast_iphone_search_phrases(query):
    words = tokenize_product_search_query(query)
    if not words:
        return []

    canonical_words = []
    has_iphone = False
    for word in words:
        variants = set(collect_product_search_terms(word))
        normalized = normalize_product_search_text(word)
        if 'iphone' in variants:
            canonical_words.append('iphone')
            has_iphone = True
        elif 'pro' in variants:
            canonical_words.append('pro')
        elif 'max' in variants:
            canonical_words.append('max')
        elif 'plus' in variants:
            canonical_words.append('plus')
        elif 'ultra' in variants:
            canonical_words.append('ultra')
        elif 'mini' in variants:
            canonical_words.append('mini')
        elif 'air' in variants:
            canonical_words.append('air')
        elif normalized:
            canonical_words.append(normalized)

    if not has_iphone or len(canonical_words) < 2:
        return []

    phrases = []
    main_phrase = ' '.join(canonical_words)
    model_phrase = ' '.join(word for word in canonical_words if word != 'iphone')
    candidate_phrases = [main_phrase]
    if model_phrase and not re.fullmatch(r'\d+', model_phrase):
        candidate_phrases.append(model_phrase)
    for phrase in candidate_phrases:
        phrase = normalize_product_search_text(phrase)
        if phrase and phrase not in phrases:
            phrases.append(phrase)
    return phrases


def has_search_alias(words, alias):
    normalized_alias = normalize_product_search_text(alias)
    return any(normalized_alias in set(collect_product_search_terms(word)) for word in (words or []))


def is_ps5_search_intent(words, raw_query=""):
    normalized_words = {
        normalize_product_search_text(word)
        for word in (words or [])
        if normalize_product_search_text(word)
    }
    normalized_query = normalize_product_search_text(raw_query)
    expanded_terms = set()
    for word in (words or []):
        expanded_terms.update(collect_product_search_terms(word))
    return (
        'ps5' in normalized_words
        or 'пс5' in normalized_words
        or 'ps5' in expanded_terms
        or 'пс5' in expanded_terms
        or re.search(r'\bps5\b', normalized_query)
        or re.search(r'\bпс5\b', normalized_query)
    )


def expand_like_case_variants(variant):
    values = [variant]
    title_value = variant.title()
    capitalize_value = variant.capitalize()
    upper_value = variant.upper()
    for value in (title_value, capitalize_value, upper_value):
        if value and value not in values:
            values.append(value)
    return values


def build_ps5_extra_search_conditions(words):
    conditions = []
    params = []
    searchable_fields = (
        'p.name',
        'p.synonyms',
        'p.brand',
        'p.model_number',
        'p.color',
        'p.storage',
    )
    for word in (words or []):
        normalized_word = normalize_product_search_text(word)
        word_terms = set(collect_product_search_terms(word))
        if normalized_word in {'ps5', 'пс5'} or 'ps5' in word_terms or 'пс5' in word_terms:
            continue
        variants = build_sql_search_variants_for_word(word, max_variants=12)
        parts = []
        for variant in variants:
            for like_variant in expand_like_case_variants(variant):
                for field in searchable_fields:
                    append_sql_search_field_match(parts, params, field, like_variant)
        if parts:
            conditions.append(f"({' OR '.join(parts)})")
    return conditions, params


def product_accessory_match_sql():
    return """
        (
            p.name LIKE '%чех%' OR p.name LIKE '%Чех%' OR LOWER(p.name) LIKE '%case%' OR LOWER(p.name) LIKE '%cover%' OR LOWER(p.name) LIKE '%sleeve%'
            OR p.name LIKE '%накладк%' OR p.name LIKE '%Накладк%' OR p.name LIKE '%стекл%' OR p.name LIKE '%Стекл%' OR LOWER(p.name) LIKE '%glass%' OR LOWER(p.name) LIKE '%film%'
            OR p.name LIKE '%кабель%' OR p.name LIKE '%Кабель%' OR LOWER(p.name) LIKE '%cable%' OR LOWER(p.name) LIKE '%cord%'
            OR p.name LIKE '%заряд%' OR p.name LIKE '%Заряд%' OR LOWER(p.name) LIKE '%charger%' OR LOWER(p.name) LIKE '%adapter%'
            OR p.name LIKE '%ремеш%' OR p.name LIKE '%Ремеш%' OR p.name LIKE '%браслет%' OR p.name LIKE '%Браслет%' OR LOWER(p.name) LIKE '%strap%' OR LOWER(p.name) LIKE '% band%' OR LOWER(p.name) LIKE '% loop%' OR LOWER(p.name) LIKE '%milanese%'
            OR LOWER(p.name) LIKE '%ssd%' OR p.name LIKE '%накопител%' OR p.name LIKE '%Накопител%' OR p.name LIKE '%дисковод%' OR p.name LIKE '%Дисковод%' OR p.name LIKE '%панел%' OR p.name LIKE '%Панел%' OR LOWER(p.name) LIKE '%panel%'
            OR p.name LIKE '%подписк%' OR p.name LIKE '%Подписк%' OR p.name LIKE '%игра для%' OR p.name LIKE '%Игра для%' OR LOWER(p.name) LIKE '% game for%'
            OR p.synonyms LIKE '%чех%' OR p.synonyms LIKE '%Чех%' OR LOWER(p.synonyms) LIKE '%case%' OR LOWER(p.synonyms) LIKE '%cover%' OR LOWER(p.synonyms) LIKE '%sleeve%'
            OR p.synonyms LIKE '%накладк%' OR p.synonyms LIKE '%Накладк%' OR p.synonyms LIKE '%стекл%' OR p.synonyms LIKE '%Стекл%' OR LOWER(p.synonyms) LIKE '%glass%' OR LOWER(p.synonyms) LIKE '%film%'
            OR p.synonyms LIKE '%кабель%' OR p.synonyms LIKE '%Кабель%' OR LOWER(p.synonyms) LIKE '%cable%' OR LOWER(p.synonyms) LIKE '%cord%'
            OR p.synonyms LIKE '%заряд%' OR p.synonyms LIKE '%Заряд%' OR LOWER(p.synonyms) LIKE '%charger%' OR LOWER(p.synonyms) LIKE '%adapter%'
            OR p.synonyms LIKE '%ремеш%' OR p.synonyms LIKE '%Ремеш%' OR p.synonyms LIKE '%браслет%' OR p.synonyms LIKE '%Браслет%' OR LOWER(p.synonyms) LIKE '%strap%' OR LOWER(p.synonyms) LIKE '% band%' OR LOWER(p.synonyms) LIKE '% loop%' OR LOWER(p.synonyms) LIKE '%milanese%'
            OR LOWER(p.synonyms) LIKE '%ssd%' OR p.synonyms LIKE '%накопител%' OR p.synonyms LIKE '%Накопител%' OR p.synonyms LIKE '%дисковод%' OR p.synonyms LIKE '%Дисковод%' OR p.synonyms LIKE '%панел%' OR p.synonyms LIKE '%Панел%' OR LOWER(p.synonyms) LIKE '%panel%'
            OR p.synonyms LIKE '%подписк%' OR p.synonyms LIKE '%Подписк%' OR p.synonyms LIKE '%игра для%' OR p.synonyms LIKE '%Игра для%' OR LOWER(p.synonyms) LIKE '% game for%'
        )
    """


def product_accessory_penalty_sql(accessory_first=False):
    accessory_rank = 0 if accessory_first else 1
    default_rank = 1 if accessory_first else 0
    return f"""
        CASE
            WHEN {product_accessory_match_sql()} THEN {accessory_rank}
            ELSE {default_rank}
        END
    """


def product_primary_device_rank_sql():
    return f"""
        CASE
            WHEN p.name LIKE 'Смартфон %' OR p.name LIKE 'смартфон %' OR LOWER(p.name) LIKE 'apple iphone%' OR LOWER(p.name) LIKE '% apple iphone%' THEN 0
            WHEN LOWER(p.name) LIKE 'apple watch%' OR LOWER(p.name) LIKE 'samsung galaxy watch%' OR LOWER(p.name) LIKE 'huawei watch%' THEN 0
            WHEN LOWER(p.name) LIKE 'sony playstation 5%' OR LOWER(p.name) LIKE 'playstation 5%' THEN 0
            WHEN LOWER(p.name) LIKE '%macbook%' OR LOWER(p.name) LIKE '%ipad%' OR p.name LIKE '%Планшет%' OR p.name LIKE '%планшет%' OR p.name LIKE '%Ноутбук%' OR p.name LIKE '%ноутбук%' THEN 0
            WHEN LOWER(p.name) LIKE '%airpods%' OR p.name LIKE '%Наушники Apple%' OR p.name LIKE '%наушники Apple%' THEN 0
            WHEN LOWER(p.name) LIKE '%dualsense%' OR LOWER(p.name) LIKE '%gamepad%' OR p.name LIKE '%Геймпад%' OR p.name LIKE '%геймпад%' THEN 1
            WHEN {product_accessory_match_sql()} THEN 3
            ELSE 2
        END
    """


def product_variant_rank_sql():
    return """
        CASE
            WHEN LOWER(p.name) LIKE '%iphone%' THEN
                CASE
                    WHEN LOWER(p.name) LIKE '% pro max%' THEN 4
                    WHEN LOWER(p.name) LIKE '% ultra%' THEN 5
                    WHEN LOWER(p.name) LIKE '% pro%' THEN 3
                    WHEN LOWER(p.name) LIKE '% plus%' THEN 2
                    WHEN LOWER(p.name) LIKE '% air%' THEN 1
                    WHEN LOWER(p.name) LIKE '% mini%' THEN 6
                    ELSE 0
                END
            ELSE 0
        END
    """


def product_storage_rank_sql():
    return """
        CASE
            WHEN p.storage LIKE '64 %' OR p.storage LIKE '64GB%' OR p.name LIKE '%64 ГБ%' OR p.name LIKE '%64 GB%' THEN 64
            WHEN p.storage LIKE '128 %' OR p.storage LIKE '128GB%' OR p.name LIKE '%128 ГБ%' OR p.name LIKE '%128 GB%' THEN 128
            WHEN p.storage LIKE '256 %' OR p.storage LIKE '256GB%' OR p.name LIKE '%256 ГБ%' OR p.name LIKE '%256 GB%' THEN 256
            WHEN p.storage LIKE '512 %' OR p.storage LIKE '512GB%' OR p.name LIKE '%512 ГБ%' OR p.name LIKE '%512 GB%' THEN 512
            WHEN p.storage LIKE '1 T%' OR p.storage LIKE '1T%' OR p.name LIKE '%1 ТБ%' OR p.name LIKE '%1 TB%' THEN 1024
            WHEN p.storage LIKE '2 T%' OR p.storage LIKE '2T%' OR p.name LIKE '%2 ТБ%' OR p.name LIKE '%2 TB%' THEN 2048
            ELSE 999999
        END
    """


@app.route('/api/products', methods=['GET', 'POST'])
@login_required
def manage_products():
    user_id = session['user_id']
    db = get_db()
    if request.method == 'GET':
        is_compact = str(request.args.get('compact', '')).strip().lower() in {'1', 'true', 'yes'}
        connected_only = str(request.args.get('connected_only', '')).strip().lower() in {'1', 'true', 'yes'}
        include_dynamic_proposed = str(request.args.get('include_dynamic_proposed', '')).strip().lower() in {'1', 'true', 'yes'}
        raw_query = str(request.args.get('q', '') or '').strip()
        page_arg = request.args.get('page')
        limit_arg = request.args.get('limit')
        use_paged_compact = is_compact and (
            bool(raw_query)
            or page_arg is not None
            or limit_arg is not None
            or bool(str(request.args.get('folder_ids', '') or '').strip())
        )
        try:
            page = int(page_arg or 1)
        except (TypeError, ValueError):
            page = 1
        try:
            limit = int(limit_arg or 20)
        except (TypeError, ValueError):
            limit = 20
        page = max(1, page)
        limit = max(1, min(limit, 100))

        folder_ids = []
        for part in str(request.args.get('folder_ids', '') or '').split(','):
            part = part.strip()
            if not part:
                continue
            try:
                folder_ids.append(int(part))
            except ValueError:
                continue
        folder_ids = list(dict.fromkeys(folder_ids))
        if len(folder_ids) > 200:
            folder_ids = []

        select_fields = """
            p.id,
            p.name,
            p.synonyms,
            p.price,
            p.folder_id,
            p.photo_url,
            p.brand,
            p.country,
            p.weight,
            p.model_number,
            p.is_on_request,
            p.color,
            p.storage,
            p.ram,
            p.warranty,
            p.sort_index
        """ if is_compact else "p.*"

        if is_compact:
            base_from = "FROM products p WHERE p.user_id = ?"
            base_params = [user_id]
        else:
            base_from = f"""
                FROM products p
                LEFT JOIN (
                    SELECT pm.product_id, COUNT(*) AS confirmed_count
                    FROM product_messages pm
                    JOIN messages m ON pm.message_id = m.id
                    WHERE pm.status = 'confirmed'
                      AND COALESCE(pm.is_actual, 1) = 1
                      AND m.user_id = ?
                    GROUP BY pm.product_id
                ) confirmed ON confirmed.product_id = p.id
                LEFT JOIN (
                    SELECT product_id, COUNT(*) AS pending_count
                    FROM product_messages
                    WHERE status = 'pending'
                    GROUP BY product_id
                ) pending ON pending.product_id = p.id
                WHERE p.user_id = ?
            """
            base_params = [user_id, user_id]

        if folder_ids:
            placeholders = ','.join('?' for _ in folder_ids)
            base_from += f" AND p.folder_id IN ({placeholders})"
            base_params.extend(folder_ids)
        if connected_only:
            if is_compact:
                base_from += """
                    AND EXISTS (
                        SELECT 1
                        FROM product_messages pm
                        JOIN messages m ON pm.message_id = m.id
                        WHERE pm.product_id = p.id
                          AND pm.status = 'confirmed'
                          AND COALESCE(pm.is_actual, 1) = 1
                          AND m.user_id = ?
                    )
                """
                base_params.append(user_id)
            else:
                base_from += " AND COALESCE(confirmed.confirmed_count, 0) > 0"

        if use_paged_compact:
            offset = (page - 1) * limit

            if raw_query:
                sql_needles = build_sql_search_needles(raw_query)
                if not sql_needles:
                    return jsonify({
                        'items': [],
                        'total': 0,
                        'page': page,
                        'limit': limit,
                    })
                query_words = tokenize_product_search_query(raw_query)
                ps5_search_intent = is_ps5_search_intent(query_words, raw_query)
                if len(query_words) == 1 and not ps5_search_intent:
                    brand_variants = build_sql_search_variants_for_word(query_words[0], max_variants=24)
                    if brand_variants:
                        brand_placeholders = ','.join('?' for _ in brand_variants)
                        brand_from = f"{base_from} AND p.brand COLLATE NOCASE IN ({brand_placeholders})"
                        brand_params = [*base_params, *brand_variants]
                        brand_total = int(db.execute(
                            f"SELECT COUNT(*) AS total {brand_from}",
                            tuple(brand_params),
                        ).fetchone()['total'] or 0)
                        if brand_total:
                            brand_accessory_intent = is_accessory_search_intent(query_words)
                            brand_query = f"""
                                SELECT {select_fields}
                                {brand_from}
                                ORDER BY
                                    {product_accessory_penalty_sql(True) if brand_accessory_intent else product_primary_device_rank_sql()},
                                    {product_accessory_penalty_sql()} ,
                                    {product_variant_rank_sql()},
                                    {product_storage_rank_sql()},
                                    CASE WHEN p.sort_index IS NULL THEN 1 ELSE 0 END,
                                    p.sort_index,
                                    p.id
                                LIMIT ? OFFSET ?
                            """
                            rows = db.execute(brand_query, (*brand_params, limit, offset)).fetchall()
                            return jsonify({
                                'items': build_compact_product_payloads(db, user_id, rows, include_dynamic_proposed=include_dynamic_proposed),
                                'total': brand_total,
                                'page': page,
                                'limit': limit,
                            })

                fast_phrase_conditions = []
                fast_phrase_params = []
                for phrase in build_fast_iphone_search_phrases(raw_query):
                    like_value = f"%{phrase}%"
                    fast_phrase_conditions.extend([
                        "p.name LIKE ? COLLATE NOCASE",
                        "p.synonyms LIKE ? COLLATE NOCASE",
                    ])
                    fast_phrase_params.extend([like_value, like_value])
                if fast_phrase_conditions:
                    phrase_from = f"{base_from} AND ({' OR '.join(fast_phrase_conditions)})"
                    phrase_params = [*base_params, *fast_phrase_params]
                    phrase_total = int(db.execute(
                        f"SELECT COUNT(*) AS total {phrase_from}",
                        tuple(phrase_params),
                    ).fetchone()['total'] or 0)
                    if phrase_total:
                        phrase_accessory_intent = is_accessory_search_intent(query_words)
                        phrase_query = f"""
                            SELECT {select_fields}
                            {phrase_from}
                            ORDER BY
                                {product_accessory_penalty_sql(True) if phrase_accessory_intent else product_primary_device_rank_sql()},
                                {product_accessory_penalty_sql()} ,
                                CASE WHEN LOWER(p.name) LIKE ? THEN 0 ELSE 1 END,
                                {product_variant_rank_sql()},
                                {product_storage_rank_sql()},
                                CASE WHEN p.sort_index IS NULL THEN 1 ELSE 0 END,
                                p.sort_index,
                                p.id
                            LIMIT ? OFFSET ?
                        """
                        phrase_rank_params = [f"%{normalize_product_search_text(raw_query)}%", limit, offset]
                        rows = db.execute(phrase_query, (*phrase_params, *phrase_rank_params)).fetchall()
                        return jsonify({
                            'items': build_compact_product_payloads(db, user_id, rows, include_dynamic_proposed=include_dynamic_proposed),
                            'total': phrase_total,
                            'page': page,
                            'limit': limit,
                        })

                if is_gamepad_search_intent(query_words) and ps5_search_intent:
                    gamepad_from = f"""
                        {base_from}
                        AND (
                            p.name LIKE ? COLLATE NOCASE
                            OR p.synonyms LIKE ? COLLATE NOCASE
                            OR p.name LIKE ? COLLATE NOCASE
                            OR p.synonyms LIKE ? COLLATE NOCASE
                            OR p.name LIKE ? COLLATE NOCASE
                            OR p.synonyms LIKE ? COLLATE NOCASE
                        )
                        AND (
                            p.name LIKE ? COLLATE NOCASE
                            OR p.synonyms LIKE ? COLLATE NOCASE
                        )
                    """
                    gamepad_params = [
                        *base_params,
                        '%DualSense%',
                        '%DualSense%',
                        '%геймпад%',
                        '%геймпад%',
                        '%controller%',
                        '%controller%',
                        '%PS5%',
                        '%PS5%',
                    ]
                    gamepad_total = int(db.execute(
                        f"SELECT COUNT(*) AS total {gamepad_from}",
                        tuple(gamepad_params),
                    ).fetchone()['total'] or 0)
                    if gamepad_total:
                        gamepad_query = f"""
                            SELECT {select_fields}
                            {gamepad_from}
                            ORDER BY
                                CASE WHEN p.name LIKE '%DualSense%' THEN 0 ELSE 1 END,
                                CASE WHEN p.sort_index IS NULL THEN 1 ELSE 0 END,
                                p.sort_index,
                                p.id
                            LIMIT ? OFFSET ?
                        """
                        rows = db.execute(gamepad_query, (*gamepad_params, limit, offset)).fetchall()
                        return jsonify({
                            'items': build_compact_product_payloads(db, user_id, rows, include_dynamic_proposed=include_dynamic_proposed),
                            'total': gamepad_total,
                            'page': page,
                                'limit': limit,
                            })

                if ps5_search_intent and not is_accessory_search_intent(query_words):
                    ps5_extra_conditions, ps5_extra_params = build_ps5_extra_search_conditions(query_words)
                    ps5_extra_sql = ''.join(f" AND {condition}" for condition in ps5_extra_conditions)
                    ps5_from = f"""
                        {base_from}
                        AND (
                            LOWER(p.name) LIKE '%playstation 5%'
                            OR LOWER(p.name) LIKE '%ps5%'
                            OR LOWER(p.model_number) LIKE '%ps5%'
                            OR LOWER(p.synonyms) LIKE '%playstation 5%'
                            OR LOWER(p.synonyms) LIKE '%ps5%'
                        )
                        {ps5_extra_sql}
                    """
                    ps5_params = [*base_params, *ps5_extra_params]
                    ps5_total = int(db.execute(
                        f"SELECT COUNT(*) AS total {ps5_from}",
                        tuple(ps5_params),
                    ).fetchone()['total'] or 0)
                    if ps5_total:
                        ps5_query = f"""
                            SELECT {select_fields}
                            {ps5_from}
                            ORDER BY
                                CASE
                                    WHEN LOWER(p.name) LIKE 'sony playstation 5%' OR LOWER(p.name) LIKE 'playstation 5%' THEN 0
                                    WHEN LOWER(p.name) LIKE '%dualsense%' THEN 1
                                    WHEN LOWER(p.name) LIKE '%gamepad%' OR p.name LIKE '%Геймпад%' OR p.name LIKE '%геймпад%' THEN 2
                                    WHEN {product_accessory_match_sql()} THEN 4
                                    ELSE 3
                                END,
                                {product_storage_rank_sql()},
                                CASE WHEN p.sort_index IS NULL THEN 1 ELSE 0 END,
                                p.sort_index,
                                p.id
                            LIMIT ? OFFSET ?
                        """
                        rows = db.execute(ps5_query, (*ps5_params, limit, offset)).fetchall()
                        return jsonify({
                            'items': build_compact_product_payloads(db, user_id, rows, include_dynamic_proposed=include_dynamic_proposed),
                            'total': ps5_total,
                            'page': page,
                            'limit': limit,
                        })

                search_from = base_from
                search_params = list(base_params)
                searchable_fields = (
                    'p.name',
                    'p.synonyms',
                    'p.brand',
                    'p.model_number',
                    'p.color',
                    'p.storage',
                )
                for variants in sql_needles:
                    parts = []
                    for variant in variants:
                        for like_variant in expand_like_case_variants(variant):
                            for field in searchable_fields:
                                append_sql_search_field_match(parts, search_params, field, like_variant)
                    if parts:
                        search_from += f" AND ({' OR '.join(parts)})"

                total_query = f"SELECT COUNT(*) AS total {search_from}"
                total = int(db.execute(total_query, tuple(search_params)).fetchone()['total'] or 0)
                if total == 0:
                    fallback_entries = (
                        get_product_search_entries_for_folder_ids(db, user_id, folder_ids)
                        if folder_ids else get_cached_product_search_entries(db, user_id)
                    )
                    fallback_ids = search_products_in_entries(
                        fallback_entries,
                        raw_query,
                        folder_ids=set(folder_ids),
                        limit=500,
                    )
                    if fallback_ids:
                        fallback_page_ids = fallback_ids[offset:offset + limit]
                        rows = fetch_ranked_product_rows_for_ids(
                            db,
                            select_fields,
                            base_from,
                            base_params,
                            fallback_page_ids,
                        )
                        return jsonify({
                            'items': build_compact_product_payloads(db, user_id, rows, include_dynamic_proposed=include_dynamic_proposed),
                            'total': len(fallback_ids),
                            'page': page,
                            'limit': limit,
                        })

                normalized_query = normalize_product_search_text(raw_query)
                first_word = tokenize_product_search_query(raw_query)[0] if tokenize_product_search_query(raw_query) else ''
                first_word = normalize_product_search_text(first_word)
                accessory_order_sql = product_accessory_penalty_sql(is_accessory_search_intent(query_words))
                query = f"""
                    SELECT {select_fields}
                    {search_from}
                    ORDER BY
                        CASE WHEN LOWER(p.brand) = ? THEN 0 ELSE 1 END,
                        {accessory_order_sql if is_accessory_search_intent(query_words) else product_primary_device_rank_sql()},
                        {product_accessory_penalty_sql()} ,
                        CASE WHEN LOWER(p.name) LIKE ? THEN 0 ELSE 1 END,
                        {product_variant_rank_sql()},
                        {product_storage_rank_sql()},
                        CASE WHEN p.sort_index IS NULL THEN 1 ELSE 0 END,
                        p.sort_index,
                        p.id
                    LIMIT ? OFFSET ?
                """
                rank_params = [
                    first_word or normalized_query,
                    f"%{normalized_query}%",
                    limit,
                    offset,
                ]
                rows = db.execute(query, (*search_params, *rank_params)).fetchall()
                items = build_compact_product_payloads(db, user_id, rows, include_dynamic_proposed=include_dynamic_proposed)
                return jsonify({
                    'items': items,
                    'total': total,
                    'page': page,
                    'limit': limit,
                })

            total_query = f"SELECT COUNT(*) AS total {base_from}"
            total = int(db.execute(total_query, tuple(base_params)).fetchone()['total'] or 0)
            query = f"""
                SELECT {select_fields}
                {base_from}
                ORDER BY CASE WHEN p.sort_index IS NULL THEN 1 ELSE 0 END, p.sort_index, p.id
                LIMIT ? OFFSET ?
            """
            rows = db.execute(query, (*base_params, limit, offset)).fetchall()
            return jsonify({
                'items': build_compact_product_payloads(db, user_id, rows, include_dynamic_proposed=include_dynamic_proposed),
                'total': total,
                'page': page,
                'limit': limit,
            })

        count_fields = "" if is_compact else """,
                   COALESCE(confirmed.confirmed_count, 0) AS confirmed_count,
                   COALESCE(pending.pending_count, 0) AS pending_count
        """
        query = f"""
            SELECT {select_fields}{count_fields}
            {base_from}
            ORDER BY CASE WHEN p.sort_index IS NULL THEN 1 ELSE 0 END, p.sort_index, p.id
        """
        products = db.execute(query, tuple(base_params)).fetchall()

        if is_compact:
            return jsonify(build_compact_product_payloads(db, user_id, products, include_dynamic_proposed=include_dynamic_proposed))

        recent_message_lines = get_recent_message_lines_for_product_matching(db, user_id)

        result = []
        for p in products:
            p_dict = dict(p)
            p_dict['proposed_count'] = 0

            if product_has_dynamic_proposed_match(p_dict, recent_message_lines):
                p_dict['proposed_count'] = 1

            if p_dict['proposed_count'] == 0 and int(p_dict.get('pending_count') or 0) > 0:
                p_dict['proposed_count'] = 1

            result.append(p_dict)

        return jsonify(result)
        
    # POST
    # POST
    # POST
    # POST
    # Так как мы передаем файлы, используем request.form, а не get_json()
    folder_id = request.form.get('folder_id')
    
    # 1. Обработка синонимов (теперь они приходят из формы)
    try:
        syns = json.loads(request.form.get('synonyms', '[]'))
    except:
        syns = []
    synonyms_str = ", ".join([s.strip() for s in syns if s.strip()])
    resolved_folder_id = resolve_product_folder_assignment(
        db,
        user_id,
        request.form.get('name'),
        folder_id,
        request.form.get('storage'),
        request.form.get('brand'),
    )
    
    # 2. ВСТАВЛЯЕМ ВАШ НОВЫЙ КОД ДЛЯ ФОТО ЗДЕСЬ:
    files = request.files.getlist('photos')
    saved_names = []
    for file in files:
        if file and file.filename:
            # Генерируем уникальное имя, чтобы не было конфликтов
            ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else 'jpg'
            filename = f"{uuid.uuid4().hex}.{ext}"
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            saved_names.append(filename)

    # Сохраняем в БД просто как строку через запятую: "id1.jpg, id2.jpg"
    photo_url_str = ",".join(saved_names) 

    # 3. Сохраняем все в базу данных
    specs_payload = parse_specs_payload(
        request.form.get('specs'),
        {
            'brand': request.form.get('brand'),
            'country': request.form.get('country'),
            'weight': request.form.get('weight'),
            'model_number': request.form.get('model_number'),
            'color': request.form.get('color'),
            'storage': request.form.get('storage'),
            'ram': request.form.get('ram'),
            'warranty': request.form.get('warranty'),
        },
    )
    specs_json = json.dumps(specs_payload, ensure_ascii=False) if specs_payload else ""

    # 3. Сохраняем все в базу данных
    db.execute("""
        INSERT INTO products 
        (user_id, name, synonyms, price, folder_id, photo_url, brand, country, weight, model_number, is_on_request, color, storage, ram, warranty, description, description_html, specs, sort_index) 
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        user_id, 
        request.form.get('name'), 
        synonyms_str, 
        float(request.form.get('price', 0.0)), 
        resolved_folder_id,
        
        photo_url_str,
        
        request.form.get('brand'),
        request.form.get('country'),
        request.form.get('weight'),
        request.form.get('model_number'),
        1 if request.form.get('is_on_request') == '1' else 0,
        
        # 🔽 ДОБАВЛЯЕМ ЭТИ СТРОКИ СЮДА 🔽
        request.form.get('color'),
        request.form.get('storage'),
        request.form.get('ram'),
        request.form.get('warranty'),
        request.form.get('description'),
        request.form.get('description_html'),
        specs_json,
        get_next_product_sort_index(db, user_id)
    ))
    db.commit()
    notify_clients()
    return jsonify({'success': True})


@app.route('/api/products/search')
@login_required
def search_products_route():
    user_id = session['user_id']
    raw_query = str(request.args.get('q', '') or '').strip()
    if not raw_query:
        return jsonify({'ids': [], 'count': 0})

    try:
        limit = int(request.args.get('limit', 200))
    except (TypeError, ValueError):
        limit = 200
    limit = max(10, min(limit, 500))

    folder_ids = set()
    for part in str(request.args.get('folder_ids', '') or '').split(','):
        part = part.strip()
        if not part:
            continue
        try:
            folder_ids.add(int(part))
        except ValueError:
            continue

    db = get_db()
    entries = get_cached_product_search_entries(db, user_id)
    result_ids = search_products_in_entries(entries, raw_query, folder_ids=folder_ids, limit=limit)
    return jsonify({
        'ids': result_ids,
        'count': len(result_ids),
    })

@app.route('/api/products/<int:prod_id>', methods=['GET', 'DELETE'])
@login_required
def delete_product(prod_id):
    if request.method == 'GET':
        db = get_db()
        user_id = session['user_id']
        product = db.execute(
            """
            SELECT p.*,
                   COALESCE(confirmed.confirmed_count, 0) AS confirmed_count,
                   COALESCE(pending.pending_count, 0) AS pending_count
            FROM products p
            LEFT JOIN (
                SELECT pm.product_id, COUNT(*) AS confirmed_count
                FROM product_messages pm
                JOIN messages m ON pm.message_id = m.id
                WHERE pm.status = 'confirmed'
                  AND COALESCE(pm.is_actual, 1) = 1
                  AND m.user_id = ?
                GROUP BY pm.product_id
            ) confirmed ON confirmed.product_id = p.id
            LEFT JOIN (
                SELECT product_id, COUNT(*) AS pending_count
                FROM product_messages
                WHERE status = 'pending'
                GROUP BY product_id
            ) pending ON pending.product_id = p.id
            WHERE p.id = ? AND p.user_id = ?
            """,
            (user_id, prod_id, user_id),
        ).fetchone()
        if not product:
            return jsonify({'error': 'Товар не найден'}), 404

        product_dict = normalize_catalog_product_snapshot(db, user_id, product)
        product_dict['proposed_count'] = 0
        if int(product_dict.get('confirmed_count') or 0) == 0:
            if int(product_dict.get('pending_count') or 0) > 0:
                product_dict['proposed_count'] = 1
            else:
                recent_message_lines = get_recent_message_lines_for_product_matching(db, user_id)
                if product_has_dynamic_proposed_match(product_dict, recent_message_lines):
                    product_dict['proposed_count'] = 1
        return jsonify(product_dict)

    db = get_db()
    db.execute("DELETE FROM products WHERE id=? AND user_id=?", (prod_id, session['user_id']))
    db.commit()
    notify_clients()
    return jsonify({'success': True})


@app.route('/api/reports/confirmed')
@login_required
def get_all_confirmed():
    user_id = session['user_id']
    db = get_db()
    fast_mode = str(request.args.get('fast', '')).strip().lower() in {'1', 'true', 'yes'}
    product_ids = []
    for part in str(request.args.get('product_ids', '') or '').split(','):
        part = part.strip()
        if not part:
            continue
        try:
            product_ids.append(int(part))
        except ValueError:
            continue
    product_ids = list(dict.fromkeys(product_ids))
    if len(product_ids) > 600:
        product_ids = []

    if AUTO_BINDING_MAINTENANCE_ENABLED and not fast_mode:
        should_notify_products = False
        with sqlite_write_lock:
            removed_duplicates = prune_duplicate_confirmed_bindings(db, user_id=user_id)
            deactivated_bindings = deactivate_confirmed_bindings_from_inactive_messages(db, user_id=user_id)
            removed_discount_bindings = remove_discount_section_confirmed_bindings(db, user_id=user_id)
            sanitized_bindings = sanitize_confirmed_binding_records(db, user_id=user_id)
            invalid_bindings = deactivate_semantically_invalid_confirmed_bindings(db, user_id=user_id)
            refreshed_bindings = refresh_confirmed_bindings_from_latest_messages(db, user_id=user_id)
            if removed_duplicates or deactivated_bindings or removed_discount_bindings or sanitized_bindings or invalid_bindings or refreshed_bindings:
                db.commit()
                should_notify_products = True
        if should_notify_products:
            notify_clients("products")
    
    query = """
        SELECT pm.id as binding_id, pm.group_id, p.id as product_id, p.name as product_name, p.sort_index as product_sort_index, pm.extracted_price, m.text, 
               m.chat_id, m.chat_title, m.sender_name, m.date, pm.line_index, pm.message_id, pm.is_actual,
               m.telegram_message_id, m.type, ec.sheet_url, tc.chat_title as tc_chat_title, tc.custom_name
        FROM product_messages pm
        JOIN products p ON pm.product_id = p.id
        JOIN messages m ON pm.message_id = m.id
        LEFT JOIN tracked_chats tc ON m.chat_id = tc.chat_id AND m.user_id = tc.user_id
        LEFT JOIN excel_configs ec ON m.chat_id = ec.chat_id AND m.type LIKE 'excel%'
        WHERE p.user_id = ?
        AND pm.status = 'confirmed'
        AND pm.is_actual = 1
        AND (m.is_blocked IS NULL OR m.is_blocked = 0)
        AND (m.is_delayed IS NULL OR m.is_delayed = 0)
    """
    params = [user_id]
    if product_ids:
        placeholders = ','.join('?' for _ in product_ids)
        query += f" AND p.id IN ({placeholders})"
        params.extend(product_ids)
    rows = db.execute(query, tuple(params)).fetchall()

    if fast_mode:
        final_result = []
        grouped_reports = {}
        for row in rows:
            d = dict(row)
            source_line, resolved_price = resolve_binding_display_line_and_price(
                d.get('text'),
                d.get('line_index'),
                d.get('extracted_price'),
                require_price=False,
            )
            if not source_line:
                continue
            if resolved_price is not None:
                d['extracted_price'] = resolved_price
            d['text'] = strip_supplier_line_noise(source_line) or source_line.strip()
            g_id = d['group_id'] if d['group_id'] else d['binding_id']
            if g_id not in grouped_reports or d['date'] > grouped_reports[g_id]['date']:
                grouped_reports[g_id] = d
        final_result = list(grouped_reports.values())
        final_result.sort(key=lambda x: x['date'], reverse=True)
        return jsonify(final_result)
    
    # Собираем актуальные версии строк (с добавленными telegram_message_id и type)
    all_msgs = db.execute("SELECT id, text, chat_id, sender_name, date, telegram_message_id, type FROM messages WHERE user_id = ? AND (is_blocked IS NULL OR is_blocked = 0) AND (is_delayed IS NULL OR is_delayed = 0) ORDER BY date DESC", (user_id,)).fetchall()
    latest_fingerprints = {}
    for row in all_msgs:
        if not row['text']: continue
        lines = row['text'].split('\n')
        in_discount_section = False
        for i, line in enumerate(lines):
            if not line.strip(): continue
            if line_is_discount_section_marker(line):
                in_discount_section = True
                continue
            if in_discount_section:
                continue
            if line_has_discount_condition_marker(line):
                continue
            fp = get_binding_fingerprint(
                row['chat_id'],
                row['sender_name'],
                row['text'],
                i,
                include_price=False,
            )
            if fp not in latest_fingerprints:
                price_val = extract_price_with_context(lines, i)

                latest_fingerprints[fp] = {
                    'message_id': row['id'],
                    'line_index': i,
                    'full_text': row['text'],
                    'text': line,
                    'date': row['date'],
                    'price': price_val,
                    'telegram_message_id': row['telegram_message_id'],
                    'type': row['type']
                }


    result_raw = []
    has_updates = False 
    delete_binding_ids = []
    for r in rows:
        d = dict(r)
        if d['line_index'] != -1 and d['text']:
            lines = d['text'].split('\n')
            if 0 <= d['line_index'] < len(lines):
                orig_line = lines[d['line_index']]
                fp = get_binding_fingerprint(
                    d['chat_id'],
                    d['sender_name'],
                    d['text'],
                    d['line_index'],
                    include_price=False,
                )
                
                latest = latest_fingerprints.get(fp)
                if latest and latest['date'] > d['date']:
                    d['text'] = latest['full_text']
                    d['extracted_price'] = latest['price']
                    d['date'] = latest['date']
                    d['telegram_message_id'] = latest['telegram_message_id']
                    d['type'] = latest['type']
                    d['message_id'] = latest['message_id']
                    d['line_index'] = latest['line_index']
                    d['is_actual'] = 1
                    
                    current_id = d.get('pm_id') or d.get('binding_id') or d.get('id')

                    try:
                        db.execute("UPDATE product_messages SET message_id=?, line_index=?, extracted_price=?, is_actual=1 WHERE id=?",
                                   (latest['message_id'], latest['line_index'], latest['price'], current_id))
                        has_updates = True
                    except sqlite3.IntegrityError:
                        if current_id:
                            db.execute("DELETE FROM product_messages WHERE id=?", (current_id,))
                            has_updates = True
                else:
                    d['text'] = r['text']
                if d.get('extracted_price') in (None, ''):
                    recovered_price = extract_price_with_context(lines, d['line_index'])
                    if recovered_price is not None:
                        d['extracted_price'] = recovered_price
                        current_id = d.get('pm_id') or d.get('binding_id') or d.get('id')
                        try:
                            db.execute("UPDATE product_messages SET extracted_price=? WHERE id=?", (recovered_price, current_id))
                            has_updates = True
                        except Exception:
                            pass
        source_line, resolved_price = resolve_binding_display_line_and_price(
            d.get('text'),
            d.get('line_index'),
            d.get('extracted_price'),
            require_price=False,
        )
        if not source_line:
            current_id = d.get('pm_id') or d.get('binding_id') or d.get('id')
            if current_id:
                delete_binding_ids.append(int(current_id))
                has_updates = True
            continue
        current_price = normalize_binding_price_value(d.get('extracted_price'))
        if resolved_price is not None and (current_price is None or abs(current_price - resolved_price) > 0.009):
            d['extracted_price'] = resolved_price
            current_id = d.get('pm_id') or d.get('binding_id') or d.get('id')
            try:
                db.execute("UPDATE product_messages SET extracted_price=? WHERE id=?", (resolved_price, current_id))
                has_updates = True
            except Exception:
                pass
        cleaned_line = strip_supplier_line_noise(source_line) or source_line.strip()
        d['match_line'] = cleaned_line
        d['text'] = cleaned_line
        result_raw.append(d)

    if delete_binding_ids:
        for offset in range(0, len(delete_binding_ids), 500):
            chunk = delete_binding_ids[offset:offset + 500]
            placeholders = ",".join("?" for _ in chunk)
            db.execute(f"DELETE FROM product_messages WHERE id IN ({placeholders})", chunk)
        
    if has_updates:
        db.commit()
        notify_clients()

    grouped_reports = {}
    for d in result_raw:
        g_id = d['group_id'] if d['group_id'] else d['binding_id']
        if g_id not in grouped_reports or d['date'] > grouped_reports[g_id]['date']:
            grouped_reports[g_id] = d
            
    final_result = list(grouped_reports.values())
    final_result.sort(key=lambda x: x['date'], reverse=True)
    return jsonify(final_result)


# --- Маршруты для Взаимодействия с ботами ---
@app.route('/api/interaction_bots', methods=['GET', 'POST'])
@login_required
def manage_interaction_bots():
    user_id = session['user_id']
    db = get_db()
    if request.method == 'GET':
        bots = db.execute("SELECT * FROM interaction_bots WHERE user_id=?", (user_id,)).fetchall()
        return jsonify([dict(b) for b in bots])
    
    # POST
    data = request.get_json()
    bot_username = data.get('bot_username').strip()
    custom_name = data.get('custom_name')
    userbot_id = data.get('userbot_id')
    interval_minutes = normalize_interval_minutes(data.get('interval_minutes'), 60)
    
    # 1. Находим запущенного юзербота (он нужен, чтобы узнать числовой ID бота)
    client_to_use = None
    client_loop_to_use = None
    
    if userbot_id in user_clients:
        _, client_to_use, _ = user_clients[userbot_id]
        client_loop_to_use = user_loops.get(userbot_id)
    else:
        active_sessions = [sid for sid, (_, _, uid) in user_clients.items() if uid == user_id]
        if active_sessions:
            sid = active_sessions[0]
            _, client_to_use, _ = user_clients[sid]
            client_loop_to_use = user_loops.get(sid)

    # 2. Превращаем текстовый @username в ЧИСЛОВОЙ ID для парсера
    numeric_chat_id = None
    if client_to_use and client_loop_to_use:
        async def get_bot_id():
            try:
                entity = await client_to_use.get_entity(bot_username)
                return entity.id
            except Exception as e:
                logging.error(f"Не удалось получить ID для {bot_username}: {e}")
                return None
        
        try:
            future = asyncio.run_coroutine_threadsafe(get_bot_id(), client_loop_to_use)
            numeric_chat_id = future.result(timeout=10)
        except Exception:
            pass

    # Если юзербот выключен, мы не сможем получить ID — выдаем ошибку
    if not numeric_chat_id:
        return jsonify({'error': 'Не удалось определить ID бота. Убедитесь, что юзербот включен и юзернейм указан верно (например, @MyBot).'}), 400

    # 3. АВТОМАТИЧЕСКОЕ ДОБАВЛЕНИЕ В ПАРСЕР (здесь используем ЧИСЛОВОЙ ID!)
    tracked_chat_id = None
    try:
        db.execute("INSERT INTO tracked_chats (user_id, chat_id, chat_title, custom_name) VALUES (?, ?, ?, ?)",
                   (user_id, numeric_chat_id, bot_username, custom_name))
    except sqlite3.IntegrityError:
        pass # Уже добавлен
    tracked_row = db.execute(
        "SELECT id FROM tracked_chats WHERE user_id=? AND chat_id=?",
        (user_id, numeric_chat_id)
    ).fetchone()
    if tracked_row:
        tracked_chat_id = tracked_row['id']

    # 4. Сохраняем автоматизацию и прямую связь с записью в отслеживаемых чатах.
    db.execute("""INSERT INTO interaction_bots 
                  (user_id, userbot_id, bot_username, custom_name, commands, interval_minutes, status, tracked_chat_id, resolved_chat_id) 
                  VALUES (?, ?, ?, ?, ?, ?, 'active', ?, ?)""",
               (user_id, userbot_id, bot_username, custom_name,
                json.dumps(data.get('commands', [])), interval_minutes, tracked_chat_id, numeric_chat_id))

    db.commit()
    notify_clients()

    # 5. ПАРСИНГ ИСТОРИИ (также по числовому ID)
    # 5. ПАРСИНГ ИСТОРИИ 
    chat_title_for_parser = custom_name if custom_name else bot_username

    return jsonify({'success': True})

@app.route('/api/interaction_bots/<int:id>', methods=['DELETE'])
@login_required
def delete_interaction_bot(id):
    db = get_db()
    user_id = session['user_id']
    
    bot = db.execute("SELECT * FROM interaction_bots WHERE id=? AND user_id=?", (id, user_id)).fetchone()
    if bot:
        tracked_ids = get_interaction_bot_tracked_chat_ids(db, user_id, [bot])
        delete_tracked_chats_by_ids(db, user_id, tracked_ids)
            
    db.execute("DELETE FROM interaction_bots WHERE id=? AND user_id=?", (id, user_id))
    db.commit()
    notify_clients()
    return jsonify({'success': True})

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

def init_db():
    if not os.path.exists('sessions'):
        os.makedirs('sessions')
    with app.app_context():
        db = get_db()
        db.execute('''
        CREATE TABLE IF NOT EXISTS pub_markups (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    pub_id INTEGER NOT NULL,
                    folder_id INTEGER DEFAULT 0,
                    markup_type TEXT DEFAULT 'percent',
                    markup_value REAL DEFAULT 0,
                    rounding INTEGER DEFAULT 100,
                    FOREIGN KEY(pub_id) REFERENCES publications(id) ON DELETE CASCADE,
                    UNIQUE(pub_id, folder_id)
                )
    ''')
        db.execute('''
        CREATE TABLE IF NOT EXISTS publications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            is_active INTEGER DEFAULT 0,
            interval_min INTEGER DEFAULT 60,
            chat_id TEXT,
            message_id TEXT,
            userbot_id INTEGER,
            template TEXT,
            allowed_items TEXT -- Здесь будет храниться JSON с выбранными галочками (как в API)
        )
    ''')
        db.execute('''
        CREATE TABLE IF NOT EXISTS api_sources (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            name TEXT,
            url TEXT,
            token TEXT,
            interval_min INTEGER DEFAULT 60,
            is_active INTEGER DEFAULT 1
        )
    ''')
        db.execute('''
        CREATE TABLE IF NOT EXISTS ai_autofill_cache (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cache_key TEXT UNIQUE NOT NULL,
            user_id INTEGER DEFAULT 0,
            product_name TEXT NOT NULL,
            article TEXT DEFAULT '',
            payload TEXT NOT NULL,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
        db.execute('''
        CREATE TABLE IF NOT EXISTS ai_repair_jobs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            product_id INTEGER NOT NULL,
            product_name TEXT NOT NULL,
            article TEXT DEFAULT '',
            preferred_folder_id INTEGER,
            status TEXT DEFAULT 'pending',
            attempts INTEGER DEFAULT 0,
            next_retry_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            last_error TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE CASCADE
        )
    ''')
        db.executescript('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                login TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                role TEXT DEFAULT 'user'
            );
            CREATE TABLE IF NOT EXISTS user_telegram_sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                api_id TEXT,
                api_hash TEXT,
                session_file TEXT,
                status TEXT DEFAULT 'inactive',
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
            );
            CREATE TABLE IF NOT EXISTS messages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                telegram_message_id INTEGER,
                type TEXT,
                text TEXT,
                date TIMESTAMP,
                chat_id INTEGER,
                chat_title TEXT,
                is_blocked INTEGER DEFAULT 0,
                sender_name TEXT,
                is_delayed INTEGER DEFAULT 0,
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE,
                UNIQUE(chat_id, telegram_message_id)
            );
            CREATE TABLE IF NOT EXISTS folders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                name TEXT NOT NULL,
                parent_id INTEGER,
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE,
                FOREIGN KEY(parent_id) REFERENCES folders(id) ON DELETE CASCADE
            );
            CREATE TABLE IF NOT EXISTS saved_messages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                message_id INTEGER,
                folder_id INTEGER,
                saved_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(message_id) REFERENCES messages(id) ON DELETE CASCADE,
                FOREIGN KEY(folder_id) REFERENCES folders(id) ON DELETE CASCADE,
                UNIQUE(message_id, folder_id)
            );
            CREATE TABLE IF NOT EXISTS tracked_chats (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                chat_id INTEGER NOT NULL,
                chat_title TEXT,
                added_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE,
                UNIQUE(user_id, chat_id)
            );
            CREATE TABLE IF NOT EXISTS products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                name TEXT NOT NULL,
                synonyms TEXT,
                price REAL,
                sort_index INTEGER,
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
            );
            CREATE TABLE IF NOT EXISTS product_messages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER NOT NULL,
                message_id INTEGER NOT NULL,
                line_index INTEGER DEFAULT -1,
                group_id INTEGER,
                status TEXT DEFAULT 'proposed',
                extracted_price REAL,
                is_actual INTEGER DEFAULT 1,
                FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE CASCADE,
                FOREIGN KEY(message_id) REFERENCES messages(id) ON DELETE CASCADE,
                UNIQUE(product_id, message_id, line_index)
            );
            
            CREATE TABLE IF NOT EXISTS interaction_bots (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                userbot_id INTEGER NOT NULL,
                bot_username TEXT NOT NULL,
                commands TEXT NOT NULL,
                interval_minutes INTEGER NOT NULL DEFAULT 60,
                last_run TIMESTAMP,
                last_command_message_id INTEGER,
                last_response_at TIMESTAMP,
                manual_intervention_at TIMESTAMP,
                status TEXT DEFAULT 'active',
                tracked_chat_id INTEGER,
                resolved_chat_id INTEGER,
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE,
                FOREIGN KEY(userbot_id) REFERENCES user_telegram_sessions(id) ON DELETE CASCADE
            );
        ''')
        # Добавляем новые колонки в существующие таблицы (игнорируем ошибку, если они уже есть)
        # 1. Безопасное добавление колонок
        columns_to_add = [
            ("products", "photo_url TEXT DEFAULT NULL"),
            ("products", "source_url TEXT DEFAULT NULL"),
            ("products", "brand TEXT DEFAULT NULL"),
            ("products", "country TEXT DEFAULT NULL"),
            ("products", "weight TEXT DEFAULT NULL"),
            ("products", "model_number TEXT DEFAULT NULL"),
            ("products", "is_on_request INTEGER DEFAULT 0"),
            ("products", "color TEXT DEFAULT NULL"),
            ("products", "storage TEXT DEFAULT NULL"),
            ("products", "ram TEXT DEFAULT NULL"),
            ("products", "warranty TEXT DEFAULT NULL"),
           ("products", "color TEXT DEFAULT NULL"),
            ("products", "storage TEXT DEFAULT NULL"),
            ("products", "ram TEXT DEFAULT NULL"),
            ("products", "warranty TEXT DEFAULT NULL"),
            ("products", "description TEXT DEFAULT NULL"),
            ("products", "description_html TEXT DEFAULT NULL"),
            ("products", "specs TEXT DEFAULT NULL"),
            ("products", "sort_index INTEGER"),
            ("products", "specs TEXT DEFAULT NULL"),
            ("products", "color TEXT DEFAULT NULL"),
            ("products", "storage TEXT DEFAULT NULL"),
            ("products", "ram TEXT DEFAULT NULL"),
            ("products", "warranty TEXT DEFAULT NULL"),
            ("products", "description TEXT DEFAULT NULL"),
            ("products", "description_html TEXT DEFAULT NULL"),
            ("products", "specs TEXT DEFAULT NULL"),
            ("product_messages", "line_index INTEGER DEFAULT -1"),
            ("product_messages", "group_id INTEGER"),
            ("product_messages", "line_index INTEGER DEFAULT -1"),
            ("product_messages", "group_id INTEGER"),
            ("product_messages", "is_actual INTEGER DEFAULT 1"), 
            ("messages", "is_blocked INTEGER DEFAULT 0"),
            ("messages", "is_delayed INTEGER DEFAULT 0"),
            ("tracked_chats", "custom_name TEXT"),
            ("interaction_bots", "custom_name TEXT"),
            ("interaction_bots", "tracked_chat_id INTEGER"),
            ("interaction_bots", "resolved_chat_id INTEGER"),
            ("interaction_bots", "last_command_message_id INTEGER"),
            ("interaction_bots", "last_response_at TIMESTAMP"),
            ("interaction_bots", "manual_intervention_at TIMESTAMP"),
            ("messages", "sender_name TEXT"),
            ("products", "folder_id INTEGER"),
            ("user_telegram_sessions", "account_name TEXT"),
            ("user_telegram_sessions", "time_start TEXT DEFAULT '00:00'"), 
            ("user_telegram_sessions", "time_end TEXT DEFAULT '23:59'"),   
            ("user_telegram_sessions", "schedule_enabled INTEGER DEFAULT 0"),
            # --- НОВЫЕ КОЛОНКИ ДЛЯ API КЛИЕНТОВ ---
            ("api_clients", "time_start TEXT DEFAULT '00:00'"),
            ("api_clients", "time_end TEXT DEFAULT '23:59'"),
            ("api_clients", "schedule_enabled INTEGER DEFAULT 0"),
            ("api_clients", "access_rules TEXT"),
            ("api_clients", "allowed_folders TEXT DEFAULT 'all'"), # <--- ДОБАВИТЬ ЭТО
            ("api_clients", "allowed_chats TEXT DEFAULT 'all'"),
            ("api_clients", "publish_chat_id INTEGER"),
            ("api_clients", "publish_message_id INTEGER"),
            ("api_clients", "publish_enabled INTEGER DEFAULT 0"),
            ("api_clients", "userbot_id INTEGER"),
            ("api_clients", "publish_template TEXT"),
            ("publications", "user_id INTEGER"),
            ("api_clients", "folders TEXT"),
            ("api_clients", "publish_interval INTEGER DEFAULT 60"),
            ("users", "session_token TEXT"),
            ("excel_configs", "is_grouped INTEGER DEFAULT 0"),
            ("publications", "markup_type TEXT DEFAULT 'percent'"), # <--- НОВОЕ
            ("publications", "markup_value REAL DEFAULT 0"),        # <--- НОВОЕ
            ("publications", "rounding INTEGER DEFAULT 100"),
            ("excel_configs", "sku_col INTEGER DEFAULT -1"),
            ("excel_configs", "source_type TEXT DEFAULT 'file'"),       # 'file' или 'google_sheet'
            ("excel_configs", "sheet_url TEXT"),                        # Ссылка на таблицу
            ("excel_configs", "parse_interval INTEGER DEFAULT 60")
        ]
        for table, col in columns_to_add:
            try:
                db.execute(f"ALTER TABLE {table} ADD COLUMN {col};")
            except sqlite3.OperationalError:
                pass # Если колонка уже есть, просто идем к следующей

        try:
            db.execute("UPDATE products SET sort_index = id WHERE sort_index IS NULL")
        except sqlite3.OperationalError:
            pass

        # 2. Обновление таблицы привязок (чтобы разные строки одного сообщения не затирали друг друга)
        # 2. Обновление таблицы привязок

        excel_configs_sql = get_table_sql(db, 'excel_configs')
        if excel_configs_sql and 'sheet_name TEXT NOT NULL' not in excel_configs_sql:
            db.execute("PRAGMA foreign_keys = OFF")
            db.executescript('''
                CREATE TABLE IF NOT EXISTS excel_configs_new (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    chat_id INTEGER NOT NULL,
                    sheet_name TEXT NOT NULL,
                    name_col INTEGER NOT NULL,
                    name_row_offset INTEGER NOT NULL,
                    price_col INTEGER NOT NULL,
                    price_row_offset INTEGER NOT NULL,
                    block_step INTEGER NOT NULL,
                    start_row INTEGER NOT NULL,
                    is_grouped INTEGER DEFAULT 0,
                    sku_col INTEGER DEFAULT -1,
                    source_type TEXT DEFAULT 'excel',
                    sheet_url TEXT,
                    parse_interval INTEGER DEFAULT 60,
                    folder_id INTEGER,
                    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE,
                    UNIQUE(chat_id, sheet_name)
                );
                INSERT OR IGNORE INTO excel_configs_new (
                    id, user_id, chat_id, sheet_name, name_col, name_row_offset,
                    price_col, price_row_offset, block_step, start_row,
                    is_grouped, sku_col, source_type, sheet_url, parse_interval, folder_id
                )
                SELECT
                    id, user_id, chat_id, 'Лист1', name_col, name_row_offset,
                    price_col, price_row_offset, block_step, start_row,
                    0, -1, 'excel', NULL, 60, NULL
                FROM excel_configs;
                DROP TABLE excel_configs;
                ALTER TABLE excel_configs_new RENAME TO excel_configs;
            ''')
            db.execute("PRAGMA foreign_keys = ON")



        try:
            # --- ТАБЛИЦА ДЛЯ НАСТРОЕК EXCEL ---
            db.executescript('''
                CREATE TABLE IF NOT EXISTS excel_missing_sheets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                chat_id INTEGER,
                chat_title TEXT,
                sheet_name TEXT,
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE,
                UNIQUE(chat_id, sheet_name)
            );
            ''')
            db.executescript('''
                CREATE TABLE IF NOT EXISTS supplier_missing_positions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    source_chat_id TEXT,
                    source_name TEXT,
                    message_id INTEGER,
                    line_index INTEGER DEFAULT -1,
                    line_text TEXT NOT NULL,
                    match_line TEXT,
                    normalized_key TEXT NOT NULL,
                    brand TEXT,
                    category TEXT,
                    price_value REAL,
                    status TEXT DEFAULT 'unmatched',
                    candidate_count INTEGER DEFAULT 0,
                    candidate_snapshot TEXT,
                    seen_count INTEGER DEFAULT 1,
                    first_seen_at TEXT,
                    last_seen_at TEXT,
                    resolved_product_id INTEGER,
                    resolved_at TEXT,
                    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE,
                    UNIQUE(user_id, normalized_key)
                );
                CREATE INDEX IF NOT EXISTS idx_supplier_missing_positions_user_status
                    ON supplier_missing_positions(user_id, status, last_seen_at);
            ''')
            db.executescript('''
                CREATE TABLE IF NOT EXISTS supplier_line_memory (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    normalized_key TEXT NOT NULL,
                    product_id INTEGER NOT NULL,
                    source_name TEXT,
                    message_id INTEGER,
                    line_text TEXT,
                    seen_count INTEGER DEFAULT 1,
                    first_confirmed_at TEXT,
                    last_confirmed_at TEXT,
                    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE,
                    UNIQUE(user_id, normalized_key)
                );
                CREATE INDEX IF NOT EXISTS idx_supplier_line_memory_user_product
                    ON supplier_line_memory(user_id, product_id);
            ''')
            db.executescript('''
                CREATE TABLE IF NOT EXISTS excel_configs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    chat_id INTEGER NOT NULL,
                    sheet_name TEXT NOT NULL DEFAULT '*',
                    name_col INTEGER NOT NULL,
                    name_row_offset INTEGER NOT NULL,
                    price_col INTEGER NOT NULL,
                    price_row_offset INTEGER NOT NULL,
                    block_step INTEGER NOT NULL,
                    start_row INTEGER NOT NULL,
                    is_grouped INTEGER DEFAULT 0,
                    sku_col INTEGER DEFAULT -1,
                    source_type TEXT DEFAULT 'excel',
                    sheet_url TEXT,
                    parse_interval INTEGER DEFAULT 60,
                    folder_id INTEGER,
                    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE,
                    UNIQUE(chat_id, sheet_name)
                );
            ''')
            # --- ТАБЛИЦЫ ДЛЯ API И НАЦЕНОК ---
            db.executescript('''
                CREATE TABLE IF NOT EXISTS api_clients (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    name TEXT NOT NULL,
                    token TEXT UNIQUE NOT NULL,
                    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
                );
                CREATE TABLE IF NOT EXISTS api_markups (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    client_id INTEGER NOT NULL,
                    folder_id INTEGER DEFAULT 0, -- 0 означает "Для всех остальных папок"
                    markup_type TEXT DEFAULT 'percent', -- 'percent' или 'fixed'
                    markup_value REAL DEFAULT 0,
                    rounding INTEGER DEFAULT 100, -- 100, 500, 1000 и тд.
                    FOREIGN KEY(client_id) REFERENCES api_clients(id) ON DELETE CASCADE,
                    UNIQUE(client_id, folder_id)
                );
            ''')
            product_messages_sql = get_table_sql(db, 'product_messages') or ""
            if 'is_actual' not in product_messages_sql:
                db.executescript('''
                    CREATE TABLE IF NOT EXISTS pm_new (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        product_id INTEGER NOT NULL,
                        message_id INTEGER NOT NULL,
                        line_index INTEGER DEFAULT -1,
                        group_id INTEGER,
                        status TEXT DEFAULT 'proposed',
                        extracted_price REAL,
                        is_actual INTEGER DEFAULT 1,
                        FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE CASCADE,
                        FOREIGN KEY(message_id) REFERENCES messages(id) ON DELETE CASCADE,
                        UNIQUE(product_id, message_id, line_index)
                    );
                    INSERT OR IGNORE INTO pm_new (id, product_id, message_id, line_index, group_id, status, extracted_price, is_actual) 
                    SELECT id, product_id, message_id, line_index, group_id, status, extracted_price, 1 FROM product_messages;
                    DROP TABLE IF EXISTS product_messages;
                    ALTER TABLE pm_new RENAME TO product_messages;
                ''')
        except Exception as e:
            pass


        messages_sql = get_table_sql(db, 'messages')
        if messages_sql and 'UNIQUE(chat_id, telegram_message_id)' not in messages_sql:
            try:
                db.execute("PRAGMA foreign_keys = OFF")
                db.executescript('''
                    CREATE TABLE IF NOT EXISTS messages_new (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        user_id INTEGER,
                        telegram_message_id INTEGER,
                        type TEXT,
                        text TEXT,
                        date TIMESTAMP,
                        chat_id INTEGER,
                        chat_title TEXT,
                        is_blocked INTEGER DEFAULT 0,
                        sender_name TEXT,
                        is_delayed INTEGER DEFAULT 0,
                        FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE,
                        UNIQUE(chat_id, telegram_message_id)
                    );
                    INSERT OR IGNORE INTO messages_new (
                        id, user_id, telegram_message_id, type, text, date,
                        chat_id, chat_title, is_blocked, sender_name, is_delayed
                    )
                    SELECT
                        id, user_id, telegram_message_id, type, text, date,
                        chat_id, chat_title, is_blocked, sender_name,
                        0
                    FROM messages;
                    DROP TABLE messages;
                    ALTER TABLE messages_new RENAME TO messages;
                ''')
                db.execute("PRAGMA foreign_keys = ON")
            except Exception as e:
                logger.error(f"Ошибка миграции сообщений: {e}")

        performance_indexes = (
            "CREATE INDEX IF NOT EXISTS idx_products_user_sort ON products(user_id, sort_index, id)",
            "CREATE INDEX IF NOT EXISTS idx_products_user_folder_sort ON products(user_id, folder_id, sort_index, id)",
            "CREATE INDEX IF NOT EXISTS idx_products_user_brand ON products(user_id, brand)",
            "CREATE INDEX IF NOT EXISTS idx_products_user_brand_nocase ON products(user_id, brand COLLATE NOCASE, sort_index, id)",
            "CREATE INDEX IF NOT EXISTS idx_product_messages_product_status ON product_messages(product_id, status)",
            "CREATE INDEX IF NOT EXISTS idx_product_messages_status_product ON product_messages(status, product_id)",
            "CREATE INDEX IF NOT EXISTS idx_product_messages_message ON product_messages(message_id)",
            "CREATE INDEX IF NOT EXISTS idx_messages_user_id ON messages(user_id, id)",
            "CREATE INDEX IF NOT EXISTS idx_folders_user_parent ON folders(user_id, parent_id, id)",
        )
        for index_sql in performance_indexes:
            try:
                db.execute(index_sql)
            except sqlite3.OperationalError as e:
                logger.warning(f"Не удалось создать индекс производительности: {e}")

        if AUTO_BINDING_MAINTENANCE_ENABLED:
            try:
                removed_confirmed_duplicates = prune_duplicate_confirmed_bindings(db)
                deactivated_confirmed_bindings = deactivate_confirmed_bindings_from_inactive_messages(db)
                invalid_confirmed_bindings = deactivate_semantically_invalid_confirmed_bindings(db)
                duplicate_actual_line_bindings = deactivate_duplicate_actual_line_bindings(db)
                if removed_confirmed_duplicates:
                    logger.info("Удалены дубли подтвержденных привязок: %s", removed_confirmed_duplicates)
                if deactivated_confirmed_bindings:
                    logger.info("Помечены неактуальными привязки из недоступных сообщений: %s", deactivated_confirmed_bindings)
                if invalid_confirmed_bindings:
                    logger.info("Помечены неактуальными семантически неверные привязки: %s", invalid_confirmed_bindings)
                if duplicate_actual_line_bindings:
                    logger.info("Помечены неактуальными дубли actual-строк: %s", duplicate_actual_line_bindings)
            except Exception as e:
                logger.warning(f"Не удалось очистить дубли подтвержденных привязок: {e}")

        if SUPPLIER_LEARNING_BACKFILL_ENABLED:
            try:
                supplier_learning_stats = backfill_supplier_learning_from_confirmed_bindings(db)
                if (
                    supplier_learning_stats.get("memory_written")
                    or supplier_learning_stats.get("missing_resolved")
                    or supplier_learning_stats.get("synonyms_added")
                ):
                    logger.info("Обучение поставщиков из подтвержденных привязок: %s", supplier_learning_stats)
            except Exception as e:
                logger.warning(f"Не удалось обучить память поставщиков из подтвержденных привязок: {e}")

        # 3. Создание учетки администратора по умолчанию
        cursor = db.execute("SELECT COUNT(*) FROM users")
        if cursor.fetchone()[0] == 0:
            pwd_hash = hashlib.sha256('admin'.encode()).hexdigest()
            db.execute("INSERT INTO users (login, password_hash, role) VALUES (?, ?, ?)",
                       ('admin', pwd_hash, 'admin'))
            db.execute("INSERT INTO folders (user_id, name, parent_id) VALUES (1, 'По умолчанию', NULL)")
            
        db.commit()
        notify_clients()

if os.environ.get("ASTORE_SKIP_INIT_DB") != "1":
    init_db()


@app.route('/api/products/<int:prod_id>/link_folder', methods=['POST'])
@login_required
def link_product_folder(prod_id):
    data = request.get_json()
    folder_id = data.get('folder_id') # Может быть None (отвязка)
    db = get_db()
    db.execute("UPDATE products SET folder_id=? WHERE id=? AND user_id=?", (folder_id, prod_id, session['user_id']))
    db.commit()
    notify_clients()
    return jsonify({'success': True})


@app.route('/api/products/<int:prod_id>', methods=['PUT'])
@login_required
def edit_product(prod_id):
    db = get_db()
    
    # 1. Синонимы (из FormData в JSON)
    try:
        syns = json.loads(request.form.get('synonyms', '[]'))
    except:
        syns = []
    synonyms_str = ", ".join([s.strip() for s in syns if s.strip()])
    resolved_folder_id = resolve_product_folder_assignment(
        db,
        session['user_id'],
        request.form.get('name'),
        request.form.get('folder_id'),
        request.form.get('storage'),
        request.form.get('brand'),
    )

    # 2. Существующие фото, которые не были удалены
    try:
        existing_photos = json.loads(request.form.get('existing_photos', '[]'))
    except:
        existing_photos = []
    try:
        photo_order = json.loads(request.form.get('photo_order', '[]'))
    except:
        photo_order = []

    # 3. Сохраняем новые фото, если они были добавлены
    files = request.files.getlist('photos')
    new_photo_tokens = request.form.getlist('new_photo_tokens')
    saved_names = []
    saved_by_token = {}
    for index, file in enumerate(files):
        if file and file.filename:
            ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else 'jpg'
            filename = f"{uuid.uuid4().hex}.{ext}"
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            saved_names.append(filename)
            if index < len(new_photo_tokens):
                saved_by_token[new_photo_tokens[index]] = filename

    # 4. Собираем финальный порядок фото
    all_photos = []
    used_existing = set()
    used_new = set()
    if isinstance(photo_order, list) and photo_order:
        for item in photo_order:
            if not isinstance(item, dict):
                continue
            item_kind = item.get('kind')
            item_value = item.get('value')
            if item_kind == 'existing' and item_value in existing_photos and item_value not in used_existing:
                all_photos.append(item_value)
                used_existing.add(item_value)
            elif item_kind == 'new' and item_value in saved_by_token and item_value not in used_new:
                all_photos.append(saved_by_token[item_value])
                used_new.add(item_value)

    for photo_name in existing_photos:
        if photo_name not in used_existing:
            all_photos.append(photo_name)
    for token, photo_name in saved_by_token.items():
        if token not in used_new:
            all_photos.append(photo_name)
    if not all_photos:
        all_photos = existing_photos + saved_names

    photo_url_str = ",".join(all_photos)

    specs_payload = parse_specs_payload(
        request.form.get('specs'),
        {
            'brand': request.form.get('brand'),
            'country': request.form.get('country'),
            'weight': request.form.get('weight'),
            'model_number': request.form.get('model_number'),
            'color': request.form.get('color'),
            'storage': request.form.get('storage'),
            'ram': request.form.get('ram'),
            'warranty': request.form.get('warranty'),
        },
    )
    specs_json = json.dumps(specs_payload, ensure_ascii=False) if specs_payload else ""

    # 5. Обновляем товар в базе
    db.execute("""
        UPDATE products 
        SET name=?, synonyms=?, photo_url=?, brand=?, country=?, weight=?, model_number=?, is_on_request=?, folder_id=?,
            color=?, storage=?, ram=?, warranty=?, description=?, description_html=?, specs=?
        WHERE id=? AND user_id=?
    """, (
        request.form.get('name'), 
        synonyms_str, 
        photo_url_str, 
        request.form.get('brand'),
        request.form.get('country'),
        request.form.get('weight'),
        request.form.get('model_number'),
        1 if request.form.get('is_on_request') == '1' else 0,
        resolved_folder_id, 
        
        # 🔽 ДОБАВЛЯЕМ ПРИЕМ ПОЛЕЙ
        request.form.get('color'),
        request.form.get('storage'),
        request.form.get('ram'),
        request.form.get('warranty'),
        request.form.get('description'),
        request.form.get('description_html'),
        specs_json,

        prod_id, 
        session['user_id']
    ))
    db.commit()
    notify_clients()
    return jsonify({'success': True})





@app.route('/api/access_tree')
@login_required
def get_access_tree():
    user_id = session['user_id']
    cached_tree = ACCESS_TREE_CACHE.get(int(user_id))
    if cached_tree is not None:
        return jsonify(cached_tree)

    db = get_db()
    folders = [dict(f) for f in db.execute("SELECT id, name, parent_id FROM folders WHERE user_id=?", (user_id,)).fetchall()]
    
    # Гениальный запрос: вытаскиваем товары сразу с их подтвержденными поставщиками
    prods_raw = db.execute("""
        SELECT p.id as product_id, p.name as product_name, p.folder_id,
               m.chat_id, tc.custom_name, tc.chat_title
        FROM products p
        LEFT JOIN product_messages pm ON p.id = pm.product_id AND pm.status = 'confirmed'
        LEFT JOIN messages m ON pm.message_id = m.id
        LEFT JOIN tracked_chats tc ON m.chat_id = tc.chat_id AND m.user_id = tc.user_id
        WHERE p.user_id = ?
        GROUP BY p.id, m.chat_id
        ORDER BY CASE WHEN p.sort_index IS NULL THEN 1 ELSE 0 END, p.sort_index, p.id
    """, (user_id,)).fetchall()
    
    products = {}
    for row in prods_raw:
        pid = row['product_id']
        if pid not in products:
            products[pid] = {'id': pid, 'name': row['product_name'], 'folder_id': row['folder_id'], 'chats': []}
        if row['chat_id']:
            c_name = row['custom_name'] if row['custom_name'] else (row['chat_title'] or 'Неизвестный поставщик')
            if not any(c['chat_id'] == row['chat_id'] for c in products[pid]['chats']):
                products[pid]['chats'].append({'chat_id': row['chat_id'], 'name': c_name})

    payload = {'folders': folders, 'products': list(products.values())}
    ACCESS_TREE_CACHE[int(user_id)] = payload
    return jsonify(payload)



def fetch_google_sheet_as_df(sheet_url):
    """
    Превращает обычную ссылку на Google Таблицу в прямую ссылку на скачивание XLSX
    и сразу читает её через pandas.
    Пример: https://docs.google.com/spreadsheets/d/1ABC123.../edit#gid=0
    """
    try:
        if "/edit" in sheet_url:
            # Заменяем концовку ссылки на команду экспорта
            export_url = sheet_url.split('/edit')[0] + '/export?format=xlsx'
        else:
            export_url = sheet_url
            
        # Читаем напрямую из интернета в DataFrame
        df = pd.read_excel(export_url)
        return df
    except Exception as e:
        logger.error(f"Ошибка при загрузке Google Таблицы: {e}")
        return None

def spreadsheet_column_label(index):
    try:
        idx = int(index)
    except (TypeError, ValueError):
        return '?'
    if idx < 0:
        return '—'
    label = ''
    idx += 1
    while idx > 0:
        idx, remainder = divmod(idx - 1, 26)
        label = chr(65 + remainder) + label
    return label

def normalize_sheet_cell_value(value):
    if value is None:
        return ''
    return str(value).strip()

def infer_simple_sheet_config(all_rows, ws=None, max_scan_rows=250):
    visible_rows = []
    for row_idx, row in enumerate(all_rows[:max_scan_rows]):
        if ws is not None and (row_idx + 1) in ws.row_dimensions and ws.row_dimensions[row_idx + 1].hidden:
            continue
        normalized_row = [normalize_sheet_cell_value(cell) for cell in row]
        visible_rows.append((row_idx, normalized_row))

    pair_matches = {}
    for row_idx, row in visible_rows:
        price_cols = []
        text_cols = []
        for col_idx, cell in enumerate(row):
            if not cell or cell.lower() in ('none', 'nan'):
                continue
            price = extract_price(cell)
            if price is not None and price > 0:
                price_cols.append(col_idx)
                continue
            if len(cell) >= 2 and not re.fullmatch(r'[\W_]+', cell):
                text_cols.append(col_idx)

        if not price_cols or not text_cols:
            continue

        for price_col in price_cols:
            ordered_text_cols = sorted(
                [col for col in text_cols if col != price_col],
                key=lambda col: (
                    0 if col < price_col else 1,
                    abs(price_col - col),
                    col
                )
            )
            if not ordered_text_cols:
                continue
            name_col = ordered_text_cols[0]
            pair_matches.setdefault((name_col, price_col), []).append(row_idx)

    if not pair_matches:
        return None

    best_pair, matched_rows = max(
        pair_matches.items(),
        key=lambda item: (
            len(item[1]),
            -abs(item[0][1] - item[0][0]),
            -item[0][0]
        )
    )

    matched_rows = sorted(matched_rows)
    step_candidates = [
        matched_rows[i + 1] - matched_rows[i]
        for i in range(len(matched_rows) - 1)
        if matched_rows[i + 1] - matched_rows[i] > 0
    ]
    block_step = 1
    if step_candidates:
        counts = {}
        for step in step_candidates:
            counts[step] = counts.get(step, 0) + 1
        block_step = max(counts.items(), key=lambda item: (item[1], -item[0]))[0]

    return {
        'name_col': best_pair[0],
        'price_col': best_pair[1],
        'start_row': matched_rows[0],
        'name_row_offset': 0,
        'price_row_offset': 0,
        'block_step': block_step,
        'sku_col': -1,
        'is_grouped': 0,
        'confidence': len(matched_rows),
        'column_labels': {
            'name': spreadsheet_column_label(best_pair[0]),
            'price': spreadsheet_column_label(best_pair[1]),
        }
    }

def parse_google_sheet_rows(all_rows, ws, config):
    parsed_lines = []
    step = max(1, int(config.get('block_step', 1) or 1))
    start = max(0, int(config.get('start_row', 0) or 0))
    n_col = int(config.get('name_col', 0) or 0)
    p_col = int(config.get('price_col', 1) or 1)
    n_off = int(config.get('name_row_offset', 0) or 0)
    p_off = int(config.get('price_row_offset', 0) or 0)
    is_grouped = int(config.get('is_grouped', 0) or 0)
    sku_col = int(config.get('sku_col', -1) or -1)

    if is_grouped == 1:
        current_group = ""
        for i in range(start, len(all_rows)):
            if i + 1 in ws.row_dimensions and ws.row_dimensions[i + 1].hidden:
                continue

            row = all_rows[i]
            col_n_val = normalize_sheet_cell_value(row[n_col]) if n_col < len(row) else ""
            col_sku_val = normalize_sheet_cell_value(row[sku_col]) if sku_col != -1 and sku_col < len(row) else ""
            col_p_val = normalize_sheet_cell_value(row[p_col]) if p_col < len(row) else ""

            has_price = bool(col_p_val and col_p_val.lower() not in ('none', 'nan'))
            extracted = extract_price(col_p_val) if has_price else None
            if extracted is not None and extracted <= 0:
                continue

            if n_col == sku_col:
                if col_n_val and not has_price:
                    current_group = col_n_val
                elif col_n_val and has_price and current_group:
                    parsed_lines.append(f"{current_group} {col_n_val} - {col_p_val}")
            else:
                if col_n_val and col_n_val.lower() != 'none' and not has_price:
                    current_group = col_n_val
                if col_sku_val and has_price and current_group:
                    parsed_lines.append(f"{current_group} {col_sku_val} - {col_p_val}")
        return parsed_lines

    for i in range(start, len(all_rows), step):
        if i + 1 in ws.row_dimensions and ws.row_dimensions[i + 1].hidden:
            continue
        name_idx = i + n_off
        price_idx = i + p_off
        if name_idx >= len(all_rows) or price_idx >= len(all_rows):
            continue
        if (name_idx + 1 in ws.row_dimensions and ws.row_dimensions[name_idx + 1].hidden) or \
           (price_idx + 1 in ws.row_dimensions and ws.row_dimensions[price_idx + 1].hidden):
            continue

        row_n = all_rows[name_idx]
        row_p = all_rows[price_idx]
        n_val = normalize_sheet_cell_value(row_n[n_col]) if n_col < len(row_n) else ""
        p_val = normalize_sheet_cell_value(row_p[p_col]) if p_col < len(row_p) else ""

        if n_val and p_val and n_val.lower() != 'none' and p_val.lower() != 'none':
            extracted = extract_price(p_val)
            if extracted is not None and extracted > 0:
                parsed_lines.append(f"{n_val} - {p_val}")
    return parsed_lines

def parse_google_sheet_with_fallback(all_rows, ws, config):
    explicit_lines = parse_google_sheet_rows(all_rows, ws, config)
    if explicit_lines or int(config.get('is_grouped', 0) or 0) == 1:
        return explicit_lines, config, False

    inferred = infer_simple_sheet_config(all_rows, ws=ws)
    if not inferred:
        return explicit_lines, config, False

    inferred_config = dict(config)
    inferred_config.update({
        'name_col': inferred['name_col'],
        'price_col': inferred['price_col'],
        'start_row': inferred['start_row'],
        'name_row_offset': inferred.get('name_row_offset', 0),
        'price_row_offset': inferred.get('price_row_offset', 0),
        'block_step': inferred.get('block_step', 1),
        'sku_col': inferred.get('sku_col', -1),
        'is_grouped': inferred.get('is_grouped', 0),
    })
    inferred_lines = parse_google_sheet_rows(all_rows, ws, inferred_config)
    return inferred_lines, inferred_config, bool(inferred_lines)

@app.route('/api/excel/google_sheet/preview', methods=['POST'])
@login_required
def preview_google_sheet_route():
    data = request.json
    sheet_url = data.get('sheet_url')
    try:
        import requests
        from io import BytesIO
        from openpyxl import load_workbook

        if "/edit" in sheet_url:
            export_url = sheet_url.split('/edit')[0] + '/export?format=xlsx'
        else:
            export_url = sheet_url
            
        resp = requests.get(export_url, timeout=30)
        if resp.status_code != 200:
            return jsonify({'error': 'Не удалось скачать таблицу'}), 400

        # Загружаем книгу ПРАВИЛЬНО (не в режиме read_only)
        xls_file = BytesIO(resp.content)
        wb = load_workbook(xls_file, data_only=True) # data_only=True читает значения, а не формулы
        
        sheets_data = {}
        sheet_suggestions = {}
        # Оставляем только видимые листы
        visible_sheets = [s.title for s in wb.worksheets if s.sheet_state == 'visible']
        
        for sn in visible_sheets:
            ws = wb[sn]
            all_rows = list(ws.iter_rows(values_only=True))
            rows_list = []
            count = 0
            
            # Итерируемся по строкам напрямую через openpyxl
            for row in all_rows:
                # Проверяем, не скрыта ли строка
                row_idx = count + 1
                if row_idx in ws.row_dimensions and ws.row_dimensions[row_idx].hidden:
                    count += 1
                    continue
                
                # Очищаем данные от None
                clean_row = [str(cell) if cell is not None else "" for cell in row]
                rows_list.append(clean_row)
                count += 1
                
                # Ограничиваем превью 15 видимыми строками
                if len(rows_list) >= 15:
                    break
            
            sheets_data[sn] = rows_list
            inferred = infer_simple_sheet_config(all_rows, ws=ws)
            if inferred:
                sheet_suggestions[sn] = inferred
            
        return jsonify({'success': True, 'sheets_data': sheets_data, 'sheet_suggestions': sheet_suggestions})
    except Exception as e:
        logger.error(f"GS Preview Error: {e}")
        return jsonify({'error': str(e)}), 500



import pdfplumber

@app.route('/api/pdf/preview', methods=['POST'])
@login_required
def preview_pdf_route():
    if 'file' not in request.files:
        return jsonify({'error': 'Нет файла'}), 400
    
    file = request.files['file']
    try:
        with pdfplumber.open(file) as pdf:
            if not pdf.pages:
                return jsonify({'error': 'PDF пустой'}), 400
            
            # Извлекаем таблицу с первой страницы
            table = pdf.pages[0].extract_table()
            if not table:
                return jsonify({'error': 'Не удалось найти чёткую таблицу в PDF (возможно, нет линий сетки).'}), 400
            
            # Очищаем таблицу от пустых значений (None) и переносов строк
            cleaned_table = []
            for row in table:
                cleaned_table.append([str(cell).replace('\n', ' ') if cell is not None else "" for cell in row])
                
        # Возвращаем в формате "1 лист" (т.к. у PDF нет листов как в Excel)
        return jsonify({'success': True, 'sheets_data': {'Страница 1': cleaned_table[:30]}}) # берем первые 30 строк для превью
    except Exception as e:
        logger.error(f"PDF Preview Error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/pdf/save_config', methods=['POST'])
@login_required
def save_pdf_config():
    data = request.json
    db = get_db()
    try:
        # Сохраняем в ту же таблицу, но с меткой source_type = 'pdf'
        db.execute("""
            INSERT INTO excel_configs 
            (user_id, chat_id, sheet_name, name_col, name_row_offset, price_col, price_row_offset, block_step, start_row, is_grouped, sku_col, source_type, parse_interval) 
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'pdf', ?)
            ON CONFLICT(chat_id, sheet_name) DO UPDATE SET 
            name_col=excluded.name_col, price_col=excluded.price_col, 
            block_step=excluded.block_step, start_row=excluded.start_row,
            is_grouped=excluded.is_grouped, sku_col=excluded.sku_col, source_type='pdf'
        """, (
            session['user_id'], data['chat_id'], data['sheet_name'], 
            data['name_col'], 0, data['price_col'], data.get('price_row_offset', 0), 
            data.get('block_step', 1), data['start_row'], data.get('is_grouped', 0), 
            data.get('sku_col', -1), data.get('parse_interval', 60)
        ))
        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/excel/google_sheet/save_config', methods=['POST'])
@login_required
def save_gs_config():
    data = request.json
    chat_id = data.get('chat_id')
    sheet_url = (data.get('sheet_url') or '').strip()
    supplier_name = (data.get('supplier_name') or '').strip()
    if not sheet_url:
        return jsonify({'error': 'Ошибка: Укажите ссылку на Google Таблицу!'}), 400
    if not chat_id and not supplier_name:
        return jsonify({'error': 'Ошибка: Укажите название поставщика!'}), 400
    db = get_db()
    try:
        if not chat_id:
            chat_id, supplier_name = ensure_google_sheet_supplier_source(db, session['user_id'], supplier_name)
        parse_interval = normalize_interval_minutes(data.get('parse_interval'), 5)
        db.execute("""
            INSERT INTO excel_configs 
            (user_id, chat_id, sheet_name, name_col, name_row_offset, price_col, price_row_offset, block_step, start_row, is_grouped, sku_col, source_type, sheet_url, parse_interval) 
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'google_sheet', ?, ?)
            ON CONFLICT(chat_id, sheet_name) DO UPDATE SET 
            name_col=excluded.name_col, name_row_offset=excluded.name_row_offset, price_col=excluded.price_col, 
            price_row_offset=excluded.price_row_offset, block_step=excluded.block_step, start_row=excluded.start_row,
            is_grouped=excluded.is_grouped, sku_col=excluded.sku_col, source_type=excluded.source_type, 
            sheet_url=excluded.sheet_url, parse_interval=excluded.parse_interval
        """, (
            session['user_id'], chat_id, data['sheet_name'], 
            data['name_col'], 0, data['price_col'], data.get('price_row_offset', 0), 
            data.get('block_step', 1), data['start_row'], data.get('is_grouped', 0), 
            data.get('sku_col', -1), sheet_url, parse_interval
        ))
        
        # Удаляем из missing_sheets (убираем плашку "Не настроен лист")
        cid_str = str(chat_id)
        core_id = cid_str[4:] if cid_str.startswith('-100') else (cid_str[1:] if cid_str.startswith('-') else cid_str)
        possible_ids = [core_id, f"-{core_id}", f"-100{core_id}"]
        for pid in possible_ids:
            if data['sheet_name'] == '*':
                db.execute("DELETE FROM excel_missing_sheets WHERE chat_id = ? AND user_id = ?", (pid, session['user_id']))
            else:
                db.execute("DELETE FROM excel_missing_sheets WHERE chat_id = ? AND sheet_name = ? AND user_id = ?", (pid, data['sheet_name'], session['user_id']))
        
        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Ошибка сохранения конфига Google Sheet: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/excel/google_sheet/parse_now', methods=['POST'])
@login_required
def parse_google_sheet_now():
    data = request.get_json() or {}
    sheet_url = (data.get('sheet_url') or '').strip()
    if not sheet_url:
        return jsonify({'error': 'Не передана ссылка на Google Таблицу'}), 400

    user_id = session['user_id']
    db = get_db()
    conf_list = db.execute(
        """
        SELECT *
        FROM excel_configs
        WHERE user_id = ?
          AND LOWER(TRIM(COALESCE(source_type, ''))) = 'google_sheet'
          AND sheet_url = ?
        ORDER BY id
        """,
        (user_id, sheet_url),
    ).fetchall()
    if not conf_list:
        logger.warning("Google Sheet parse_now: configs not found for user %s and url %s", user_id, sheet_url)
        return jsonify({'error': 'Для этой Google Таблицы нет сохранённых настроек'}), 404

    try:
        stats = process_google_sheet_configs(db, sheet_url, conf_list, ignore_schedule=True)
        total_lines = sum(int(item.get('lines_count') or 0) for item in stats)
        logger.info("Google Sheet parse_now finished: %s", stats)
        return jsonify({
            'success': True,
            'stats': stats,
            'messages_count': len(stats),
            'lines_count': total_lines,
        })
    except Exception as e:
        logger.exception("Google Sheet parse_now failed for %s", sheet_url)
        return jsonify({'error': str(e)}), 500

def process_google_sheet_configs(db, url, conf_list, *, ignore_schedule=False):
    from io import BytesIO
    from openpyxl import load_workbook

    export_url = url.split('/edit')[0] + '/export?format=xlsx' if "/edit" in url else url
    resp = requests.get(export_url, timeout=30)
    if resp.status_code != 200:
        raise RuntimeError(f"Не удалось скачать таблицу: {resp.status_code}")

    xls_file = BytesIO(resp.content)
    wb = load_workbook(xls_file, data_only=True)
    visible_sheets = [s.title for s in wb.worksheets if s.sheet_state == 'visible']
    parse_stats = []
    pending_auto_confirm = []

    for config in conf_list:
        c_dict = dict(config)
        sheet_target = c_dict['sheet_name']
        user_id = c_dict['user_id']
        chat_id = c_dict['chat_id']

        if sheet_target != '*' and sheet_target not in visible_sheets:
            continue

        sheets_to_parse = visible_sheets if sheet_target == '*' else [sheet_target]
        parsed_lines = []
        fallback_used = False

        for sn in sheets_to_parse:
            if sn not in wb.sheetnames:
                continue
            ws = wb[sn]
            all_rows = list(ws.iter_rows(values_only=True))
            sheet_lines, effective_config, used_fallback = parse_google_sheet_with_fallback(all_rows, ws, c_dict)
            if used_fallback:
                fallback_used = True
                logger.info(
                    "Google Sheet %s / %s: fallback columns %s -> %s",
                    url,
                    sn,
                    spreadsheet_column_label(effective_config.get('name_col')),
                    spreadsheet_column_label(effective_config.get('price_col')),
                )
            parsed_lines.extend(sheet_lines)

        if not parsed_lines:
            continue

        msg_type = f"excel_{c_dict['id']}"
        msg_text = normalize_incoming_text("📊 [GOOGLE ТАБЛИЦА]:\n" + "\n".join(parsed_lines))

        def persist_sheet_snapshot():
            tc = db.execute(
                "SELECT custom_name, chat_title FROM tracked_chats WHERE chat_id = ? AND user_id = ?",
                (chat_id, user_id)
            ).fetchone()
            chat_title = tc['custom_name'] if tc and tc['custom_name'] else (tc['chat_title'] if tc else str(chat_id))
            now_str = (datetime.now(timezone.utc).replace(tzinfo=None) + timedelta(hours=3)).strftime("%Y-%m-%d %H:%M:%S")

            old_bindings = db.execute("""
                SELECT pm.id, pm.product_id, pm.line_index, p.name, p.country, p.synonyms, p.brand, p.model_number, p.color, p.storage
                FROM product_messages pm
                JOIN products p ON pm.product_id = p.id
                JOIN messages m ON pm.message_id = m.id
                WHERE m.chat_id = ? AND m.user_id = ? AND m.type = ?
                  AND COALESCE(pm.is_actual, 1) = 1
            """, (chat_id, user_id, msg_type)).fetchall()

            cleanup_source_type_before_new_snapshot(db, user_id, chat_id, msg_type)
            cursor = db.execute(
                "INSERT INTO messages (user_id, type, text, date, chat_id, chat_title, sender_name) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (user_id, msg_type, msg_text, now_str, chat_id, chat_title, f"Лист: {sheet_target}")
            )
            new_msg_id = cursor.lastrowid

            if old_bindings:
                lines = msg_text.split('\n')
                for b in old_bindings:
                    match_found = False
                    new_price = None
                    new_line_idx = 0
                    for i, line in enumerate(lines):
                        if (
                            binding_product_row_matches_line(b, line, require_known_category=True)
                        ):
                            match_found = True
                            new_price = extract_price(line)
                            new_line_idx = i
                            break
                    if match_found and new_price and new_price > 0:
                        occupied = db.execute(
                            """
                            SELECT id FROM product_messages
                            WHERE message_id = ?
                              AND line_index = ?
                              AND product_id <> ?
                              AND status = 'confirmed'
                              AND COALESCE(is_actual, 1) = 1
                            LIMIT 1
                            """,
                            (new_msg_id, new_line_idx, b['product_id']),
                        ).fetchone()
                        if occupied:
                            db.execute("UPDATE product_messages SET is_actual = 0 WHERE id = ?", (b['id'],))
                            continue
                        db.execute(
                            """
                            DELETE FROM product_messages
                            WHERE product_id = ? AND message_id = ? AND line_index = ? AND id != ?
                            """,
                            (b['product_id'], new_msg_id, new_line_idx, b['id']),
                        )
                        db.execute(
                            "UPDATE product_messages SET message_id = ?, line_index = ?, extracted_price = ?, is_actual = 1 WHERE id = ?",
                            (new_msg_id, new_line_idx, new_price, b['id'])
                        )
                    else:
                        db.execute("UPDATE product_messages SET is_actual = 0 WHERE id = ?", (b['id'],))

            db.execute("""
                DELETE FROM messages
                WHERE chat_id = ? AND user_id = ? AND id != ? AND type = ?
                  AND id NOT IN (SELECT message_id FROM product_messages)
            """, (chat_id, user_id, new_msg_id, msg_type))
            return {
                'config_id': c_dict['id'],
                'message_id': new_msg_id,
                'sheet_name': sheet_target,
                'lines_count': len(parsed_lines),
                'fallback_used': fallback_used,
                'auto_confirm_queued': True,
                'auto_confirm_inserted': 0,
                'auto_confirm_updated': 0,
                'auto_confirm_ambiguous': 0,
                'auto_confirm_unmatched': 0,
                'synonyms_added': 0,
            }

        try:
            result = run_db_write_transaction(
                db,
                persist_sheet_snapshot,
                retries=6,
                retry_delay=0.75,
                use_lock=True,
            )
            if result:
                parse_stats.append(result)
                pending_auto_confirm.append((user_id, result.get('message_id'), sheet_target))
                logger.info(
                    "Google Sheet %s / %s: saved %s lines as message %s",
                    url,
                    sheet_target,
                    result.get('lines_count'),
                    result.get('message_id'),
                )
        except Exception:
            logger.exception("Google Sheet %s / %s: failed to persist parsed snapshot", url, sheet_target)
            continue

    for user_id, message_id, sheet_target in pending_auto_confirm:
        enqueue_auto_confirm_source_message(
            user_id,
            message_id,
            f"Google Sheet {sheet_target}",
        )

    if parse_stats:
        notify_clients()
    return parse_stats


def build_google_sheet_source_chat_id(user_id, supplier_name):
    normalized = re.sub(r'\s+', ' ', str(supplier_name or '').strip()).lower()
    digest = hashlib.sha1(f"{user_id}:{normalized}".encode('utf-8')).hexdigest()[:16]
    return f"gs_supplier_{digest}"


def ensure_google_sheet_supplier_source(db, user_id, supplier_name):
    display_name = re.sub(r'\s+', ' ', str(supplier_name or '').strip())
    if not display_name:
        raise ValueError("Название поставщика обязательно")

    synthetic_chat_id = build_google_sheet_source_chat_id(user_id, display_name)
    existing = db.execute(
        "SELECT id FROM tracked_chats WHERE user_id = ? AND chat_id = ?",
        (user_id, synthetic_chat_id),
    ).fetchone()
    if existing:
        db.execute(
            "UPDATE tracked_chats SET custom_name = ?, chat_title = COALESCE(chat_title, ?) WHERE id = ?",
            (display_name, display_name, existing['id']),
        )
    else:
        db.execute(
            """
            INSERT INTO tracked_chats (user_id, chat_id, chat_title, custom_name)
            VALUES (?, ?, ?, ?)
            """,
            (user_id, synthetic_chat_id, display_name, display_name),
        )
    return synthetic_chat_id, display_name


def format_price_list(template, products):
    """Вставляет товары в шаблон между маркерами START_PRICE_LIST и END_PRICE_LIST."""
    lines = []
    for p in products:
        lines.append(f"• {p['name']} — {p['price']} руб.")
    price_list = "\n".join(lines)
    
    if "<!-- START_PRICE_LIST -->" in template and "<!-- END_PRICE_LIST -->" in template:
        # Заменяем только часть между маркерами
        parts = template.split("<!-- START_PRICE_LIST -->")
        before = parts[0]
        after = parts[1].split("<!-- END_PRICE_LIST -->")[1]
        return before + "<!-- START_PRICE_LIST -->\n" + price_list + "\n<!-- END_PRICE_LIST -->" + after
    else:
        # Иначе заменяем всё сообщение
        return price_list

# Словарь в памяти для отслеживания времени последней публикации
# Словарь в памяти для отслеживания времени последней публикации
last_publish_times = {}

def publish_scheduler():
    """Фоновый поток для публикации каталогов в Telegram."""
    time.sleep(10)
    with app.app_context():
        while True:
            try:
                db = get_db()
                # Берем все активные публикации
                pubs = db.execute("SELECT * FROM publications WHERE is_active = 1").fetchall()
                from datetime import timedelta
                now = datetime.now(timezone.utc).replace(tzinfo=None) + timedelta(hours=3)
                
                for pub in pubs:
                    pub_id = pub['id']
                    interval_minutes = pub['interval_min'] or 60
                    
                    # 1. ПРОВЕРКА ИНТЕРВАЛА
                    last_run = last_publish_times.get(pub_id)
                    if last_run:
                        diff_minutes = (now - last_run).total_seconds() / 60
                        if diff_minutes < interval_minutes:
                            continue  # Время публикации еще не пришло
                    
                    # 2. ПОЛУЧЕНИЕ ЦЕН И ТОВАРОВ
                    userbot_id = pub['userbot_id']
                    if userbot_id not in user_clients: continue
                    
                    # Извлекаем user_id из юзербота
                    _, tg_client, user_id = user_clients[userbot_id] 
                    
                    allowed_items = {}
                    try: allowed_items = json.loads(pub['allowed_items']) 
                    except: pass
                    
                    if not allowed_items: continue
                    
                    # ---- НОВОЕ: ДОСТАЕМ НАЦЕНКИ ----
                    markups_raw = db.execute("SELECT folder_id, markup_type, markup_value, rounding FROM pub_markups WHERE pub_id=?", (pub_id,)).fetchall()
                    markups = {}
                    default_markup = {'markup_type': 'percent', 'markup_value': 0, 'rounding': 100}
                    for m in markups_raw:
                        if m['folder_id'] == 0: default_markup = dict(m)
                        else: markups[m['folder_id']] = dict(m)
                    # --------------------------------
                    
                    final_products = []
                    
                    # Перебираем товары и разрешенных поставщиков из JSON
                    # Перебираем товары и разрешенных поставщиков из JSON
                    for p_id_str, allowed_suppliers in allowed_items.items():
                        if not allowed_suppliers: 
                            continue
                        
                        # --- НОВАЯ ЛОГИКА ДЛЯ "ЦЕНА ПО ЗАПРОСУ" ---
                        if '0' in allowed_suppliers or 0 in allowed_suppliers:
                            # У товара нет цен, достаем из базы только его название
                            p_info = db.execute("SELECT name FROM products WHERE id = ?", (int(p_id_str),)).fetchone()
                            if p_info:
                                final_products.append({
                                    'name': p_info['name'], 
                                    'price': 'Цена по запросу'
                                })
                            continue  # Пропускаем запросы цен, идем к следующему товару
                        # -----------------------------------------

                        placeholders = ','.join(['?'] * len(allowed_suppliers))
                        query_params = [int(p_id_str)] + allowed_suppliers
                        
                        # ВАЖНО: Добавлено p.folder_id в SELECT
                        price_row = db.execute(f"""
                            SELECT pm.extracted_price, p.name, p.folder_id 
                            FROM product_messages pm 
                            JOIN messages m ON pm.message_id = m.id 
                            JOIN products p ON p.id = pm.product_id
                            WHERE pm.product_id = ? 
                              AND pm.status = 'confirmed' 
                              AND pm.is_actual = 1 
                              AND m.chat_id IN ({placeholders})
                            ORDER BY pm.extracted_price ASC, m.date DESC LIMIT 1
                        """, query_params).fetchone()
                        
                        if price_row and price_row['extracted_price'] is not None:
                            # -------- ПРИМЕНЯЕМ НАЦЕНКУ ПО ПАПКАМ ---------
                            base_price = float(price_row['extracted_price'])
                            f_id = price_row['folder_id'] or 0
                            mk = markups.get(f_id, default_markup)
                            
                            if mk['markup_type'] == 'percent':
                                final_price = base_price * (1 + mk['markup_value'] / 100)
                            else:
                                final_price = base_price + mk['markup_value']
                                
                            rnd = mk['rounding']
                            if rnd > 0:
                                final_price = math.ceil(final_price / rnd) * rnd

                            final_products.append({
                                'name': price_row['name'], 
                                'price': int(final_price)
                            })
                    
                    if not final_products: continue

                    template = pub['template'] or "🔥 Актуальные цены:\n\n{prices}\n\nДля заказа пишите менеджеру!"
                    
                    # 1. Красиво собираем товары в текстовый список
                    prices_lines = []
                    for p in final_products:
                        if p['price'] == 'Цена по запросу':
                            prices_lines.append(f"▪️ {p['name']} — {p['price']}")
                        else:
                            prices_lines.append(f"▪️ {p['name']} — {p['price']} ₽")
                            
                    prices_text = "\n".join(prices_lines)
                    
                    # 2. Подставляем список вместо метки {prices}
                    if "{prices}" in template:
                        message_text = template.replace("{prices}", prices_text)
                    else:
                        # Если вы случайно забыли написать {prices} в шаблоне, 
                        # бот просто аккуратно приклеит прайс-лист в самый низ сообщения
                        message_text = f"{template}\n\n{prices_text}"

                    loop = user_loops.get(userbot_id)
                    if not loop or not loop.is_running(): continue

                    chat_id = pub['chat_id']
                    message_id = pub['message_id']
                    
                    # 3. ОТПРАВКА В ТЕЛЕГРАМ
                    # 3. ОТПРАВКА В ТЕЛЕГРАМ
                    # 3. ОТПРАВКА В ТЕЛЕГРАМ
                    async def update_or_create(cid, mid, txt, p_id):
                        try:
                            cid = int(cid)
                        except ValueError:
                            pass

                        # --- ЛОГИКА РАЗДЕЛЕНИЯ ТЕКСТА ---
                        parts = []
                        max_len = 4000 # Оставляем небольшой запас до 4096
                        temp_txt = txt
                        
                        while len(temp_txt) > max_len:
                            # Ищем последний перенос строки в пределах лимита, чтобы не рвать строку пополам
                            split_idx = temp_txt.rfind('\n', 0, max_len)
                            if split_idx == -1:
                                split_idx = max_len # Если нет переносов, режем жестко
                                
                            parts.append(temp_txt[:split_idx])
                            temp_txt = temp_txt[split_idx:].strip()
                            
                        if temp_txt:
                            parts.append(temp_txt)
                        # --------------------------------

                        if mid:
                            try:
                                # ВАЖНЫЙ НЮАНС ПРИ РЕДАКТИРОВАНИИ: 
                                # Telegram не позволяет превратить 1 старое сообщение в 2 новых.
                                # Поэтому при редактировании обновится только первая часть, а остаток обрежется.
                                await tg_client.edit_message(cid, int(mid), parts[0])
                                if len(parts) > 1:
                                    logger.warning(f"Прайс слишком большой для обновления одного сообщения (ID {mid}). Хвост обрезан.")
                            except Exception as e:
                                logger.error(f"Ошибка редактирования: {e}")
                        else:
                            try:
                                # В режиме "отправки нового" просто шлем все куски друг за другом
                                first_msg = None
                                for part in parts:
                                    msg = await tg_client.send_message(cid, part)
                                    if not first_msg:
                                        first_msg = msg # Запоминаем только самое первое сообщение

                                # Сохраняем в базу ID первого сообщения (если захотите его потом удалять/редактировать)
                                if first_msg:
                                    with app.app_context():
                                        local_db = get_db()
                                        local_db.execute("UPDATE publications SET message_id = ? WHERE id = ?", (str(first_msg.id), p_id))
                                        local_db.commit()
                            except Exception as e:
                                logger.error(f"Ошибка отправки разделенного сообщения: {e}")
                                
                    try:
                        asyncio.run_coroutine_threadsafe(update_or_create(chat_id, message_id, message_text, pub_id), loop)
                        last_publish_times[pub_id] = now
                    except Exception as e:
                        logger.error(f"Ошибка передачи в цикл публикации: {e}")

            except Exception as e:
                logging.error(f"Ошибка в планировщике публикации: {e}")
            time.sleep(60)


def get_catalog_for_client(client_id):
    """Возвращает каталог для указанного клиента (как в /api/v1/catalog)."""
    db = get_db()
    client = db.execute("SELECT * FROM api_clients WHERE id=?", (client_id,)).fetchone()
    if not client:
        return None

    user_id = client['user_id']
    
    if client['schedule_enabled']:
        from datetime import timedelta
        now = (datetime.now(timezone.utc).replace(tzinfo=None) + timedelta(hours=3)).time()
        start_time = datetime.strptime(client['time_start'], '%H:%M').time()
        end_time = datetime.strptime(client['time_end'], '%H:%M').time()
        
        is_active = start_time <= now <= end_time if start_time <= end_time else (now >= start_time or now <= end_time)
        if not is_active: return {'categories': [], 'products': []}
        
    user_id = client['user_id']
    client_id = client['id']
    
    allowed_folders = client['allowed_folders'] if 'allowed_folders' in client.keys() else 'all'
    allowed_chats = client['allowed_chats'] if 'allowed_chats' in client.keys() else 'all'

    folders_list = allowed_folders.split(',') if allowed_folders not in ('all', '', None) else allowed_folders
    chats_list = allowed_chats.split(',') if allowed_chats not in ('all', '', None) else allowed_chats

    markups_raw = db.execute("SELECT folder_id, markup_type, markup_value, rounding FROM api_markups WHERE client_id=?", (client_id,)).fetchall()
    markups = {}
    default_markup = {'markup_type': 'percent', 'markup_value': 0, 'rounding': 100}
    for m in markups_raw:
        if m['folder_id'] == 0: default_markup = dict(m)
        else: markups[m['folder_id']] = dict(m)
            
    products_raw = db.execute("""
        SELECT p.id, p.name, p.folder_id, p.price as manual_price, p.photo_url, p.brand, p.country, p.weight, p.model_number, p.is_on_request,
               p.color, p.storage, p.ram, p.warranty, p.description, p.description_html, p.specs,
               (SELECT pm.extracted_price FROM product_messages pm JOIN messages m ON pm.message_id = m.id WHERE pm.product_id = p.id AND pm.status = 'confirmed' AND pm.is_actual = 1 AND pm.extracted_price IS NOT NULL ORDER BY pm.extracted_price ASC, m.date DESC LIMIT 1) as parsed_price,
               (SELECT pm.is_actual FROM product_messages pm JOIN messages m ON pm.message_id = m.id WHERE pm.product_id = p.id AND pm.status = 'confirmed' AND pm.is_actual = 1 AND pm.extracted_price IS NOT NULL ORDER BY pm.extracted_price ASC, m.date DESC LIMIT 1) as is_actual,
               (SELECT m.chat_id FROM product_messages pm JOIN messages m ON pm.message_id = m.id WHERE pm.product_id = p.id AND pm.status = 'confirmed' AND pm.is_actual = 1 AND pm.extracted_price IS NOT NULL ORDER BY pm.extracted_price ASC, m.date DESC LIMIT 1) as latest_chat_id
        FROM products p
        WHERE p.user_id = ?
    """, (user_id,)).fetchall()
    
    products = []
    active_folder_ids = set()
    
    try:
        access_rules = json.loads(client['access_rules']) if 'access_rules' in client.keys() and client['access_rules'] else {}
    except:
        access_rules = {}
        
    for p in products_raw:
        if not access_rules: continue
        
        pid_str = str(p['id'])
        if pid_str not in access_rules: continue
        
        allowed_chats = access_rules[pid_str]
        chat_id_str = str(p['latest_chat_id'])
        
        if chat_id_str != 'None':
            if allowed_chats != ['all'] and chat_id_str not in allowed_chats: 
                continue
        
        base_price = float(p['parsed_price'] if (p['parsed_price'] is not None and p['is_actual'] == 1) else (p['manual_price'] or 0))
        
        if base_price > 0 and not p['is_on_request']:
            mk = markups.get(p['folder_id'], default_markup)
            if mk['markup_type'] == 'percent':
                final_price = base_price * (1 + mk['markup_value'] / 100)
            else:
                final_price = base_price + mk['markup_value']
                
            rnd = mk['rounding']
            if rnd > 0:
                final_price = math.ceil(final_price / rnd) * rnd
            final_price = int(final_price)
        else:
            final_price = "По запросу"
            
        photo_list = []
        if p['photo_url']:
            for url in p['photo_url'].split(','):
                url = url.strip()
                if url:
                    if url.startswith('http'):
                        photo_list.append(url)
                    else:
                        photo_list.append(f"https://engine.astoredirect.ru/uploads/{url}")

        # Единый, правильный парсинг specs
        specs_data = parse_specs_payload(p['specs'], p)

        products.append({
            'id': p['id'],
            'name': p['name'],
            'category_id': p['folder_id'],
            'price': final_price,
            'price2': "По запросу" ,
            'photo_url': photo_list, 
            'brand': p['brand'],
            'country': p['country'],
            'weight': p['weight'],
            'model_number': p['model_number'],
            'is_on_request': bool(p['is_on_request']),
            
            # Все системные поля и объект характеристик
            'color': p['color'],
            'storage': p['storage'],
            'ram': p['ram'],
            'warranty': p['warranty'],
            'description': p['description'],
            'description_html': p['description_html'],
            'specs': specs_data
        })
        
        if p['folder_id']:
            active_folder_ids.add(p['folder_id'])
            
    folders_raw = db.execute("SELECT id, name, parent_id FROM folders WHERE user_id=?", (user_id,)).fetchall()
    
    parent_map = {f['id']: f['parent_id'] for f in folders_raw}
    folders_to_keep = set(active_folder_ids)
    
    for fid in active_folder_ids:
        current = fid
        while current in parent_map and parent_map[current] is not None:
            current = parent_map[current]
            folders_to_keep.add(current)
            
    categories = [{'id': f['id'], 'name': f['name'], 'parent_id': f['parent_id']} 
                  for f in folders_raw if f['id'] in folders_to_keep]
        
    return {
        'categories': categories,
        'products': products
    }

def google_sheets_scheduler():
    """Фоновый поток для парсинга Гугл Таблиц по расписанию."""
    time.sleep(5)
    last_parsed = {}
    
    with app.app_context():
        while True:
            try:
                db = get_db()
                configs = db.execute(
                    "SELECT * FROM excel_configs WHERE LOWER(TRIM(COALESCE(source_type, ''))) = 'google_sheet'"
                ).fetchall()
                current_time = time.time()
                
                url_configs = {}
                for c in configs:
                    url = c['sheet_url']
                    if not url:
                        continue
                    if url not in url_configs:
                        url_configs[url] = []
                    url_configs[url].append(c)

                for url, conf_list in url_configs.items():
                    first_conf = conf_list[0]
                    interval_sec = int(first_conf['parse_interval'] or 60) * 60
                    last_time = last_parsed.get(url)
                    if last_time and (current_time - last_time) < interval_sec:
                        continue
                    logger.info(f"Парсинг Гугл Таблицы {url} по расписанию...")
                    try:
                        process_google_sheet_configs(db, url, conf_list, ignore_schedule=False)
                        last_parsed[url] = current_time
                    except Exception:
                        logger.exception("Ошибка планового парсинга Google Таблицы %s", url)

            except Exception as e:
                logger.error(f"Глобальная ошибка планировщика GS: {e}", exc_info=True)
                
            time.sleep(60) # Проверка раз в минуту

# Запускаем поток вместе с сервером. Тесты и CLI-аудиты импортируют модуль без фоновых задач.
if os.environ.get("ASTORE_DISABLE_BACKGROUND_THREADS") != "1" and GOOGLE_SHEETS_SCHEDULER_ENABLED:
    threading.Thread(target=google_sheets_scheduler, daemon=True).start()



@app.route('/api/messages/<int:msg_id>/block', methods=['POST'])
@login_required
def block_message(msg_id):
    db = get_db()
    # Устанавливаем флаг блокировки для сообщения
    db.execute("UPDATE messages SET is_blocked = 1 WHERE id = ? AND user_id = ?", (msg_id, session['user_id']))
    db.commit()
    notify_clients()
    return jsonify({'success': True})


@app.route('/api/product_messages/merge', methods=['POST'])
@login_required
def merge_product_messages():
    data = request.get_json()
    pm_ids = data.get('pm_ids', [])
    if len(pm_ids) < 2:
        return jsonify({'error': 'Выберите минимум 2 сообщения'}), 400
    
    db = get_db()
    # Берем минимальный ID как главный номер группы
    group_id = min(map(int, pm_ids))
    placeholders = ','.join('?' for _ in pm_ids)
    
    db.execute(f"UPDATE product_messages SET group_id = ? WHERE id IN ({placeholders})", [group_id] + pm_ids)
    db.commit()
    notify_clients()
    return jsonify({'success': True})


@app.route('/api/products/<int:source_id>/merge', methods=['POST'])
@login_required
def merge_product(source_id):
    target_id = request.get_json()['target_id']
    user_id = session['user_id']
    db = get_db()
    
    # Получаем синонимы обоих товаров
    src = db.execute("SELECT synonyms FROM products WHERE id=? AND user_id=?", (source_id, user_id)).fetchone()
    tgt = db.execute("SELECT synonyms FROM products WHERE id=? AND user_id=?", (target_id, user_id)).fetchone()
    
    if src and tgt:
        src_syns = [s.strip() for s in src['synonyms'].split(',') if s.strip()]
        tgt_syns = [s.strip() for s in tgt['synonyms'].split(',') if s.strip()]
        combined = list(set(src_syns + tgt_syns)) # Объединяем без дубликатов
        
        # Обновляем синонимы целевого товара
        db.execute("UPDATE products SET synonyms=? WHERE id=?", (', '.join(combined), target_id))
        
        # Переносим привязанные сообщения (игнорируем ошибку, если такое сообщение уже привязано)
        try:
            db.execute("UPDATE OR IGNORE product_messages SET product_id=? WHERE product_id=?", (target_id, source_id))
        except:
            pass
            
        # Удаляем старый товар
        db.execute("DELETE FROM products WHERE id=?", (source_id,))
        db.commit()
        notify_clients()
    return jsonify({'success': True})

import re
@app.route('/api/profile/change_password', methods=['POST'])
@login_required
def change_password():
    user_id = session['user_id']
    data = request.get_json()
    
    old_password = data.get('old_password')
    new_password = data.get('new_password')

    if not old_password or not new_password:
        return jsonify({'error': 'Заполните все поля'}), 400

    db = get_db()
    
    # Хэшируем введённый старый пароль и проверяем его
    old_pwd_hash = hashlib.sha256(old_password.encode()).hexdigest()
    user = db.execute("SELECT id FROM users WHERE id = ? AND password_hash = ?", (user_id, old_pwd_hash)).fetchone()

    if not user:
        return jsonify({'error': 'Неверный старый пароль'}), 400

    # Если всё верно, хэшируем новый пароль и записываем в БД 
    new_pwd_hash = hashlib.sha256(new_password.encode()).hexdigest()
    db.execute("UPDATE users SET password_hash = ? WHERE id = ?", (new_pwd_hash, user_id))
    db.commit()

    return jsonify({'success': True})


async def fetch_chat_history(client, target_entity, chat_title, user_id, db_chat_id=None):
    """Парсинг старых сообщений с подробной статистикой"""
    if getattr(target_entity, 'bot', False):
        logger.info(f"История бота {chat_title} не импортируется: принимаем только свежие ответы на команды.")
        return
    if db_chat_id is None:
        db_chat_id = getattr(target_entity, 'id', target_entity)

    try:
        with app.app_context():
            db = get_db()
            added = 0
            duplicates = 0
            no_text = 0
            errors = 0
            
            async for message in client.iter_messages(target_entity): 
                # Пропускаем пустые (например, просто фото без подписи)
                if (not message.text and not message.document):
                    no_text += 1
                    continue
                
                try:
                    sender = await message.get_sender()
                    sender_name = f"{getattr(sender, 'first_name', '')} {getattr(sender, 'last_name', '')}".strip() or "Unknown"
                    msg_type = 'channel' if message.is_channel else 'group' if message.is_group else 'private'
                    
                    parsed_text = await parse_excel_message(client, message, db_chat_id, user_id)
                    if not parsed_text:
                        no_text += 1
                        continue

                    # ИСПРАВЛЕНИЕ: Форматируем дату вручную, чтобы не было DeprecationWarning
                    from datetime import timedelta
                    msg_date = (message.date + timedelta(hours=3)).strftime("%Y-%m-%d %H:%M:%S")

                    db.execute(
                        'INSERT INTO messages (user_id, telegram_message_id, type, text, date, chat_id, chat_title, sender_name) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
                        (user_id, message.id, msg_type, parsed_text, msg_date, db_chat_id, chat_title, sender_name)
                    )
                    db.commit()
                    added += 1
                    
                    if added % 50 == 0:
                        
                        notify_clients()
                        
                except sqlite3.IntegrityError:
                    duplicates += 1  # Сообщение уже есть в базе
                except Exception as msg_e:
                    errors += 1
                    logger.warning(f"Ошибка в сообщении {message.id}: {msg_e}")
                    
            db.commit()
            notify_clients()
            
            # ВЫВОДИМ ПОДРОБНЫЙ ОТЧЕТ В КОНСОЛЬ
            logger.info(f"Парсинг [{chat_title}]: Добавлено {added} | Уже были в базе: {duplicates} | Без текста/фото: {no_text} | Ошибок: {errors}")
            
    except Exception as e:
        logger.error(f"Глобальная ошибка парсинга истории для {chat_title}: {e}")

@app.route('/api/api_clients/<int:client_id>/publish', methods=['POST'])
@login_required
def update_publish_settings(client_id):
    data = request.json
    db = get_db()
    # Проверяем, что клиент принадлежит текущему пользователю
    client = db.execute("SELECT id FROM api_clients WHERE id=? AND user_id=?", 
                        (client_id, session['user_id'])).fetchone()
    if not client:
        return jsonify({'error': 'Клиент не найден'}), 404

    db.execute("""
        UPDATE api_clients 
        SET publish_enabled=?, publish_chat_id=?, userbot_id=?, publish_template=?, folders=?, publish_interval=?
        WHERE id=?
    """, (data.get('publish_enabled', 0), data.get('publish_chat_id'), data.get('userbot_id'),
          data.get('publish_template'), data.get('folders'), data.get('publish_interval', 60), client_id))
    db.commit()
    return jsonify({'success': True})

# ---------- Декораторы авторизации ----------
def login_required(f):
    @functools.wraps(f)
    def wrapped(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login_page'))
        user = get_session_user_record()
        if has_invalid_session_token(user):
            return build_expired_session_response()
        return f(*args, **kwargs)
    return wrapped

def admin_required(f):
    @functools.wraps(f)
    def wrapped(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login_page'))
        row = get_session_user_record()
        if has_invalid_session_token(row):
            return build_expired_session_response()
        if not row or row['role'] != 'admin':
            return "Access denied", 403
        return f(*args, **kwargs)
    return wrapped

# ---------- Фоновый слушатель сообщений ----------
def stop_message_listener(session_id):
    client_info = user_clients.get(session_id)
    if client_info:
        _, client, _ = client_info
        loop = user_loops.get(session_id)
        if loop and loop.is_running():
            async def safe_disconnect():
                try:
                    await client.disconnect()
                except:
                    pass
            try:
                asyncio.run_coroutine_threadsafe(safe_disconnect(), loop)
            except:
                pass

    # ВАЖНО: Мы БОЛЬШЕ НЕ удаляем user_clients.pop() здесь!
    # Это сделает сам фоновый поток перед своей смертью.

    try:
        with app.app_context():
            db = get_db()
            db.execute("UPDATE user_telegram_sessions SET status = 'inactive' WHERE id = ?", (session_id,))
            db.commit()
            if 'notify_clients' in globals(): notify_clients()
    except Exception as e:
        logging.error(f"Ошибка БД при остановке сессии {session_id}: {e}")

@app.route('/logo.svg')
def serve_logo():
    # Отдаем логотип из директории приложения, а не из текущего cwd процесса
    response = send_from_directory(BASE_DIR, 'logo.svg', max_age=31536000)
    response.cache_control.public = True
    response.cache_control.max_age = 31536000
    return response

# Было: @app.route('/api/excel/configs/chat/<int:chat_id>', methods=['DELETE'])
@app.route('/api/excel/configs/chat/<chat_id>', methods=['DELETE'])
def delete_excel_chat_configs(chat_id):
    try:
        chat_id = int(chat_id)
        user_id = session.get('user_id')
        db = get_db()
        
        # 1. Удаляем привязки к товарам
        db.execute("""
            DELETE FROM product_messages 
            WHERE message_id IN (SELECT id FROM messages WHERE chat_id = ? AND user_id = ? AND type LIKE 'excel%')
        """, (chat_id, user_id))
        
        # 2. Удаляем сами сообщения (прайс-листы)
        db.execute("DELETE FROM messages WHERE chat_id = ? AND user_id = ? AND type LIKE 'excel%'", (chat_id, user_id))
        
        # 3. Удаляем конфигурацию парсера
        db.execute("DELETE FROM excel_configs WHERE chat_id = ? AND user_id = ?", (chat_id, user_id))
        
        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        print(f"Ошибка при удалении файла: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/publications/<int:pub_id>/markups', methods=['GET', 'POST'])
@login_required
def manage_pub_markups(pub_id):
    db = get_db()
    if request.method == 'GET':
        markups = db.execute("""
            SELECT pm.*, f.name as folder_name 
            FROM pub_markups pm 
            LEFT JOIN folders f ON pm.folder_id = f.id 
            WHERE pm.pub_id=?
        """, (pub_id,)).fetchall()
        return jsonify([dict(m) for m in markups])
    
    data = request.json
    folder_id = data.get('folder_id') or 0
    db.execute("""
        INSERT INTO pub_markups (pub_id, folder_id, markup_type, markup_value, rounding) 
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(pub_id, folder_id) DO UPDATE SET 
        markup_type=excluded.markup_type, markup_value=excluded.markup_value, rounding=excluded.rounding
    """, (pub_id, folder_id, data['markup_type'], data['markup_value'], data['rounding']))
    db.commit()
    return jsonify({'success': True})

@app.route('/api/pub_markups/<int:markup_id>', methods=['DELETE'])
@login_required
def delete_pub_markup(markup_id):
    db = get_db()
    db.execute("DELETE FROM pub_markups WHERE id=? AND folder_id != 0", (markup_id,))
    db.commit()
    return jsonify({'success': True})


def start_message_listener(session_id, user_id, api_id, api_hash):
    with listener_locks.setdefault(session_id, threading.Lock()):
        if session_id in background_tasks:
            logging.info(f"Слушатель для сессии {session_id} уже запущен")
            return

        def listener():
            loop = asyncio.new_event_loop()
         
            asyncio.set_event_loop(loop)
            loop.set_exception_handler(custom_exception_handler)
            session_file = f'sessions/user_{user_id}_{session_id}'
            client = TelegramClient(session_file, int(api_id), api_hash)
            
            user_clients[session_id] = (threading.current_thread(), client, user_id)
            user_loops[session_id] = loop

           

   
            # --- НОВЫЙ БЛОК: СЛУШАЕМ УДАЛЕНИЯ ---
            @client.on(MessageDeleted)
            async def deleted_handler(event):
                if not is_parsing_allowed(session_id):
                    return
                with app.app_context():
                    db = get_db()
                    try:
                        with sqlite_write_lock:
                            for msg_id in event.deleted_ids:
                                # 1. Отвязываем товары
                                db.execute("""
                                    UPDATE product_messages 
                                    SET is_actual = 0 
                                    WHERE message_id IN (SELECT id FROM messages WHERE telegram_message_id = ?)
                                """, (msg_id,))
                                
                                # 2. Сообщение оставляем для истории карточки, но скрываем из парсера.
                                db.execute("""
                                    UPDATE messages
                                    SET is_blocked = 1
                                    WHERE telegram_message_id = ? AND user_id = ?
                                """, (msg_id, user_id))
                                
                            db.commit()
                        notify_clients()
                    except sqlite3.OperationalError as e:
                        logger.warning(f"База занята, не удалось обновить удаленные сообщения: {e}")

            @client.on(events.MessageEdited)
            async def edit_handler(event):
                parsing_allowed_now = is_parsing_allowed(session_id)
                message = event.message
                if (not message.text and not message.document):
                    return
                message_event_date = getattr(message, 'edit_date', None) or message.date
                
                chat = await event.get_chat()

                # --- ИГНОРИРУЕМ ИСХОДЯЩИЕ ТОЛЬКО ДЛЯ БОТОВ ---
                if getattr(message, 'out', False) and event.is_private and getattr(chat, 'bot', False):
                    chat_id = chat.id
                    bot_un_clean = normalize_telegram_ref(getattr(chat, 'username', ''))
                    with app.app_context():
                        db = get_db()
                        interaction_row = get_interaction_bot_for_userbot(
                            db,
                            user_id,
                            session_id,
                            chat_id,
                            bot_un_clean,
                        )
                        if interaction_row and not is_automated_outgoing_interaction_command(interaction_row, message):
                            with sqlite_write_lock:
                                mark_manual_interaction_intervention(db, interaction_row, message_event_date)
                                db.commit()
                            logger.info(
                                "Ручное сообщение в чат бота %s через юзербот %s: ответы после него не попадут в парсер",
                                bot_un_clean or chat_id,
                                session_id,
                            )
                    return
                # ---------------------------------------------
                
                chat_id = chat.id
                if event.is_private:
                    first = getattr(chat, 'first_name', '') or ''
                    last = getattr(chat, 'last_name', '') or ''
                    chat_title = f"{first} {last}".strip() or "Unknown"
                    msg_type = 'private'
                elif event.is_group:
                    chat_title = getattr(chat, 'title', 'Untitled Group')
                    msg_type = 'group'
                elif event.is_channel:
                    chat_title = getattr(chat, 'title', 'Untitled Channel')
                    msg_type = 'channel'
                else:
                    chat_title = "Unknown"
                    msg_type = 'unknown'
                
              
                # --- УМНАЯ ИЗОЛЯЦИЯ ПАРСИНГА БОТОВ (ФИНАЛ) ---
                sender = await event.get_sender()
                is_bot = getattr(sender, 'bot', False)
                accepted_interaction_bot = False
                interaction_bot_row = None
                bot_un_clean = ""
                sender_name = getattr(sender, 'username', '') if sender else ''
                if not sender_name:
                    sender_name = getattr(sender, 'first_name', 'Бот') if sender else 'Бот'
	                
                if is_bot:
                    bot_un = getattr(sender, 'username', '')
                    if not bot_un:
                        try:
                            entity = await client.get_entity(chat_id)
                            bot_un = getattr(entity, 'username', '')
                        except: pass
                        
                    bot_un_clean = bot_un.lower().replace('@', '').strip() if bot_un else ''
                    
                    print(f"\n[БОТ ДЕБАГ] 🤖 Пришло сообщение от бота: '{bot_un_clean}'. Chat ID: {chat_id}")
                    
                    with app.app_context():
                        local_db = get_db()
                        check_bot = should_accept_interaction_bot_message(
                            local_db,
                            user_id,
                            session_id,
                            chat_id,
                            bot_un_clean,
                            message_event_date,
                            getattr(message, 'reply_to_msg_id', None),
                            getattr(message, 'id', None),
                        )
                        
                        if not check_bot:
                            print(f"[БОТ ДЕБАГ] ❌ ОТКАЗ: Бот '{bot_un_clean}' не является свежим ответом этого юзербота. Игнорируем.")
                            return
                        else:
                            print(f"[БОТ ДЕБАГ] ✅ БОТ ОДОБРЕН: '{bot_un_clean}' найден в базе. Идем дальше...")
                            
                            tracked_row = ensure_tracked_source_row(
                                local_db,
                                user_id,
                                chat_id,
                                chat_title,
                                check_bot['custom_name'] if 'custom_name' in check_bot.keys() else (sender_name or bot_un_clean),
                            )
                            if tracked_row:
                                local_db.execute(
                                    """
                                    UPDATE interaction_bots
                                    SET tracked_chat_id = ?,
                                        resolved_chat_id = ?
                                    WHERE id = ?
                                    """,
                                    (tracked_row['id'], chat_id, check_bot['id']),
                                )
                            local_db.commit()

                            if not tracked_row:
                                print(f"[БОТ ДЕБАГ] ❌ ОТКАЗ: Бот одобрен, но источник не удалось создать!")
                                return
                            else:
                                print(f"[БОТ ДЕБАГ] ✅ ПОЛНЫЙ УСПЕХ: Источник бота готов, начинаем парсить!")
                                accepted_interaction_bot = True
                                interaction_bot_row = check_bot
	                # -------------------------------------------------------------

                if not parsing_allowed_now and not accepted_interaction_bot:
                    return

                parsed_text = await parse_excel_message(client, message, chat_id, user_id)
                if not parsed_text:
                    return
                parsed_text_has_price_lines = source_text_has_supplier_price_lines(parsed_text)
                with app.app_context():
                    db = get_db()
                    try:
                        with sqlite_write_lock:
                            if (
                                accepted_interaction_bot
                                and interaction_bot_row
                                and parsed_text_has_price_lines
                                and should_cleanup_interaction_response_snapshot(interaction_bot_row, user_id, chat_id)
                            ):
                                cleanup_source_before_new_snapshot(db, user_id, chat_id)
                            if accepted_interaction_bot and interaction_bot_row:
                                mark_interaction_bot_response_seen(db, interaction_bot_row, message_event_date)

                            msg_date = telegram_message_date_to_local_naive(message.date).strftime("%Y-%m-%d %H:%M:%S")
                            existing_message = db.execute(
                                "SELECT id FROM messages WHERE telegram_message_id = ? AND chat_id = ? AND user_id = ?",
                                (message.id, chat_id, user_id),
                            ).fetchone()
                            if existing_message:
                                db.execute(
                                    """
                                    UPDATE messages
                                    SET text = ?, date = ?, chat_title = ?, sender_name = ?, type = ?, is_blocked = 0, is_delayed = 0
                                    WHERE id = ?
                                    """,
                                    (parsed_text, msg_date, chat_title, sender_name, msg_type, existing_message['id'])
                                )
                                edited_message_db_id = existing_message['id']
                            else:
                                cursor = db.execute(
                                    """
                                    INSERT INTO messages
                                    (user_id, telegram_message_id, type, text, date, chat_id, chat_title, sender_name, is_delayed, is_blocked)
                                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, 0, 0)
                                    """,
                                    (user_id, message.id, msg_type, parsed_text, msg_date, chat_id, chat_title, sender_name),
                                )
                                edited_message_db_id = cursor.lastrowid

                            bindings = db.execute("""
                                SELECT pm.id, pm.product_id, pm.line_index, p.name, p.country, p.synonyms, p.brand, p.model_number, p.color, p.storage
                                FROM product_messages pm
                                JOIN products p ON pm.product_id = p.id
                                JOIN messages m ON pm.message_id = m.id
                                WHERE m.telegram_message_id = ? AND m.chat_id = ? AND m.user_id = ?
                            """, (message.id, chat_id, user_id)).fetchall()
                            lines = parsed_text.split('\n') if parsed_text else []

                            should_revalidate_bindings = parsed_text_has_price_lines or not accepted_interaction_bot
                            if should_revalidate_bindings:
                                for b in bindings:
                                    line_idx = b['line_index']
                                    is_actual = 0
                                    new_price = None

                                    if line_idx == -1:
                                        if binding_product_row_matches_line(b, parsed_text, require_known_category=True):
                                            is_actual = 1
                                            new_price = extract_price(parsed_text)
                                    elif 0 <= line_idx < len(lines):
                                        edited_line = lines[line_idx]
                                        if binding_product_row_matches_line(b, edited_line, require_known_category=True):
                                            is_actual = 1
                                            new_price = extract_price(edited_line)

                                    if is_actual:
                                        db.execute("""
                                            DELETE FROM product_messages
                                            WHERE product_id = ? AND message_id = ? AND line_index = ? AND id != ?
                                        """, (b['product_id'], edited_message_db_id, line_idx, b['id']))
                                    db.execute("""
                                        UPDATE product_messages
                                        SET is_actual = ?, extracted_price = ?
                                        WHERE id = ?
                                    """, (is_actual, new_price, b['id']))

                            db.commit()
                        if parsed_text_has_price_lines:
                            enqueue_auto_confirm_source_message(user_id, edited_message_db_id, f"Edited {chat_id}")
                        notify_clients()
                    except sqlite3.IntegrityError:
                        pass
                    except sqlite3.OperationalError as e:
                        logger.warning("Не удалось сохранить edit сообщения %s/%s: %s", chat_id, message.id, e)

            # --------------------------------------------------
            # --------------------------------------------------

            

            @client.on(events.NewMessage)
            async def handler(event):
                message = event.message
                
                if (not message.text and not message.document):
                    return

                chat = await event.get_chat()

                # --- ИГНОРИРУЕМ ИСХОДЯЩИЕ ТОЛЬКО ДЛЯ БОТОВ ---
                if getattr(message, 'out', False) and event.is_private and getattr(chat, 'bot', False):
                    chat_id = chat.id
                    bot_un_clean = normalize_telegram_ref(getattr(chat, 'username', ''))
                    with app.app_context():
                        db = get_db()
                        interaction_row = get_interaction_bot_for_userbot(
                            db,
                            user_id,
                            session_id,
                            chat_id,
                            bot_un_clean,
                        )
                        if interaction_row and not is_automated_outgoing_interaction_command(interaction_row, message):
                            with sqlite_write_lock:
                                mark_manual_interaction_intervention(db, interaction_row, message.date)
                                db.commit()
                            logger.info(
                                "Ручное сообщение в чат бота %s через юзербот %s: ответы после него не попадут в парсер",
                                bot_un_clean or chat_id,
                                session_id,
                            )
                    return
                # ---------------------------------------------

                parsing_allowed_now = is_parsing_allowed(session_id)
                
                chat_id = chat.id

                # Проверка: отслеживаем ли мы этот чат?
                with app.app_context():
                    db = get_db()
                    cursor = db.execute(
                        "SELECT COUNT(*) FROM tracked_chats WHERE user_id = ? AND chat_id = ?",
                        (user_id, chat_id)
                    )
                    count = cursor.fetchone()[0]
                    
                    # Обычные чаты должны быть добавлены вручную. Ответы активных private-ботов
                    # пропускаем дальше: ниже они будут проверены по interaction_bots и созданы как источник.
                    if count == 0 and not (event.is_private and getattr(chat, 'bot', False)):
                        return

                # Определяем тип и название чата
                if event.is_private:
                    first = getattr(chat, 'first_name', '') or ''
                    last = getattr(chat, 'last_name', '') or ''
                    chat_title = f"{first} {last}".strip() or "Unknown"
                    msg_type = 'private'
                elif event.is_group:
                    chat_title = getattr(chat, 'title', 'Untitled Group')
                    msg_type = 'group'
                elif event.is_channel:
                    chat_title = getattr(chat, 'title', 'Untitled Channel')
                    msg_type = 'channel'
                else:
                    chat_title = "Unknown"
                    msg_type = 'unknown'

                # --- УМНАЯ ИЗОЛЯЦИЯ ПАРСИНГА БОТОВ (ФИНАЛ) ---
                sender = await event.get_sender()
                is_bot = getattr(sender, 'bot', False)
                sender_name = getattr(sender, 'username', '')
                if not sender_name:
                    sender_name = getattr(sender, 'first_name', 'Бот')
                accepted_interaction_bot = False
                interaction_bot_row = None

                if is_bot:
                    bot_un = getattr(sender, 'username', '')
                    if not bot_un:
                        try:
                            entity = await client.get_entity(chat_id)
                            bot_un = getattr(entity, 'username', '')
                        except: pass
                        
                    bot_un_clean = bot_un.lower().replace('@', '').strip() if bot_un else ''
                    
                    # Генерируем все возможные варианты того, как бот мог быть записан в базу
                    bot_un_at = f"@{bot_un_clean}"
                    bot_link_1 = f"https://t.me/{bot_un_clean}"
                    bot_link_2 = f"t.me/{bot_un_clean}"
                    
                    print(f"\n[БОТ ДЕБАГ] 🤖 Пришло сообщение от бота: '{bot_un_clean}'. Chat ID: {chat_id}")
                    
                    with app.app_context():
                        local_db = get_db()
                        check_bot = should_accept_interaction_bot_message(
                            local_db,
                            user_id,
                            session_id,
                            chat_id,
                            bot_un_clean,
                            message.date,
                            getattr(message, 'reply_to_msg_id', None),
                        )
                        
                        if not check_bot:
                            print(f"[БОТ ДЕБАГ] ❌ ОТКАЗ: Бот '{bot_un_clean}' не является свежим ответом на наш запрос. Игнорируем.")
                            return
                        else:
                            print(f"[БОТ ДЕБАГ] ✅ БОТ ОДОБРЕН: '{bot_un_clean}' найден в базе. Идем дальше...")
                            
                            tracked_row = ensure_tracked_source_row(
                                local_db,
                                user_id,
                                chat_id,
                                chat_title,
                                check_bot['custom_name'] if 'custom_name' in check_bot.keys() else (sender_name or bot_un_clean),
                            )
                            if tracked_row:
                                local_db.execute(
                                    """
                                    UPDATE interaction_bots
                                    SET tracked_chat_id = ?,
                                        resolved_chat_id = ?
                                    WHERE id = ?
                                    """,
                                    (tracked_row['id'], chat_id, check_bot['id']),
                                )
                            local_db.commit()

                            if not tracked_row:
                                print(f"[БОТ ДЕБАГ] ❌ ОТКАЗ: Бот одобрен, но источник не удалось создать!")
                                return
                            else:
                                print(f"[БОТ ДЕБАГ] ✅ ПОЛНЫЙ УСПЕХ: Источник бота готов, начинаем парсить!")
                                accepted_interaction_bot = True
                                interaction_bot_row = check_bot
                # -------------------------------------------------------------

                if not parsing_allowed_now and not accepted_interaction_bot:
                    return

                parsed_text = await parse_excel_message(client, message, chat_id, user_id)
                if not parsed_text:
                    return
                parsed_text_has_price_lines = source_text_has_supplier_price_lines(parsed_text)
                with app.app_context():
                    db = get_db()
                    try:
                        with sqlite_write_lock:
                            if (
                                accepted_interaction_bot
                                and parsed_text_has_price_lines
                                and should_cleanup_interaction_response_snapshot(interaction_bot_row, user_id, chat_id)
                            ):
                                cleanup_source_before_new_snapshot(db, user_id, chat_id)
                            if accepted_interaction_bot:
                                mark_interaction_bot_response_seen(db, interaction_bot_row, message.date)
                            msg_date = telegram_message_date_to_local_naive(message.date).strftime("%Y-%m-%d %H:%M:%S")
                            cursor = db.execute(
                                'INSERT INTO messages (user_id, telegram_message_id, type, text, date, chat_id, chat_title, sender_name, is_delayed) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 0)',
                                (user_id, message.id, msg_type, parsed_text, msg_date, chat_id, chat_title, sender_name)
                            )
                            message_db_id = cursor.lastrowid
                            if chat_title:
                                db.execute(
                                    "UPDATE tracked_chats SET chat_title = ? WHERE user_id = ? AND chat_id = ? AND chat_title IS NULL",
                                    (chat_title, user_id, chat_id)
                                )
                            db.commit()
                        if parsed_text_has_price_lines:
                            enqueue_auto_confirm_source_message(user_id, message_db_id, sender_name)
                        notify_clients()

                    except sqlite3.IntegrityError:
                        existing = db.execute(
                            """
                            SELECT id
                            FROM messages
                            WHERE telegram_message_id = ? AND chat_id = ? AND user_id = ?
                            """,
                            (message.id, chat_id, user_id),
                        ).fetchone()
                        if existing and source_text_has_supplier_price_lines(parsed_text):
                            enqueue_auto_confirm_source_message(user_id, existing['id'], sender_name)
                    except sqlite3.OperationalError as e:
                        logger.warning("Не удалось сохранить новое сообщение %s/%s: %s", chat_id, message.id, e)

            async def main():
                await client.connect()
            
                # --- ОБРАБОТКА СЛЕТЕВШЕЙ АВТОРИЗАЦИИ ---
                if not await client.is_user_authorized():
                    logging.error(f"Сессия {session_id}: Клиент не авторизован!")
                    try:
                        await client.disconnect()
                    except Exception:
                        pass
                
                    with app.app_context():
                        db = get_db()
                        # Устанавливаем статус unauthorized
                        db.execute("UPDATE user_telegram_sessions SET status = 'unauthorized' WHERE id = ?", (session_id,))
                        db.commit()
                        if 'notify_clients' in globals(): notify_clients()
                    return
                # ---------------------------------------

                me = await client.get_me()
            
                # --- Получаем @username или Имя юзербота ---
                account_name = getattr(me, 'username', None)
                if account_name:
                    account_name = f"@{account_name}"
                else:
                    account_name = getattr(me, 'first_name', f"ID: {me.id}")
                # --------------------------------------------------
            
                logging.info(f"Сессия {session_id}: слушатель запущен для {me.first_name}")
            
                with app.app_context():
                    db = get_db()
                    try:
                        db.execute(
                            "UPDATE user_telegram_sessions SET status = 'active', session_file = ?, account_name = ? WHERE id = ?",
                            (session_file, account_name, session_id)
                        )
                    except sqlite3.OperationalError:
                        db.execute(
                            "UPDATE user_telegram_sessions SET status = 'active', session_file = ? WHERE id = ?",
                            (session_file, session_id)
                        )
                    db.commit()
                    if 'notify_clients' in globals(): notify_clients()
                
                # Блокирующий вызов: клиент слушает события до отключения
                # Блокирующий вызов: клиент слушает события до отключения
                try:
                    await client.run_until_disconnected()
                except AuthKeyUnregisteredError:
                    logging.error(f"Сессия {session_id}: Ключ авторизации отозван сервером Telegram (слетела сессия).")
                    with app.app_context():
                        db = get_db()
                        db.execute("UPDATE user_telegram_sessions SET status = 'unauthorized' WHERE id = ?", (session_id,))
                        db.commit()
                        if 'notify_clients' in globals(): notify_clients()


            # --- ЗАПУСК И БЕЗОПАСНАЯ ОЧИСТКА В ЦИКЛЕ ---
            try:
                loop.run_until_complete(main())
            except Exception as e:
                logging.exception(f"Сессия {session_id}: ошибка в слушателе - {e}")
            finally:
                # 1. Корректное отключение от серверов Telegram
                try:
                    if client and client.is_connected():
                        loop.run_until_complete(client.disconnect())
                except Exception:
                    pass
            
                # 2. Очищаем глобальные словари ВНУТРИ родного потока.
                # Это полностью исключает панику сборщика мусора (GC) в Flask!
                user_clients.pop(session_id, None)
                user_loops.pop(session_id, None)
                background_tasks.pop(session_id, None)

                # 3. Отменяем оставшиеся фоновые задачи Telethon и закрываем цикл
                try:
                    tasks = asyncio.all_tasks(loop)
                    for t in tasks:
                        t.cancel()
                
                    if tasks:
                        # Даем задачам шанс чисто закрыться (игнорируя ошибки отмены)
                        loop.run_until_complete(asyncio.gather(*tasks, return_exceptions=True))
                
                    loop.close()
                except Exception:
                    pass
                
                # 4. Не гасим статус active при штатном завершении потока.
                # PM2/redeploy останавливает слушатель перед перезапуском процесса; если здесь записать
                # inactive, следующий старт уже не поднимет Telegram-бота. Ручное выключение через UI
                # по-прежнему выставляет inactive в toggle_userbot(), а слетевшая авторизация остается unauthorized.
                try:
                    with app.app_context():
                        local_db = get_db()
                        current_row = local_db.execute(
                            "SELECT status FROM user_telegram_sessions WHERE id = ?",
                            (session_id,)
                        ).fetchone()
                        current_status = current_row['status'] if current_row else None
                        if current_status == 'active':
                            logging.info(
                                f"Сессия {session_id}: слушатель остановлен, статус active сохранен для автозапуска"
                            )
                except Exception as e:
                    logging.error(f"Не удалось проверить статус сессии {session_id} в БД: {e}")
                # ===========================================================================

                logging.info(f"Сессия {session_id}: слушатель полностью остановлен")

    # --- СОЗДАНИЕ И ЗАПУСК ПОТОКА ---
    # Даем потоку имя для удобства дебага
    thread = threading.Thread(target=listener, daemon=True, name=f"UserbotThread-{session_id}")
    thread.start()
    background_tasks[session_id] = thread
    logging.info(f"Поток слушателя для сессии {session_id} создан")

# ---------- Маршруты Flask ----------
@app.route('/')
@login_required
def index():
    return render_template('store.html')

@app.route('/api/profile/logout_others', methods=['POST'])
@login_required
def logout_others():
    user_id = session['user_id']
    new_token = secrets.token_hex(16)
    
    db = get_db()
    db.execute("UPDATE users SET session_token = ? WHERE id = ?", (new_token, user_id))
    db.commit()
    
    session['session_token'] = new_token
    
    # --- НОВАЯ СТРОКА: Рассылаем сигнал всем подключенным клиентам ---
    socketio.emit('force_logout')
    
    return jsonify({'success': True})

@app.route('/login', methods=['GET', 'POST'])
def login_page():
    if request.method == 'POST':
        login = request.form.get('login')
        password = request.form.get('password')
        pwd_hash = hashlib.sha256(password.encode()).hexdigest()
        
        db = get_db()
        user = db.execute("SELECT * FROM users WHERE login = ? AND password_hash = ?", (login, pwd_hash)).fetchone()
        
        if user:
            # --- ГЕНЕРАЦИЯ И СОХРАНЕНИЕ ТОКЕНА ---
            token = user['session_token'] if 'session_token' in user.keys() else None
            if not token:
                token = secrets.token_hex(16)
                try:
                    db.execute("UPDATE users SET session_token = ? WHERE id = ?", (token, user['id']))
                    db.commit()
                except sqlite3.OperationalError:
                    pass # На случай, если колонка еще не создалась
            
            session['user_id'] = user['id']
            session['role'] = user['role']
            session['login'] = user['login']
            session['session_token'] = token # Обязательно записываем в браузер
            # -------------------------------------
            
            return redirect(url_for('index'))
        return render_template_string(LOGIN_TEMPLATE, error="Неверный логин или пароль")
    return render_template_string(LOGIN_TEMPLATE)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login_page'))

@app.route('/api/publish_tree_data', methods=['GET'])
@login_required
def get_publish_tree_data():
    db = get_db()
    user_id = session.get('user_id')
    
    folders = db.execute("SELECT id, name, parent_id FROM folders WHERE user_id=?", (user_id,)).fetchall()
    products = db.execute("""
        SELECT id, name, folder_id, sort_index
        FROM products
        WHERE user_id=?
        ORDER BY CASE WHEN sort_index IS NULL THEN 1 ELSE 0 END, sort_index, id
    """, (user_id,)).fetchall()
    
    # Получаем поставщиков, у которых есть актуальные цены на товары
    # ВАЖНО: Я использую m.chat_id и m.chat_title. Если в вашей базе таблица messages 
    # использует sender_id или channel_name, просто замените эти слова в запросе ниже!
    suppliers = db.execute("""
        SELECT pm.product_id, m.chat_id as supplier_id, m.chat_title as supplier_name, pm.extracted_price
        FROM product_messages pm
        JOIN messages m ON pm.message_id = m.id
        WHERE pm.status = 'confirmed' AND pm.is_actual = 1 AND m.user_id = ?
        GROUP BY pm.product_id, m.chat_id
    """, (user_id,)).fetchall()
    
    return jsonify({
        'folders': [dict(f) for f in folders],
        'products': [dict(p) for p in products],
        'suppliers': [dict(s) for s in suppliers]
    })

@app.route('/api/tracked_chats/<int:chat_id>', methods=['PUT'])
@login_required
def update_tracked_chat(chat_id):
    data = request.get_json()
    custom_name = data.get('custom_name')
    user_id = session['user_id']
    db = get_db()
    # Проверяем, принадлежит ли чат пользователю
    cur = db.execute("SELECT id FROM tracked_chats WHERE id = ? AND user_id = ?", (chat_id, user_id))
    if not cur.fetchone():
        return jsonify({'error': 'Not found'}), 404
    # Разрешаем пустую строку, сохраняем как NULL
    if custom_name == '':
        custom_name = None
    db.execute("UPDATE tracked_chats SET custom_name = ? WHERE id = ?", (custom_name, chat_id))
    db.commit()
    notify_clients()
    return jsonify({'success': True})


@app.route('/api/get_qr', methods=['POST'])
@login_required
def get_qr():
    ensure_event_loop()
    data = request.get_json()
    api_id = data.get('api_id')
    api_hash = data.get('api_hash')
    session_id = data.get('session_id') # <--- ВОТ ЗДЕСЬ ПРИНИМАЕМ ID СТАРОЙ СЕССИИ
    user_id = session['user_id']

    if not api_id or not api_hash:
        return jsonify({'success': False, 'error': 'Missing fields'}), 400

    db = get_db()
    
    # --- ИСПРАВЛЕННАЯ ЛОГИКА ---
    if session_id:
        # Если передали ID старой сессии, просто обновляем её статус
        db.execute("UPDATE user_telegram_sessions SET status = 'pending' WHERE id = ? AND user_id = ?", (session_id, user_id))
    else:
        # Иначе создаем новую
        cursor = db.execute(
            "INSERT INTO user_telegram_sessions (user_id, api_id, api_hash, status) VALUES (?, ?, ?, 'pending')",
            (user_id, api_id, api_hash)
        )
        session_id = cursor.lastrowid
    # ---------------------------
        
    db.commit()
    if 'notify_clients' in globals(): notify_clients()
    
    session_file = f'sessions/user_{user_id}_{session_id}'
    q = queue.Queue()

    def qr_worker():
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        client = TelegramClient(session_file, int(api_id), api_hash)

        async def _run():
            await client.connect()
            if await client.is_user_authorized():
                await client.disconnect()
                q.put({'already_authorized': True})
                return

            try:
                qr_login = await client.qr_login()
                q.put({'qr_url': qr_login.url})
                
                await qr_login.wait(timeout=300)
                await client.disconnect()
                
                with app.app_context():
                    local_db = get_db()
                    local_db.execute("UPDATE user_telegram_sessions SET status = 'active' WHERE id = ?", (session_id,))
                    local_db.commit()
                    if 'notify_clients' in globals(): notify_clients()
                logging.info(f"Сессия {session_id}: статус обновлён на active")
                start_message_listener(session_id, user_id, api_id, api_hash)

            except asyncio.TimeoutError:
                logging.error(f"Сессия {session_id}: таймаут ожидания QR")
                await client.disconnect()
                with app.app_context():
                    local_db = get_db()
                    # Если ошибка, возвращаем статус unauthorized вместо удаления
                    local_db.execute("UPDATE user_telegram_sessions SET status = 'unauthorized' WHERE id = ?", (session_id,))
                    local_db.commit()
                    if 'notify_clients' in globals(): notify_clients()
            except Exception as e:
                logging.exception(f"Сессия {session_id}: ошибка при ожидании QR")
                await client.disconnect()
                with app.app_context():
                    local_db = get_db()
                    local_db.execute("UPDATE user_telegram_sessions SET status = 'unauthorized' WHERE id = ?", (session_id,))
                    local_db.commit()
                    if 'notify_clients' in globals(): notify_clients()
            finally:
                qr_sessions.pop(session_id, None)

        loop.run_until_complete(_run())
        loop.close()

    thread = threading.Thread(target=qr_worker, daemon=True)
    thread.start()

    try:
        res = q.get(timeout=15)
    except queue.Empty:
        db.execute("UPDATE user_telegram_sessions SET status = 'unauthorized' WHERE id = ?", (session_id,))
        db.commit()
        if 'notify_clients' in globals(): notify_clients()
        return jsonify({'success': False, 'error': 'Таймаут генерации QR'}), 500

    if res.get('already_authorized'):
        start_message_listener(session_id, user_id, api_id, api_hash)
        db.execute("UPDATE user_telegram_sessions SET status = 'active' WHERE id = ?", (session_id,))
        db.commit()
        if 'notify_clients' in globals(): notify_clients()
        return jsonify({'success': True, 'already_authorized': True})

    qr_url = res.get('qr_url')
    if not qr_url:
        db.execute("UPDATE user_telegram_sessions SET status = 'unauthorized' WHERE id = ?", (session_id,))
        db.commit()
        if 'notify_clients' in globals(): notify_clients()
        return jsonify({'success': False, 'error': 'Не удалось сгенерировать QR'}), 500

    qr_sessions[session_id] = True 

    img = qrcode.make(qr_url)
    buffer = BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    img_base64 = base64.b64encode(buffer.getvalue()).decode()

    return jsonify({'success': True, 'qr': img_base64, 'session_id': session_id})

@app.route('/api/check_login')
@login_required
def check_login():
    session_id = request.args.get('session_id')
    if not session_id:
        return jsonify({'success': False, 'error': 'Missing session_id'}), 400
    session_id = int(session_id)

    db = get_db()
    row = db.execute("SELECT status FROM user_telegram_sessions WHERE id = ?", (session_id,)).fetchone()
    if row and row['status'] == 'active':
        return jsonify({'success': True})
    return jsonify({'success': False})

@app.route('/api/check_session')
@login_required
def check_session():
    user_id = session['user_id']
    active = any(uid == user_id for _, _, uid in user_clients.values())
    return jsonify({'authorized': active})

import requests

# --- РОУТЫ ДЛЯ НАСТРОЙКИ API-ПАРСЕРА ---
@app.route('/api/api_sources', methods=['GET', 'POST'])
@login_required
def handle_api_sources():
    db = get_db()
    user_id = session.get('user_id')
    
    if request.method == 'GET':
        sources = db.execute("SELECT * FROM api_sources WHERE user_id=?", (user_id,)).fetchall()
        return jsonify([dict(s) for s in sources])
    else:
        data = request.json
        db.execute("""
            INSERT INTO api_sources (user_id, name, url, token, interval_min)
            VALUES (?, ?, ?, ?, ?)
        """, (user_id, data.get('name'), data.get('url'), data.get('token'), data.get('interval_min', 60)))
        db.commit()
        return jsonify({'success': True})

@app.route('/api/api_sources/<int:source_id>', methods=['DELETE'])
def delete_api_source(source_id):
    try:
        db = get_db()
        user_id = session.get('user_id')
        dummy_chat_id = f"api_src_{source_id}"
        
        # 1. Удаляем привязки цен к товарам
        db.execute("DELETE FROM product_messages WHERE message_id IN (SELECT id FROM messages WHERE chat_id=? AND user_id=?)", (dummy_chat_id, user_id))
        
        # 2. Удаляем сами сохраненные прайсы/сообщения
        db.execute("DELETE FROM messages WHERE chat_id=? AND user_id=?", (dummy_chat_id, user_id))
        
        # 3. Удаляем само подключение API
        db.execute("DELETE FROM api_sources WHERE id=? AND user_id=?", (source_id, user_id))
        
        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        print(f"Ошибка при удалении API: {e}")
        return jsonify({'error': str(e)}), 500

# --- ФОНОВЫЙ ПАРСЕР ЧУЖИХ API ---
def api_parser_scheduler():
    """Раз в минуту проверяет чужие API и заносит товары в базу как поставщик"""
    time.sleep(5)
    last_sync_times = {}
    
    with app.app_context():
        while True:
            try:
                db = get_db()
                sources = db.execute("SELECT * FROM api_sources WHERE is_active=1").fetchall()
                from datetime import timedelta
                now = datetime.now(timezone.utc).replace(tzinfo=None) + timedelta(hours=3)
                
                for src in sources:
                    src_id = src['id']
                    user_id = src['user_id']
                    interval = src['interval_min']

                    if not is_user_parsing_allowed(user_id):
                        last_sync_times[src_id] = now
                        continue
                    
                    # Проверяем, пришло ли время опрашивать этот API
                    last_run = last_sync_times.get(src_id)
                    if last_run and (now - last_run).total_seconds() / 60 < interval:
                        continue
                        
                    # Формируем URL с токеном (как в вашей системе)
                    url = f"{src['url']}?token={src['token']}"
                    
                    try:
                        resp = requests.get(url, timeout=15)
                        if resp.status_code == 200:
                            data = resp.json()
                            products = data.get('products', [])
                            
                            # Отбираем только валидные товары
                            valid_products = [p for p in products if p.get('name') and p.get('price') is not None]
                            
                            # Создаем "фейковый чат", чтобы API выглядел как поставщик
                            dummy_chat_id = f"api_src_{src_id}"
                            dummy_chat_title = f"🌐 API: {src['name']}"
                            
                            # 1. Собираем все товары в единый текст (построчно)
                            lines = [f"Авто-обновление из API ({len(valid_products)} товаров)"]
                            for p in valid_products:
                                lines.append(f"{p['name']} — {p['price']}")
                                
                            msg_text = normalize_incoming_text("\n".join(lines))
                            sender_name = f"🌐 API: {src['name']}"
                            
                            cleanup_source_before_new_snapshot(db, user_id, dummy_chat_id)
                            # Вставляем "сообщение" от этого поставщика с полным списком
                            cursor = db.execute(
                                "INSERT INTO messages (user_id, chat_id, chat_title, text, date, sender_name) VALUES (?, ?, ?, ?, ?, ?) RETURNING id",
                                (user_id, dummy_chat_id, dummy_chat_title, msg_text, now, sender_name)
                            )
                            dummy_msg_id = cursor.fetchone()['id']
                            
                            # 2. Привязываем товары с указанием конкретной строки (line_index)
                            for i, p in enumerate(valid_products):
                                name = p.get('name')
                                price = p.get('price')
                                
                                # Ищем товар в базе или создаем новый (без папки)
                                prod_row = db.execute("SELECT id FROM products WHERE name=? AND user_id=?", (name, user_id)).fetchone()
                                if prod_row:
                                    prod_id = prod_row['id']
                                else:
                                    cur = db.execute(
                                        "INSERT INTO products (user_id, name, sort_index) VALUES (?, ?, ?) RETURNING id",
                                        (user_id, name, build_product_sort_index(i + 1, name)),
                                    )
                                    prod_id = cur.fetchone()['id']
                                update_product_sort_index(
                                    db,
                                    user_id,
                                    prod_id,
                                    build_product_sort_index(i + 1, name),
                                )
                                    
                                # line_index = i + 1, так как 0-я строка - это заголовок "Авто-обновление..."
                                db.execute("""
                                    INSERT INTO product_messages (product_id, message_id, extracted_price, status, is_actual, line_index)
                                    VALUES (?, ?, ?, 'pending', 1, ?)
                                """, (prod_id, dummy_msg_id, price, i + 1))
                                
                            db.commit()
                            notify_clients()
                            print(f"✅ Спарсили {len(valid_products)} товаров из {src['name']}")
                    except Exception as req_err:
                        print(f"❌ Ошибка запроса к API {src['name']}: {req_err}")
                        
                    last_sync_times[src_id] = now
            except Exception as e:
                print(f"Ошибка в планировщике парсинга API: {e}")
                
            time.sleep(60) # Проверяем таймеры каждую минуту



import math
from flask import request, jsonify, session

import math
from flask import request, jsonify, session

def get_sender_filter_label(msg):
    custom_name = str(msg.get('custom_name') or '').strip()
    if custom_name:
        return custom_name
    msg_type_value = str(msg.get('type') or '').strip().lower()
    if msg_type_value in {'channel', 'group'}:
        return str(msg.get('chat_title') or 'Чат').strip() or 'Чат'
    sender_name_value = str(msg.get('sender_name') or '').replace(' None', '').strip()
    if not sender_name_value or sender_name_value == 'Unknown':
        fallback_title = str(msg.get('chat_title') or '').strip()
        return fallback_title or 'Неизвестно'
    return sender_name_value


def compose_excel_source_display_label(base_label, sender_name_value):
    base_value = str(base_label or '').strip()
    sender_value = str(sender_name_value or '').replace(' None', '').strip()
    if base_value and sender_value and sender_value.lower().startswith('лист:') and sender_value != base_value:
        return f"{base_value} · {sender_value}"
    return base_value or sender_value


def get_message_source_display_label(msg):
    base_label = get_sender_filter_label(msg)
    msg_type_value = str(msg.get('type') or '').strip().lower()
    sender_name_value = str(msg.get('sender_name') or '').replace(' None', '').strip()
    if msg_type_value.startswith('excel_') and sender_name_value:
        return compose_excel_source_display_label(base_label, sender_name_value)
    return base_label


def get_message_source_bucket_key(msg):
    chat_id = str(msg.get('chat_id') or '').strip()
    msg_type_value = str(msg.get('type') or '').strip().lower()
    if msg_type_value.startswith('excel_'):
        return f"{chat_id}:{msg_type_value}"
    return chat_id


def enrich_message_row(msg):
    row = dict(msg)
    row['sender_filter_label'] = get_sender_filter_label(row)
    row['source_display_name'] = get_message_source_display_label(row)
    return row


def attach_message_linked_lines(db, user_id, messages):
    message_ids = []
    for msg in messages or []:
        try:
            message_ids.append(int(msg.get('id') or 0))
        except (TypeError, ValueError):
            continue
    message_ids = [message_id for message_id in dict.fromkeys(message_ids) if message_id > 0]
    if not message_ids:
        return messages

    placeholders = ",".join("?" for _ in message_ids)
    rows = db.execute(
        f"""
        SELECT
            pm.message_id,
            COALESCE(pm.line_index, -1) AS line_index,
            pm.status,
            COALESCE(pm.is_actual, 1) AS is_actual,
            pm.extracted_price,
            p.id AS product_id,
            p.name AS product_name
        FROM product_messages pm
        JOIN products p ON p.id = pm.product_id
        WHERE p.user_id = ?
          AND pm.status = 'confirmed'
          AND pm.message_id IN ({placeholders})
        ORDER BY pm.message_id, COALESCE(pm.line_index, -1), COALESCE(pm.is_actual, 1) DESC, p.name
        """,
        (user_id, *message_ids),
    ).fetchall()

    linked_by_message = {}
    for row in rows:
        linked_by_message.setdefault(int(row['message_id']), []).append({
            'line_index': int(row['line_index'] if row['line_index'] is not None else -1),
            'status': row['status'],
            'is_actual': int(row['is_actual'] or 0),
            'extracted_price': row['extracted_price'],
            'product_id': row['product_id'],
            'product_name': row['product_name'],
        })

    for msg in messages or []:
        try:
            msg_id = int(msg.get('id') or 0)
        except (TypeError, ValueError):
            msg_id = 0
        msg['linked_lines'] = linked_by_message.get(msg_id, [])
    return messages


def message_matches_sender_filter(msg, sender_filter):
    target = str(sender_filter or '').strip()
    if not target or target == 'all':
        return True
    candidates = set()
    for value in (
        msg.get('source_display_name'),
        msg.get('sender_filter_label'),
        msg.get('custom_name'),
        msg.get('chat_title'),
        msg.get('sender_name'),
    ):
        clean_value = str(value or '').replace(' None', '').strip()
        if not clean_value:
            continue
        candidates.add(clean_value)
        if ' · Лист:' in clean_value:
            candidates.add(clean_value.split(' · Лист:', 1)[0].strip())
    return target in candidates


def get_message_sender_option_labels(msg):
    msg_type_value = str(msg.get('type') or '').strip().lower()
    source_display_name = str(msg.get('source_display_name') or '').strip()
    sender_filter_label = str(msg.get('sender_filter_label') or '').strip()
    custom_name = str(msg.get('custom_name') or '').strip()
    chat_title = str(msg.get('chat_title') or '').strip()

    if msg_type_value.startswith('excel_'):
        for label in (custom_name, chat_title, sender_filter_label, source_display_name):
            label = str(label or '').strip()
            if not label or label == 'Неизвестно' or label.lower().startswith('лист:'):
                continue
            if ' · Лист:' in label:
                label = label.split(' · Лист:', 1)[0].strip()
            if label:
                return {label}
        return set()

    return {
        label for label in (source_display_name, sender_filter_label)
        if label and label != 'Неизвестно'
    }


def load_active_google_sheet_sender_options(db, user_id):
    rows = db.execute(
        """
        SELECT
            ec.sheet_name,
            COALESCE(tc.custom_name, tc.chat_title, CAST(ec.chat_id AS TEXT)) AS source_name,
            (
                SELECT m.sender_name
                FROM messages m
                WHERE m.user_id = ec.user_id
                  AND CAST(m.chat_id AS TEXT) = CAST(ec.chat_id AS TEXT)
                  AND m.type = ('excel_' || ec.id)
                ORDER BY m.date DESC, m.id DESC
                LIMIT 1
            ) AS db_sender_name
        FROM excel_configs ec
        LEFT JOIN tracked_chats tc
               ON CAST(ec.chat_id AS TEXT) = CAST(tc.chat_id AS TEXT)
              AND ec.user_id = tc.user_id
        WHERE ec.user_id = ?
          AND LOWER(TRIM(COALESCE(ec.source_type, ''))) = 'google_sheet'
        ORDER BY ec.id DESC
        """,
        (user_id,),
    ).fetchall()

    labels = set()
    for row in rows:
        source_name = str(row['source_name'] or '').strip()
        if not source_name or source_name == 'Неизвестно' or source_name.startswith('gs_supplier_'):
            continue
        labels.add(source_name)

    return labels


def connected_message_source_clause(message_alias="m"):
    return f"""
        (
            EXISTS (
                SELECT 1
                FROM tracked_chats tc_allowed
                WHERE tc_allowed.user_id = {message_alias}.user_id
                  AND CAST(tc_allowed.chat_id AS TEXT) = CAST({message_alias}.chat_id AS TEXT)
            )
            OR (
                CAST({message_alias}.chat_id AS TEXT) LIKE 'api_src_%'
                AND EXISTS (
                    SELECT 1
                    FROM api_sources api_allowed
                    WHERE api_allowed.user_id = {message_alias}.user_id
                      AND api_allowed.is_active = 1
                      AND CAST({message_alias}.chat_id AS TEXT) = ('api_src_' || api_allowed.id)
                )
            )
            OR (
                {message_alias}.type LIKE 'excel_%'
                AND EXISTS (
                    SELECT 1
                    FROM excel_configs ec_allowed
                    WHERE ec_allowed.user_id = {message_alias}.user_id
                      AND CAST(ec_allowed.chat_id AS TEXT) = CAST({message_alias}.chat_id AS TEXT)
                      AND {message_alias}.type = ('excel_' || ec_allowed.id)
                )
            )
        )
    """


def purge_unconnected_source_messages(db, user_id):
    message_alias = "m_purge"
    stale_rows = db.execute(
        f"""
        SELECT {message_alias}.id
        FROM messages {message_alias}
        WHERE {message_alias}.user_id = ?
          AND NOT {connected_message_source_clause(message_alias)}
        """,
        (user_id,),
    ).fetchall()
    stale_ids = [int(row['id']) for row in stale_rows]
    if not stale_ids:
        return 0

    def writer():
        removed = 0
        for offset in range(0, len(stale_ids), 500):
            chunk = stale_ids[offset:offset + 500]
            placeholders = ",".join("?" for _ in chunk)
            if table_exists(db, "saved_messages"):
                db.execute(f"DELETE FROM saved_messages WHERE message_id IN ({placeholders})", chunk)
            db.execute(f"DELETE FROM product_messages WHERE message_id IN ({placeholders})", chunk)
            db.execute(f"DELETE FROM messages WHERE id IN ({placeholders})", chunk)
            removed += len(chunk)
        return removed

    removed = run_db_write_transaction(db, writer, retries=3, retry_delay=0.5, use_lock=True)
    logger.info("Удалены сообщения неподключённых источников: %s", removed)
    return removed


def load_grouped_messages_for_filters(db, user_id, target_where_str, target_params, *, dedupe_by_first_product=True):
    data_query = f"""
        SELECT m.*, tc.custom_name
        FROM messages m
        LEFT JOIN tracked_chats tc
               ON CAST(m.chat_id AS TEXT) = CAST(tc.chat_id AS TEXT)
              AND m.user_id = tc.user_id
        WHERE {target_where_str}
        ORDER BY m.date DESC
        LIMIT 2000
    """
    cursor = db.execute(data_query, target_params)
    all_messages = [dict(row) for row in cursor.fetchall()]

    if not dedupe_by_first_product:
        return sort_message_rows_preserving_bot_order(all_messages, db, user_id)

    grouped_messages = []
    seen_products_per_seller = {}
    seller_msg_count = {}

    def clean_line(line):
        line = re.sub(r'\b\d{1,3}(?:[ .,]\d{3})+\b|\b\d{4,}\b', '', str(line or '').lower())
        line = re.sub(r'[₽$€рRkKкК]', '', line)
        return line.strip()

    for msg in all_messages:
        source_bucket_key = get_message_source_bucket_key(msg)
        if source_bucket_key not in seen_products_per_seller:
            seen_products_per_seller[source_bucket_key] = set()
            seller_msg_count[source_bucket_key] = 0

        seller_msg_count[source_bucket_key] += 1
        if seller_msg_count[source_bucket_key] > 1000:
            continue

        text = msg['text'] if msg['text'] else ''
        lines = [line.strip() for line in text.split('\n') if line.strip()]
        if not lines:
            continue

        first_product = ""
        for line in lines:
            cleaned = clean_line(line)
            if len(cleaned) > 3:
                first_product = cleaned
                break

        if first_product and first_product in seen_products_per_seller[source_bucket_key]:
            continue

        grouped_messages.append(msg)
        for line in lines:
            cleaned = clean_line(line)
            if len(cleaned) > 3:
                seen_products_per_seller[source_bucket_key].add(cleaned)

    return sort_message_rows_preserving_bot_order(grouped_messages, db, user_id)


def build_messages_payload(db, user_id, args, *, paginate=True):
    try:
        purge_unconnected_source_messages(db, user_id)
    except sqlite3.OperationalError as e:
        logger.warning("Не удалось удалить сообщения неподключённых источников перед выдачей: %s", e)

    msg_type = args.get('type', 'all')
    search = args.get('search', '').lower()
    exclude = args.get('exclude', '').lower()
    sender = args.get('sender', '')
    msg_source_type = args.get('msg_type', '')

    try:
        page = int(args.get('page', 1))
        if page < 1:
            page = 1
    except (TypeError, ValueError):
        page = 1

    try:
        limit = int(args.get('limit', 10))
        if limit < 5:
            limit = 5
        elif limit > 100:
            limit = 100
    except (TypeError, ValueError):
        limit = 10

    offset = (page - 1) * limit

    where_clauses = [
        "m.user_id = ?",
        "(m.is_blocked IS NULL OR m.is_blocked = 0)",
        "(m.is_delayed IS NULL OR m.is_delayed = 0)",
        connected_message_source_clause("m"),
    ]
    params = [user_id]

    if msg_type == 'favorites':
        where_clauses.append("m.is_favorite = 1")
    elif str(msg_type).startswith('catalog_folder_'):
        try:
            catalog_folder_id = int(str(msg_type).replace('catalog_folder_', '', 1))
        except (TypeError, ValueError):
            catalog_folder_id = 0

        folder_rows = []
        if catalog_folder_id:
            folder_rows = db.execute(
                """
                WITH RECURSIVE folder_tree(id) AS (
                    SELECT id
                    FROM folders
                    WHERE id = ? AND user_id = ?
                    UNION ALL
                    SELECT f.id
                    FROM folders f
                    JOIN folder_tree ft ON f.parent_id = ft.id
                    WHERE f.user_id = ?
                )
                SELECT id FROM folder_tree
                """,
                (catalog_folder_id, user_id, user_id),
            ).fetchall()

        catalog_folder_ids = [int(row['id']) for row in folder_rows if row['id'] is not None]
        if catalog_folder_ids:
            placeholders = ','.join('?' for _ in catalog_folder_ids)
            where_clauses.append(f"""
                EXISTS (
                    SELECT 1
                    FROM product_messages pm
                    JOIN products p ON p.id = pm.product_id
                    WHERE pm.message_id = m.id
                      AND pm.status = 'confirmed'
                      AND p.user_id = ?
                      AND p.folder_id IN ({placeholders})
                )
            """)
            params.extend([user_id, *catalog_folder_ids])
        else:
            where_clauses.append("1 = 0")
    elif str(msg_type).startswith('folder_'):
        folder_id = str(msg_type).replace('folder_', '')
        where_clauses.append("m.id IN (SELECT message_id FROM saved_messages WHERE folder_id = ?)")
        params.append(folder_id)
    elif str(msg_type).startswith('chat_'):
        chat_id = str(msg_type).replace('chat_', '')
        where_clauses.append("m.chat_id = ?")
        params.append(chat_id)

    search_terms = [term.strip() for term in re.split(r"[\s|,]+", search) if term.strip()]
    for term in search_terms:
        where_clauses.append("LOWER(m.text) LIKE ?")
        params.append(f"%{term}%")

    exclude_terms = [term.strip() for term in re.split(r"[|,]+", exclude) if term.strip()]
    if not exclude_terms and exclude:
        exclude_terms = [exclude.strip()]
    for term in exclude_terms:
        where_clauses.append("LOWER(m.text) NOT LIKE ?")
        params.append(f"%{term}%")

    base_where_clauses = list(where_clauses)
    base_params = list(params)

    if msg_source_type and msg_source_type != 'all':
        where_clauses.append("m.type = ?")
        params.append(msg_source_type)

    dedupe_messages = paginate and not (sender and sender != 'all')
    grouped_messages = load_grouped_messages_for_filters(
        db,
        user_id,
        " AND ".join(where_clauses),
        params,
        dedupe_by_first_product=dedupe_messages,
    )
    sender_source_messages = load_grouped_messages_for_filters(db, user_id, " AND ".join(base_where_clauses), base_params)

    grouped_messages = [enrich_message_row(msg) for msg in grouped_messages]
    sender_source_messages = [enrich_message_row(msg) for msg in sender_source_messages]

    if sender and sender != 'all':
        grouped_messages = [msg for msg in grouped_messages if message_matches_sender_filter(msg, sender)]

    sender_options = set()
    for msg in sender_source_messages:
        sender_options.update(get_message_sender_option_labels(msg))
    sender_options.update(load_active_google_sheet_sender_options(db, user_id))
    sender_options = sorted(sender_options)

    total_count = len(grouped_messages)
    total_pages = math.ceil(total_count / limit) if total_count > 0 else 1
    messages_page = grouped_messages[offset: offset + limit] if paginate else grouped_messages
    attach_message_linked_lines(db, user_id, messages_page)

    return {
        'messages': messages_page,
        'senders': sender_options,
        'total': total_count,
        'page': page,
        'limit': limit,
        'total_pages': total_pages,
    }


@app.route('/api/messages')
def get_messages():
    if 'user_id' not in session:
        return jsonify({'messages': [], 'error': 'Not authorized'}), 401

    user_id = session['user_id']
    db = get_db()

    try:
        return jsonify(build_messages_payload(db, user_id, request.args, paginate=True))
    except Exception as e:
        page = request.args.get('page', 1)
        limit = request.args.get('limit', 10)
        print(f"Ошибка в API сообщений: {e}")
        return jsonify({'messages': [], 'senders': [], 'total': 0, 'page': page, 'limit': limit, 'total_pages': 1, 'error': str(e)})


def build_telegram_message_url(msg):
    msg_type = str(msg.get('type') or '').strip().lower()
    telegram_message_id = msg.get('telegram_message_id')
    chat_id = str(msg.get('chat_id') or '').strip()
    if telegram_message_id and msg_type in {'channel', 'group'} and chat_id:
        clean_chat_id = re.sub(r'^-100', '', chat_id)
        clean_chat_id = clean_chat_id.lstrip('-')
        if clean_chat_id:
            return f"https://t.me/c/{clean_chat_id}/{telegram_message_id}"
    return ''


def normalize_excel_sheet_name(raw_name, used_names):
    base = re.sub(r'[\[\]:*?/\\\\]', ' ', str(raw_name or '').strip()) or 'Лист'
    base = re.sub(r'\s+', ' ', base).strip()[:31] or 'Лист'
    candidate = base
    suffix = 2
    while candidate in used_names:
        tail = f" {suffix}"
        candidate = f"{base[:31-len(tail)]}{tail}".strip()
        suffix += 1
    used_names.add(candidate)
    return candidate


def build_message_export_rows(msg):
    sender_label = str(msg.get('source_display_name') or get_message_source_display_label(msg)).strip()
    telegram_link = build_telegram_message_url(msg)
    base = {
        'Дата': msg.get('date') or '',
        'Тип': msg.get('type') or '',
        'Поставщик': sender_label,
        'Чат': msg.get('chat_title') or '',
        'Имя в панели': msg.get('custom_name') or '',
        'Ссылка Telegram': telegram_link,
        'ID в БД': msg.get('id') or '',
        'Telegram message id': msg.get('telegram_message_id') or '',
        'Chat ID': msg.get('chat_id') or '',
    }

    raw_text = normalize_incoming_text(msg.get('text') or '')
    lines = [str(line or '').strip() for line in raw_text.split('\n') if str(line or '').strip()]
    if not lines:
        return [{
            **base,
            'Раздел': '',
            'Строка прайса': '',
            'Контекст': '',
            'Цена': '',
            'Сообщение целиком': '',
        }]

    rows = []
    context_headers = []
    in_discount_section = False
    price_line_found = False

    for raw_line in lines:
        line = raw_line.strip()
        if not line:
            continue
        if line_is_discount_section_marker(line):
            in_discount_section = True
            context_headers = update_supplier_context_headers(context_headers, line)
            continue

        if in_discount_section or line_has_discount_condition_marker(line):
            continue

        if supplier_line_has_price_marker(line) and extract_price(line):
            clean_line = strip_supplier_line_noise(line)
            contextual_line = make_supplier_contextual_line(context_headers, clean_line)
            current_section = context_headers[-1] if context_headers else ''
            rows.append({
                **base,
                'Раздел': current_section,
                'Строка прайса': clean_line,
                'Контекст': contextual_line if contextual_line != clean_line else '',
                'Цена': extract_price(line) or '',
                'Сообщение целиком': raw_text,
            })
            price_line_found = True
            continue

        context_headers = update_supplier_context_headers(context_headers, line)

    if price_line_found:
        return rows

    return [{
        **base,
        'Раздел': context_headers[-1] if context_headers else '',
        'Строка прайса': lines[0][:500],
        'Контекст': '',
        'Цена': '',
        'Сообщение целиком': raw_text,
    }]


def style_export_worksheet(worksheet):
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    worksheet.freeze_panes = 'A2'
    worksheet.auto_filter.ref = worksheet.dimensions

    widths = {
        'A': 18,  # Дата
        'B': 12,  # Тип
        'C': 20,  # Поставщик
        'D': 22,  # Чат
        'E': 22,  # Имя в панели
        'F': 24,  # Раздел
        'G': 58,  # Строка прайса
        'H': 58,  # Контекст
        'I': 14,  # Цена
        'J': 40,  # Ссылка
        'K': 12,  # id db
        'L': 18,  # tg id
        'M': 18,  # chat id
        'N': 18,  # raw
    }

    header_font = Font(bold=True)
    wrap_alignment = Alignment(vertical='top', wrap_text=True)
    top_alignment = Alignment(vertical='top')

    for cell in worksheet[1]:
        cell.font = header_font
        cell.alignment = wrap_alignment

    max_row = worksheet.max_row
    max_col = worksheet.max_column
    for col_idx in range(1, max_col + 1):
        column_letter = get_column_letter(col_idx)
        worksheet.column_dimensions[column_letter].width = widths.get(column_letter, 18)

    for row in worksheet.iter_rows(min_row=2, max_row=max_row):
        worksheet.row_dimensions[row[0].row].height = 34
        for cell in row:
            if cell.column in (7, 8, 10, 14):
                cell.alignment = wrap_alignment
            else:
                cell.alignment = top_alignment


@app.route('/api/messages/export.xlsx')
@login_required
def export_messages_excel():
    user_id = session['user_id']
    db = get_db()
    payload = build_messages_payload(db, user_id, request.args, paginate=False)
    messages = payload.get('messages') or []
    sender_filter = str(request.args.get('sender', 'all') or 'all').strip()

    grouped_rows = {}
    for msg in messages:
        sender_label = get_sender_filter_label(msg)
        grouped_rows.setdefault(sender_label, []).extend(build_message_export_rows(msg))

    buffer = BytesIO()
    used_sheet_names = set()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        if not grouped_rows:
            pd.DataFrame([{'Статус': 'Нет сообщений по выбранным фильтрам'}]).to_excel(writer, sheet_name='Пусто', index=False)
        else:
            sender_names = sorted(grouped_rows)
            for sender_name in sender_names:
                rows = grouped_rows[sender_name]
                df = pd.DataFrame(rows)
                sheet_name = normalize_excel_sheet_name(sender_name, used_sheet_names)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                worksheet = writer.sheets[sheet_name]
                style_export_worksheet(worksheet)

    buffer.seek(0)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    if sender_filter and sender_filter != 'all':
        safe_sender = re.sub(r'[^A-Za-z0-9А-Яа-я._-]+', '_', sender_filter).strip('_') or 'sender'
        filename = f"messages_{safe_sender}_{timestamp}.xlsx"
    else:
        filename = f"messages__{timestamp}.xlsx"

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# --- МАРШРУТЫ ДЛЯ АВТОПУБЛИКАЦИЙ ---

@app.route('/api/publications', methods=['GET'])
@login_required
def get_publications():
    db = get_db()
    user_id = session['user_id'] # Получаем ID текущего пользователя
    pubs = db.execute("SELECT * FROM publications WHERE user_id = ?", (user_id,)).fetchall() # Фильтруем
    return jsonify([dict(p) for p in pubs])

@app.route('/api/publications', methods=['POST'])
@login_required
def save_publication():
    data = request.json
    db = get_db()
    pub_id = data.get('id')
    
    allowed_items_json = json.dumps(data.get('allowed_items', {}))
    markup_type = data.get('markup_type', 'percent')
    markup_value = float(data.get('markup_value', 0))
    rounding = int(data.get('rounding', 100))
    
    if pub_id:
        db.execute("""
            UPDATE publications 
            SET name=?, is_active=?, interval_min=?, chat_id=?, message_id=?, userbot_id=?, template=?, allowed_items=?, markup_type=?, markup_value=?, rounding=?
            WHERE id=?
        """, (data.get('name'), data.get('is_active', 0), data.get('interval_min', 60), 
              data.get('chat_id'), data.get('message_id'), data.get('userbot_id'), 
              data.get('template'), allowed_items_json, markup_type, markup_value, rounding, pub_id))
    else:
        db.execute("""
    INSERT INTO publications (name, is_active, interval_min, chat_id, message_id, userbot_id, template, allowed_items, markup_type, markup_value, rounding, user_id)
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (data.get('name'), data.get('is_active', 0), data.get('interval_min', 60), 
              data.get('chat_id'), data.get('message_id'), data.get('userbot_id'), 
              data.get('template'), allowed_items_json, markup_type, markup_value, rounding,session['user_id']))
        # Получаем ID только что созданной публикации
        pub_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        
    # --- СОХРАНЕНИЕ НАЦЕНОК, КОТОРЫЕ БЫЛИ ДОБАВЛЕНЫ ДО СОЗДАНИЯ ПУБЛИКАЦИИ ---
    temp_markups = data.get('temp_markups', [])
    for m in temp_markups:
        db.execute("""
            INSERT INTO pub_markups (pub_id, folder_id, markup_type, markup_value, rounding) 
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(pub_id, folder_id) DO UPDATE SET 
            markup_type=excluded.markup_type, markup_value=excluded.markup_value, rounding=excluded.rounding
        """, (pub_id, m['folder_id'], m['markup_type'], m['markup_value'], m['rounding']))
        
    db.commit()
    notify_clients()
    return jsonify({'success': True})

@app.route('/api/publications/<int:pub_id>', methods=['DELETE'])
@login_required
def delete_publication(pub_id):
    db = get_db()
    user_id = session['user_id']
    existing = db.execute(
        "SELECT id FROM publications WHERE id=? AND user_id=?",
        (pub_id, user_id),
    ).fetchone()
    if not existing:
        return jsonify({'error': 'Publication not found'}), 404
    db.execute("DELETE FROM pub_markups WHERE pub_id=?", (pub_id,))
    db.execute("DELETE FROM publications WHERE id=? AND user_id=?", (pub_id, user_id))
    db.commit()
    notify_clients()
    return jsonify({'success': True})
    
@app.route('/api/publications/<int:pub_id>/toggle', methods=['POST'])
@login_required
def toggle_publication(pub_id):
    data = request.json
    db = get_db()
    db.execute("UPDATE publications SET is_active=? WHERE id=?", (data.get('is_active', 0), pub_id))
    db.commit()
    notify_clients()
    return jsonify({'success': True})

import re


def smart_folder_sort_key(folder):
    """
    Умная сортировка:
    1. 'По умолчанию' всегда сверху.
    2. Обычные папки идут по алфавиту.
    3. Папки с гигабайтами и терабайтами идут в конце и сортируются по объему.
    """
    name = folder.get('name', '').strip()
    name_lower = name.lower()
    
    # Папка "По умолчанию" всегда будет самой первой (-1)
    if name_lower == 'по умолчанию':
        return (-1, 0, name)
        
    # Ищем названия, состоящие из цифр (64, 128) и опционально "гб" или "тб"
    match = re.fullmatch(r'(\d+)\s*(тб|tb|гб|gb)?', name_lower)
    
    if match:
        val = int(match.group(1))
        unit = match.group(2)
        
        # Переводим терабайты в гигабайты (1 ТБ = 1024 ГБ) для правильной последовательности
        if unit in ['тб', 'tb']:
            val *= 1024
            
        # (1, ...) отправляет такие папки в конец списка, затем сортирует по числу
        return (1, val, name)
        
    # (0, ...) оставляет обычные папки в начале списка и сортирует по алфавиту
    return (0, 0, name)

def sort_folders_recursive(folders_list):
    """Рекурсивно сортирует каждую папку и её подпапки"""
    folders_list.sort(key=smart_folder_sort_key)
    for f in folders_list:
        if f.get('children'):
            sort_folders_recursive(f['children'])


@app.route('/api/folders/tree')
@login_required
def get_folder_tree():
    user_id = session['user_id']
    db = get_db()
    
    # Извлекаем все папки из БД
    cursor = db.execute("SELECT id, name, parent_id FROM folders WHERE user_id = ?", (user_id,))
    rows = cursor.fetchall()

    direct_product_counts = {
        int(row['folder_id']): int(row['count'] or 0)
        for row in db.execute(
            """
            SELECT folder_id, COUNT(*) AS count
            FROM products
            WHERE user_id = ? AND folder_id IS NOT NULL
            GROUP BY folder_id
            """,
            (user_id,),
        ).fetchall()
        if row['folder_id'] is not None
    }
    direct_connected_counts = {
        int(row['folder_id']): int(row['count'] or 0)
        for row in db.execute(
            """
            SELECT p.folder_id, COUNT(*) AS count
            FROM products p
            WHERE p.user_id = ?
              AND p.folder_id IS NOT NULL
              AND EXISTS (
                  SELECT 1
                  FROM product_messages pm
                  JOIN messages m ON pm.message_id = m.id
                  WHERE pm.product_id = p.id
                    AND pm.status = 'confirmed'
                    AND COALESCE(pm.is_actual, 1) = 1
                    AND m.user_id = ?
              )
            GROUP BY p.folder_id
            """,
            (user_id, user_id),
        ).fetchall()
        if row['folder_id'] is not None
    }
    
    folders = {}
    roots = []
    
    # 1. Заполняем словарь всеми папками
    for row in rows:
        folders[row['id']] = {
            'id': row['id'],
            'name': row['name'],
            'parent_id': row['parent_id'],
            'direct_product_count': direct_product_counts.get(int(row['id']), 0),
            'direct_connected_product_count': direct_connected_counts.get(int(row['id']), 0),
            'product_count': 0,
            'connected_product_count': 0,
            'children': []
        }
        
    # 2. Выстраиваем дерево (кто в какой папке лежит)
    for fid, f in folders.items():
        if f['parent_id'] is None:
            roots.append(f)
        else:
            if f['parent_id'] in folders:
                folders[f['parent_id']]['children'].append(f)
                
    # 3. Применяем умную сортировку ко всему дереву перед отправкой на фронтенд
    sort_folders_recursive(roots)

    def populate_counts(items):
        total_products = 0
        total_connected = 0
        for item in items:
            child_products, child_connected = populate_counts(item.get('children') or [])
            item['product_count'] = int(item.get('direct_product_count') or 0) + child_products
            item['connected_product_count'] = int(item.get('direct_connected_product_count') or 0) + child_connected
            total_products += item['product_count']
            total_connected += item['connected_product_count']
        return total_products, total_connected

    populate_counts(roots)
                
    return jsonify(roots)

@app.route('/api/folders', methods=['POST'])
@login_required
def create_folder():
    user_id = session['user_id']
    data = request.get_json()
    name = data.get('name')
    parent_id = data.get('parent_id')
    if not name:
        return jsonify({'error': 'Name required'}), 400
    db = get_db()
    try:
        db.execute("INSERT INTO folders (user_id, name, parent_id) VALUES (?, ?, ?)",
                   (user_id, name, parent_id))
        db.commit()
        notify_clients()
        return jsonify({'success': True})
    except sqlite3.IntegrityError:
        return jsonify({'error': 'Folder exists'}), 400

@app.route('/api/folders/<int:folder_id>/rename', methods=['POST'])
@login_required
def rename_folder_route(folder_id):
    user_id = session['user_id']
    data = request.get_json()
    new_name = data.get('name')
    
    if not new_name:
        return jsonify({'error': 'Имя не может быть пустым'}), 400
        
    db = get_db()
    
    row = db.execute("SELECT name FROM folders WHERE id = ? AND user_id = ?", (folder_id, user_id)).fetchone()
    if not row:
        return jsonify({'error': 'Папка не найдена'}), 404
    if row['name'] == 'По умолчанию':
        return jsonify({'error': 'Нельзя переименовать папку по умолчанию'}), 400
        
    try:
        db.execute("UPDATE folders SET name = ? WHERE id = ? AND user_id = ?", (new_name, folder_id, user_id))
        db.commit()
        notify_clients() # Уведомляем клиентов для автообновления
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Ошибка переименования: {e}")
        return jsonify({'error': 'Системная ошибка при переименовании'}), 500

@app.route('/api/folders/<int:folder_id>/move', methods=['POST'])
@login_required
def move_folder_route(folder_id):
    user_id = session['user_id']
    data = request.get_json() or {}
    parent_id = data.get('parent_id')

    db = get_db()
    folder = db.execute(
        "SELECT id, name, parent_id FROM folders WHERE id = ? AND user_id = ?",
        (folder_id, user_id)
    ).fetchone()
    if not folder:
        return jsonify({'error': 'Папка не найдена'}), 404
    if folder['name'] == 'По умолчанию':
        return jsonify({'error': 'Нельзя перемещать папку по умолчанию'}), 400

    if parent_id in ('', None):
        parent_id = None
    else:
        try:
            parent_id = int(parent_id)
        except (TypeError, ValueError):
            return jsonify({'error': 'Некорректная целевая папка'}), 400

    if parent_id == folder_id:
        return jsonify({'error': 'Нельзя переместить папку в саму себя'}), 400

    if parent_id is not None:
        target = db.execute(
            "SELECT id FROM folders WHERE id = ? AND user_id = ?",
            (parent_id, user_id)
        ).fetchone()
        if not target:
            return jsonify({'error': 'Целевая папка не найдена'}), 404

        folders_raw = db.execute(
            "SELECT id, parent_id FROM folders WHERE user_id = ?",
            (user_id,)
        ).fetchall()
        parent_map = {row['id']: row['parent_id'] for row in folders_raw}
        cursor = parent_id
        while cursor is not None:
            if cursor == folder_id:
                return jsonify({'error': 'Нельзя переместить папку внутрь самой себя или своей подпапки'}), 400
            cursor = parent_map.get(cursor)

    db.execute(
        "UPDATE folders SET parent_id = ? WHERE id = ? AND user_id = ?",
        (parent_id, folder_id, user_id)
    )
    db.commit()
    notify_clients()
    return jsonify({'success': True})

@app.route('/api/folders/<int:folder_id>', methods=['DELETE'])
@login_required
def delete_folder(folder_id):
    user_id = session['user_id']
    db = get_db()
    row = db.execute("SELECT name FROM folders WHERE id = ? AND user_id = ?", (folder_id, user_id)).fetchone()
    if not row:
        return jsonify({'error': 'Folder not found'}), 404
    if row['name'] == 'По умолчанию':
        return jsonify({'error': 'Cannot delete default folder'}), 400
        
    db.execute("DELETE FROM folders WHERE id = ?", (folder_id,))
    db.commit()
    notify_clients()
    return jsonify({'success': True})

@app.route('/api/folder/<int:folder_id>/messages')
@login_required
def get_folder_messages(folder_id):
    user_id = session['user_id']
    db = get_db()
    cursor = db.execute("""
        SELECT m.*, tc.custom_name, tc.chat_title as tc_chat_title
        FROM messages m
        JOIN saved_messages sm ON m.id = sm.message_id
        LEFT JOIN tracked_chats tc ON m.chat_id = tc.chat_id AND m.user_id = tc.user_id
        WHERE sm.folder_id = ? AND m.user_id = ?
        AND (m.is_blocked IS NULL OR m.is_blocked = 0)
        ORDER BY sm.saved_date DESC
    """, (folder_id, user_id))
    # ...
    rows = cursor.fetchall()
    messages = [{
        'id': row['id'],
        'telegram_message_id': row['telegram_message_id'], 
        'chat_id': row['chat_id'],                         
        'type': row['type'],
        'text': row['text'],
        'date': row['date'],
        'chat_title': row['chat_title'],
        'sender_name': row['sender_name'] if 'sender_name' in row.keys() else 'Unknown',
        'custom_name': row['custom_name'] if 'custom_name' in row.keys() else None,
        'tc_chat_title': row['tc_chat_title'] if 'tc_chat_title' in row.keys() else None # <--- ДОБАВИТЬ ЭТУ СТРОКУ
    } for row in rows]
    messages = [enrich_message_row(row) for row in messages]
    return jsonify(messages)


@app.route('/api/save_messages', methods=['POST'])
@login_required
def save_messages():
    user_id = session['user_id']
    data = request.get_json()
    message_ids = data.get('message_ids', [])
    folder_id = data.get('folder_id')
    if not message_ids or not folder_id:
        return jsonify({'error': 'Missing data'}), 400
    db = get_db()
    cursor = db.execute("SELECT id FROM folders WHERE id = ? AND user_id = ?", (folder_id, user_id))
    if not cursor.fetchone():
        return jsonify({'error': 'Folder not found'}), 404
    for mid in message_ids:
        try:
            db.execute("INSERT INTO saved_messages (message_id, folder_id) VALUES (?, ?)", (mid, folder_id))
        except sqlite3.IntegrityError:
            pass
    db.commit()
    notify_clients()
    return jsonify({'success': True})

@app.route('/api/userbots', methods=['GET'])
@login_required
def get_userbots():
    user_id = session['user_id']
    db = get_db()
    # ДОБАВИЛИ time_start, time_end, schedule_enabled в SQL-запрос
    try:
        cursor = db.execute("SELECT id, api_id, api_hash, status, account_name, time_start, time_end, schedule_enabled FROM user_telegram_sessions WHERE user_id=?", (user_id,))
    except sqlite3.OperationalError:
        cursor = db.execute("SELECT id, api_id, api_hash, status, time_start, time_end, schedule_enabled FROM user_telegram_sessions WHERE user_id=?", (user_id,))
        
    bots = [dict(row) for row in cursor.fetchall()]
    return jsonify(bots)

# --- НОВЫЙ МАРШРУТ ДЛЯ СОХРАНЕНИЯ РАСПИСАНИЯ ---
@app.route('/api/userbots/<int:bot_id>/schedule', methods=['POST'])
@login_required
def update_userbot_schedule(bot_id):
    user_id = session['user_id']
    data = request.get_json(silent=True) or request.form.to_dict() or {}
    raw_enabled = data.get('schedule_enabled', data.get('enabled', 0))
    if isinstance(raw_enabled, str):
        schedule_enabled = 1 if raw_enabled.strip().lower() in {'1', 'true', 'yes', 'on', 'checked'} else 0
    else:
        schedule_enabled = 1 if raw_enabled else 0
    time_start = str(data.get('time_start') or '00:00').strip()[:5]
    time_end = str(data.get('time_end') or '23:59').strip()[:5]
    try:
        datetime.strptime(time_start, '%H:%M')
        datetime.strptime(time_end, '%H:%M')
    except ValueError:
        return jsonify({'success': False, 'error': 'Некорректное время расписания'}), 400
    db = get_db()
    cur = db.execute("""
        UPDATE user_telegram_sessions 
        SET time_start = ?, time_end = ?, schedule_enabled = ? 
        WHERE id = ? AND user_id = ?
    """, (time_start, time_end, schedule_enabled, bot_id, user_id))
    if cur.rowcount == 0:
        return jsonify({'success': False, 'error': 'Юзербот не найден'}), 404
    db.commit()
    notify_clients()
    return jsonify({
        'success': True,
        'id': bot_id,
        'schedule_enabled': schedule_enabled,
        'time_start': time_start,
        'time_end': time_end,
    })
# -----------------------------------------------

@app.route('/api/userbots/<int:bot_id>/toggle', methods=['POST'])
@login_required
def toggle_userbot(bot_id):
    user_id = session['user_id']
    db = get_db()
    cursor = db.execute("SELECT user_id, api_id, api_hash, status FROM user_telegram_sessions WHERE id=?", (bot_id,))
    row = cursor.fetchone()
    if not row or row['user_id'] != user_id:
        return jsonify({'error': 'Not found'}), 404

    if row['status'] == 'active':
        stop_message_listener(bot_id)
        new_status = 'inactive'
    else:
        start_message_listener(bot_id, user_id, row['api_id'], row['api_hash'])
        new_status = 'active'

    db.execute("UPDATE user_telegram_sessions SET status = ? WHERE id = ?", (new_status, bot_id))
    db.commit()
    notify_clients()
    return jsonify({'success': True, 'status': new_status})

@app.route('/api/userbots/<int:bot_id>', methods=['DELETE'])
@login_required
def delete_userbot(bot_id):
    user_id = session['user_id']
    db = get_db()
    cursor = db.execute("SELECT user_id, session_file FROM user_telegram_sessions WHERE id=?", (bot_id,))
    row = cursor.fetchone()
    if not row or row['user_id'] != user_id:
        return jsonify({'error': 'Not found'}), 404

    if bot_id in user_clients:
        stop_message_listener(bot_id)

    session_file = row['session_file']
    if session_file and os.path.exists(session_file):
        os.remove(session_file)

    interaction_rows = db.execute(
        "SELECT * FROM interaction_bots WHERE user_id=? AND userbot_id=?",
        (user_id, bot_id)
    ).fetchall()
    tracked_ids = get_interaction_bot_tracked_chat_ids(db, user_id, interaction_rows)
    delete_tracked_chats_by_ids(db, user_id, tracked_ids)
    db.execute("DELETE FROM interaction_bots WHERE user_id=? AND userbot_id=?", (user_id, bot_id))
    db.execute("DELETE FROM user_telegram_sessions WHERE id = ?", (bot_id,))

    remaining_accounts = db.execute(
        "SELECT COUNT(*) FROM user_telegram_sessions WHERE user_id=?",
        (user_id,)
    ).fetchone()[0]
    if not remaining_accounts:
        db.execute("DELETE FROM interaction_bots WHERE user_id=?", (user_id,))
        delete_all_tracked_chats_for_user(db, user_id)

    db.commit()
    notify_clients()
    return jsonify({'success': True})



@app.route('/api/excel/missing_sheets', methods=['GET'])
@login_required
def get_missing_sheets():
    db = get_db()
    missing = db.execute("SELECT * FROM excel_missing_sheets WHERE user_id = ?", (session['user_id'],)).fetchall()
    return jsonify([dict(m) for m in missing])

@app.route('/api/excel/missing_sheets/<int:m_id>', methods=['DELETE'])
@login_required
def delete_missing_sheet(m_id):
    db = get_db()
    db.execute("DELETE FROM excel_missing_sheets WHERE id = ? AND user_id = ?", (m_id, session['user_id']))
    db.commit()
    return jsonify({'success': True})


@app.route('/api/source/missing_positions', methods=['GET'])
@login_required
def get_source_missing_positions():
    db = get_db()
    ensure_supplier_missing_positions_table(db)
    include_resolved = request.args.get('include_resolved', '').strip().lower() in {'1', 'true', 'yes'}
    limit_raw = request.args.get('limit', '').strip()
    try:
        limit = max(1, min(1000, int(limit_raw))) if limit_raw else 200
    except ValueError:
        limit = 200

    if include_resolved:
        rows = db.execute(
            """
            SELECT *
            FROM supplier_missing_positions
            WHERE user_id = ?
            ORDER BY
                CASE WHEN COALESCE(status, '') = 'resolved' THEN 1 ELSE 0 END,
                COALESCE(seen_count, 0) DESC,
                COALESCE(last_seen_at, '') DESC,
                id DESC
            LIMIT ?
            """,
            (session['user_id'], limit),
        ).fetchall()
    else:
        rows = db.execute(
            """
            SELECT *
            FROM supplier_missing_positions
            WHERE user_id = ?
              AND COALESCE(status, '') != 'resolved'
            ORDER BY COALESCE(seen_count, 0) DESC, COALESCE(last_seen_at, '') DESC, id DESC
            LIMIT ?
            """,
            (session['user_id'], limit),
        ).fetchall()
    return jsonify([dict(row) for row in rows])

@app.route('/api/excel/parse_latest/<chat_id>', methods=['POST'])
@login_required
def parse_latest_excel_route(chat_id):
    chat_id = int(chat_id)
    user_id = session['user_id']    
    
    # 1. Собираем ВСЕ активные сессии текущего пользователя
    active_sessions = [sid for sid, (_, _, uid) in user_clients.items() if uid == user_id]
    if not active_sessions:
        return jsonify({'error': 'Юзерботы не запущены. Включите хотя бы один аккаунт во вкладке "Подключение".'}), 400

    def background_parser():
        with app.app_context():
            db = get_db()
            tc = db.execute("SELECT chat_title, custom_name FROM tracked_chats WHERE chat_id = ? AND user_id = ?", (chat_id, user_id)).fetchone()
            chat_title = tc['custom_name'] if tc and tc['custom_name'] else (tc['chat_title'] if tc else str(chat_id))

        success = False
        
        # 2. ПЕРЕБИРАЕМ ВСЕ АККАУНТЫ ПО ОЧЕРЕДИ
        for sid in active_sessions:
            if sid not in user_clients: continue
            thread, client, uid = user_clients[sid]
            client_loop = user_loops.get(sid)
            
            if not client_loop: continue

            async def try_parse_for_account():
                try:
                    # Пытаемся получить доступ к каналу
                    entity = await client.get_entity(chat_id)
                    
                    # Ищем последние сообщения
                    async for message in client.iter_messages(entity, limit=100):
                        if message.document and getattr(message.file, 'ext', '').lower() in ['.xlsx', '.xls', '.pdf']:
                            
                            parsed_text = await parse_excel_message(client, message, chat_id, user_id)
                            
                            if parsed_text:
                                msg_type = 'channel' if message.is_channel else 'group' if message.is_group else 'private'
                                sender = await message.get_sender()
                                sender_name = f"{getattr(sender, 'first_name', '')} {getattr(sender, 'last_name', '')}".strip() or "Unknown"
                                
                                from datetime import timedelta
                                msg_date = (message.date + timedelta(hours=3)).strftime("%Y-%m-%d %H:%M:%S")

                                # 3. Записываем в БД для вывода в СКЛАД
                                with app.app_context():
                                    db = get_db()
                                    try:
                                        db.execute(
                                            'INSERT INTO messages (user_id, telegram_message_id, type, text, date, chat_id, chat_title, sender_name) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
                                            (user_id, message.id, msg_type, parsed_text, msg_date, chat_id, chat_title, sender_name)
                                        )
                                    except sqlite3.IntegrityError:
                                        db.execute(
                                            "UPDATE messages SET text = ? WHERE telegram_message_id = ? AND user_id = ?", 
                                            (parsed_text, message.id, user_id)
                                        )
                                    db.commit()
                                    # Обновляем Склад в реальном времени
                                    notify_clients()
                                    logger.info(f"✅ Успешно спарсили файл из {chat_title} через аккаунт сессии {sid}")
                                return True # Сообщаем об успехе
                    return False # Файл не найден в последних 100 сообщениях
                except Exception as e:
                    # У этого аккаунта нет доступа к каналу или произошла другая ошибка
                    logger.warning(f"Аккаунт (сессия {sid}) не имеет доступа к чату {chat_id} или произошла ошибка: {e}")
                    return False

            try:
                # Передаем задачу в цикл событий конкретного Telethon-клиента
                future = asyncio.run_coroutine_threadsafe(try_parse_for_account(), client_loop)
                is_parsed = future.result(timeout=60)
                
                if is_parsed:
                    success = True
                    break # 🔥 КАК ТОЛЬКО НАШЛИ И СПАРСИЛИ — ОСТАНАВЛИВАЕМ ПЕРЕБОР
                    
            except Exception as e:
                logger.error(f"Таймаут или критическая ошибка при попытке парсинга через аккаунт {sid}: {e}")
                continue # Пробуем следующий аккаунт

        if not success:
            logger.error(f"❌ Ни один из {len(active_sessions)} аккаунтов не смог спарсить файл для {chat_id}. Возможно, нет доступа или файла нет в последних 100 сообщениях.")

    # Запускаем логику перебора в отдельном потоке, чтобы не подвешивать сервер Flask
    threading.Thread(target=background_parser, daemon=True).start()
    
    return jsonify({'success': True})


@app.route('/api/tracked_chats', methods=['GET'])
@login_required
def get_tracked_chats():
    user_id = session['user_id']
    db = get_db()
    # ДОБАВЛЕНО: custom_name в выборку
    cursor = db.execute("SELECT id, chat_id, chat_title, custom_name FROM tracked_chats WHERE user_id=?", (user_id,))
    chats = [dict(row) for row in cursor.fetchall()]
    return jsonify(chats)

@app.route('/api/tracked_chats', methods=['POST'])
@login_required
def add_tracked_chat():
    data = request.get_json()
    raw_chat_id = data.get('chat_id')
    custom_name = data.get('custom_name') # Получаем кастомное название
    user_id = session['user_id']
  
    if not raw_chat_id:
        return jsonify({'error': 'chat_id required'}), 400

    chat_title = custom_name # Изначально ставим кастомное имя (если оно пустое, ниже получим из ТГ)
    numeric_chat_id = None
    client_to_use = None
    client_loop_to_use = None

    try:
        entity_query = int(raw_chat_id)
    except ValueError:
        entity_query = raw_chat_id

    # Ищем активную сессию юзербота для текущего пользователя
    # Ищем активные сессии юзерботов для текущего пользователя
    active_sessions = [sid for sid, (_, _, uid) in user_clients.items() if uid == user_id]
    
    # ПЕРЕБИРАЕМ ВСЕХ БОТОВ, ПОКА НЕ НАЙДЕМ ТОГО, У КОТОРОГО ЕСТЬ ДОСТУП
    for sid in active_sessions:
        thread, client, uid = user_clients[sid]
        client_loop = user_loops.get(sid)
        
        async def fetch_chat():
            return await client.get_entity(entity_query)
            
        if client_loop:
            try:
                future = asyncio.run_coroutine_threadsafe(fetch_chat(), client_loop) 
                entity = future.result(timeout=5)
                numeric_chat_id = entity.id 
                
                # Если доступ есть, запоминаем этого бота как рабочего для данного чата
                client_to_use = client
                client_loop_to_use = client_loop
                
                # Записываем название чата
                if not chat_title:
                    if hasattr(entity, 'title'):
                        chat_title = entity.title
                    else:
                        chat_title = f"{entity.first_name or ''} {entity.last_name or ''}".strip()
                        
                break # Успешно нашли нужного бота, прерываем цикл перебора!
                
            except Exception as e:
                logging.warning(f"Бот сессии {sid} не имеет доступа к чату или произошла ошибка: {e}")
                continue # Пробуем следующего бота
                
                # Если кастомное имя не ввели, берем оригинальное название из Telegram
                if not chat_title:
                    if hasattr(entity, 'title'):
                        chat_title = entity.title
                    else:
                        chat_title = f"{entity.first_name or ''} {entity.last_name or ''}".strip()
            except Exception as e:
                logging.warning(f"Не удалось получить информацию о чате: {e}")

    # Фолбэк, если юзербот не запущен или произошла ошибка
    if not numeric_chat_id:
        try:
            numeric_chat_id = int(raw_chat_id)
        except ValueError:
            return jsonify({'error': 'Запустите юзербота, чтобы добавлять чаты по ссылке, или введите точный числовой ID'}), 400

    # Если название всё ещё пустое (например, юзербот выключен и ввели просто ID без кастомного имени)
    if not chat_title:
        chat_title = str(numeric_chat_id)

    db = get_db()
    try:
        db.execute(
            "INSERT INTO tracked_chats (user_id, chat_id, chat_title, custom_name) VALUES (?, ?, ?, ?)",
            (user_id, numeric_chat_id, chat_title, custom_name)
        )
        db.commit()
        notify_clients()
        
        # Запускаем парсинг истории сообщений в фоне, если юзербот активен
        # Запускаем парсинг истории сообщений в фоне, если юзербот активен
        if client_to_use and client_loop_to_use and numeric_chat_id:
            target_to_parse = entity if 'entity' in locals() else numeric_chat_id
            asyncio.run_coroutine_threadsafe(
                fetch_chat_history(client_to_use, target_to_parse, chat_title, user_id, numeric_chat_id), 
                client_loop_to_use
            )

        return jsonify({'success': True, 'chat_title': chat_title})
    except sqlite3.IntegrityError:
        return jsonify({'error': 'Этот чат уже добавлен в список отслеживаемых'}), 400

@app.route('/api/tracked_chats/<item_id>', methods=['DELETE'])
@login_required
def delete_tracked_chat(item_id):
    try:
        db = get_db()
        user_id = session.get('user_id')
        delete_tracked_chats_by_ids(db, user_id, [item_id])
        db.commit()
        notify_clients()
        return jsonify({'success': True})
    except Exception as e:
        db.rollback()
        print(f"Ошибка при удалении чата: {e}")
        return jsonify({'error': str(e)}), 500

# ---------- Административные маршруты ----------
@app.route('/api/admin/users')
@admin_required
def admin_users():
    db = get_db()
    cursor = db.execute("SELECT id, login, role FROM users")
    users = [{'id': row['id'], 'login': row['login'], 'role': row['role']} for row in cursor.fetchall()]
    return jsonify(users)

@app.route('/api/admin/add_user', methods=['POST'])
@admin_required
def admin_add_user():
    data = request.get_json()
    login = data.get('login')
    password = data.get('password')
    if not login or not password:
        return jsonify({'error': 'Login and password required'}), 400
    pwd_hash = hashlib.sha256(password.encode()).hexdigest()
    db = get_db()
    try:
        db.execute("INSERT INTO users (login, password_hash, role) VALUES (?, ?, 'user')", (login, pwd_hash))
        new_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        db.execute("INSERT INTO folders (user_id, name, parent_id) VALUES (?, 'По умолчанию', NULL)", (new_id,))
        db.commit()
        notify_clients()
        return jsonify({'success': True})
    except sqlite3.IntegrityError:
        return jsonify({'error': 'Login already exists'}), 400


@app.route('/api/admin/users/<int:user_id>/password', methods=['POST'])
@admin_required
def admin_reset_user_password(user_id):
    data = request.get_json() or {}
    password = (data.get('password') or '').strip()
    if not password:
        return jsonify({'error': 'Password required'}), 400

    db = get_db()
    user = db.execute("SELECT id FROM users WHERE id = ?", (user_id,)).fetchone()
    if not user:
        return jsonify({'error': 'User not found'}), 404

    pwd_hash = hashlib.sha256(password.encode()).hexdigest()
    db.execute("UPDATE users SET password_hash = ? WHERE id = ?", (pwd_hash, user_id))
    db.commit()
    return jsonify({'success': True})


@app.route('/api/admin/users/<int:user_id>', methods=['DELETE'])
@admin_required
def admin_delete_user(user_id):
    current_user_id = session.get('user_id')
    if current_user_id == user_id:
        return jsonify({'error': 'Нельзя удалить текущего администратора'}), 400

    db = get_db()
    user = db.execute("SELECT id FROM users WHERE id = ?", (user_id,)).fetchone()
    if not user:
        return jsonify({'error': 'User not found'}), 404

    try:
        db.execute("DELETE FROM saved_messages WHERE message_id IN (SELECT id FROM messages WHERE user_id = ?)", (user_id,))
        db.execute("DELETE FROM saved_messages WHERE folder_id IN (SELECT id FROM folders WHERE user_id = ?)", (user_id,))
        db.execute("DELETE FROM product_messages WHERE message_id IN (SELECT id FROM messages WHERE user_id = ?)", (user_id,))
        db.execute("DELETE FROM product_messages WHERE product_id IN (SELECT id FROM products WHERE user_id = ?)", (user_id,))

        if table_exists(db, "pub_markups"):
            db.execute("DELETE FROM pub_markups WHERE pub_id IN (SELECT id FROM publications WHERE user_id = ?)", (user_id,))
        if table_exists(db, "publication_items"):
            db.execute("DELETE FROM publication_items WHERE publication_id IN (SELECT id FROM publications WHERE user_id = ?)", (user_id,))
        if table_exists(db, "api_markups"):
            db.execute("DELETE FROM api_markups WHERE client_id IN (SELECT id FROM api_clients WHERE user_id = ?)", (user_id,))

        if table_exists(db, "excel_missing_sheets"):
            db.execute("DELETE FROM excel_missing_sheets WHERE user_id = ?", (user_id,))
        if table_exists(db, "excel_configs"):
            db.execute("DELETE FROM excel_configs WHERE user_id = ?", (user_id,))
        if table_exists(db, "interaction_bots"):
            db.execute("DELETE FROM interaction_bots WHERE user_id = ?", (user_id,))
        if table_exists(db, "api_sources"):
            db.execute("DELETE FROM api_sources WHERE user_id = ?", (user_id,))
        if table_exists(db, "api_clients"):
            db.execute("DELETE FROM api_clients WHERE user_id = ?", (user_id,))

        db.execute("DELETE FROM publications WHERE user_id = ?", (user_id,))
        db.execute("DELETE FROM tracked_chats WHERE user_id = ?", (user_id,))
        db.execute("DELETE FROM user_telegram_sessions WHERE user_id = ?", (user_id,))
        db.execute("DELETE FROM products WHERE user_id = ?", (user_id,))
        db.execute("DELETE FROM messages WHERE user_id = ?", (user_id,))
        db.execute("DELETE FROM folders WHERE user_id = ?", (user_id,))
        db.execute("DELETE FROM users WHERE id = ?", (user_id,))
        db.commit()
    except Exception:
        db.rollback()
        raise

    socketio.emit('force_logout')
    notify_clients()
    return jsonify({'success': True})




@app.route('/api/admin/clear_messages', methods=['POST'])
@admin_required
def admin_clear_messages():
    data = request.get_json()
    period = data.get('period', 'all')
    db = get_db()
    
    try:
        if period == 'all':
            db.execute("DELETE FROM messages")
        else:
            days = int(period)
            cutoff_date = datetime.now() - timedelta(days=days)
            cutoff_str = cutoff_date.strftime("%Y-%m-%d %H:%M:%S")
            db.execute("DELETE FROM messages WHERE date < ?", (cutoff_str,))
        
        db.commit()
        notify_clients() 
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Ошибка при очистке сообщений: {e}")
        return jsonify({'error': str(e)}), 500



# ---------- Шаблоны страниц ----------
LOGIN_TEMPLATE = """
<!DOCTYPE html>
<html lang="ru">
<head>
<link rel="icon" type="image/svg+xml" href="/logo.svg">
<script src="https://cdnjs.cloudflare.com/ajax/libs/socket.io/4.0.1/socket.io.js"></script>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>a:store direct</title>
    <style>
        :root {
            --accent-blue: #b8d8f5;
            --panel-border: rgba(150, 180, 214, 0.18);
            --text-soft: #bcc6d2;
            --text-faint: #90a0b4;
            --footer-bg:
                radial-gradient(circle at center top, rgba(104, 107, 114, 0.16) 0%, rgba(39, 41, 47, 0.08) 28%, rgba(12, 13, 17, 0) 58%),
                linear-gradient(90deg, #0a0b0f 0%, #171920 50%, #262932 100%);
        }
        html {
            min-height: 100%;
            background: #0b111a;
        }
        * { box-sizing: border-box; }
        body {
            margin: 0;
            min-height: 100vh;
            font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif;
            color: #e0e0e0;
            background-color: #0b111a;
            background:
                radial-gradient(circle at top center, rgba(72, 106, 156, 0.22) 0%, rgba(22, 35, 54, 0.90) 34%, rgba(10,16,25,1) 100%),
                linear-gradient(180deg, #111820 0%, #0f1724 42%, #0b111a 100%);
            background-repeat: no-repeat, no-repeat;
            background-size: 160% 130%, 100% 100%;
            background-position: center top, center top;
            background-attachment: fixed, fixed;
            display: flex;
            flex-direction: column;
            overflow-x: hidden;
        }
        .login-topbar {
            background:
                radial-gradient(circle at center top, rgba(104, 107, 114, 0.20) 0%, rgba(39, 41, 47, 0.12) 28%, rgba(12, 13, 17, 0) 58%),
                linear-gradient(90deg, #0a0b0f 0%, #171920 50%, #262932 100%);
            background-repeat: no-repeat, no-repeat;
            background-size: 140% 120%, 100% 100%;
            background-position: center top, center top;
            border-bottom: 1px solid rgba(255,255,255,0.08);
            box-shadow: inset 0 -1px 0 rgba(255,255,255,0.04), 0 14px 30px rgba(0,0,0,0.34);
        }
        .login-topbar-inner {
            max-width: 1400px;
            margin: 0 auto;
            min-height: 84px;
            padding: 0 28px;
            display: flex;
            align-items: center;
            gap: 14px;
        }
        .login-topbar img {
            display: block;
            width: 192px;
            height: 46px;
            object-fit: contain;
            flex: 0 0 192px;
        }
        .login-topbar-text {
            color: #c9cacf;
            font-size: 17px;
            font-weight: 700;
            letter-spacing: 0;
            line-height: 1;
            white-space: nowrap;
        }
        .login-shell {
            flex: 1;
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 40px 20px;
            min-height: calc(100vh - 84px - 64px);
        }
        .login-box {
            width: 100%;
            max-width: 420px;
            padding: 30px 30px 26px;
            border-radius: 20px;
            border: 1px solid var(--panel-border);
            background: linear-gradient(180deg, rgba(29, 42, 61, 0.98) 0%, rgba(20, 31, 47, 0.98) 100%);
            box-shadow: 0 24px 60px rgba(0,0,0,0.30);
        }
        .login-box h2 {
            margin: 0 0 8px;
            color: var(--accent-blue);
            font-size: 22px;
            font-weight: 800;
            text-align: center;
        }
        .login-subtitle {
            margin: 0 0 22px;
            color: var(--text-soft);
            font-size: 14px;
            text-align: center;
            font-weight: 600;
        }
        input {
            width: 100%;
            padding: 13px 14px;
            margin: 8px 0 16px 0;
            background: rgba(255,255,255,0.04);
            border: 1px solid rgba(255,255,255,0.10);
            color: #fff;
            border-radius: 10px;
            font-size: 14px;
            outline: none;
        }
        input::placeholder {
            color: var(--text-faint);
        }
        input:focus {
            border-color: rgba(184, 216, 245, 0.42);
            box-shadow: 0 0 0 3px rgba(184, 216, 245, 0.10);
        }
        button {
            width: 100%;
            padding: 13px 14px;
            background:
                radial-gradient(circle at center top, rgba(104, 107, 114, 0.12) 0%, rgba(39, 41, 47, 0.08) 28%, rgba(12, 13, 17, 0) 58%),
                linear-gradient(90deg, #171920 0%, #22252d 50%, #2b2f38 100%);
            color: #cfd3da;
            border: 1px solid rgba(255,255,255,0.10);
            border-radius: 10px;
            cursor: pointer;
            font-weight: 700;
            font-size: 15px;
            transition: background 0.18s ease, border-color 0.18s ease, color 0.18s ease;
        }
        button:hover {
            background:
                radial-gradient(circle at center top, rgba(116, 120, 128, 0.14) 0%, rgba(48, 50, 58, 0.10) 28%, rgba(12, 13, 17, 0) 58%),
                linear-gradient(90deg, #1b1d24 0%, #262932 50%, #323742 100%);
            border-color: rgba(255,255,255,0.16);
            color: #dce0e7;
        }
        .error {
            color: #ffd2d0;
            background: rgba(231, 98, 95, 0.14);
            padding: 10px 12px;
            border-radius: 10px;
            margin-bottom: 15px;
            font-size: 14px;
            border: 1px solid rgba(231, 98, 95, 0.24);
        }
        .password-container { position: relative; margin: 8px 0 16px 0; }
        .password-container input { margin: 0; padding-right: 40px; }
        .toggle-password {
            position: absolute;
            right: 12px;
            top: 50%;
            transform: translateY(-50%);
            cursor: pointer;
            color: var(--text-faint);
            font-size: 18px;
            user-select: none;
        }
        .login-footer {
            background: var(--footer-bg);
            background-repeat: no-repeat, no-repeat;
            background-size: 140% 120%, 100% 100%;
            background-position: center top, center top;
            border-top: 1px solid rgba(255,255,255,0.08);
            box-shadow: inset 0 1px 0 rgba(255,255,255,0.04);
        }
        .login-footer-inner {
            max-width: 1400px;
            min-height: 64px;
            margin: 0 auto;
            padding: 0 28px;
            display: flex;
            align-items: center;
            justify-content: center;
            color: var(--text-faint);
            font-size: 14px;
            text-align: center;
        }
    </style>
</head>
<body>
    <div class="login-topbar">
        <div class="login-topbar-inner">
            <img src="/logo.svg?v=7" alt="a:store">
            <span class="login-topbar-text">direct</span>
        </div>
    </div>
    <div class="login-shell">
        <div class="login-box">
            <h2>Вход в систему</h2>
            <p class="login-subtitle">Авторизация в рабочем интерфейсе a:store direct</p>
            {% if error %}
                <div class="error">{{ error }}</div>
            {% endif %}
            <form action="/login" method="POST">
                <input type="text" name="login" placeholder="Логин" required>
                
                <div class="password-container">
                    <input type="password" id="login-password" name="password" placeholder="Пароль" required>
                    <span class="toggle-password" onclick="togglePassword(this)"></span>
                </div>

                <button type="submit">Вход</button>
            </form>
        </div>
    </div>
    <div class="login-footer">
        <div class="login-footer-inner">© astore, 2022-2026. Все права защищены.</div>
    </div>

    <script>
        function togglePassword(iconElement) {
            const passwordInput = document.getElementById('login-password');
            if (passwordInput.type === 'password') {
                passwordInput.type = 'text';
                iconElement.innerText = '🙈'; // Меняем иконку на закрытый глаз
            } else {
                passwordInput.type = 'password';
                iconElement.innerText = '👁️'; // Меняем иконку на открытый глаз
            }
        }
    </script>
</body>
</html>
"""


@app.route('/api/interaction_bots/<int:id>', methods=['PUT'])
@login_required
def update_interaction_bot(id):
    data = request.get_json()
    user_id = session['user_id']
    db = get_db()
    
    bot_username = data.get('bot_username').strip()
    custom_name = data.get('custom_name')
    userbot_id = data.get('userbot_id')
    commands = json.dumps(data.get('commands', []))
    interval_minutes = normalize_interval_minutes(data.get('interval_minutes'), 60)
    
    # Обновляем настройки бота
    db.execute("""
        UPDATE interaction_bots 
        SET userbot_id=?, bot_username=?, custom_name=?, commands=?, interval_minutes=?
        WHERE id=? AND user_id=?
    """, (userbot_id, bot_username, custom_name, commands, interval_minutes, id, user_id))
    
    db.commit()
    notify_clients()
    return jsonify({'success': True})


@app.route('/api/fix_db')
@login_required
def fix_db():
    db = get_db()
    try:
        # Отключаем проверки ключей на секунду, чтобы SQLite разрешил замену
        db.execute("PRAGMA foreign_keys = OFF")
        db.executescript('''
            CREATE TABLE IF NOT EXISTS messages_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                telegram_message_id INTEGER,
                type TEXT,
                text TEXT,
                date TIMESTAMP,
                chat_id INTEGER,
                chat_title TEXT,
                is_blocked INTEGER DEFAULT 0,
                sender_name TEXT,
                is_delayed INTEGER DEFAULT 0,
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE,
                UNIQUE(chat_id, telegram_message_id)
            );
            INSERT OR IGNORE INTO messages_new (id, user_id, telegram_message_id, type, text, date, chat_id, chat_title, is_blocked, sender_name, is_delayed)
            SELECT id, user_id, telegram_message_id, type, text, date, chat_id, chat_title, is_blocked, sender_name, 0 FROM messages;
            DROP TABLE messages;
            ALTER TABLE messages_new RENAME TO messages;
        ''')
        db.execute("PRAGMA foreign_keys = ON")
        db.commit()
        return "<h1>✅ База данных успешно исправлена!</h1><p>Глобальная блокировка сообщений снята. Теперь парсер будет выкачивать всё.</p><br><a href='/'>Вернуться в парсер</a>"
    except Exception as e:
        return f"<h1>❌ Ошибка:</h1><p>{e}</p>"

@app.route('/api/cleanup_old_bug')
@login_required
def cleanup_old_bug():
    user_id = session['user_id']
    db = get_db()
    
    # Считаем, сколько сломанных сообщений мы найдем
    cursor = db.execute("SELECT COUNT(*) FROM messages WHERE telegram_message_id IS NULL AND user_id = ?", (user_id,))
    bad_count = cursor.fetchone()[0]
    
    if bad_count > 0:
        # Удаляем их (каскадное удаление само почистит привязки к товарам, если они были)
        db.execute("DELETE FROM messages WHERE telegram_message_id IS NULL AND user_id = ?", (user_id,))
        db.commit()
        notify_clients()
        return f"<h1>Уборка завершена! 🧹</h1><p>Удалено сломанных сообщений: <b>{bad_count}</b>.</p><br><a href='/'>Вернуться в парсер</a>"
    else:
        return "<h1>Всё чисто! ✨</h1><p>Сломанных сообщений в базе не найдено.</p><br><a href='/'>Вернуться в парсер</a>"






# ================= УПРАВЛЕНИЕ API И КЛИЕНТАМИ =================

# ✅ А ЭТОТ БЛОК ДОЛЖЕН ОСТАТЬСЯ ✅
@app.route('/api/api_clients/<int:client_id>/filters', methods=['POST'])
@login_required
def update_api_client_filters(client_id):
    data = request.json
    access_rules = json.dumps(data.get('access_rules', {})) # Конвертируем в JSON
    db = get_db()
    db.execute("UPDATE api_clients SET access_rules=? WHERE id=? AND user_id=?", 
               (access_rules, client_id, session['user_id']))
    db.commit()
    return jsonify({'success': True})


@app.route('/api/api_clients', methods=['GET', 'POST'])
@login_required
def manage_api_clients():
    user_id = session['user_id']
    db = get_db()
    if request.method == 'GET':
        # Теперь достаем все поля, включая тайминги
        clients = db.execute("SELECT * FROM api_clients WHERE user_id=?", (user_id,)).fetchall()
        return jsonify([dict(c) for c in clients])
    
    # POST: Создаем нового клиента
    data = request.json
    name = data.get('name')
    schedule_enabled = int(data.get('schedule_enabled', 0))
    time_start = data.get('time_start', '00:00')
    time_end = data.get('time_end', '23:59')
    token = secrets.token_hex(24) 
    
    db.execute("""
        INSERT INTO api_clients (user_id, name, token, schedule_enabled, time_start, time_end) 
        VALUES (?, ?, ?, ?, ?, ?)
    """, (user_id, name, token, schedule_enabled, time_start, time_end))
    
    client_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    
    db.execute("INSERT INTO api_markups (client_id, folder_id, markup_type, markup_value, rounding) VALUES (?, 0, 'percent', 0, 100)", (client_id,))
    db.commit()
    return jsonify({'success': True})

# --- ДОБАВЬ НОВЫЙ МАРШРУТ ДЛЯ РЕДАКТИРОВАНИЯ ТАЙМИНГОВ ---
@app.route('/api/api_clients/<int:client_id>/schedule', methods=['POST'])
@login_required
def update_api_client_schedule(client_id):
    data = request.json
    db = get_db()
    db.execute("""
        UPDATE api_clients 
        SET schedule_enabled=?, time_start=?, time_end=? 
        WHERE id=? AND user_id=?
    """, (int(data['schedule_enabled']), data['time_start'], data['time_end'], client_id, session['user_id']))
    db.commit()
    return jsonify({'success': True})

@app.route('/api/api_clients/<int:client_id>', methods=['DELETE'])
@login_required
def delete_api_client(client_id):
    db = get_db()
    db.execute("DELETE FROM api_clients WHERE id=? AND user_id=?", (client_id, session['user_id']))
    db.commit()
    return jsonify({'success': True})

@app.route('/api/api_clients/<int:client_id>/markups', methods=['GET', 'POST'])
@login_required
def manage_markups(client_id):
    db = get_db()
    if request.method == 'GET':
        markups = db.execute("""
            SELECT am.*, f.name as folder_name 
            FROM api_markups am 
            LEFT JOIN folders f ON am.folder_id = f.id 
            WHERE am.client_id=?
        """, (client_id,)).fetchall()
        return jsonify([dict(m) for m in markups])
    
    data = request.json
    folder_id = data.get('folder_id') or 0
    db.execute("""
        INSERT INTO api_markups (client_id, folder_id, markup_type, markup_value, rounding) 
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(client_id, folder_id) DO UPDATE SET 
        markup_type=excluded.markup_type, markup_value=excluded.markup_value, rounding=excluded.rounding
    """, (client_id, folder_id, data['markup_type'], data['markup_value'], data['rounding']))
    db.commit()
    return jsonify({'success': True})

@app.route('/api/api_markups/<int:markup_id>', methods=['DELETE'])
@login_required
def delete_markup(markup_id):
    db = get_db()
    # Запрещаем удалять базовую настройку (folder_id = 0)
    db.execute("DELETE FROM api_markups WHERE id=? AND folder_id != 0 AND client_id IN (SELECT id FROM api_clients WHERE user_id=?)", (markup_id, session['user_id']))
    db.commit()
    return jsonify({'success': True})


# ================= ПУБЛИЧНЫЙ API ДЛЯ ВЫДАЧИ КАТАЛОГА =================
@app.route('/api/v1/catalog', methods=['GET'])
def public_api_catalog():
    token = request.args.get('token')
    if not token:
        return jsonify({'error': 'Токен не предоставлен'}), 401
        
    db = get_db()
    client = db.execute("SELECT * FROM api_clients WHERE token=?", (token,)).fetchone()
    if not client:
        return jsonify({'error': 'Неверный токен'}), 401
        
    # --- ПРОВЕРКА РАСПИСАНИЯ ---
    if client['schedule_enabled']:
        from datetime import timedelta
        # МСК время (UTC+3)
        now = (datetime.now(timezone.utc).replace(tzinfo=None) + timedelta(hours=3)).time()
        start_time = datetime.strptime(client['time_start'], '%H:%M').time()
        end_time = datetime.strptime(client['time_end'], '%H:%M').time()
        
        is_active = start_time <= now <= end_time if start_time <= end_time else (now >= start_time or now <= end_time)
        if not is_active: 
            return jsonify({'categories': [], 'products': []})
            
    user_id = client['user_id']
    client_id = client['id']

    # --- ПОЛУЧАЕМ НАЦЕНКИ ---
    markups_raw = db.execute("SELECT folder_id, markup_type, markup_value, rounding FROM api_markups WHERE client_id=?", (client_id,)).fetchall()
    markups = {}
    default_markup = {'markup_type': 'percent', 'markup_value': 0, 'rounding': 100}
    for m in markups_raw:
        if m['folder_id'] == 0: 
            default_markup = dict(m)
        else: 
            markups[m['folder_id']] = dict(m)

    # --- ПОЛУЧАЕМ ТОВАРЫ (С ПОЛНЫМ НАБОРОМ ПОЛЕЙ И ЦЕНАМИ) ---
    # Используем подзапросы, чтобы сразу вытащить последнюю подтвержденную цену
    products_raw = db.execute("""
        SELECT p.id, p.name, p.folder_id, p.price as manual_price, p.photo_url, p.brand, p.country, p.weight, p.model_number, p.is_on_request,
               p.color, p.storage, p.ram, p.warranty, p.description, p.description_html, p.specs,
               (SELECT pm.extracted_price FROM product_messages pm JOIN messages m ON pm.message_id = m.id 
                WHERE pm.product_id = p.id AND pm.status = 'confirmed' AND pm.is_actual = 1 AND pm.extracted_price IS NOT NULL ORDER BY pm.extracted_price ASC, m.date DESC LIMIT 1) as parsed_price,
               (SELECT pm.is_actual FROM product_messages pm JOIN messages m ON pm.message_id = m.id 
                WHERE pm.product_id = p.id AND pm.status = 'confirmed' AND pm.is_actual = 1 AND pm.extracted_price IS NOT NULL ORDER BY pm.extracted_price ASC, m.date DESC LIMIT 1) as is_actual,
               (SELECT m.chat_id FROM product_messages pm JOIN messages m ON pm.message_id = m.id 
                WHERE pm.product_id = p.id AND pm.status = 'confirmed' AND pm.is_actual = 1 AND pm.extracted_price IS NOT NULL ORDER BY pm.extracted_price ASC, m.date DESC LIMIT 1) as latest_chat_id
        FROM products p
        WHERE p.user_id = ?
    """, (user_id,)).fetchall()
    
    products = []
    active_folder_ids = set()
    
    # Правила доступа (какие товары и каких поставщиков видит этот клиент)
    try:
        access_rules = json.loads(client['access_rules']) if 'access_rules' in client.keys() and client['access_rules'] else {}
    except:
        access_rules = {}

    for p in products_raw:
        # 1. Фильтр доступа: если товар не отмечен в access_rules - пропускаем
        pid_str = str(p['id'])
        if not access_rules or pid_str not in access_rules: 
            continue
            
        # 2. Фильтр поставщиков (чатов)
        allowed_chats = access_rules[pid_str]
        chat_id_str = str(p['latest_chat_id'])
        
        # Если есть цена от конкретного чата, проверяем, разрешен ли он (или разрешены 'all')
        if chat_id_str != 'None':
            if allowed_chats != ['all'] and chat_id_str not in allowed_chats:
                # Цена от этого поставщика запрещена, но мы можем использовать ручную цену p['manual_price']
                # Если ручной цены нет, товар станет "По запросу"
                current_price_source = None 
            else:
                current_price_source = p['parsed_price'] if p['is_actual'] == 1 else None
        else:
            current_price_source = None

        # 3. Расчет базовой цены
        base_price = float(current_price_source if current_price_source is not None else (p['manual_price'] or 0))
        
        # 4. Применение наценок и округления
        if base_price > 0 and not p['is_on_request']:
            mk = markups.get(p['folder_id'], default_markup)
            if mk['markup_type'] == 'percent':
                final_price = base_price * (1 + mk['markup_value'] / 100)
            else:
                final_price = base_price + mk['markup_value']
                
            rnd = mk['rounding']
            if rnd > 0:
                final_price = math.ceil(final_price / rnd) * rnd
            final_price = int(final_price)
        else:
            final_price = "По запросу"

        # 5. Обработка фотографий (формируем полные ссылки)
        photo_links = []
        if p['photo_url']:
            # Базовый URL вашего сервера
            base_img_url = "https://engine.astoredirect.ru/uploads/"
            photo_links = [
                (f"{base_img_url}{link.strip()}" if not link.strip().startswith('http') else link.strip())
                for link in p['photo_url'].split(',') if link.strip()
            ]

        # 6. Обработка технических характеристик (specs)
        specs_data = parse_specs_payload(p['specs'], p)

        # 7. Формируем финальный объект товара
        products.append({
            'id': p['id'],
            'name': p['name'],
            'category_id': p['folder_id'],
            'price': final_price, 
            'price2': "По запросу" ,
            "photo_url": photo_links,
            'brand': p['brand'],
            'country': p['country'],
            'weight': p['weight'],
            'model_number': p['model_number'],
            'is_on_request': bool(p['is_on_request']),
            
            # Новые поля
            'color': p['color'],
            'storage': p['storage'],
            'ram': p['ram'],
            'warranty': p['warranty'],
            'description': p['description'],
            'description_html': p['description_html'],
            'specs': specs_data
        })
        
        # Запоминаем ID папки, чтобы потом показать её в списке категорий
        if p['folder_id']:
            active_folder_ids.add(p['folder_id'])

    # --- ОБРАБОТКА КАТЕГОРИЙ (только те, в которых есть товары) ---
    folders_raw = db.execute("SELECT id, name, parent_id FROM folders WHERE user_id=?", (user_id,)).fetchall()
    parent_map = {f['id']: f['parent_id'] for f in folders_raw}
    folders_to_keep = set(active_folder_ids)
    
    # Добавляем всех родителей активных папок (чтобы дерево не рассыпалось)
    for fid in active_folder_ids:
        current = fid
        while current in parent_map and parent_map[current] is not None:
            current = parent_map[current]
            folders_to_keep.add(current)
            
    categories = [
        {'id': f['id'], 'name': f['name'], 'parent_id': f['parent_id']} 
        for f in folders_raw if f['id'] in folders_to_keep
    ]
                  
    # --- СОРТИРОВКА (А-Я) ---
    products.sort(key=lambda x: (x['name'] or '').lower())
    categories.sort(key=lambda x: (x['name'] or '').lower())
        
    return jsonify({
        'categories': categories,
        'products': products
    })


# ================= ПАРСИНГ EXCEL ТАБЛИЦ =================
def safe_spreadsheet_cell_text(value):
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def generic_file_name_cell_is_noise(value):
    text = safe_spreadsheet_cell_text(value)
    normalized = normalize_match_text(text)
    if not normalized:
        return True
    if extract_price(text):
        return True
    if normalized in {
        "name", "название", "наименование", "товар", "модель", "model", "product",
        "price", "цена", "стоимость", "остаток", "qty", "кол во", "количество",
    }:
        return True
    if re.fullmatch(r"[\d\s.,]+", normalized):
        return True
    return False


def generic_file_row_to_price_line(cells, context_headers):
    cleaned = [safe_spreadsheet_cell_text(cell) for cell in cells]
    if not any(cleaned):
        return None

    price_candidates = []
    for index, cell in enumerate(cleaned):
        price = extract_price(cell)
        if price is not None:
            price_candidates.append((index, price, cell))
    if not price_candidates:
        return None

    price_index, price, price_cell = price_candidates[-1]
    name_parts = [
        cell for cell in cleaned[:price_index]
        if cell and not generic_file_name_cell_is_noise(cell)
    ]
    if not name_parts:
        name_parts = [
            cell for index, cell in enumerate(cleaned)
            if index != price_index and cell and not generic_file_name_cell_is_noise(cell)
        ]
    name = strip_supplier_line_noise(" ".join(name_parts))
    if not name or len(normalize_match_text(name)) < 3:
        return None

    contextual_name = make_supplier_contextual_line(context_headers, name)
    price_display = price_cell.strip() if price_cell else str(price)
    if extract_price(price_display) is None:
        price_display = str(price)
    return normalize_incoming_text(f"{contextual_name} - {price_display}")


def parse_generic_rows_to_price_lines(rows, max_lines=GENERIC_FILE_PARSE_MAX_LINES):
    parsed_lines = []
    context_headers = []
    for row in rows:
        cells = [safe_spreadsheet_cell_text(cell) for cell in row]
        if not any(cells):
            continue
        line = generic_file_row_to_price_line(cells, context_headers)
        if line:
            parsed_lines.append(line)
            if len(parsed_lines) >= max_lines:
                break
            continue
        header_candidate = " ".join(cell for cell in cells if cell)
        context_headers = update_supplier_context_headers(context_headers, header_candidate)
    return parsed_lines


def parse_generic_excel_file(tmp_path):
    parsed_lines = []
    xls = pd.ExcelFile(tmp_path)
    for sheet_name in xls.sheet_names:
        try:
            df = pd.read_excel(xls, sheet_name=sheet_name, header=None, dtype=object)
        except Exception as e:
            logger.warning("Generic Excel fallback skipped sheet %s: %s", sheet_name, e)
            continue
        df = df.fillna("")
        parsed_lines.extend(
            parse_generic_rows_to_price_lines(
                df.values.tolist(),
                max_lines=max(1, GENERIC_FILE_PARSE_MAX_LINES - len(parsed_lines)),
            )
        )
        if len(parsed_lines) >= GENERIC_FILE_PARSE_MAX_LINES:
            break
    return parsed_lines


def parse_generic_pdf_file(tmp_path):
    import pdfplumber

    parsed_lines = []
    with pdfplumber.open(tmp_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables() or []
            for table in tables:
                rows = [
                    [safe_spreadsheet_cell_text(cell) for cell in row]
                    for row in table
                ]
                parsed_lines.extend(
                    parse_generic_rows_to_price_lines(
                        rows,
                        max_lines=max(1, GENERIC_FILE_PARSE_MAX_LINES - len(parsed_lines)),
                    )
                )
                if len(parsed_lines) >= GENERIC_FILE_PARSE_MAX_LINES:
                    return parsed_lines

            if parsed_lines:
                continue
            text = page.extract_text() or ""
            for raw_line in text.splitlines():
                line = normalize_incoming_text(raw_line)
                if line and supplier_line_has_price_marker(line) and extract_price(line):
                    parsed_lines.append(line)
                    if len(parsed_lines) >= GENERIC_FILE_PARSE_MAX_LINES:
                        return parsed_lines
    return parsed_lines


async def parse_excel_message(client, message, chat_id, user_id):
    text_to_save = normalize_incoming_text(message.text or "")
    if not message.document:
        return text_to_save
        
    ext = getattr(message.file, 'ext', '').lower()
    if ext not in ['.xlsx', '.xls', '.pdf']:
        return text_to_save

    # Функция для приведения ID к единому формату (без -100)
    def clean_id(cid):
        s = str(cid)
        if s.startswith('-100'): return s[4:]
        if s.startswith('-'): return s[1:]
        return s

    with app.app_context():
        db = get_db()
        target_id = clean_id(chat_id)
        
        # Загружаем все конфиги для этого пользователя
        all_configs = db.execute("SELECT * FROM excel_configs WHERE user_id = ?", (user_id,)).fetchall()
        configs = {}
        for c in all_configs:
            if clean_id(c['chat_id']) == target_id:
                configs[c['sheet_name']] = dict(c)
        
        has_explicit_configs = bool(configs)
        if not has_explicit_configs:
            logger.info("Нет настроек парсинга для чата %s, включаю авто-fallback по файлу", chat_id)

        tmp_path = None
        try:
            # Скачиваем файл во временное хранилище
            suffix = ext
            with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
                await client.download_media(message, tmp.name)
                tmp_path = tmp.name
                
            parsed_lines = []
            unknown_sheets = []
            wildcard_config = configs.get('*')

            # --- ОБРАБОТКА PDF ---
            if ext == '.pdf':
                import pdfplumber
                # Для PDF ищем: 'Страница 1', затем '*', если нет - берем первый попавшийся
                config = configs.get('Страница 1') or wildcard_config
                if not config and configs:
                    config = list(configs.values())[0]
                
                if config:
                    with pdfplumber.open(tmp_path) as pdf:
                        for page in pdf.pages:
                            table = page.extract_table()
                            if not table: continue
                            
                            # Превращаем таблицу в безопасный список строк
                            clean_rows = []
                            for row in table:
                                clean_rows.append([str(cell).replace('\n', ' ') if cell is not None else "" for cell in row])
                            
                            step = int(config['block_step'])
                            start = int(config['start_row'])
                            n_col = int(config['name_col'])
                            p_col = int(config['price_col'])
                            n_off = int(config['name_row_offset'])
                            p_off = int(config['price_row_offset'])
                            is_grouped = config.get('is_grouped', 0)
                            sku_col = config.get('sku_col', -1)

                            if is_grouped == 1:
                                current_group = ""
                                for i in range(start, len(clean_rows)):
                                    row = clean_rows[i]
                                    # Проверка: хватает ли колонок в этой строке PDF?
                                    if len(row) <= max(n_col, p_col, sku_col): continue
                                    
                                    col_n_val = row[n_col].strip()
                                    col_sku_val = row[sku_col].strip() if sku_col != -1 else ""
                                    col_p_val = row[p_col].strip()
                                    
                                    if col_n_val and not col_p_val: current_group = col_n_val
                                    if col_p_val and current_group:
                                        name = col_sku_val if col_sku_val else col_n_val
                                        parsed_lines.append(f"{current_group} {name} - {col_p_val}")
                            else:
	                                for i in range(start, len(clean_rows), step):
	                                    name_idx = i + n_off
	                                    price_idx = i + p_off
	                                    if name_idx < len(clean_rows) and price_idx < len(clean_rows):
	                                        row_n = clean_rows[name_idx]
	                                        row_p = clean_rows[price_idx]
	                                        if len(row_n) > n_col and len(row_p) > p_col:
	                                            n_val = row_n[n_col].strip()
	                                            p_val = row_p[p_col].strip()
	                                            if n_val and p_val:
	                                                parsed_lines.append(f"{n_val} - {p_val}")
                else:
                    parsed_lines.extend(parse_generic_pdf_file(tmp_path))

            # --- ОБРАБОТКА EXCEL ---
            else:
                if has_explicit_configs:
                    xls = pd.ExcelFile(tmp_path)
                    for sheet_name in xls.sheet_names:
                        config = configs.get(sheet_name) or wildcard_config
                        if not config:
                            unknown_sheets.append(sheet_name)
                            continue            
                    
                            
                        df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
                        df = df.fillna("")
                        
                        step = config['block_step']
                        start = config['start_row']
                        n_col = config['name_col']
                        p_col = config['price_col']
                        n_off = config['name_row_offset']
                        p_off = config['price_row_offset']
                        is_grouped = config.get('is_grouped', 0)
                        sku_col = config.get('sku_col', -1)
                        
                        if is_grouped == 1:
                            current_group = ""
                            for i in range(start, len(df)):
                                try:
                                    col_n_val = str(df.iat[i, n_col]).strip()
                                    col_sku_val = str(df.iat[i, sku_col]).strip() if sku_col != -1 else ""
                                    col_p_val = str(df.iat[i, p_col]).strip()
                                    
                                    is_empty_name = (not col_n_val or col_n_val == 'nan')
                                    has_price = (col_p_val and col_p_val != 'nan')
                                    
                                    if n_col == sku_col:
                                        if col_n_val and not has_price: current_group = col_n_val
                                        elif col_n_val and has_price and current_group:
                                            parsed_lines.append(f"{current_group} {col_n_val} - {col_p_val}")
                                    else:
                                        if col_n_val and not is_empty_name: current_group = col_n_val
                                        if col_sku_val and col_sku_val != 'nan' and has_price and current_group:
                                            parsed_lines.append(f"{current_group} {col_sku_val} - {col_p_val}")
                                except Exception:
                                    continue
                        else:
                            for i in range(start, len(df), step):
                                try:
                                    name_row = i + n_off
                                    price_row = i + p_off
                                    if name_row < len(df) and price_row < len(df):
                                        name_val = str(df.iat[name_row, n_col]).strip()
                                        price_val = str(df.iat[price_row, p_col]).strip()
                                        
                                        if name_val and name_val != 'nan' and price_val and price_val != 'nan':
                                            parsed_lines.append(f"{name_val} - {price_val}")
                                except Exception:
                                    continue
                else:
                    parsed_lines.extend(parse_generic_excel_file(tmp_path))

            if not parsed_lines and has_explicit_configs:
                try:
                    fallback_lines = parse_generic_pdf_file(tmp_path) if ext == '.pdf' else parse_generic_excel_file(tmp_path)
                    if fallback_lines:
                        parsed_lines.extend(fallback_lines)
                        logger.info("Авто-fallback извлек %s строк из файла чата %s", len(fallback_lines), chat_id)
                except Exception as fallback_e:
                    logger.warning("Авто-fallback файла для чата %s не сработал: %s", chat_id, fallback_e)
            
            # --- ФОРМИРОВАНИЕ ИТОГОВОГО ТЕКСТА ---
            if parsed_lines:
                doc_type = "PDF" if ext == '.pdf' else "EXCEL"
                text_to_save = normalize_incoming_text(f"📊 [{doc_type} ДАННЫЕ]:\n" + "\n".join(parsed_lines))
                
                if unknown_sheets and ext != '.pdf':
                    with app.app_context():
                        db = get_db()
                        all_chats = db.execute("SELECT chat_id, chat_title, custom_name FROM tracked_chats WHERE user_id = ?", (user_id,)).fetchall()
                        chat_title = str(chat_id)
                        for tc in all_chats:
                            if clean_id(tc['chat_id']) == target_id:
                                chat_title = tc['custom_name'] if tc['custom_name'] else tc['chat_title']
                                break
                                
                        for us in unknown_sheets:
                            db.execute("INSERT OR IGNORE INTO excel_missing_sheets (user_id, chat_id, chat_title, sheet_name) VALUES (?, ?, ?, ?)", (user_id, chat_id, chat_title, us))
                        db.commit()
                        
                    warning_text = f"\n\n⚠️ Внимание! В файле найдены ненастроенные листы: {', '.join(unknown_sheets)}. Перейдите в раздел «Парсинг Excel», чтобы их настроить."
                    if text_to_save:
                        text_to_save += warning_text
                    else:
                        text_to_save = warning_text.strip()
                        
        except Exception as e:
            logger.error(f"Ошибка парсинга файла: {e}")
        finally:
            if 'tmp_path' in locals() and os.path.exists(tmp_path):
                os.remove(tmp_path)
                
    return normalize_incoming_text(text_to_save)


@app.route('/api/excel/preview_latest/<chat_id>', methods=['GET'])
@login_required
def preview_latest_excel(chat_id):
    chat_id = int(chat_id) # Добавляем преобразование здесь
    user_id = session['user_id']
    
    # Ищем активную сессию юзербота
    client_to_use = None
    client_loop_to_use = None
    for sid, (thread, client, uid) in user_clients.items():
        if uid == user_id:
            client_to_use = client
            client_loop_to_use = user_loops.get(sid)
            break
            
    if not client_to_use:
        return jsonify({'error': 'Юзербот не запущен. Запустите его во вкладке "Юзерботы".'}), 400

    async def fetch_and_preview():
        try:
            entity = await client_to_use.get_entity(chat_id)
            # Ищем последние 50 сообщений
            async for message in client_to_use.iter_messages(entity, limit=50):
                if message.document and getattr(message.file, 'ext', '').lower() in ['.xlsx', '.xls', '.pdf']:
                    ext = getattr(message.file, 'ext', '').lower()
                    
                    # Скачиваем файл во временную директорию
                    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
                        await client_to_use.download_media(message, tmp.name)
                        tmp_path = tmp.name
                        
                    try:
                        xls = pd.ExcelFile(tmp_path)
                        data = {}
                        # Читаем по 15 строк для предпросмотра
                        for sheet in xls.sheet_names:
                            df = pd.read_excel(xls, sheet_name=sheet, header=None, nrows=15)
                            df = df.fillna("")
                            data[sheet] = df.values.tolist()
                        return {'success': True, 'sheets_data': data}
                    finally:
                        if os.path.exists(tmp_path):
                            os.remove(tmp_path)
                            
            return {'success': False, 'error': 'В последних 50 сообщениях чата не найден Excel файл'}
        except Exception as e:
            return {'success': False, 'error': str(e)}

    # Запускаем асинхронную задачу в цикле юзербота и ждем результат
    future = asyncio.run_coroutine_threadsafe(fetch_and_preview(), client_loop_to_use)
    try:
        result = future.result(timeout=30)
        if result.get('success'):
            return jsonify(result)
        else:
            return jsonify({'error': result.get('error')}), 400
    except Exception as e:
        return jsonify({'error': 'Таймаут или ошибка при скачивании файла: ' + str(e)}), 500


@app.route('/api/excel/preview', methods=['POST'])
@login_required
def preview_excel():
    if 'file' not in request.files:
        return jsonify({'error': 'Нет файла'}), 400
    file = request.files['file']
    try:
        xls = pd.ExcelFile(file)
        data = {}
        # Читаем по 15 строк с КАЖДОГО листа
        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet, header=None, nrows=15)
            df = df.fillna("")
            data[sheet] = df.values.tolist()
        return jsonify({'success': True, 'sheets_data': data})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/excel/save_config', methods=['POST'])
@login_required
def save_excel_config():
    data = request.json
    db = get_db()
    
    chat_id = int(data['chat_id'])
    sheet_name = data['sheet_name']
    is_grouped = data.get('is_grouped', 0)
    sku_col = data.get('sku_col', -1)
    
    db.execute("""
        INSERT INTO excel_configs (user_id, chat_id, sheet_name, name_col, name_row_offset, price_col, price_row_offset, block_step, start_row, is_grouped, sku_col) 
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(chat_id, sheet_name) DO UPDATE SET 
        name_col=excluded.name_col, name_row_offset=excluded.name_row_offset, price_col=excluded.price_col, 
        price_row_offset=excluded.price_row_offset, block_step=excluded.block_step, start_row=excluded.start_row,
        is_grouped=excluded.is_grouped, sku_col=excluded.sku_col
    """, (session['user_id'], chat_id, sheet_name, data['name_col'], data['name_row_offset'], data['price_col'], data['price_row_offset'], data['block_step'], data['start_row'], is_grouped, sku_col))

    cid_str = str(chat_id)
    core_id = cid_str[4:] if cid_str.startswith('-100') else (cid_str[1:] if cid_str.startswith('-') else cid_str)
    possible_ids = [core_id, f"-{core_id}", f"-100{core_id}"]
    
    for pid in possible_ids:
        if sheet_name == '*':
            db.execute("DELETE FROM excel_missing_sheets WHERE chat_id = ? AND user_id = ?", (pid, session['user_id']))
        else:
            db.execute("DELETE FROM excel_missing_sheets WHERE chat_id = ? AND sheet_name = ? AND user_id = ?", (pid, sheet_name, session['user_id']))
    
    db.commit()
    return jsonify({'success': True})



@app.route('/api/excel/configs', methods=['GET'])
@login_required
def get_excel_configs():
    db = get_db()
    query = """
        SELECT ec.id, ec.chat_id, ec.sheet_name, ec.name_col, ec.price_col, ec.parse_interval,
               ec.sheet_url, ec.source_type, ec.folder_id, ec.block_step, ec.price_row_offset,
               ec.name_row_offset, ec.is_grouped, ec.sku_col,
               COALESCE(tc.custom_name, tc.chat_title, ec.chat_id) AS source_name,
               COALESCE(tc.custom_name, tc.chat_title, ec.chat_id) AS chat_title,
               (
                   SELECT COUNT(*)
                   FROM messages m
                   WHERE m.user_id = ec.user_id
                     AND CAST(m.chat_id AS TEXT) = CAST(ec.chat_id AS TEXT)
                     AND m.type = ('excel_' || ec.id)
               ) AS message_count,
               (
                   SELECT m.date
                   FROM messages m
                   WHERE m.user_id = ec.user_id
                     AND CAST(m.chat_id AS TEXT) = CAST(ec.chat_id AS TEXT)
                     AND m.type = ('excel_' || ec.id)
                   ORDER BY m.date DESC, m.id DESC
                   LIMIT 1
               ) AS last_message_date,
               (
                   SELECT m.sender_name
                   FROM messages m
                   WHERE m.user_id = ec.user_id
                     AND CAST(m.chat_id AS TEXT) = CAST(ec.chat_id AS TEXT)
                     AND m.type = ('excel_' || ec.id)
                   ORDER BY m.date DESC, m.id DESC
                   LIMIT 1
               ) AS db_sender_name
        FROM excel_configs ec
        LEFT JOIN tracked_chats tc ON ec.chat_id = tc.chat_id AND ec.user_id = tc.user_id
        WHERE ec.user_id = ?
        ORDER BY ec.source_type, ec.chat_id, ec.sheet_url, ec.sheet_name, ec.id
    """
    configs = db.execute(query, (session['user_id'],)).fetchall()
    return jsonify([dict(c) for c in configs])

@app.route('/api/excel/configs/<int:config_id>', methods=['DELETE'])
@login_required
def delete_excel_config(config_id):
    db = get_db()
    config = db.execute(
        "SELECT chat_id, source_type FROM excel_configs WHERE id = ? AND user_id = ?",
        (config_id, session['user_id'])
    ).fetchone()
    db.execute("DELETE FROM excel_configs WHERE id = ? AND user_id = ?", (config_id, session['user_id']))
    if config and config['source_type'] == 'google_sheet':
        remaining = db.execute(
            "SELECT 1 FROM excel_configs WHERE user_id = ? AND CAST(chat_id AS TEXT) = CAST(? AS TEXT) LIMIT 1",
            (session['user_id'], config['chat_id'])
        ).fetchone()
        if not remaining:
            db.execute("""
                DELETE FROM product_messages
                WHERE message_id IN (
                    SELECT id FROM messages
                    WHERE user_id = ? AND CAST(chat_id AS TEXT) = CAST(? AS TEXT)
                )
            """, (session['user_id'], config['chat_id']))
            db.execute(
                "DELETE FROM messages WHERE user_id = ? AND CAST(chat_id AS TEXT) = CAST(? AS TEXT)",
                (session['user_id'], config['chat_id'])
            )
    db.commit()
    return jsonify({'success': True})


import sqlite3

def update_database():
    conn = sqlite3.connect(app.config['DATABASE'])
    cursor = conn.cursor()
    
    print("Начинаю проверку базы данных...")

    # --- ДОБАВЛЯЕМ КОЛОНКУ IS_DELAYED ---
    try:
        cursor.execute("ALTER TABLE messages ADD COLUMN is_delayed INTEGER DEFAULT 0")
        print("✅ Колонка is_delayed успешно добавлена в таблицу messages!")
    except sqlite3.OperationalError:
        pass

    # Полный список всех колонок для работы парсера (Google, PDF, Excel)
    columns_to_add = {
        "source_type": "TEXT DEFAULT 'excel'",
        "sheet_url": "TEXT",
        "parse_interval": "INTEGER DEFAULT 60",
        "is_grouped": "INTEGER DEFAULT 0",
        "sku_col": "INTEGER DEFAULT -1",
        "name_col": "INTEGER DEFAULT -1",
        "name_row_offset": "INTEGER DEFAULT 0",
        "price_row_offset": "INTEGER DEFAULT 0",
        "block_step": "INTEGER DEFAULT 1",
        "sheet_name": "TEXT DEFAULT '*'"
    }

    for col, col_type in columns_to_add.items():
        try:
            cursor.execute(f"ALTER TABLE excel_configs ADD COLUMN {col} {col_type}")
            print(f"✅ Колонка {col} успешно добавлена.")
        except sqlite3.OperationalError:
            print(f"⏩ Колонка {col} уже существует, пропускаем.")

    # Устанавливаем тип 'excel' для всех записей, где он еще не задан
    cursor.execute("UPDATE excel_configs SET source_type = 'excel' WHERE source_type IS NULL OR source_type = ''")

    # --- ДОБАВЛЯЕМ КОЛОНКИ FOLDER_ID ---
    try:
        cursor.execute("ALTER TABLE target_chats ADD COLUMN folder_id INTEGER;")
        print("✅ Колонка folder_id добавлена в таблицу target_chats")
    except sqlite3.OperationalError:
        pass

    try:
        cursor.execute("ALTER TABLE excel_configs ADD COLUMN folder_id INTEGER;")
        print("✅ Колонка folder_id добавлена в таблицу excel_configs")
    except sqlite3.OperationalError:
        pass

    # Фиксируем все изменения ОДИН РАЗ и закрываем базу
    conn.commit()
    conn.close()
    
    print("✅ Данные обновлены, старые конфиги теперь помечены как excel")
    print("🎉 Полное обновление базы данных завершено!")




@app.route('/api/fix_delayed')
def fix_delayed():
    try:
        db = get_db()
        db.execute("ALTER TABLE messages ADD COLUMN is_delayed INTEGER DEFAULT 0")
        db.commit()
        return "<h1>✅ База данных успешно обновлена! Колонка is_delayed добавлена.</h1>"
    except Exception as e:
        return f"<h1>Уже добавлена или произошла ошибка:</h1><p>{e}</p>"

from datetime import datetime

def is_parsing_allowed(session_id):
    """
    Проверяет, попадает ли текущее время в разрешенный интервал тайминга.
    """
    try:
        with app.app_context():
            db = get_db()
            # Добавили проверку schedule_enabled
            row = db.execute("SELECT schedule_enabled, time_start, time_end FROM user_telegram_sessions WHERE id = ?", (session_id,)).fetchone()
            
            # Если галочка "Расписание" не стоит или данных нет — разрешаем 24/7
            if not row or not row['schedule_enabled']:
                return True 
            
            # Сдвигаем время сервера на +3 часа (Москва).
            current_time = (datetime.now(timezone.utc).replace(tzinfo=None) + timedelta(hours=3)).time()
            
            start_time = datetime.strptime(row['time_start'], '%H:%M').time()
            end_time = datetime.strptime(row['time_end'], '%H:%M').time()

            if start_time <= end_time:
                return start_time <= current_time <= end_time
            else:
                return current_time >= start_time or current_time <= end_time

    except Exception as e:
        logger.error(f"Ошибка проверки тайминга для сессии {session_id}: {e}")
        return True

def is_user_parsing_allowed(user_id):
    """
    Общий стоп-кран для файлов/API: если у пользователя есть активные Telegram-аккаунты
    с расписанием, новые данные принимаются только в разрешенное время хотя бы одного из них.
    """
    try:
        with app.app_context():
            db = get_db()
            rows = db.execute("""
                SELECT id, schedule_enabled
                FROM user_telegram_sessions
                WHERE user_id = ? AND status = 'active'
            """, (user_id,)).fetchall()
            scheduled_rows = [row for row in rows if row['schedule_enabled']]
            if not scheduled_rows:
                return True
            return any(is_parsing_allowed(row['id']) for row in scheduled_rows)
    except Exception as e:
        logger.error(f"Ошибка проверки общего тайминга для пользователя {user_id}: {e}")
        return True

def delayed_messages_scheduler():
    while True:
        # Старое поведение оживляло накопленные вне расписания сообщения позже.
        # Теперь вне расписания новые данные вообще не сохраняются, чтобы не сбивать актуальные карточки.
        time.sleep(60)
        continue
        try:
            with app.app_context():
                db = get_db()
                # Проверяем все активные аккаунты
                cursor = db.execute("SELECT id as session_id, user_id FROM user_telegram_sessions WHERE status = 'active'")
                for session_row in cursor.fetchall():
                    session_id = session_row['session_id']
                    user_id = session_row['user_id']
                    
                    # Если для этого аккаунта СЕЙЧАС рабочее время
                    if is_parsing_allowed(session_id):
                        # Достаем все его скрытые сообщения
                        msg_cursor = db.execute("SELECT * FROM messages WHERE is_delayed = 1 AND user_id = ? ORDER BY date ASC", (user_id,))
                        delayed_messages = msg_cursor.fetchall()
                        
                        for msg in delayed_messages:
                            msg_id = msg['id']
                            parsed_text = msg['text']
                            chat_id = msg['chat_id']
                            
                            # 1. Снимаем флаг 'скрыто'
                            db.execute("UPDATE messages SET is_delayed = 0 WHERE id = ?", (msg_id,))
                            
                            # 2. ЗАПУСКАЕМ ОБРАБОТКУ (миграция привязок к товарам)
                            old_bindings = db.execute("""
                                SELECT pm.id, pm.product_id, pm.line_index, p.name, p.country, p.synonyms, p.brand, p.model_number, p.color, p.storage
                                FROM product_messages pm
                                JOIN products p ON pm.product_id = p.id
                                JOIN messages m ON pm.message_id = m.id
                                WHERE m.chat_id = ? AND m.user_id = ? AND m.id != ?
                            """, (chat_id, user_id, msg_id)).fetchall()

                            if old_bindings and parsed_text:
                                lines = parsed_text.split('\n')
                                
                                for b in old_bindings:
                                    line_idx = b['line_index']
                                    match_found = False
                                    new_price = None
                                    
                                    if line_idx == -1:
                                        if (
                                            binding_product_row_matches_line(b, parsed_text, require_known_category=True)
                                        ):
                                            match_found = True
                                            new_price = extract_price(parsed_text)
                                    else:
                                        for i, line in enumerate(lines):
                                            if (
                                                binding_product_row_matches_line(b, line, require_known_category=True)
                                            ):
                                                match_found = True
                                                new_price = extract_price(line)
                                                if not new_price and i + 1 < len(lines):
                                                    new_price = extract_price(lines[i+1])
                                                if not new_price and i + 2 < len(lines):
                                                    new_price = extract_price(lines[i+2])
                                                line_idx = i
                                                break
                                     
                                    if match_found:
                                        db.execute("""
                                            DELETE FROM product_messages
                                            WHERE product_id = ? AND message_id = ? AND line_index = ? AND id != ?
                                        """, (b['product_id'], msg_id, line_idx, b['id']))
                                        db.execute("""
                                            UPDATE product_messages 
                                            SET message_id = ?, line_index = ?, extracted_price = ?, is_actual = 1 
                                            WHERE id = ?
                                        """, (msg_id, line_idx, new_price, b['id']))
                            
                            # Очистка старых сообщений (зазор 5 минут)
                            cutoff_time = (datetime.now(timezone.utc).replace(tzinfo=None) + timedelta(hours=3) - timedelta(minutes=5)).strftime("%Y-%m-%d %H:%M:%S")
                            
                            # 1. Помечаем товары как неактуальные
                            db.execute("""
                                UPDATE product_messages 
                                SET is_actual = 0 
                                WHERE message_id IN (
                                    SELECT id FROM messages 
                                    WHERE chat_id = ? AND user_id = ? AND id != ? AND date < ?
                                )
                            """, (chat_id, user_id, msg_id, cutoff_time))
                            
                            # 2. Удаляем старые сообщения
                            db.execute("""
                                DELETE FROM messages 
                                WHERE chat_id = ? AND user_id = ? AND id != ? AND date < ?
                                  AND id NOT IN (SELECT message_id FROM product_messages WHERE is_actual = 1)
                            """, (chat_id, user_id, msg_id, cutoff_time))

                            db.commit()
                            
                            # 3. Отправляем уведомление на фронтенд
                            notify_clients()
                            
        except Exception as e:
            print(f"Ошибка в планировщике отложенных сообщений: {e}")
            
        time.sleep(60) # Проверяем тайминги каждую минуту


def process_pending_ai_repair_jobs():
    with app.app_context():
        db = get_db()
        jobs = db.execute(
            """
            SELECT *
            FROM ai_repair_jobs
            WHERE status IN ('pending', 'retry_wait')
              AND next_retry_at <= CURRENT_TIMESTAMP
            ORDER BY created_at
            LIMIT ?
            """,
            (AI_REPAIR_JOB_BATCH_SIZE,),
        ).fetchall()

        for job in jobs:
            db.execute(
                "UPDATE ai_repair_jobs SET status = 'processing', updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                (job["id"],),
            )
            db.commit()

            product_row = db.execute(
                "SELECT * FROM products WHERE id = ? AND user_id = ?",
                (job["product_id"], job["user_id"]),
            ).fetchone()
            if not product_row:
                db.execute(
                    """
                    UPDATE ai_repair_jobs
                    SET status = 'failed', last_error = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                    """,
                    ("Товар удален или недоступен", job["id"]),
                )
                db.commit()
                continue

            product_name = str(job["product_name"] or product_row["name"] or "").strip()
            article = str(job["article"] or product_row["model_number"] or "").strip()
            preferred_folder_id = normalize_folder_id(job["preferred_folder_id"])

            try:
                payload, _ = get_or_generate_autofill_payload(
                    db,
                    job["user_id"],
                    product_name,
                    article,
                    use_cache=True,
                )
                upsert_product_from_autofill(
                    db,
                    job["user_id"],
                    product_row,
                    payload,
                    preferred_folder_id if preferred_folder_id is not None else product_row["folder_id"],
                )
                db.execute(
                    """
                    UPDATE ai_repair_jobs
                    SET status = 'completed', attempts = attempts + 1, last_error = NULL, updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                    """,
                    (job["id"],),
                )
                db.commit()
                notify_clients()
            except Exception as e:
                attempts = int(job["attempts"] or 0) + 1
                if is_retryable_ai_error(e):
                    retry_delay_seconds = compute_ai_retry_delay_seconds(attempts)
                    db.execute(
                        """
                        UPDATE ai_repair_jobs
                        SET status = 'retry_wait',
                            attempts = ?,
                            last_error = ?,
                            next_retry_at = datetime('now', '+' || ? || ' seconds'),
                            updated_at = CURRENT_TIMESTAMP
                        WHERE id = ?
                        """,
                        (attempts, str(e)[:1000], retry_delay_seconds, job["id"]),
                    )
                else:
                    db.execute(
                        """
                        UPDATE ai_repair_jobs
                        SET status = 'failed', attempts = ?, last_error = ?, updated_at = CURRENT_TIMESTAMP
                        WHERE id = ?
                        """,
                        (attempts, str(e)[:1000], job["id"]),
                    )
                db.commit()


def ai_repair_scheduler():
    while True:
        try:
            process_pending_ai_repair_jobs()
        except Exception as e:
            logger.error(f"Ошибка AI-планировщика исправлений: {e}")
        time.sleep(60)


def warm_product_search_cache():
    try:
        with app.app_context():
            db = get_db()
            rows = db.execute("SELECT id FROM users ORDER BY id").fetchall()
            for row in rows:
                get_cached_product_search_entries(db, int(row['id']))
            logger.info("Product search cache warmed")
    except Exception as e:
        logger.error(f"Ошибка прогрева кеша поиска товаров: {e}")


_background_schedulers_started = False


def start_background_schedulers():
    global _background_schedulers_started
    if _background_schedulers_started:
        return
    threading.Thread(target=publish_scheduler, daemon=True).start()
    threading.Thread(target=api_parser_scheduler, daemon=True).start()
    threading.Thread(target=channel_snapshot_scheduler, daemon=True).start()
    threading.Thread(target=delayed_messages_scheduler, daemon=True).start()
    threading.Thread(target=ai_repair_scheduler, daemon=True).start()
    threading.Thread(target=bot_interaction_scheduler, daemon=True).start()
    threading.Thread(target=auto_binding_sweep_scheduler, daemon=True).start()
    threading.Thread(target=warm_product_search_cache, daemon=True).start()
    _background_schedulers_started = True

# ---------- Запуск приложения ----------

if __name__ == '__main__':
    update_database()
    start_background_schedulers()
    
    with app.app_context():
        db = get_db()
        cursor = db.execute("SELECT id, user_id, api_id, api_hash FROM user_telegram_sessions WHERE status = 'active'")
        for row in cursor.fetchall():
            start_message_listener(row['id'], row['user_id'], row['api_id'], row['api_hash'])
    port = int(os.environ.get('PORT', '5001'))
    socketio.run(app, debug=True, host='0.0.0.0', port=port, use_reloader=False, allow_unsafe_werkzeug=True)
